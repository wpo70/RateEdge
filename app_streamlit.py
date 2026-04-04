
import math
import os
import json

# Load .env file for local development (ignored in Streamlit Cloud)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed   —   use environment variables directly
from dataclasses import dataclass
from datetime import date, timedelta
try:
    from zoneinfo import ZoneInfo
    SYDNEY_TZ = ZoneInfo("Australia/Sydney")
    WELLINGTON_TZ = ZoneInfo("Pacific/Auckland")
    NEW_YORK_TZ = ZoneInfo("America/New_York")
except ImportError:
    from datetime import timezone, timedelta
    SYDNEY_TZ = timezone(timedelta(hours=11))      # AEDT approx
    WELLINGTON_TZ = timezone(timedelta(hours=13))  # NZDT approx
    NEW_YORK_TZ = timezone(timedelta(hours=-4))    # EDT approx

# Canonical EOD close times per currency (local time)
CCY_TZ = {
    "AUD": SYDNEY_TZ,
    "NZD": WELLINGTON_TZ,
    "USD": NEW_YORK_TZ,
}
CCY_EOD = {
    "AUD": (16, 30),   # 4:30pm Sydney
    "NZD": (17, 0),    # 5:00pm Wellington
    "USD": (16, 30),   # 4:30pm New York
}
from typing import Optional, List, Tuple, Dict

import numpy as np
import pandas as pd
import streamlit as st
from statistics import NormalDist
import plotly.graph_objects as go
import requests
import scipy.optimize
from scipy.interpolate import PchipInterpolator

# ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
# AFMA AUSTRALIAN BUSINESS DAY CALENDAR
# Sydney/NSW calendar   —   national + NSW state holidays.
# Used throughout all date calculations: spot lags, roll dates, resets.
# AFMA convention: Modified Following (moves to next BD, unless that
# crosses month end in which case use preceding BD).
# ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë

def _easter(year: int):
    """Return (Good Friday, Easter Monday) for year   —   Anonymous Gregorian."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    month = (h + l - 7*m + 114) // 31
    day = (h + l - 7*m + 114) % 31 + 1
    easter_sunday = date(year, month, day)
    return easter_sunday - timedelta(days=2), easter_sunday + timedelta(days=1)

def _au_holidays(year: int) -> set:
    """
    Australian national + NSW public holidays for a given year.
    Follows AFMA/ASX 24 Sydney financial centre calendar.
    """
    h = set()

    def _sub_mon(d):
        """Substitute Monday if date falls on weekend."""
        wd = d.weekday()
        if wd == 5: return d + timedelta(days=2)
        if wd == 6: return d + timedelta(days=1)
        return d

    # New Year's Day
    h.add(_sub_mon(date(year, 1, 1)))
    # Australia Day   —   26 Jan
    h.add(_sub_mon(date(year, 1, 26)))
    # Good Friday + Easter Saturday (NSW bank holiday) + Easter Monday
    gf, em = _easter(year)
    h.add(gf)
    h.add(gf + timedelta(days=1))   # Easter Saturday
    h.add(em)
    # Anzac Day   —   25 Apr (NSW: if Sunday → Monday sub)
    anzac = date(year, 4, 25)
    h.add(anzac + timedelta(days=1) if anzac.weekday() == 6 else anzac)
    # King's/Queen's Birthday   —   2nd Monday in June (NSW)
    june_mons = [date(year, 6, d) for d in range(1, 31) if date(year, 6, d).weekday() == 0]
    if len(june_mons) >= 2:
        h.add(june_mons[1])
    # Bank Holiday   —   1st Monday in August (NSW)
    aug_mons = [date(year, 8, d) for d in range(1, 32) if date(year, 8, d).weekday() == 0]
    if aug_mons:
        h.add(aug_mons[0])
    # Christmas Day + Boxing Day (with substitution)
    xmas = date(year, 12, 25)
    box  = date(year, 12, 26)
    xwd  = xmas.weekday()
    if xwd == 5:    # Sat: Mon=Xmas sub, Tue=Boxing sub
        h.add(date(year, 12, 27)); h.add(date(year, 12, 28))
    elif xwd == 6:  # Sun: Mon=Boxing, Tue=Xmas sub
        h.add(date(year, 12, 26)); h.add(date(year, 12, 27))
    else:
        h.add(xmas)
        bwd = box.weekday()
        if bwd == 5:   h.add(date(year, 12, 28))
        elif bwd == 6: h.add(date(year, 12, 27))
        else:          h.add(box)
    return h

_AU_HOL_CACHE: dict = {}

def is_au_bd(d: date) -> bool:
    """True if d is an Australian (Sydney/NSW) business day."""
    if d.weekday() >= 5:
        return False
    if d.year not in _AU_HOL_CACHE:
        _AU_HOL_CACHE[d.year] = _au_holidays(d.year)
    return d not in _AU_HOL_CACHE[d.year]

def next_au_bd(d: date) -> date:
    """Return d if it is a business day, else advance to next business day."""
    while not is_au_bd(d):
        d += timedelta(days=1)
    return d

def prev_au_bd(d: date) -> date:
    """Return d if it is a business day, else go back to preceding business day."""
    while not is_au_bd(d):
        d -= timedelta(days=1)
    return d

def modified_following(d: date) -> date:
    """
    AFMA Modified Following: advance to next BD.
    If that crosses to the next calendar month, use preceding BD instead.
    """
    orig_month = d.month
    candidate = next_au_bd(d)
    if candidate.month != orig_month:
        return prev_au_bd(d)
    return candidate

def add_au_bds(d: date, n: int) -> date:
    """Add n Australian business days to date d."""
    step = 1 if n >= 0 else -1
    remaining = abs(n)
    while remaining > 0:
        d += timedelta(days=step)
        if is_au_bd(d):
            remaining -= 1
    return d

def au_spot_date(trade_date: date, ccy: str = "AUD") -> date:
    """
    Return spot (effective) date from trade date.
    AUD/NZD: T+1 BD (AFMA convention for IRS/ZCS).
    USD/EUR/GBP/JPY/CAD: T+2 BD.
    """
    lag = 1 if ccy in ("AUD",) else 2
    return add_au_bds(trade_date, lag)

def au_end_date(start: date, tenor_y: float) -> date:
    """
    Compute end date from start + tenor (in years), Modified Following.
    Uses anniversary date rule: same day/month, n years forward.
    """
    n = int(tenor_y)
    try:
        raw = start.replace(year=start.year + n)
    except ValueError:  # Feb 29 in non-leap year
        raw = start.replace(year=start.year + n, day=28)
    return modified_following(raw)

def build_au_reset_schedule(start: date, end: date, freq_months: int) -> list:
    """
    Build a list of (reset_date, start_date, end_date, days, accrual_frac) tuples
    for an AUD swap/cap/ZCS floating leg. Modified Following throughout.
    freq_months: 3 = quarterly, 6 = semi-annual.
    """
    periods = []
    current = start
    while current < end:
        # Next period end: add freq_months calendar months
        m = current.month + freq_months
        y = current.year + (m - 1) // 12
        m = (m - 1) % 12 + 1
        try:
            raw_next = date(y, m, start.day)
        except ValueError:
            raw_next = date(y, m, 28)
        next_end = modified_following(raw_next)
        if next_end > end:
            next_end = end
        days = (next_end - current).days
        periods.append({
            "reset_date": current,
            "start_date": current,
            "end_date": next_end,
            "accrual_days": days,
            "accrual_frac": days / 365.0,
        })
        current = next_end
        if current >= end:
            break
    return periods

# ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë

# ============================
# RateEdge Authentication
# ============================
# ── RateEdge Auth API ─────────────────────────────────────────────
SITE_ID = "options"

def _auth_api():
    try:
        return st.secrets["AUTH_API"]
    except Exception:
        return os.environ.get("AUTH_API", "https://rateedge-auth.onrender.com")

def request_otp(email):
    """Send OTP via RateEdge Auth API."""
    try:
        resp = requests.post(f"{_auth_api()}/api/auth/request-otp",
            json={"email": email, "site": SITE_ID}, timeout=15)
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, {"error": "Auth server waking up - try again in 30s"}
    except Exception as e:
        return 500, {"error": f"Cannot reach auth server: {str(e)}"}

def verify_otp(email, code):
    """Verify OTP via RateEdge Auth API."""
    try:
        resp = requests.post(f"{_auth_api()}/api/auth/verify-otp",
            json={"email": email, "site": SITE_ID, "code": code}, timeout=15)
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, {"error": "Auth server error - try again"}
    except Exception as e:
        return 500, {"error": str(e)}

try:
    from streamlit_plotly_events import plotly_events
except ImportError:  # optional dependency
    plotly_events = None

# Import the new hybrid vol surface editor
try:
    from vol_editor import render_vol_surface_editor, render_bulk_adjustment_tools, render_conventions_tab
    render_vol_surface_editor_unified = render_vol_surface_editor
    render_vol_surface_editor_3d = render_vol_surface_editor
    HAS_3D_DRAG = True
    HAS_VOL_EDITOR = True
except ImportError:
    HAS_VOL_EDITOR = False
    render_conventions_tab = None
    render_vol_surface_editor_unified = None
    render_bulk_adjustment_tools = None

try:
    import psycopg2
    from psycopg2.extras import Json
    HAS_POSTGRES = True
except ImportError:
    HAS_POSTGRES = False

SUPPORTED_CURRENCIES = ["AUD", "NZD", "USD"]
ALL_CURRENCIES = ["AUD", "NZD", "USD", "EUR (PENDING)", "GBP (PENDING)", "JPY (PENDING)", "CAD (PENDING)"]


# ============================
# Database Functions
# ============================

def get_db_url():
    """
    Get Supabase PostgreSQL connection URL from environment variable.
    Set RATEEDGE_DB_URL in .env (local) or Streamlit secrets (cloud).
    Format: postgresql://postgres.[project-ref]:[password]@aws-0-ap-southeast-2.pooler.supabase.com:5432/postgres
    """
    import os
    # 1. Try Streamlit secrets (cloud deployment)
    try:
        return st.secrets["RATEEDGE_DB_URL"]
    except Exception:
        pass
    # 2. Try environment variable (local .env)
    url = os.environ.get("RATEEDGE_DB_URL", "")
    return url if url else None

@st.cache_resource(show_spinner=False)
def _get_db_params():
    """Parse and cache DB connection params once."""
    db_url = get_db_url()
    if not db_url:
        return None
    try:
        from urllib.parse import urlparse, parse_qs
        r = urlparse(db_url)
        qs = parse_qs(r.query)
        return {
            "host": r.hostname, "port": r.port or 5432,
            "dbname": r.path.lstrip('/'), "user": r.username,
            "password": r.password,
            "sslmode": qs.get('sslmode', ['require'])[0]
        }
    except:
        return None

def get_db_connection():
    """
    Get Supabase PostgreSQL connection via shared pooler URL.
    """
    if not HAS_POSTGRES:
        return None
    params = _get_db_params()
    if not params:
        return None
    try:
        conn = psycopg2.connect(
            host=params["host"], port=params["port"],
            dbname=params["dbname"], user=params["user"],
            password=params["password"], sslmode=params["sslmode"],
            connect_timeout=10,
            keepalives=1, keepalives_idle=30,
            keepalives_interval=10, keepalives_count=3,
            options="-c statement_timeout=15000"  # 15s max per query
        )
        return conn
    except Exception:
        return None


def init_database():
    """Create tables if they don't exist"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_configs (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(50) NOT NULL,
                config_type VARCHAR(50) NOT NULL,
                currency VARCHAR(3) NOT NULL,
                data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(user_id, config_type, currency)
            );
            
            CREATE INDEX IF NOT EXISTS idx_user_configs_user 
            ON user_configs(user_id);
            
            CREATE TABLE IF NOT EXISTS vol_history (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(50) NOT NULL,
                currency VARCHAR(3) NOT NULL,
                snapshot_date TIMESTAMP NOT NULL,
                label VARCHAR(100) NOT NULL,
                atm_vols JSONB NOT NULL,
                atm_prems JSONB,
                sabr_alpha JSONB,
                sabr_beta JSONB,
                sabr_rho JSONB,
                sabr_nu JSONB,
                notes TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_vol_history_user_ccy 
            ON vol_history(user_id, currency);
            
            CREATE INDEX IF NOT EXISTS idx_vol_history_date 
            ON vol_history(snapshot_date DESC);
            
            ALTER TABLE vol_history ADD COLUMN IF NOT EXISTS atm_prems JSONB;

            CREATE TABLE IF NOT EXISTS sod_reports (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(50) NOT NULL,
                report_date DATE NOT NULL,
                created_at TIMESTAMP DEFAULT NOW(),
                usd_t1_label TEXT,
                usd_t2_label TEXT,
                aud_snap_label TEXT,
                report_data JSONB,
                notes TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_sod_reports_user_date
            ON sod_reports(user_id, report_date DESC);

            CREATE TABLE IF NOT EXISTS blotter_mids (
                id SERIAL PRIMARY KEY,
                ccy VARCHAR(3) NOT NULL,
                key VARCHAR(50) NOT NULL,
                value NUMERIC,
                label TEXT,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(ccy, key)
            );
            CREATE INDEX IF NOT EXISTS idx_blotter_mids_ccy ON blotter_mids(ccy, key);
        """)
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Database init failed: {e}")
        return False


def save_user_config(user_id: str, config_type: str, currency: str, data: dict, _conn=None):
    """Save user config to database. Pass _conn to reuse an existing connection."""
    own_conn = _conn is None
    conn = get_db_connection() if own_conn else _conn
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO user_configs (user_id, config_type, currency, data, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (user_id, config_type, currency)
            DO UPDATE SET data = %s, updated_at = NOW()
        """, (user_id, config_type, currency, Json(data), Json(data)))
        if own_conn:
            conn.commit()
            cur.close()
            conn.close()
        else:
            cur.close()
        return True
    except Exception as e:
        if own_conn:
            try: conn.close()
            except: pass
        return False


@st.cache_data(ttl=60, show_spinner=False)
def load_user_config(user_id: str, config_type: str, currency: str) -> Optional[dict]:
    """Load user config from database"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT data FROM user_configs 
            WHERE user_id = %s AND config_type = %s AND currency = %s
        """, (user_id, config_type, currency))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0] if row else None
    except Exception as e:
        return None


def load_all_user_configs(user_id: str) -> dict:
    """Load all configs for a user"""
    conn = get_db_connection()
    if not conn:
        return {}
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT config_type, currency, data, updated_at 
            FROM user_configs 
            WHERE user_id = %s
            ORDER BY updated_at DESC
        """, (user_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        configs = {}
        for config_type, currency, data, updated_at in rows:
            if config_type not in configs:
                configs[config_type] = {}
            configs[config_type][currency] = {
                "data": data,
                "updated_at": updated_at.isoformat() if updated_at else None
            }
        return configs
    except Exception as e:
        return {}


def save_all_session_data(user_id: str):
    """Save all current session data using a single DB connection."""
    conn = get_db_connection()
    if not conn:
        return 0
    saved = 0
    try:
        def _save(config_type, currency, data):
            nonlocal saved
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO user_configs (user_id, config_type, currency, data, updated_at)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (user_id, config_type, currency)
                DO UPDATE SET data = %s, updated_at = NOW()
            """, (user_id, config_type, currency, Json(data), Json(data)))
            cur.close()
            saved += 1

        _debug_msgs = []
        for ccy in SUPPORTED_CURRENCIES:
            # ATM vols are NOT saved to user_configs — they live in vol_history only
            # Saving here caused stale/wrong surfaces to override vol_history on load
            vol_data = st.session_state.get("vol_data", {}).get(ccy, {})

            for param in ["alpha", "beta", "rho", "nu"]:
                val = vol_data.get(param)
                if val is not None:
                    val_save = val.copy()
                    if val_save.index.name == "Expiry":
                        val_save = val_save.reset_index()
                    elif "Expiry" not in val_save.columns:
                        first_idx = val_save.index[0] if len(val_save) > 0 else None
                        if isinstance(first_idx, str) and first_idx.lower().endswith(('w','m','y')):
                            val_save = val_save.reset_index()
                            val_save.columns = ["Expiry"] + list(val_save.columns[1:])
                    records = []
                    for _, row in val_save.iterrows():
                        rec = {"Expiry": row.get("Expiry", "")}
                        for col in val_save.columns:
                            if col != "Expiry": rec[col] = row[col]
                        records.append(rec)
                    _save(f"sabr_{param}", ccy, {"values": records})

            basis = st.session_state.get("basis_curves", {}).get(ccy, {})
            for basis_type in ["6v3", "3v1", "ois"]:
                bc = basis.get(basis_type)
                if bc is not None:
                    bc_save = bc.copy()
                    if bc_save.index.name is not None:
                        bc_save = bc_save.reset_index()
                    _save(f"basis_{basis_type}", ccy, {"values": bc_save.to_dict(orient="records")})

        # AUD par rates — needed to rebuild QQ/SS zero curves after DB load
        _par_qq = st.session_state.get("_aud_par_qq", {})
        _par_ss = st.session_state.get("_aud_par_ss", {})
        if _par_qq and _par_ss:
            _save("aud_par_rates", "AUD", {"par_qq": _par_qq, "par_ss": _par_ss})

        # CFS spreads — save all 9 wedge spread values
        _cf_spread_keys = ["cf_spr_3m1y","cf_spr_1y1y","cf_spr_2y1y","cf_spr_3y1y",
                           "cf_spr_4y1y","cf_spr_5y2y","cf_spr_7y3y","cf_spr_10y2y","cf_spr_12y3y"]
        _cf_spread_data = {k: float(st.session_state.get(k, 0)) for k in _cf_spread_keys
                           if k in st.session_state}
        if _cf_spread_data:
            _save("cf_spreads", "AUD", _cf_spread_data)

        # CFS table data — manually entered straddle overrides per wedge
        _cfs_tdata = st.session_state.get("cfs_table_data", {})
        if _cfs_tdata:
            _save("cfs_table_data", "AUD", _cfs_tdata)

        # FWD prefs
        _fwd_prefs = {
            "irs_sp_list":  [list(x) for x in st.session_state.get("irs_sp_list", [])],
            "irs_fl_list":  [list(x) for x in st.session_state.get("irs_fl_list", [])],
            "fvfv_list":    [list(x) for x in st.session_state.get("fvfv_list", [])],
            "b6_list":      list(st.session_state.get("b6_list", [])),
            "fv6_list":     [list(x) for x in st.session_state.get("fv6_list", [])],
            "bsp_list":     [list(x) for x in st.session_state.get("bsp_list", [])],
        }
        if any(_fwd_prefs.values()):
            _save("fwd_analysis_prefs", "GLB", _fwd_prefs)

        conn.commit()
        try:
            import streamlit as _st
            _st.session_state["_save_debug_msgs"] = _debug_msgs
        except: pass
    except Exception as _e:
        try: conn.rollback()
        except: pass
        try:
            import streamlit as _st
            _st.session_state["_save_last_error"] = str(_e)
        except: pass
    finally:
        try: conn.close()
        except: pass
    return saved


def load_all_session_data(user_id: str, load_date: str = None) -> int:
    """Load all saved data into session state. load_date: specific date string (YYYY-MM-DD) or None for latest."""
    configs = load_all_user_configs(user_id)
    # If nothing found, try the other admin email as fallback
    # (desktop may have saved as wpo@rateedge.au, phone logs in as wpo70@icloud.com or vice versa)
    if not configs:
        _ADMIN_ALIASES = {"wpo@rateedge.au": "wpo70@icloud.com", "wpo70@icloud.com": "wpo@rateedge.au"}
        _alt = _ADMIN_ALIASES.get(user_id)
        if _alt:
            configs = load_all_user_configs(_alt)
    if not configs:
        return 0
    
    loaded = 0
    
    # Initialize session state containers
    if "curves" not in st.session_state:
        st.session_state["curves"] = {}
    if "vol_data" not in st.session_state:
        st.session_state["vol_data"] = {}
    if "basis_curves" not in st.session_state:
        st.session_state["basis_curves"] = {}
    if "vol_editor" not in st.session_state:
        st.session_state["vol_editor"] = {"working": {}, "base": {}, "history": {}, "redo_stack": {}, "selected_cell": {}}
    
    for ccy in SUPPORTED_CURRENCIES:
        # Initialize vol_data for this currency
        if ccy not in st.session_state["vol_data"]:
            st.session_state["vol_data"][ccy] = {}
        
        # NOTE: Curves are NOT auto-loaded from DB on login.
        # Curves must be committed from BBG_Feed via Vol/Upload tab.
        # This ensures curves always reflect the live BBG rates, not stale DB zeros.

        # ATM vols are NOT loaded from user_configs — they come from vol_history only
        # (see auto-load block at login which pulls latest snapshot per currency)
        
        # Load SABR params into vol_data
        for param in ["alpha", "beta", "rho", "nu"]:
            key = f"sabr_{param}"
            if key in configs and ccy in configs[key]:
                try:
                    df = pd.DataFrame(configs[key][ccy]["data"]["values"])
                    # Reorder columns to ensure Expiry is first if it exists
                    if "Expiry" in df.columns:
                        cols = ["Expiry"] + [c for c in df.columns if c != "Expiry"]
                        df = df[cols]
                    st.session_state["vol_data"][ccy][param] = df
                    loaded += 1
                except:
                    pass
        
        # Load basis curves
        if ccy not in st.session_state["basis_curves"]:
            st.session_state["basis_curves"][ccy] = {}
        
        for basis_type in ["6v3", "3v1", "ois"]:
            key = f"basis_{basis_type}"
            if key in configs and ccy in configs[key]:
                try:
                    df = pd.DataFrame(configs[key][ccy]["data"]["values"])
                    st.session_state["basis_curves"][ccy][basis_type] = df
                    if "config_basis" not in st.session_state:
                        st.session_state["config_basis"] = {}
                    if ccy not in st.session_state["config_basis"]:
                        st.session_state["config_basis"][ccy] = {}
                    st.session_state["config_basis"][ccy][basis_type] = df
                    loaded += 1
                except:
                    pass

    # AUD: restore par rates and rebuild zero curves from aud_par_rates in user_configs
    if "aud_par_rates" in configs and "AUD" in configs["aud_par_rates"]:
        try:
            _pdata = configs["aud_par_rates"]["AUD"]["data"]
            _par_qq = {float(k): float(v) for k, v in _pdata.get("par_qq", {}).items()}
            _par_ss = {float(k): float(v) for k, v in _pdata.get("par_ss", {}).items()}
            if _par_qq and _par_ss:
                st.session_state["_aud_par_qq"] = _par_qq
                st.session_state["_aud_par_ss"] = _par_ss

                # Rebuild _irs_par_rates["AUD"] so IRS Par chart shows real par rates
                _par_rows = []
                for _t, _r in sorted(_par_qq.items()): _par_rows.append({"Tenor": f"{_t}Y", "Par Rate (%)": _r, "Conv": "Q/Q"})
                for _t, _r in sorted(_par_ss.items()): _par_rows.append({"Tenor": f"{_t}Y", "Par Rate (%)": _r, "Conv": "S/S"})
                if "_irs_par_rates" not in st.session_state: st.session_state["_irs_par_rates"] = {}
                st.session_state["_irs_par_rates"]["AUD"] = pd.DataFrame(_par_rows)

                # Rebuild pure QQ and SS zero curves using OIS from session state
                _ois_df = st.session_state.get("config_basis", {}).get("AUD", {}).get("ois")
                _basis_df = st.session_state.get("config_basis", {}).get("AUD", {}).get("6v3")

                _ois_rates_rebuild = {}
                if _ois_df is not None and not _ois_df.empty:
                    for _, _row in _ois_df.iterrows():
                        _t_ois = float(_row["MaturityY"])
                        if _t_ois <= 3.01:  # only seed OIS up to 3Y
                            _ois_rates_rebuild[_t_ois] = float(_row["ZeroRatePct"])

                _bx2 = _by2 = None
                if _basis_df is not None and not _basis_df.empty:
                    _bx2 = _basis_df["MaturityY"].to_numpy().astype(float)
                    _by2 = _basis_df["BasisBp"].to_numpy().astype(float)
                def _b2(t):
                    if _bx2 is None: return 0.0
                    return float(np.interp(t, _bx2, _by2))

                _SPOT2 = 1.0 / 252.0

                def _rebuild_zero(par_inputs, all_qq):
                    _dfs2 = {0.0: 1.0}
                    for _t2, _r2 in sorted(_ois_rates_rebuild.items()):
                        _dfs2[_t2] = math.exp(-_r2 / 100.0 * _t2)
                    def _dfi2(t):
                        _ts2 = sorted(_dfs2.keys()); _dfv2 = [_dfs2[x] for x in _ts2]
                        if t <= _ts2[0]: return _dfv2[0]
                        if t >= _ts2[-1]:
                            _z2 = -math.log(_dfv2[-1]) / _ts2[-1]; return math.exp(-_z2 * t)
                        for _i2 in range(len(_ts2) - 1):
                            if _ts2[_i2] <= t <= _ts2[_i2+1]:
                                _w2 = (t - _ts2[_i2]) / (_ts2[_i2+1] - _ts2[_i2])
                                return math.exp((1-_w2)*math.log(_dfv2[_i2]) + _w2*math.log(_dfv2[_i2+1]))
                        return _dfv2[-1]
                    _freq2 = 0.25 if all_qq else 0.50
                    for _tenor2 in sorted(par_inputs.keys()):
                        _c2 = par_inputs[_tenor2] / 100.0
                        _swap_end2 = _SPOT2 + _tenor2
                        _times2 = []; _t2i = _SPOT2 + _freq2
                        while _t2i <= _swap_end2 + 1e-9:
                            _times2.append(round(min(_t2i, _swap_end2), 8)); _t2i += _freq2
                        if not _times2: continue
                        _ann2 = sum(_dfi2(_ti2) * _freq2 for _ti2 in _times2[:-1])
                        _df_end2 = (_dfi2(_SPOT2) - _c2 * _ann2) / (1.0 + _c2 * _freq2)
                        if _df_end2 > 0 and not math.isnan(_df_end2):
                            _dfs2[_swap_end2] = _df_end2
                    _MATS2 = [0.25,0.5,0.75,1.0,1.5,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,10.0,12.0,15.0,20.0,25.0,30.0,40.0,50.0]
                    _zc2 = {}
                    for _m2 in _MATS2:
                        _d2 = _dfi2(_m2)
                        if _d2 > 0 and not math.isnan(_d2):
                            _zc2[_m2] = -math.log(_d2) / _m2 * 100
                    return _zc2

                _par_qq_full2 = dict(_par_qq)
                for _t2, _r2 in _par_ss.items(): _par_qq_full2[_t2] = _r2 - _b2(_t2) / 100.0
                _par_ss_full2 = dict(_par_ss)
                for _t2, _r2 in _par_qq.items(): _par_ss_full2[_t2] = _r2 + _b2(_t2) / 100.0

                st.session_state["_aud_zc_qq"] = _rebuild_zero(_par_qq_full2, all_qq=True)
                st.session_state["_aud_zc_ss"] = _rebuild_zero(_par_ss_full2, all_qq=False)
                loaded += 1
        except Exception as _pex:
            pass

    # Restore CFS wedge spreads
    if "cf_spreads" in configs and "AUD" in configs["cf_spreads"]:
        try:
            _cf_d = configs["cf_spreads"]["AUD"]["data"]
            for _k, _v in _cf_d.items():
                st.session_state[_k] = float(_v)
            loaded += 1
        except Exception:
            pass

    # Restore CFS table data (manually entered straddle overrides)
    if "cfs_table_data" in configs and "AUD" in configs["cfs_table_data"]:
        try:
            _ctd = configs["cfs_table_data"]["AUD"]["data"]
            if not isinstance(_ctd, dict):
                _ctd = {}
            if "cfs_table_data" not in st.session_state:
                st.session_state["cfs_table_data"] = {}
            st.session_state["cfs_table_data"].update(_ctd)
            loaded += 1
        except Exception:
            pass

    # Restore FWD Analysis series lists   —   DB always wins, overrides tab defaults
    if "fwd_analysis_prefs" in configs and "GLB" in configs["fwd_analysis_prefs"]:
        try:
            _p = configs["fwd_analysis_prefs"]["GLB"]["data"]
            if "irs_sp_list" in _p:
                st.session_state["irs_sp_list"] = [tuple(x) for x in _p["irs_sp_list"]]
            if "irs_fl_list" in _p:
                st.session_state["irs_fl_list"] = [tuple(x) for x in _p["irs_fl_list"]]
            if "fvfv_list" in _p:
                st.session_state["fvfv_list"] = [tuple(x) for x in _p["fvfv_list"]]
            if "b6_list" in _p:
                st.session_state["b6_list"] = list(_p["b6_list"])
            if "fv6_list" in _p:
                st.session_state["fv6_list"] = [tuple(x) for x in _p["fv6_list"]]
            if "bsp_list" in _p:
                st.session_state["bsp_list"] = [tuple(x) for x in _p["bsp_list"]]
            loaded += 1
        except:
            pass

    return loaded


# ============================
# Historical Vol Storage
# ============================

def save_vol_snapshot(user_id: str, currency: str, label: str, notes: str = ""):
    # Normalise admin email so snapshots are always under one canonical user_id
    _ADMIN_ALIASES = {"wpo70@icloud.com": "wpo@rateedge.au"}
    user_id = _ADMIN_ALIASES.get(user_id, user_id)
    """Save current vol matrix and SABR params as historical snapshot"""
    from datetime import datetime
    
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        # Get current vol data from session
        vol_data = st.session_state.get("vol_data", {}).get(currency, {})
        atm = vol_data.get("atm")
        
        if atm is None:
            st.error(f"No ATM vol data loaded for {currency}")
            return False
        
        # Convert DataFrame to JSON
        atm_json = atm.to_dict(orient="records")
        
        # Get SABR params if available
        sabr_alpha = vol_data.get("alpha")
        sabr_beta = vol_data.get("beta")
        sabr_rho = vol_data.get("rho")
        sabr_nu = vol_data.get("nu")
        
        sabr_alpha_json = sabr_alpha.to_dict(orient="records") if sabr_alpha is not None else None
        sabr_beta_json = sabr_beta.to_dict(orient="records") if sabr_beta is not None else None
        sabr_rho_json = sabr_rho.to_dict(orient="records") if sabr_rho is not None else None
        sabr_nu_json = sabr_nu.to_dict(orient="records") if sabr_nu is not None else None
        
        # Use the pricer's own Forward Premium (bp) matrix from Curves tab → Generate ATM Matrix
        # This is st.session_state["atm_prem_matrix"][currency]["prem"]
        # If not generated yet, atm_prems is saved as NULL
        atm_prems_json = None
        try:
            _prem_df = st.session_state.get("atm_prem_matrix", {}).get(currency, {}).get("prem")
            if _prem_df is not None:
                _prem_save = _prem_df.copy()
                if _prem_save.index.name == "Expiry":
                    _prem_save = _prem_save.reset_index()
                elif "Expiry" not in _prem_save.columns:
                    _first = _prem_save.index[0] if len(_prem_save) > 0 else None
                    if isinstance(_first, str):
                        _prem_save = _prem_save.reset_index()
                        _prem_save.columns = ["Expiry"] + list(_prem_save.columns[1:])
                atm_prems_json = _prem_save.to_dict(orient="records")
        except Exception:
            atm_prems_json = None

        # Insert into vol_history table
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO vol_history 
            (user_id, currency, snapshot_date, label, atm_vols, atm_prems, sabr_alpha, sabr_beta, sabr_rho, sabr_nu, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            user_id, 
            currency, 
            datetime.now(),
            label,
            Json({"values": atm_json}),
            Json({"values": atm_prems_json}) if atm_prems_json else None,
            Json({"values": sabr_alpha_json}) if sabr_alpha_json else None,
            Json({"values": sabr_beta_json}) if sabr_beta_json else None,
            Json({"values": sabr_rho_json}) if sabr_rho_json else None,
            Json({"values": sabr_nu_json}) if sabr_nu_json else None,
            notes
        ))
        
        snapshot_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        
        return snapshot_id
        
    except Exception as e:
        st.error(f"Failed to save snapshot: {e}")
        return False


@st.cache_data(ttl=60, show_spinner=False)
def list_vol_snapshots(user_id: str, currency: str = None):
    # Normalise: both admin emails share the same snapshots
    _ADMIN_ALIASES = {"wpo70@icloud.com": "wpo@rateedge.au", "wpo@rateedge.au": "wpo@rateedge.au"}
    user_id = _ADMIN_ALIASES.get(user_id, user_id)
    """List all historical vol snapshots for a user"""
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        cur = conn.cursor()
        if currency:
            cur.execute("""
                SELECT id, currency, snapshot_date, label, notes, created_at
                FROM vol_history
                WHERE user_id = %s AND currency = %s
                ORDER BY snapshot_date DESC
            """, (user_id, currency))
        else:
            cur.execute("""
                SELECT id, currency, snapshot_date, label, notes, created_at
                FROM vol_history
                WHERE user_id = %s
                ORDER BY snapshot_date DESC
            """, (user_id,))
        
        snapshots = cur.fetchall()
        cur.close()
        conn.close()
        
        return [
            {
                "id": s[0],
                "currency": s[1],
                "snapshot_date": s[2],
                "label": s[3],
                "notes": s[4],
                "created_at": s[5]
            }
            for s in snapshots
        ]
        
    except Exception as e:
        st.error(f"Failed to list snapshots: {e}")
        return []


def load_vol_snapshot(snapshot_id: int):
    """Load a historical vol snapshot"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT currency, atm_vols, sabr_alpha, sabr_beta, sabr_rho, sabr_nu, label, snapshot_date
            FROM vol_history
            WHERE id = %s
        """, (snapshot_id,))
        
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if not row:
            return None
        
        currency, atm_vols, sabr_alpha, sabr_beta, sabr_rho, sabr_nu, label, snapshot_date = row
        
        # Convert JSON to DataFrames
        atm_df = pd.DataFrame(atm_vols["values"]) if atm_vols else None
        sabr_alpha_df = pd.DataFrame(sabr_alpha["values"]) if sabr_alpha else None
        sabr_beta_df = pd.DataFrame(sabr_beta["values"]) if sabr_beta else None
        sabr_rho_df = pd.DataFrame(sabr_rho["values"]) if sabr_rho else None
        sabr_nu_df = pd.DataFrame(sabr_nu["values"]) if sabr_nu else None
        
        return {
            "currency": currency,
            "label": label,
            "snapshot_date": snapshot_date,
            "atm": atm_df,
            "alpha": sabr_alpha_df,
            "beta": sabr_beta_df,
            "rho": sabr_rho_df,
            "nu": sabr_nu_df
        }
        
    except Exception as e:
        st.error(f"Failed to load snapshot: {e}")
        return None


def delete_vol_snapshot(snapshot_id: int):
    """Delete a historical vol snapshot"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM vol_history WHERE id = %s", (snapshot_id,))
        conn.commit()
        cur.close()
        conn.close()
        return True
        
    except Exception as e:
        st.error(f"Failed to delete snapshot: {e}")
        return False


def save_sod_report(user_id: str, report_date, usd_t1_label: str, usd_t2_label: str,
                    aud_snap_label: str, report_data: dict, notes: str = "") -> Optional[int]:
    """Save a SOD report to the database."""
    conn = get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO sod_reports (user_id, report_date, usd_t1_label, usd_t2_label, aud_snap_label, report_data, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (user_id, report_date, usd_t1_label, usd_t2_label, aud_snap_label,
              Json(report_data), notes))
        report_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return report_id
    except Exception as e:
        st.error(f"Failed to save SOD report: {e}")
        return None


def list_sod_reports(user_id: str, limit: int = 30) -> list:
    """List saved SOD reports for a user."""
    conn = get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, report_date, created_at, usd_t1_label, usd_t2_label, aud_snap_label, notes
            FROM sod_reports
            WHERE user_id = %s
            ORDER BY report_date DESC, created_at DESC
            LIMIT %s
        """, (user_id, limit))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [{"id": r[0], "report_date": r[1], "created_at": r[2],
                 "usd_t1": r[3], "usd_t2": r[4], "aud_snap": r[5], "notes": r[6]}
                for r in rows]
    except Exception as e:
        return []


def load_sod_report(report_id: int) -> Optional[dict]:
    """Load a SOD report from the database."""
    conn = get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT report_date, usd_t1_label, usd_t2_label, aud_snap_label, report_data, notes, created_at
            FROM sod_reports WHERE id = %s
        """, (report_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None
        return {"report_date": row[0], "usd_t1": row[1], "usd_t2": row[2],
                "aud_snap": row[3], "data": row[4], "notes": row[5], "created_at": row[6]}
    except Exception as e:
        return None


def delete_sod_report(report_id: int) -> bool:
    """Delete a SOD report."""
    conn = get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM sod_reports WHERE id = %s", (report_id,))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception:
        return False


def publish_blotter_mids(ccy: str, mids: dict) -> int:
    """Publish mid market values to blotter_mids table for the React blotter to consume.
    mids = {key: {"value": float, "label": str}}
    Returns number of rows upserted. Uses executemany for performance.
    """
    conn = get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        rows = []
        for key, data in mids.items():
            val = data.get("value")
            lbl = data.get("label", "")
            if val is None:
                continue
            rows.append((ccy, key, float(val), lbl))
        if not rows:
            conn.close()
            return 0
        # True batch upsert via execute_values
        from psycopg2.extras import execute_values as _ev
        _ev(cur, """
            INSERT INTO blotter_mids (ccy, key, value, label, updated_at)
            VALUES %s
            ON CONFLICT (ccy, key) DO UPDATE
                SET value = EXCLUDED.value,
                    label = EXCLUDED.label,
                    updated_at = NOW()
        """, rows, template="(%s, %s, %s, %s, NOW())")
        conn.commit()
        cur.close()
        conn.close()
        return len(rows)
    except Exception as e:
        st.error(f"Failed to publish blotter mids: {e}")
        return 0


def export_vol_surface_to_excel(currency: str, include_sabr: bool = True) -> Optional[bytes]:
    """Export vol surface (ATM + optionally SABR) to Excel file"""
    from io import BytesIO
    
    try:
        # Get vol data
        vol_data = st.session_state.get("vol_data", {}).get(currency, {})
        atm = vol_data.get("atm")
        
        if atm is None or atm.empty:
            return None
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write ATM vol surface
            atm.to_excel(writer, sheet_name=f"ATM_Vols_{currency}", index=False)
            
            # Write SABR parameters if requested and available
            if include_sabr:
                sabr_alpha = vol_data.get("alpha")
                sabr_beta = vol_data.get("beta")
                sabr_rho = vol_data.get("rho")
                sabr_nu = vol_data.get("nu")
                
                if sabr_alpha is not None:
                    sabr_alpha.to_excel(writer, sheet_name=f"SABR_Alpha_{currency}", index=False)
                if sabr_beta is not None:
                    sabr_beta.to_excel(writer, sheet_name=f"SABR_Beta_{currency}", index=False)
                if sabr_rho is not None:
                    sabr_rho.to_excel(writer, sheet_name=f"SABR_Rho_{currency}", index=False)
                if sabr_nu is not None:
                    sabr_nu.to_excel(writer, sheet_name=f"SABR_Nu_{currency}", index=False)
            
            # Add metadata sheet
            metadata = pd.DataFrame({
                "Property": ["Currency", "Export Date", "Export Time", "Includes SABR"],
                "Value": [
                    currency,
                    pd.Timestamp.now().strftime('%Y-%m-%d'),
                    pd.Timestamp.now().strftime('%H:%M:%S'),
                    "Yes" if include_sabr else "No"
                ]
            })
            metadata.to_excel(writer, sheet_name="Metadata", index=False)
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Export failed: {e}")
        return None


# ============================
# Data structures
# ============================

@dataclass
class SwaptionTicket:
    side: str
    payoff_type: str  # "vanilla", "digital", "straddle"
    notional: float
    currency: str
    expiry_years: float
    swap_tenor_years: float
    forward: float
    strike: float
    vol: float
    discount_rate: float
    annuity: float
    model: str  # "Black" or "Normal"
    payout_bp: float = 1.0
    label: str = ""
    use_curve: bool = False

    def df(self, t: Optional[float] = None) -> float:
        if t is None:
            t = self.expiry_years
        return math.exp(-self.discount_rate * t)


# ============================
# Helper functions
# ============================

def label_to_years(lbl: str) -> float:
    from datetime import date as _date, datetime as _dt
    try:
        _f = float(lbl)
        if _f >= 0:
            return _f
    except (TypeError, ValueError):
        pass
    if lbl is None or str(lbl).strip() == "" or str(lbl).strip().lower() in ("none","nan","expiry"):
        return 0.0
    x = str(lbl).strip()
    # Handle DD/MM/YYYY or YYYY-MM-DD date strings
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            target = _dt.strptime(x, fmt).date()
            days = (target - _date.today()).days
            return max(days / 365.0, 1/365.0)
        except ValueError:
            pass
    xl = x.lower()
    if xl.endswith("d"):
        return float(xl[:-1]) / 365.0
    if xl.endswith("w"):
        return float(xl[:-1]) / 52.0
    if xl.endswith("m"):
        return float(xl[:-1]) / 12.0
    if xl.endswith("y"):
        return float(xl[:-1])
    return float(xl)


def load_atm_surface(df: pd.DataFrame, name: str) -> pd.DataFrame:
    if "Expiry" not in df.columns:
        raise ValueError(f"{name} must have an 'Expiry' column")
    return df.copy()


def ensure_sabr_matrix(df: pd.DataFrame, name: str) -> pd.DataFrame:
    if "Expiry" not in df.columns:
        raise ValueError(f"{name} must have an 'Expiry' column")
    return df.copy()


def get_matrix_value(mat: Optional[pd.DataFrame],
                     expiry_label: str,
                     tenor_years: float) -> Optional[float]:
    if mat is None or mat.empty:
        return None

    _tcols = [c for c in mat.columns if c != "Expiry"]
    if not _tcols:
        return None

    _ten_yrs = np.array([label_to_years(str(c)) for c in _tcols], dtype=float)
    _exp_yrs = mat["Expiry"].apply(lambda x: label_to_years(str(x).strip().lower())).values.astype(float)
    _exp_y_req = label_to_years(str(expiry_label).strip().lower())

    # Fast path: exact expiry row match
    _lbl = str(expiry_label).strip().lower()
    _row_match = mat[mat["Expiry"].astype(str).str.strip().str.lower() == _lbl]
    if not _row_match.empty:
        _row_vals = pd.to_numeric(_row_match.iloc[0][_tcols], errors='coerce').values.astype(float)
        _mask = ~np.isnan(_row_vals)
        if _mask.sum() == 0: return None
        return float(np.interp(tenor_years, _ten_yrs[_mask], _row_vals[_mask]))

    # Slow path: interpolate across expiry axis too
    _col_vals = []
    for _i in range(len(mat)):
        _row = pd.to_numeric(mat[_tcols].iloc[_i], errors='coerce').values.astype(float)
        _mask = ~np.isnan(_row)
        if _mask.sum() == 0: _col_vals.append(np.nan); continue
        _col_vals.append(float(np.interp(tenor_years, _ten_yrs[_mask], _row[_mask])))
    _col_vals = np.array(_col_vals, dtype=float)
    _valid = ~np.isnan(_col_vals)
    if _valid.sum() == 0: return None
    return float(np.interp(_exp_y_req, _exp_yrs[_valid], _col_vals[_valid]))


def get_sabr_params_from_matrices(a: Optional[pd.DataFrame],
                                  b: Optional[pd.DataFrame],
                                  r: Optional[pd.DataFrame],
                                  n: Optional[pd.DataFrame],
                                  expiry_label: str,
                                  tenor_years: float) -> Optional[dict]:
    alpha = get_matrix_value(a, expiry_label, tenor_years)
    beta = get_matrix_value(b, expiry_label, tenor_years)
    rho = get_matrix_value(r, expiry_label, tenor_years)
    nu = get_matrix_value(n, expiry_label, tenor_years)
    if any(x is None for x in (alpha, beta, rho, nu)):
        return None
    return dict(alpha=alpha, beta=beta, rho=rho, nu=nu)


def sabr_implied_vol_black(F: float, K: float, T: float,
                           alpha: float, beta: float,
                           rho: float, nu: float,
                           eps: float = 1e-7) -> float:
    """Hagan SABR approximation for Black vols."""
    F = float(F)
    K = float(K)
    alpha = float(alpha)
    beta = float(beta)
    rho = float(rho)
    nu = float(nu)

    if T <= 0 or alpha <= 0:
        return 0.0

    # ATM case
    if abs(F - K) < eps:
        Fmid = F
        term1 = alpha / (Fmid ** (1 - beta))
        term2 = (
            ((1 - beta) ** 2 / 24.0) * (alpha ** 2 / (Fmid ** (2 - 2 * beta))) +
            0.25 * rho * beta * nu * alpha / (Fmid ** (1 - beta)) +
            (2 - 3 * rho * rho) * nu * nu / 24.0
        ) * T
        return term1 * (1 + term2)

    # Away from ATM
    FK = F * K
    logFK = math.log(F / K)
    
    # Leading order term
    FK_beta = FK ** ((1 - beta) / 2.0)
    
    # z and chi(z) for the smile
    z = (nu / alpha) * FK_beta * logFK
    
    if abs(z) < eps:
        x_z = 1.0
    else:
        sqrt_term = math.sqrt(1 - 2 * rho * z + z * z)
        x_z = z / math.log((sqrt_term + z - rho) / (1 - rho))
    
    # Numerator correction for logFK
    num_corr = 1 + ((1 - beta) ** 2 / 24.0) * (logFK ** 2) + ((1 - beta) ** 4 / 1920.0) * (logFK ** 4)
    
    # Denominator (time-dependent correction)
    den_corr = 1 + (
        ((1 - beta) ** 2 / 24.0) * (alpha ** 2 / (FK ** (1 - beta))) +
        0.25 * rho * beta * nu * alpha / FK_beta +
        (2 - 3 * rho * rho) * nu * nu / 24.0
    ) * T
    
    vol = (alpha / FK_beta) * x_z * num_corr / den_corr
    return max(vol, 0.0)


def sabr_normal_atm_vol(F: float, T: float, alpha: float, beta: float, rho: float, nu: float) -> float:
    """Normal (Bachelier) SABR ATM vol approximation."""
    if T <= 0 or alpha <= 0 or F <= 0:
        return 0.0
    return alpha * (F ** beta) * (1.0 + (2.0 - 3.0 * rho ** 2) / 24.0 * nu ** 2 * T)


def sabr_normal_vol_smile(F: float, K: float, T: float,
                           alpha: float, beta: float, rho: float, nu: float) -> float:
    """Full Normal SABR smile vol (Antonov et al.). Returns vol in same units as alpha.
    For ATM (K≈F) reduces to sabr_normal_atm_vol.
    Formula: σ_N = alpha × (FK)^(β/2) × z/χ(z) × B
    """
    if T <= 0 or alpha <= 0 or F <= 0 or K <= 0:
        return sabr_normal_atm_vol(F, T, alpha, beta, rho, nu)
    if abs(K - F) < 1e-8:
        return sabr_normal_atm_vol(F, T, alpha, beta, rho, nu)
    try:
        FK = F * K
        FK_mid = FK ** (beta / 2.0)          # (FK)^(β/2)  — Normal SABR uses multiply
        FK_mid_inv = FK ** ((1 - beta) / 2.0) # (FK)^((1-β)/2) — for B correction terms
        log_FK = math.log(F / K)
        z = (nu / alpha) * FK_mid_inv * log_FK
        if abs(z) < 1e-7:
            x_z = 1.0
        else:
            _sq = math.sqrt(1 - 2 * rho * z + z ** 2)
            _denom = math.log((_sq + z - rho) / (1 - rho))
            x_z = z / _denom if abs(_denom) > 1e-10 else 1.0
        A = alpha * FK_mid  # Normal SABR: α × (FK)^(β/2)
        B = 1 + ((1 - beta) ** 2 / 24 * alpha ** 2 / FK ** (1 - beta) +
                  rho * beta * nu * alpha / (4 * FK ** ((1 - beta) / 2)) +
                  (2 - 3 * rho ** 2) / 24 * nu ** 2) * T
        return A * x_z * B
    except Exception:
        return sabr_normal_atm_vol(F, T, alpha, beta, rho, nu)


def sabr_implied_alpha_from_atm(atm_vol_normal: float, F: float, T: float,
                                  beta: float, rho: float, nu: float) -> float:
    """Back out alpha from ATM normal vol, given fixed beta/rho/nu.
    Solves: atm_vol = alpha * F^beta * (1 + (2-3ρ,ν~)/24 * ×~ * T)
    """
    if T <= 0 or F <= 0:
        return 0.0
    denom = (F ** beta) * (1.0 + (2.0 - 3.0 * rho ** 2) / 24.0 * nu ** 2 * T)
    if abs(denom) < 1e-12:
        return 0.0
    return atm_vol_normal / denom


# ============================
# Curve construction
# ============================

def load_curve(df: pd.DataFrame, name: str) -> pd.DataFrame:
    if {"Maturity (Years)", "Zero Rate (%)"}.issubset(df.columns):
        out = df[["Maturity (Years)", "Zero Rate (%)"]].copy()
        out.rename(columns={"Maturity (Years)": "MaturityY", "Zero Rate (%)": "ZeroRatePct"}, inplace=True)
        out = out.sort_values("MaturityY").reset_index(drop=True)
        return out
    raise ValueError(f"{name}: expected columns 'Maturity (Years)', 'Zero Rate (%)'")


def interpolate_zero(curve: pd.DataFrame, t: float) -> float:
    xs = curve["MaturityY"].to_numpy().astype(float)
    ys = curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
    if t <= xs[0]:
        return float(ys[0])
    if t >= xs[-1]:
        return float(ys[-1])
    dfs = np.exp(-ys * xs)
    df_t = np.exp(np.interp(t, xs, np.log(dfs)))
    return -math.log(df_t) / t


def df_from_curve(curve: pd.DataFrame, t: float) -> float:
    z = interpolate_zero(curve, t)
    return math.exp(-z * t)


def _next_bd(d: "date") -> "date":
    """Next business day (Mon-Fri, no holiday calendar)."""
    from datetime import date as _date, timedelta as _td
    while d.weekday() >= 5:
        d += _td(days=1)
    return d

def _add_months(d: "date", months: int) -> "date":
    """Add whole months, clamping to end-of-month."""
    import calendar as _cal
    from datetime import date as _date
    m = d.month - 1 + months
    y = d.year + m // 12
    m = m % 12 + 1
    return _date(y, m, min(d.day, _cal.monthrange(y, m)[1]))

def _add_years(d: "date", years: int) -> "date":
    from datetime import date as _date
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return _date(d.year + years, d.month, 28)

def _mod_fol(d: "date") -> "date":
    """Modified following: if adjusted date falls in next month, go backward."""
    from datetime import timedelta as _td
    nd = _next_bd(d)
    if nd.month != d.month:
        pd_ = d
        while pd_.weekday() >= 5:
            pd_ -= _td(days=1)
        return pd_
    return nd

def _act365(d1: "date", d2: "date") -> float:
    return (d2 - d1).days / 365.0

def _pricing_date() -> "date":
    """Today's date for schedule generation."""
    from datetime import date as _date
    import streamlit as _st
    try:
        d = _st.session_state.get("pricing_date")
        if d: return d
    except Exception:
        pass
    return _date.today()

def _spot_date(spot_lag_bd: int) -> "date":
    """Spot date = today + spot_lag business days."""
    from datetime import timedelta as _td
    d = _pricing_date()
    count = 0
    while count < spot_lag_bd:
        d += _td(days=1)
        if d.weekday() < 5:
            count += 1
    return d

def _fwd_start_date(expiry_years: float, spot_lag_bd: int) -> "date":
    """Forward start date: spot + expiry (mod-fol). Uses days for <1m, months otherwise."""
    from datetime import timedelta as _td
    spot = _spot_date(spot_lag_bd)
    total_days = expiry_years * 365.25
    total_months = int(round(expiry_years * 12))
    if total_days < 27:
        # Sub-monthly: add whole days then mod-fol
        raw = spot + _td(days=int(round(total_days)))
    else:
        raw = _add_months(spot, total_months)
    return _mod_fol(raw)

def _build_date_schedule(fwd_start: "date", tenor_years: float, months_per_period: int) -> List[Tuple[float, float]]:
    """
    Build actual payment schedule using mod-fol date arithmetic.
    Returns list of (time_in_years_from_today, act365_accrual).
    Final cashflow uses total months (not rounded years) to handle 18m, 1.5Y etc correctly.
    """
    today = _pricing_date()
    total_months = int(round(tenor_years * 12))
    n = int(round(tenor_years * (12 / months_per_period)))
    schedule = []
    prev = fwd_start
    for i in range(1, n + 1):
        # Always use months arithmetic - last cashflow uses total_months exactly
        raw = _add_months(fwd_start, i * months_per_period if i < n else total_months)
        pay = _mod_fol(raw)
        accrual = _act365(prev, pay)
        t_years = _act365(today, pay)
        schedule.append((t_years, accrual))
        prev = pay
    return schedule

def build_aud_schedule(expiry: float, tenor: float) -> List[Tuple[float, float]]:
    """AUD: T+1BD spot, mod-fol, Act/365. Q/Q (3m) for ≤3Y, S/S (6m) for >3Y."""
    months_per = 3 if tenor <= 3.0 else 6
    fwd_start = _fwd_start_date(expiry, spot_lag_bd=1)
    return _build_date_schedule(fwd_start, tenor, months_per)


def build_generic_schedule(expiry: float, tenor: float, freq: float = 0.5, spot_lag: float = 1.0) -> List[Tuple[float, float]]:
    """T+2BD spot (NZD/USD), mod-fol, Act/365. freq: 0.25=Q/Q, 0.5=S/S."""
    months_per = int(round(freq * 12))
    fwd_start = _fwd_start_date(expiry, spot_lag_bd=int(round(spot_lag)))
    return _build_date_schedule(fwd_start, tenor, months_per)


def forward_and_annuity_from_curve(curve: pd.DataFrame,
                                   ccy: str,
                                   expiry: float,
                                   tenor: float,
                                   ois_curve: Optional[pd.DataFrame] = None,
                                   freq_override: Optional[float] = None) -> Tuple[float, float, List[Tuple[float, float]]]:
    """
    Calculate forward swap rate and annuity.
    Uses LINEAR zero-rate interpolation (same as fast_forward_rate / matrix) so
    the swaption pricer forward exactly matches the Rate/Vol Matrix.
    freq_override: 0.25 = Q/Q, 0.5 = S/S, None = market convention
    """
    # Fast session-level cache keyed by curve commit ID (no hashing)
    try:
        _cid = st.session_state.get("_curve_commit_ids", {}).get(ccy, 0)
        _oid = st.session_state.get("_curve_commit_ids", {}).get(f"{ccy}_ois", 0) if ois_curve is not None else -1
        _ck = (ccy, _cid, _oid, round(expiry, 6), round(tenor, 6), round(freq_override or -1, 6))
        _fc = st.session_state.setdefault("_fwd_ann_cache", {})
        if _ck in _fc:
            return _fc[_ck]
    except Exception:
        _ck = None; _fc = {}

    if freq_override is not None:
        # T+2 BD for NZD/USD, T+1 BD for AUD (AFMA calendar   —   year frac approx here)
        spot_lag = 2.0 / 252.0 if ccy in ["NZD", "USD"] else 1.0 / 252.0
        sched = build_generic_schedule(expiry, tenor, freq=freq_override, spot_lag=spot_lag * 252)
    elif ccy == "AUD":
        sched = build_aud_schedule(expiry, tenor)
    elif ccy == "NZD":
        freq_nzd = 0.25 if tenor <= 2.0 else 0.5
        sched = build_generic_schedule(expiry, tenor, freq=freq_nzd, spot_lag=2.0)
    elif ccy == "USD":
        sched = build_generic_schedule(expiry, tenor, freq=0.5, spot_lag=2.0)
    else:
        sched = build_generic_schedule(expiry, tenor, freq=0.5, spot_lag=1.0)

    if not sched:
        return 0.0, 0.0, []

    # IRS for projection, OIS for annuity discounting (dual-curve)
    # AUD: use 3M BBSW curve for Q/Q, 6M BBSW curve for S/S
    basis_6v3 = get_basis_curve(ccy, "6v3") if ccy == "AUD" else None

    def _df_proj(crv: pd.DataFrame, t: float, freq: float) -> float:
        """Projection discount factor with convention-aware basis adjustment for AUD."""
        xs = crv["MaturityY"].to_numpy().astype(float)
        ys = crv["ZeroRatePct"].to_numpy().astype(float) / 100.0
        z = float(np.interp(t, xs, ys))
        if ccy == "AUD" and basis_6v3 is not None and not basis_6v3.empty:
            try:
                # Handle different column name formats
                _b6_cols = basis_6v3.columns.tolist()
                _mat_col = next((c for c in _b6_cols if "matur" in c.lower() or "tenor" in c.lower() or c in ("MaturityY","Tenor","tenor_years")), _b6_cols[0])
                _bp_col  = next((c for c in _b6_cols if "basis" in c.lower() or "bp" in c.lower() or "spread" in c.lower()), _b6_cols[1])
                bx = basis_6v3[_mat_col].to_numpy().astype(float)
                by = basis_6v3[_bp_col].to_numpy().astype(float) / 10000.0
                if freq == 0.25 and t > 3.0:
                    z = z - float(np.interp(t, bx, by))
                elif freq == 0.5 and t <= 3.0:
                    z = z + float(np.interp(t, bx, by))
            except Exception:
                pass
        return math.exp(-z * t)

    def _df_disc(crv: pd.DataFrame, t: float) -> float:
        xs = crv["MaturityY"].to_numpy().astype(float)
        ys = crv["ZeroRatePct"].to_numpy().astype(float) / 100.0
        z = float(np.interp(t, xs, ys))
        return math.exp(-z * t)

    disc_curve = ois_curve if ois_curve is not None else curve

    # Determine effective frequency from schedule (periods per year → years per period)
    _n_periods = len(sched)
    _total_time = sched[-1][0] - (sched[0][0] - sched[0][1]) if sched else 0.5
    _sched_freq = round(_total_time / _n_periods * 4) / 4 if _n_periods > 0 else 0.5  # round to nearest 0.25

    ann = 0.0
    for T_i, accrual in sched:
        ann += _df_disc(disc_curve, T_i) * accrual

    swap_start = sched[0][0] - sched[0][1]
    df_start = _df_proj(curve, swap_start, _sched_freq)
    df_end   = _df_proj(curve, sched[-1][0], _sched_freq)
    fwd = (df_start - df_end) / ann if ann > 0 else 0.0
    _result = (fwd, ann, sched)
    try:
        if _ck is not None and len(_fc) < 5000:
            _fc[_ck] = _result
    except Exception:
        pass
    return _result


# ============================
# Swaption pricing
# ============================

def black_swaption_vanilla(ticket: SwaptionTicket) -> dict:
    F = ticket.forward
    K = ticket.strike
    sigma = max(ticket.vol, 1e-8)
    T = max(ticket.expiry_years, 1e-8)
    df = ticket.df()
    annuity = ticket.annuity if ticket.annuity > 0 else ticket.swap_tenor_years

    if F <= 0 or K <= 0 or annuity <= 0:
        bpv = df * annuity * ticket.notional * 0.0001
        return {"pv": 0.0, "pv_bp": 0.0, "delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0, "bpv": bpv}

    lnFK = math.log(F / K)
    vol_sqrt_t = sigma * math.sqrt(T)
    d1 = (lnFK + 0.5 * sigma * sigma * T) / vol_sqrt_t
    d2 = d1 - vol_sqrt_t
    N = NormalDist().cdf
    phi = lambda x: math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)

    if ticket.side.lower().startswith("payer"):
        price_rate = F * N(d1) - K * N(d2)
        delta_ratio = N(d1)          # hedge ratio: 0.5 ATM = 50mm hedge per 100mm notional
    else:
        price_rate = K * N(-d2) - F * N(-d1)
        delta_ratio = -N(-d1)

    pv = df * annuity * ticket.notional * price_rate
    bpv = df * annuity * ticket.notional * 0.0001
    delta = delta_ratio * ticket.notional   # notional-equivalent swap hedge
    delta_dv01 = delta_ratio * bpv          # dollar DV01
    pv_bp_spot = pv / (ticket.notional * 0.0001) if ticket.notional > 0 else 0.0
    pv_bp_fwd  = pv_bp_spot / df if df > 0 else pv_bp_spot
    pv_bp = pv_bp_fwd
    vega  = df * annuity * ticket.notional * F * phi(d1) * math.sqrt(T) * 0.0001
    gamma = df * annuity * ticket.notional * phi(d1) / (F * sigma * math.sqrt(T)) * 0.0001
    theta = -0.5 * df * annuity * ticket.notional * F * sigma * phi(d1) / 365.0

    return {"pv": pv, "pv_bp": pv_bp, "pv_bp_spot": pv_bp_spot, "pv_bp_fwd": pv_bp_fwd,
            "delta": delta, "delta_dv01": delta_dv01, "delta_ratio": delta_ratio,
            "gamma": gamma, "vega": vega, "theta": theta, "bpv": bpv}


def bachelier_swaption_vanilla(ticket: SwaptionTicket) -> dict:
    F = ticket.forward
    K = ticket.strike
    sigma_n = max(ticket.vol, 1e-8)
    T = max(ticket.expiry_years, 1e-8)
    df = ticket.df()
    annuity = ticket.annuity if ticket.annuity > 0 else ticket.swap_tenor_years

    if annuity <= 0:
        bpv = df * annuity * ticket.notional * 0.0001
        return {"pv": 0.0, "pv_bp": 0.0, "delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0, "bpv": bpv}

    d = (F - K) / (sigma_n * math.sqrt(T))
    phi = math.exp(-0.5 * d * d) / math.sqrt(2 * math.pi)
    N = NormalDist().cdf(d)

    if ticket.side.lower().startswith("payer"):
        price_rate = (F - K) * N + sigma_n * math.sqrt(T) * phi
        delta_ratio = N
    else:
        price_rate = (K - F) * (1 - N) + sigma_n * math.sqrt(T) * phi
        delta_ratio = -(1 - N)

    pv = df * annuity * ticket.notional * price_rate
    bpv = df * annuity * ticket.notional * 0.0001
    delta = delta_ratio * ticket.notional   # notional-equivalent swap hedge
    delta_dv01 = delta_ratio * bpv          # dollar DV01
    # pv_bp = premium in bp of notional (spot). Fwd = divide by df.
    pv_bp_spot = pv / (ticket.notional * 0.0001) if ticket.notional > 0 else 0.0
    pv_bp_fwd  = pv_bp_spot / df if df > 0 else pv_bp_spot
    pv_bp = pv_bp_fwd  # default to fwd (market convention)
    vega  = df * annuity * ticket.notional * math.sqrt(T) * phi * 0.0001
    gamma = df * annuity * ticket.notional * phi / (sigma_n * math.sqrt(T)) * 0.0001
    theta = -0.5 * df * annuity * ticket.notional * sigma_n * phi / math.sqrt(T) / 365.0

    return {"pv": pv, "pv_bp": pv_bp, "pv_bp_spot": pv_bp_spot, "pv_bp_fwd": pv_bp_fwd,
            "delta": delta, "delta_dv01": delta_dv01, "delta_ratio": delta_ratio,
            "gamma": gamma, "vega": vega, "theta": theta, "bpv": bpv}


def black_swaption_digital(ticket: SwaptionTicket) -> dict:
    F = ticket.forward
    K = ticket.strike
    sigma = max(ticket.vol, 1e-8)
    T = max(ticket.expiry_years, 1e-8)
    df = ticket.df()
    payout_rate = ticket.payout_bp * 0.0001
    annuity = ticket.annuity if ticket.annuity > 0 else ticket.swap_tenor_years

    lnFK = math.log(F / K)
    vol_sqrt_t = sigma * math.sqrt(T)
    d2 = (lnFK - 0.5 * sigma * sigma * T) / vol_sqrt_t
    N = NormalDist().cdf
    phi = lambda x: math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)

    if ticket.side.lower().startswith("payer"):
        prob = N(d2)
        delta_prob = phi(d2) / (F * sigma * math.sqrt(T))
    else:
        prob = N(-d2)
        delta_prob = -phi(d2) / (F * sigma * math.sqrt(T))

    pv = df * annuity * ticket.notional * payout_rate * prob
    delta = df * annuity * ticket.notional * payout_rate * delta_prob
    bpv = df * annuity * ticket.notional * 0.0001
    pv_bp = pv / bpv if bpv != 0 else 0.0

    return {"pv": pv, "pv_bp": pv_bp, "delta": delta, "gamma": 0.0, "vega": 0.0, "theta": 0.0, "bpv": bpv}


def bachelier_swaption_digital(ticket: SwaptionTicket) -> dict:
    F = ticket.forward
    K = ticket.strike
    sigma_n = max(ticket.vol, 1e-8)
    T = max(ticket.expiry_years, 1e-8)
    df = ticket.df()
    payout_rate = ticket.payout_bp * 0.0001
    annuity = ticket.annuity if ticket.annuity > 0 else ticket.swap_tenor_years

    d = (F - K) / (sigma_n * math.sqrt(T))
    phi = math.exp(-0.5 * d * d) / math.sqrt(2 * math.pi)
    N = NormalDist().cdf(d)

    if ticket.side.lower().startswith("payer"):
        prob = N
        delta_prob = phi / (sigma_n * math.sqrt(T))
    else:
        prob = 1 - N
        delta_prob = -phi / (sigma_n * math.sqrt(T))

    pv = df * annuity * ticket.notional * payout_rate * prob
    delta = df * annuity * ticket.notional * payout_rate * delta_prob
    bpv = df * annuity * ticket.notional * 0.0001
    pv_bp = pv / bpv if bpv != 0 else 0.0

    return {"pv": pv, "pv_bp": pv_bp, "delta": delta, "gamma": 0.0, "vega": 0.0, "theta": 0.0, "bpv": bpv}


def price_swaption(ticket: SwaptionTicket) -> dict:
    if ticket.model == "Black":
        if ticket.payoff_type == "vanilla":
            return black_swaption_vanilla(ticket)
        elif ticket.payoff_type == "digital":
            return black_swaption_digital(ticket)
        elif ticket.payoff_type == "straddle":
            t_p = SwaptionTicket(**{**ticket.__dict__, "side": "Payer", "payoff_type": "vanilla"})
            t_r = SwaptionTicket(**{**ticket.__dict__, "side": "Receiver", "payoff_type": "vanilla"})
            p1 = black_swaption_vanilla(t_p)
            p2 = black_swaption_vanilla(t_r)
            return {k: p1.get(k, 0.0) + p2.get(k, 0.0) for k in p1.keys()}
    else:
        if ticket.payoff_type == "vanilla":
            return bachelier_swaption_vanilla(ticket)
        elif ticket.payoff_type == "digital":
            return bachelier_swaption_digital(ticket)
        elif ticket.payoff_type == "straddle":
            t_p = SwaptionTicket(**{**ticket.__dict__, "side": "Payer", "payoff_type": "vanilla"})
            t_r = SwaptionTicket(**{**ticket.__dict__, "side": "Receiver", "payoff_type": "vanilla"})
            p1 = bachelier_swaption_vanilla(t_p)
            p2 = bachelier_swaption_vanilla(t_r)
            return {k: p1.get(k, 0.0) + p2.get(k, 0.0) for k in p1.keys()}
    return bachelier_swaption_vanilla(ticket)


# ============================
# Caps / Floors
# ============================

def black_caplet(notional: float, accrual: float,
                 F: float, K: float, sigma: float, T: float,
                 r: float, is_cap: bool = True) -> dict:
    if T <= 0 or sigma <= 0:
        return {"pv": 0.0, "delta": 0.0, "vega": 0.0, "gamma": 0.0}
    df = math.exp(-r * T)
    if F <= 0 or K <= 0:
        return {"pv": 0.0, "delta": 0.0, "vega": 0.0, "gamma": 0.0}
    lnFK = math.log(F / K)
    vol_sqrt_t = sigma * math.sqrt(T)
    d1 = (lnFK + 0.5 * sigma * sigma * T) / vol_sqrt_t
    d2 = d1 - vol_sqrt_t
    N = NormalDist().cdf
    phi = lambda x: math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)
    if is_cap:
        price_rate = F * N(d1) - K * N(d2)
        delta_rate = N(d1)
    else:
        price_rate = K * N(-d2) - F * N(-d1)
        delta_rate = -N(-d1)
    vega_rate = F * phi(d1) * math.sqrt(T)
    gamma_rate = phi(d1) / (F * sigma * math.sqrt(T))
    pv = notional * accrual * df * price_rate
    delta = notional * accrual * df * delta_rate * 0.0001   # DV01 contribution per caplet
    vega  = notional * accrual * df * vega_rate  * 0.0001
    gamma = notional * accrual * df * gamma_rate * 0.0001
    return {"pv": pv, "delta": delta, "vega": vega, "gamma": gamma}


def bachelier_caplet(notional: float, accrual: float,
                     F: float, K: float, sigma_n: float, T: float,
                     r: float, is_cap: bool = True) -> dict:
    if T <= 0 or sigma_n <= 0:
        return {"pv": 0.0, "delta": 0.0, "vega": 0.0, "gamma": 0.0}
    df = math.exp(-r * T)
    d = (F - K) / (sigma_n * math.sqrt(T))
    phi = math.exp(-0.5 * d * d) / math.sqrt(2 * math.pi)
    N = NormalDist().cdf(d)
    if is_cap:
        price_rate = (F - K) * N + sigma_n * math.sqrt(T) * phi
        delta_rate = N
    else:
        price_rate = (K - F) * (1 - N) + sigma_n * math.sqrt(T) * phi
        delta_rate = N - 1
    pv = notional * accrual * df * price_rate
    delta = notional * accrual * df * delta_rate * 0.0001      # DV01 contribution per caplet
    vega  = notional * accrual * df * math.sqrt(T) * phi       * 0.0001
    gamma = notional * accrual * df * phi / (sigma_n * math.sqrt(T)) * 0.0001
    return {"pv": pv, "delta": delta, "vega": vega, "gamma": gamma}


# ============================
# XVA helpers
# ============================

def simple_exposure_profile(pv0: float,
                            vega: float,
                            vol: float,
                            horizon_years: float,
                            n_steps: int = 20) -> Tuple[np.ndarray, np.ndarray]:
    ts = np.linspace(0.0, horizon_years, n_steps)
    ee = pv0 + vega * vol * np.sqrt(np.maximum(ts, 1e-8))
    return ts, np.maximum(ee, 0.0)


def cva_from_hazard(ee_times: np.ndarray,
                    ee_values: np.ndarray,
                    hazard_rate: float,
                    lgd: float,
                    discount_rate: float) -> float:
    h = max(hazard_rate, 0.0)
    ts = ee_times
    cvs = 0.0
    prev_pd = 0.0
    for t, ee in zip(ts, ee_values):
        pd_t = 1.0 - math.exp(-h * t)
        dPD = max(pd_t - prev_pd, 0.0)
        df = math.exp(-discount_rate * t)
        cvs += ee * dPD * df
        prev_pd = pd_t
    return -lgd * cvs


# ============================
# Caplet Vol Term Structure Builder
# Advanced Bootstrap Solver with Monotone Cubic Interpolation
# ============================

def price_caplets_with_vol_curve(ccy, tenor_y, caplet_vol_dict, notional_mm=1.0, expiry_y=0.0):
    """
    Price caplets using vol curve. Used by both bootstrap solver and pricer.
    expiry_y: forward start (0 = spot cap). Cap runs from expiry_y to expiry_y+tenor_y.
    The first fixing at expiry_y is known (forward rate) and skipped.
    Returns premium in bp per leg (not straddle).
    """
    _cc = st.session_state.get("config_curves", {}).get(ccy)
    curve = _cc if _cc is not None else get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    if ois_curve is None:
        ois_curve = curve
    
    # Build quarterly schedule from expiry_y to expiry_y + tenor_y
    cap_start = expiry_y if expiry_y > 0 else 1.0 / 252.0
    cap_end   = cap_start + tenor_y
    
    sched = []
    t = cap_start
    while t < cap_end - 1e-8:
        t_next = min(t + 0.25, cap_end)
        accrual = t_next - t
        sched.append((t_next, accrual))
        t = t_next
    
    total_pv = 0.0
    
    for i, (T_fix, accrual) in enumerate(sched):
        if i == 0:
            continue  # Skip first fixing
        
        # Get vol from curve - use exact T_fix, don't round
        if T_fix in caplet_vol_dict:
            vol_bp = caplet_vol_dict[T_fix]
        else:
            # Linear interpolation (same logic as get_caplet_vol_for_fixing)
            mats = sorted(caplet_vol_dict.keys())
            if T_fix < mats[0]:
                vol_bp = caplet_vol_dict[mats[0]]
            elif T_fix > mats[-1]:
                vol_bp = caplet_vol_dict[mats[-1]]
            else:
                for j in range(len(mats)-1):
                    if mats[j] <= T_fix <= mats[j+1]:
                        alpha = (T_fix - mats[j]) / (mats[j+1] - mats[j])
                        vol_bp = caplet_vol_dict[mats[j]] + alpha * (caplet_vol_dict[mats[j+1]] - caplet_vol_dict[mats[j]])
                        break
        
        sigma = vol_bp / 10000.0
        
        # Individual caplet forward
        period_start = max(T_fix - 0.25, 0.001)
        period_tenor = 0.25
        F_caplet, _, _ = forward_and_annuity_from_curve(curve, ccy, period_start, period_tenor, ois_curve)
        
        disc_rate = interpolate_zero(ois_curve, T_fix)
        
        # Price with F=K (ATM)
        res = bachelier_caplet(notional_mm * 1e6, accrual, F_caplet, F_caplet, sigma, T_fix, disc_rate, is_cap=True)
        total_pv += res["pv"]
    
    # Return premium in bp
    premium_bp = (total_pv / (notional_mm * 1e6)) * 10000.0
    return premium_bp


def build_caplet_vol_curve_from_surface(ccy: str, atm_surface):
    """
    Build caplet vol curve using SWAPTION SURFACE ONLY - no user spreads.
    CFS targets are derived from cumulative swaption premiums:
      - 1Y CFS target = 3m1y swaption premium (spot cap to 1Y)
      - 2Y CFS target = 1Y CFS + 1y1y swaption premium (wedge)
      - etc.
    This is the pure surface-implied caplet vol term structure.
    """
    import scipy.optimize as opt
    if atm_surface is None or atm_surface.empty:
        return None

    _cc = st.session_state.get("config_curves", {}).get(ccy)
    curve = _cc if _cc is not None else get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    if curve is None:
        return None
    if ois_curve is None:
        ois_curve = curve

    def get_swpt_prem(expiry_label, tenor_y):
        try:
            vol_bp = get_matrix_value(atm_surface, expiry_label, tenor_y)
            if vol_bp is None: return None
            exp_y = label_to_years(expiry_label)
            _, ann, _ = forward_and_annuity_from_curve(curve, ccy, exp_y, tenor_y, None)
            sigma_n = vol_bp / 10000.0
            sqrt_t = math.sqrt(max(exp_y, 0.001))
            return 2 * 0.3989 * sigma_n * sqrt_t * ann * 10000
        except:
            return None

    # Build cumulative CFS targets from swaption surface alone
    # Each CFS = sum of all swaption wedge premiums up to that maturity
    wedge_chain = [
        ("3m1y", "3m", 1.0, 1.0),    # 3m1y swaption → 1Y CFS
        ("1y1y", "1y", 1.0, 2.0),    # + 1y1y wedge  → 2Y CFS
        ("2y1y", "2y", 1.0, 3.0),    # + 2y1y wedge  → 3Y CFS
        ("3y1y", "3y", 1.0, 4.0),
        ("4y1y", "4y", 1.0, 5.0),
        ("5y2y", "5y", 2.0, 7.0),
        ("7y3y", "7y", 3.0, 10.0),
        ("10y2y","10y",2.0, 12.0),
        ("12y3y","12y",3.0, 15.0),
    ]

    cumulative_leg_prems = {}
    cumulative_prem = 0.0
    for lbl, exp, tenor, result_mat in wedge_chain:
        p = get_swpt_prem(exp, tenor)
        if p is None:
            break
        cumulative_prem += p / 2.0   # leg premium (straddle/2)
        cumulative_leg_prems[result_mat] = cumulative_prem

    if not cumulative_leg_prems:
        return None

    # Now bootstrap exactly as WEDGES mode, using these surface-derived targets
    caplet_vols = {}

    def price_caplets_flat_vol_inner(vol_bp, final_maturity_y):
        flat_curve = {round(t, 2): vol_bp for t in np.arange(0.25, final_maturity_y + 0.26, 0.25)}
        return price_caplets_with_vol_curve(ccy, final_maturity_y, flat_curve, notional_mm=1.0)

    # Step 1: solve 1Y flat vol
    if 1.0 in cumulative_leg_prems:
        try:
            vol_1y = opt.brentq(
                lambda v: price_caplets_flat_vol_inner(v, 1.0) - cumulative_leg_prems[1.0],
                1.0, 200.0, xtol=0.001
            )
            for t in [0.25, 0.5, 0.75, 1.0]:
                caplet_vols[t] = max(vol_1y, 1.0)
        except:
            for t in [0.25, 0.5, 0.75, 1.0]:
                caplet_vols[t] = max(cumulative_leg_prems[1.0] * 1.58, 1.0)

    # Step 2: initial guesses for anchors
    anchor_mats = sorted([m for m in cumulative_leg_prems if m > 1.0])
    for m in anchor_mats:
        caplet_vols[m] = caplet_vols.get(1.0, 75.0)

    # Step 3: simultaneous solve
    def price_with_interp(anchor_vols_array):
        from scipy.interpolate import CubicSpline
        temp_vols = dict(caplet_vols)
        for i, m in enumerate(anchor_mats):
            temp_vols[m] = max(anchor_vols_array[i], 1.0)
        all_mats = np.array(sorted([m for m in temp_vols if m >= 1.0 and m == int(m)]))
        all_vs   = np.array([temp_vols[m] for m in all_mats])
        cs = CubicSpline(all_mats, all_vs)
        interp = {}
        for t in [0.25, 0.5, 0.75, 1.0]:
            interp[t] = temp_vols.get(t, temp_vols.get(1.0, 75.0))
        t = 1.25
        while t <= all_mats[-1] + 1e-6:
            interp[round(t, 2)] = max(float(cs(t)), 1.0)
            t += 0.25
        errors = []
        for m in anchor_mats:
            errors.append(price_caplets_with_vol_curve(ccy, m, interp, notional_mm=1.0) - cumulative_leg_prems[m])
        return np.array(errors)

    if anchor_mats:
        from scipy.optimize import least_squares
        try:
            res = least_squares(price_with_interp, [caplet_vols[m] for m in anchor_mats],
                                ftol=1e-4, xtol=1e-4, gtol=1e-4, max_nfev=200)
            if res.success:
                for i, m in enumerate(anchor_mats):
                    caplet_vols[m] = max(res.x[i], 1.0)
        except:
            pass

    # Final interpolation
    from scipy.interpolate import CubicSpline
    anch = np.array(sorted([m for m in caplet_vols if m >= 1.0 and m == int(m)]))
    if len(anch) < 2:
        return caplet_vols
    cs = CubicSpline(anch, [caplet_vols[m] for m in anch])
    final = {}
    for t in [0.25, 0.5, 0.75, 1.0]:
        final[t] = caplet_vols.get(t, caplet_vols.get(1.0, 75.0))
    t = 1.25
    while t <= anch[-1] + 1e-6:
        final[round(t, 2)] = max(float(cs(t)), 1.0)
        t += 0.25
    return final


def build_caplet_vol_curve(ccy: str, atm_surface, sabr_params=None, 
                          spread_3m1y=-3.0, spread_1y1y=12.0, spread_2y1y=15.0, 
                          spread_3y1y=19.0, spread_4y1y=22.0, spread_5y2y=40.0, spread_7y3y=60.0,
                          spread_10y2y=50.0, spread_12y3y=70.0):
    """
    Build caplet vol curve using cumulative premium method with proper solving.
    """
    if atm_surface is None or atm_surface.empty:
        return None
    
    import scipy.optimize as opt
    
    caplet_vols = {}
    cumulative_leg_prems = {}
    
    def get_swaption_premium(expiry_label, tenor_y):
        """Get swaption PREMIUM - calculate directly from vol surface"""
        try:
            # Get curve and vol
            _cc = st.session_state.get("config_curves", {}).get(ccy)
            curve = _cc if _cc is not None else get_ccy_curve(ccy)
            _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
            ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
            if atm_surface is None or curve is None:
                return None
            
            vol_bp = get_matrix_value(atm_surface, expiry_label, tenor_y)
            if vol_bp is None:
                return None
            
            # Calculate premium using SAME formula as table
            exp_y = label_to_years(expiry_label)
            _, ann, _ = forward_and_annuity_from_curve(curve, ccy, exp_y, tenor_y, None)
            
            sigma_n = vol_bp / 10000.0
            sqrt_t = math.sqrt(max(exp_y, 0.001))
            fwd_premium = 2 * 0.3989 * sigma_n * sqrt_t * ann
            premium_bp = fwd_premium * 10000
            
            return premium_bp  # Full precision, no rounding
        except:
            return None
    
    def price_caplets_flat_vol(vol_bp, final_maturity_y):
        """
        Price caplets using flat vol.
        Returns total LEG premium in bp.
        """
        # Create flat vol curve
        flat_curve = {}
        t = 0.25
        while t <= final_maturity_y + 1e-6:
            flat_curve[round(t, 2)] = vol_bp
            t += 0.25
        
        # Use shared pricing function
        return price_caplets_with_vol_curve(ccy, final_maturity_y, flat_curve, notional_mm=1.0)
    
    # === STEP 1: 1Y CFS ===
    # 1Y CFS = 3m start to 1Y maturity = 9 months = 3 quarterly fixings
    # Get EXACT CFS straddle from table OR calculate directly
    table_data_1y = st.session_state.get("cfs_table_data", {}).get("3m1y", {})
    cfs_1y_straddle = table_data_1y.get("cfs_straddle", None)
    
    if cfs_1y_straddle is None or cfs_1y_straddle <= 0:
        # Calculate directly: swaption premium + spread
        swaption_1y_straddle = get_swaption_premium("3m", 1.0)
        if swaption_1y_straddle is not None:
            cfs_1y_straddle = swaption_1y_straddle + spread_3m1y
    
    if cfs_1y_straddle and cfs_1y_straddle > 0:
        cfs_1y_leg = cfs_1y_straddle / 2.0
        cumulative_leg_prems[1.0] = cfs_1y_leg
        
        # Solve for FLAT vol to 1Y
        def objective_1y(vol_bp):
            return price_caplets_flat_vol(vol_bp, 1.0) - cfs_1y_leg
        
        try:
            vol_1y = opt.brentq(objective_1y, 1.0, 200.0, xtol=0.001)
            # Store flat vol at quarterly points for final interpolation
            for t in [0.25, 0.5, 0.75, 1.0]:
                caplet_vols[t] = max(vol_1y, 1.0)
        except:
            vol_fallback = max(cfs_1y_leg * 1.58, 1.0)
            for t in [0.25, 0.5, 0.75, 1.0]:
                caplet_vols[t] = vol_fallback
    
    # === STEP 2: BOOTSTRAP EACH 1Y GAP SEPARATELY ===
    # Helper: price ONLY caplets in a specific gap
    def price_gap_caplets(vol_bp, gap_start_y, gap_end_y):
        """
        Price ONLY the caplets from gap_start_y to gap_end_y using flat vol.
        Returns premium in bp contributed by this gap.
        """
        _cc = st.session_state.get("config_curves", {}).get(ccy)
        curve = _cc if _cc is not None else get_ccy_curve(ccy)
        _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
        ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
        if ois_curve is None:
            ois_curve = curve
        
        # Build schedule for the GAP only
        swap_start = 0.0 + 1.0 / 252.0
        gap_start_abs = swap_start + gap_start_y
        gap_end_abs = swap_start + gap_end_y
        
        sched = []
        t = gap_start_abs
        while t < gap_end_abs - 1e-8:
            t_next = min(t + 0.25, gap_end_abs)
            accrual = t_next - t
            # Store relative time from today for T_fix
            T_fix = t_next
            sched.append((T_fix, accrual))
            t = t_next
        
        sigma = vol_bp / 10000.0
        total_prem_dollars = 0.0
        
        for T_fix, accrual in sched:
            df = df_from_curve(ois_curve, T_fix)
            phi_zero = 1.0 / math.sqrt(2.0 * math.pi)
            price_rate = sigma * math.sqrt(T_fix) * phi_zero
            pv_dollars = 1e6 * accrual * df * price_rate
            total_prem_dollars += pv_dollars
        
        return (total_prem_dollars / 1e6) * 10000.0
    
    wedges = [
        ("1y1y", "1y", 1.0, spread_1y1y, 2.0, 1.0),
        ("2y1y", "2y", 1.0, spread_2y1y, 3.0, 2.0),
        ("3y1y", "3y", 1.0, spread_3y1y, 4.0, 3.0),
        ("4y1y", "4y", 1.0, spread_4y1y, 5.0, 4.0),
        ("5y2y", "5y", 2.0, spread_5y2y, 7.0, 5.0),
        ("7y3y", "7y", 3.0, spread_7y3y, 10.0, 7.0),
        ("10y2y", "10y", 2.0, spread_10y2y, 12.0, 10.0),
        ("12y3y", "12y", 3.0, spread_12y3y, 15.0, 12.0),
    ]
    
    for table_label, expiry, tenor, spread, result_mat, prior_mat in wedges:
        if prior_mat not in cumulative_leg_prems:
            continue
        
        wedge_data = st.session_state.get("cfs_table_data", {}).get(table_label, {})
        wedge_straddle = wedge_data.get("cfs_straddle", None)
        
        if wedge_straddle is None or wedge_straddle <= 0:
            wedge_swaption_straddle = get_swaption_premium(expiry, tenor)
            if wedge_swaption_straddle is not None:
                wedge_straddle = wedge_swaption_straddle + spread
        
        if wedge_straddle is None or wedge_straddle <= 0:
            continue
        
        wedge_leg = wedge_straddle / 2.0
        
        # This wedge_leg is the INCREMENTAL premium for the gap (prior_mat to result_mat)
        gap_premium_target = wedge_leg
        
        # Bootstrap: find flat vol for THIS GAP ONLY
        def objective(vol_bp):
            return price_gap_caplets(vol_bp, prior_mat, result_mat) - gap_premium_target
        
        wedge_leg = wedge_straddle / 2.0
        cumulative_leg_prem = cumulative_leg_prems[prior_mat] + wedge_leg
        cumulative_leg_prems[result_mat] = cumulative_leg_prem
        
        # Store initial guess for anchor vol (will be refined later)
        gap_premium = wedge_leg
        initial_vol_guess = max(gap_premium * 1.5, 50.0)
        caplet_vols[result_mat] = initial_vol_guess
    
    # === STEP 3: SOLVE FOR ALL ANCHOR VOLS SIMULTANEOUSLY ===
    # Must solve all at once because cubic spline shape depends on ALL anchors
    
    def price_with_interp_curve(anchor_vols_array):
        """
        Given vols at anchor points, cubic spline interpolate and price all maturities.
        Returns array of pricing errors vs targets.
        """
        # Build vol dict from array
        temp_vols = dict(caplet_vols)  # Start with 1Y vols
        anchor_mats_to_solve = sorted([m for m in cumulative_leg_prems.keys() if m > 1.0])
        
        for i, mat in enumerate(anchor_mats_to_solve):
            temp_vols[mat] = max(anchor_vols_array[i], 1.0)
        
        # Cubic spline - only use integer anchors
        from scipy.interpolate import CubicSpline
        all_anchor_mats = np.array(sorted([m for m in temp_vols.keys() if m >= 1.0 and m == int(m)]))
        all_anchor_vols = np.array([temp_vols[m] for m in all_anchor_mats])
        
        cs = CubicSpline(all_anchor_mats, all_anchor_vols)
        
        # Build interpolated curve
        interp_curve = {}
        for t in [0.25, 0.5, 0.75, 1.0]:
            interp_curve[t] = temp_vols.get(t, temp_vols.get(1.0, 75.0))
        
        t = 1.25
        max_mat = all_anchor_mats[-1]
        while t <= max_mat + 1e-6:
            interp_curve[round(t, 2)] = max(float(cs(t)), 1.0)
            t += 0.25
        
        # Price each maturity using SHARED pricing function
        errors = []
        for check_mat in anchor_mats_to_solve:
            target_prem = cumulative_leg_prems[check_mat]
            actual_prem = price_caplets_with_vol_curve(ccy, check_mat, interp_curve, notional_mm=1.0)
            errors.append(actual_prem - target_prem)
        
        return np.array(errors)
    
    # Solve for all anchor vols simultaneously
    anchor_mats_to_solve = sorted([m for m in cumulative_leg_prems.keys() if m > 1.0])
    
    if len(anchor_mats_to_solve) > 0:
        initial_guess = np.array([caplet_vols[m] for m in anchor_mats_to_solve])
        
        from scipy.optimize import least_squares
        try:
            result = least_squares(price_with_interp_curve, initial_guess,
                                   ftol=1e-4, xtol=1e-4, gtol=1e-4,
                                   max_nfev=200)
            
            if result.success:
                for i, mat in enumerate(anchor_mats_to_solve):
                    caplet_vols[mat] = max(result.x[i], 1.0)
        except:
            pass
    
    # Final cubic spline interpolation with solved anchors
    from scipy.interpolate import CubicSpline
    anchor_mats = np.array(sorted([m for m in caplet_vols.keys() if m >= 1.0 and m == int(m)]))
    anchor_vols = np.array([caplet_vols[m] for m in anchor_mats])
    
    if len(anchor_mats) >= 2:
        cs = CubicSpline(anchor_mats, anchor_vols)
        
        caplet_vols_final = {}
        for t in [0.25, 0.5, 0.75, 1.0]:
            caplet_vols_final[t] = caplet_vols.get(t, caplet_vols.get(1.0, 75.0))
        
        t = 1.25
        max_mat = anchor_mats[-1]
        while t <= max_mat + 1e-6:
            caplet_vols_final[round(t, 2)] = max(float(cs(t)), 1.0)
            t += 0.25  # Caps are always quarterly
        
        return caplet_vols_final
    
    return caplet_vols if caplet_vols else None



def get_caplet_vol_for_fixing(caplet_curve, fixing_time_y):
    """
    Get caplet vol for a specific fixing time from the term structure.
    Uses linear interpolation if needed.
    """
    if caplet_curve is None or len(caplet_curve) == 0:
        return None
    
    # Exact match
    if fixing_time_y in caplet_curve:
        return caplet_curve[fixing_time_y]
    
    # Interpolate
    maturities = sorted(caplet_curve.keys())
    
    # Extrapolate flat before first point
    if fixing_time_y < maturities[0]:
        return caplet_curve[maturities[0]]
    
    # Extrapolate flat after last point
    if fixing_time_y > maturities[-1]:
        return caplet_curve[maturities[-1]]
    
    # Linear interpolation
    for i in range(len(maturities) - 1):
        t1, t2 = maturities[i], maturities[i + 1]
        if t1 <= fixing_time_y <= t2:
            v1, v2 = caplet_curve[t1], caplet_curve[t2]
            alpha = (fixing_time_y - t1) / (t2 - t1)
            return v1 + alpha * (v2 - v1)
    
    return caplet_curve[maturities[-1]]


# ============================
# Theme
# ============================

def apply_rateedge_theme(theme_name: str):
    """
    RateEdge brand themes using official colors from brand guide.
    Navy: #1e3a5f, Blue: #2563eb, Red: #dc2626, Slate: #0f172a
    """
    if theme_name == "Clean Light":
        bg = "#f1f5f9"       # gray-100
        card = "#ffffff"
        border = "#e2e8f0"   # gray-200
        text = "#1e3a5f"     # navy
        accent = "#dc2626"   # red (brand)
        accent2 = "#2563eb"  # blue (brand)
        muted = "#64748b"    # gray-500
        tab_text = "#1e3a5f"
        sidebar_arrow = "#1e3a5f"
    else:  # Dealer Dark (default)
        bg = "#0f172a"       # slate dark
        card = "#1e293b"
        border = "#334155"   # gray-700
        text = "#f1f5f9"     # gray-100
        accent = "#ef4444"   # red-light
        accent2 = "#3b82f6"  # blue-light
        muted = "#94a3b8"    # gray-400
        tab_text = "#f1f5f9"
        sidebar_arrow = "#f1f5f9"

    st.markdown(
        f"""
        <style>
        .stApp {{
            background-color: {bg};
            color: {text};
        }}
        [data-testid="stHeader"] {{
            background: transparent;
        }}
        /* Sidebar */
        [data-testid="stSidebar"] {{
            background-color: {card};
            border-right: 1px solid {border};
        }}
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stMarkdown {{
            color: {text} !important;
        }}
        /* Sidebar collapse/expand button - make arrows visible */
        [data-testid="stSidebar"] button[kind="header"],
        [data-testid="collapsedControl"] {{
            color: {sidebar_arrow} !important;
        }}
        [data-testid="stSidebar"] svg,
        [data-testid="collapsedControl"] svg {{
            fill: {sidebar_arrow} !important;
            stroke: {sidebar_arrow} !important;
        }}
        button[kind="headerNoPadding"] svg {{
            fill: {sidebar_arrow} !important;
            stroke: {sidebar_arrow} !important;
        }}
        /* Collapsed sidebar button */
        .css-1rs6os {{
            color: {sidebar_arrow} !important;
        }}
        .css-1rs6os svg {{
            fill: {sidebar_arrow} !important;
        }}
        /* Alternative selectors for sidebar toggle */
        [data-testid="baseButton-headerNoPadding"] {{
            color: {sidebar_arrow} !important;
        }}
        [data-testid="baseButton-headerNoPadding"] svg {{
            fill: {sidebar_arrow} !important;
            stroke: {sidebar_arrow} !important;
        }}
        /* Tabs - readable text */
        .stTabs [data-baseweb="tab-list"] {{
            gap: 4px;
            background-color: {card};
            padding: 0.5rem;
            border-radius: 8px;
            border: 1px solid {border};
        }}
        .stTabs [data-baseweb="tab"] {{
            color: {tab_text} !important;
            background-color: transparent;
            border-radius: 6px;
            padding: 0.5rem 0.75rem;
            font-weight: 500;
            font-size: 0.85rem;
        }}
        .stTabs [data-baseweb="tab"]:hover {{
            background-color: {border};
        }}
        .stTabs [aria-selected="true"] {{
            background-color: {accent} !important;
            color: white !important;
        }}
        /* Tab indicator bar - hide default and use background instead */
        .stTabs [data-baseweb="tab-highlight"] {{
            background-color: transparent !important;
        }}
        .stTabs [data-baseweb="tab-border"] {{
            background-color: transparent !important;
        }}
        /* Cards */
        .rateedge-card {{
            background-color: {card};
            padding: 1.0rem 1.2rem;
            border-radius: 14px;
            border: 1px solid {border};
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }}
        .rateedge-header {{
            font-size: 1.6rem;
            font-weight: 650;
            color: {text};
            margin-bottom: 0.2rem;
        }}
        .rateedge-sub {{
            font-size: 0.9rem;
            color: {muted};
        }}
        .rateedge-badge {{
            display:inline-block;
            padding:0.15rem 0.5rem;
            font-size:0.72rem;
            border-radius:999px;
            border:1px solid {accent};
            color:{accent};
            margin-left:0.5rem;
        }}
        /* Text */
        label, .stMarkdown, .stText {{
            color: {text} !important;
        }}
        h1, h2, h3, h4 {{
            color: {text} !important;
        }}
        /* Buttons */
        .stButton>button {{
            background-color: {accent};
            color: white;
            border-radius: 8px;
            border: 1px solid {accent};
            font-weight: 600;
        }}
        .stButton>button:hover {{
            background-color: {accent2};
            color: white;
            border-color: {accent2};
        }}
        /* Download buttons - make them visible */
        .stDownloadButton>button {{
            background-color: {accent};
            color: white !important;
            border-radius: 8px;
            border: 1px solid {accent};
            font-weight: 600;
        }}
        .stDownloadButton>button:hover {{
            background-color: {accent2};
            color: white !important;
            border-color: {accent2};
        }}
        /* Inputs */
        .stTextInput input, .stNumberInput input {{
            background-color: {card} !important;
            color: {text} !important;
            border: 1px solid {border} !important;
        }}
        .stSelectbox > div > div {{
            background-color: {card} !important;
            color: {text} !important;
        }}
        /* DataFrames */
        .stDataFrame {{
            color: {text};
        }}
        /* Landing page cards */
        .landing-card {{
            background: {card};
            border: 1px solid {border};
            border-radius: 16px;
            padding: 2rem;
            text-align: center;
        }}
        .feature-card {{
            background: {card};
            border: 1px solid {border};
            border-radius: 12px;
            padding: 1.5rem;
        }}
        /* Status cards */
        .status-card {{
            background: {card};
            border: 1px solid {border};
            border-radius: 10px;
            padding: 1rem;
            margin: 0.5rem 0;
        }}
        .status-ok {{
            color: #22c55e;
        }}
        .status-no {{
            color: #ef4444;
        }}
        .status-time {{
            color: {muted};
            font-size: 0.75rem;
        }}
        /* NUCLEAR FIX: Force all form label text to be visible */
        /* Target the actual rendered text elements in Streamlit */
        .stRadio > label,
        .stRadio label p,
        .stRadio label span,
        .stRadio div[role="radiogroup"] label,
        .stRadio div[role="radiogroup"] label p,
        .stRadio div[role="radiogroup"] label span,
        .stRadio div[data-testid="stMarkdownContainer"] p,
        .stCheckbox > label,
        .stCheckbox label p,
        .stCheckbox label span,
        .stCheckbox div[data-testid="stMarkdownContainer"] p,
        div[data-baseweb="radio"] ~ div,
        div[data-baseweb="radio"] ~ div p,
        div[data-baseweb="checkbox"] ~ div,
        div[data-baseweb="checkbox"] ~ div p {{
            color: {text} !important;
            -webkit-text-fill-color: {text} !important;
        }}
        /* Target the markdown text inside radio/checkbox labels */
        [data-testid="stRadio"] [data-testid="stMarkdownContainer"],
        [data-testid="stRadio"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stCheckbox"] [data-testid="stMarkdownContainer"],
        [data-testid="stCheckbox"] [data-testid="stMarkdownContainer"] p {{
            color: {text} !important;
            -webkit-text-fill-color: {text} !important;
        }}
        /* Disabled state - still readable but dimmed */
        [data-testid="stCheckbox"][aria-disabled="true"] label,
        [data-testid="stCheckbox"][aria-disabled="true"] p,
        [data-testid="stRadio"][aria-disabled="true"] label,
        [data-testid="stRadio"][aria-disabled="true"] p {{
            color: {muted} !important;
            -webkit-text-fill-color: {muted} !important;
            opacity: 0.7;
        }}
        /* Selectbox dropdown text */
        .stSelectbox label,
        .stSelectbox [data-baseweb="select"] span {{
            color: {text} !important;
        }}
        /* Number input labels */
        .stNumberInput label {{
            color: {text} !important;
        }}
        /* Caption text */
        .stCaption, [data-testid="stCaptionContainer"] {{
            color: {muted} !important;
        }}
        /* Metric values - make readable in dark mode */
        [data-testid="stMetricValue"] {{
            color: {text} !important;
        }}
        [data-testid="stMetricLabel"] {{
            color: {text} !important;
        }}
        [data-testid="stMetricDelta"] {{
            color: {text} !important;
        }}
        /* DataFrame text colors */
        .stDataFrame, .stDataFrame td, .stDataFrame th {{
            color: {text} !important;
        }}
        [data-testid="stDataFrame"] {{
            color: {text} !important;
        }}
        [data-testid="stDataFrame"] td {{
            color: {text} !important;
        }}
        [data-testid="stDataFrame"] th {{
            color: {text} !important;
        }}
        /* Table cells in dataframes */
        .dvn-scroller td, .dvn-scroller th {{
            color: {text} !important;
        }}
        /* Expander text */
        .streamlit-expanderHeader {{
            color: {text} !important;
        }}
        [data-testid="stExpander"] summary {{
            color: {text} !important;
        }}
        /* Radio buttons - WHITE circles */
        div[data-testid="stRadio"] label[data-baseweb="radio"] > div:first-child {{
            background-color: #ffffff !important;
            border-color: #ffffff !important;
        }}
        .stRadio > div[role="radiogroup"] label[data-baseweb="radio"] > div:first-child > div {{
            background-color: #ffffff !important;
        }}
        [data-testid="stRadio"] [data-baseweb="radio"] input:checked + div {{
            background-color: #ffffff !important;
            border-color: #ffffff !important;
        }}
        /* NUCLEAR - ALL TEXT INSIDE RADIO/CHECKBOX */
        [data-testid="stRadio"] * {{
            color: #fbbf24 !important;
            -webkit-text-fill-color: #fbbf24 !important;
        }}
        [data-testid="stCheckbox"] * {{
            color: #ffffff !important;
            -webkit-text-fill-color: #ffffff !important;
        }}
        /* Also target by class */
        .stRadio * {{
            color: #fbbf24 !important;
            -webkit-text-fill-color: #fbbf24 !important;
        }}
        .stCheckbox * {{
            color: #ffffff !important;
            -webkit-text-fill-color: #ffffff !important;
        }}
        /* DISABLED checkboxes - still visible but dimmed */
        [data-testid="stCheckbox"][aria-disabled="true"] *,
        .stCheckbox[aria-disabled="true"] *,
        div[data-testid="stCheckbox"] input:disabled ~ label,
        div[data-testid="stCheckbox"] input:disabled ~ label * {{
            color: #94a3b8 !important;
            -webkit-text-fill-color: #94a3b8 !important;
            opacity: 0.7;
        }}
        /* Target the actual label text element directly */
        [data-baseweb="checkbox"] + div,
        [data-baseweb="checkbox"] ~ div,
        [data-baseweb="checkbox"] ~ div p,
        [data-baseweb="checkbox"] ~ div span {{
            color: #ffffff !important;
            -webkit-text-fill-color: #ffffff !important;
        }}
        /* Hide ALL Streamlit chrome — code must not be visible */
        [data-testid="manage-app-button"] {{display: none !important;}}
        [data-testid="stToolbar"] {{display: none !important;}}
        [data-testid="stDecoration"] {{display: none !important;}}
        [data-testid="stStatusWidget"] {{display: none !important;}}
        button[kind="managedApp"] {{display: none !important;}}
        .stAppDeployButton {{display: none !important;}}
        .stDeployButton {{display: none !important;}}
        [title="Manage app"] {{display: none !important;}}
        [title="View app on Streamlit Community Cloud"] {{display: none !important;}}
        a[href*="github.com"] {{display: none !important;}}
        a[href*="streamlit.io"] {{display: none !important;}}
        .viewerBadge_container__r5tak {{display: none !important;}}
        .viewerBadge_link__qRIco {{display: none !important;}}
        header[data-testid="stHeader"] > div:last-child {{display: none !important;}}
        header[data-testid="stHeader"] .stToolbarActions {{display: none !important;}}
        footer {{visibility: hidden !important; display: none !important;}}
        #MainMenu {{visibility: hidden !important; display: none !important;}}
        header {{visibility: hidden !important;}}
        </style>""",
        unsafe_allow_html=True,
    )
    
    # Force radio/checkbox colors with JavaScript (runs after render)
    import streamlit.components.v1 as components
    components.html("""
    <script>
    function fixColors() {
        const p = window.parent.document;
        // Checkboxes - WHITE
        p.querySelectorAll('[data-testid="stCheckbox"] label, [data-testid="stCheckbox"] span, [data-testid="stCheckbox"] p, [data-testid="stCheckbox"] div').forEach(el => {
            el.style.setProperty('color', '#ffffff', 'important');
            el.style.setProperty('-webkit-text-fill-color', '#ffffff', 'important');
        });
        // Radio buttons - YELLOW
        p.querySelectorAll('[data-testid="stRadio"] label, [data-testid="stRadio"] span, [data-testid="stRadio"] p, [data-testid="stRadio"] div').forEach(el => {
            el.style.setProperty('color', '#fbbf24', 'important');
            el.style.setProperty('-webkit-text-fill-color', '#fbbf24', 'important');
        });
        // Baseweb checkbox labels - WHITE
        p.querySelectorAll('[data-baseweb="checkbox"] ~ div').forEach(el => {
            el.style.setProperty('color', '#ffffff', 'important');
            el.style.setProperty('-webkit-text-fill-color', '#ffffff', 'important');
        });
        // Baseweb radio labels - YELLOW
        p.querySelectorAll('[data-baseweb="radio"] ~ div').forEach(el => {
            el.style.setProperty('color', '#fbbf24', 'important');
            el.style.setProperty('-webkit-text-fill-color', '#fbbf24', 'important');
        });
    }
    fixColors();
    setInterval(fixColors, 2000);
    </script>
    """, height=0)


def show_header(ccy: str):
    st.markdown(
        f"""
        <div class="rateedge-header">
           RateEdge Options  {ccy}
           <span class="rateedge-badge">Swaptions  Caps/Floors  Exotics  CVA  RV</span>
        </div>
        <div class="rateedge-sub">
           Premium at expiry  swaps start T+1BD  SABR smiles  CVA / horizon / RV tools.
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================
# Session state helpers
# ============================

_PORTFOLIO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "swaption_portfolio.json")
_CORR_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exotics_config.json")

# Tenors in the correlation matrix
_CORR_TENORS = ["2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "30Y"]

# Default correlation matrix (upper-triangular stored as dict "Xa_Xb")
_CORR_DEFAULTS = {
    "2Y_3Y": 0.988, "2Y_5Y": 0.968, "2Y_7Y": 0.945, "2Y_10Y": 0.918,
    "2Y_15Y": 0.890, "2Y_20Y": 0.872, "2Y_30Y": 0.850,
    "3Y_5Y": 0.982, "3Y_7Y": 0.961, "3Y_10Y": 0.936,
    "3Y_15Y": 0.910, "3Y_20Y": 0.893, "3Y_30Y": 0.872,
    "5Y_7Y": 0.985, "5Y_10Y": 0.965, "5Y_15Y": 0.942,
    "5Y_20Y": 0.926, "5Y_30Y": 0.907,
    "7Y_10Y": 0.983, "7Y_15Y": 0.963, "7Y_20Y": 0.948, "7Y_30Y": 0.931,
    "10Y_15Y": 0.985, "10Y_20Y": 0.973, "10Y_30Y": 0.958,
    "15Y_20Y": 0.993, "15Y_30Y": 0.982,
    "20Y_30Y": 0.994,
}

# Default CMS convexity bumps (bp) keyed by CMS tenor e.g. "5Y"
_CMS_BUMP_DEFAULTS = {t: 0.0 for t in _CORR_TENORS}


def _save_exotics_config():
    """Persist correlation matrix and CMS bumps to disk."""
    try:
        data = {
            "correlations": {k: st.session_state.get(f"corr_{k}", v)
                             for k, v in _CORR_DEFAULTS.items()},
            "cms_bumps": {t: st.session_state.get(f"cms_bump_{t}", 0.0)
                          for t in _CORR_TENORS},
        }
        with open(_CORR_FILE, "w") as _f:
            json.dump(data, _f, indent=2)
    except Exception:
        pass


def _load_exotics_config():
    """Load correlation matrix and CMS bumps from disk."""
    try:
        with open(_CORR_FILE, "r") as _f:
            return json.load(_f)
    except Exception:
        return {}


def get_correlation(t1: str, t2: str) -> float:
    """Return correlation between two swap rate tenors (e.g. '10Y', '30Y')."""
    if t1 == t2:
        return 1.0
    key = f"{t1}_{t2}" if f"{t1}_{t2}" in _CORR_DEFAULTS else f"{t2}_{t1}"
    return float(st.session_state.get(f"corr_{key}", _CORR_DEFAULTS.get(key, 0.90)))


def get_cms_bump(tenor: str) -> float:
    """Return CMS convexity adjustment bump in bp for a given tenor."""
    return float(st.session_state.get(f"cms_bump_{tenor}", 0.0))

def _save_portfolio():
    """Save full portfolio to Supabase user_configs (per user scratchpad)."""
    try:
        _uid = st.session_state.get("username", "")
        _port = st.session_state.get("portfolio", [])
        if _uid and HAS_POSTGRES:
            _conn = get_db_connection()
            if _conn:
                _cur = _conn.cursor()
                _cur.execute("""
                    INSERT INTO user_configs (user_id, config_type, config_data, updated_at)
                    VALUES (%s, 'portfolio', %s, NOW())
                    ON CONFLICT (user_id, config_type)
                    DO UPDATE SET config_data = EXCLUDED.config_data, updated_at = NOW()
                """, (_uid, json.dumps({"portfolio": _port}, default=str)))
                _conn.commit()
                _cur.close()
                _conn.close()
                return
    except Exception:
        pass
    # Fallback to file
    try:
        with open(_PORTFOLIO_FILE, "w") as _f:
            json.dump(st.session_state.get("portfolio", []), _f, default=str)
    except: pass

def _load_portfolio() -> list:
    """Load portfolio from Supabase per user, fallback to file."""
    try:
        _uid = st.session_state.get("username", "")
        if _uid and HAS_POSTGRES:
            _conn = get_db_connection()
            if _conn:
                _cur = _conn.cursor()
                _cur.execute("""
                    SELECT config_data FROM user_configs
                    WHERE user_id = %s AND config_type = 'portfolio'
                """, (_uid,))
                _row = _cur.fetchone()
                _cur.close()
                _conn.close()
                if _row and _row[0]:
                    _data = _row[0] if isinstance(_row[0], dict) else json.loads(_row[0])
                    _port = _data.get("portfolio", [])
                    for entry in _port:
                        for fld in ("pv","pv_bp","delta","gamma","vega","theta","bpv","strike","forward","notional_mm"):
                            if fld in entry:
                                try: entry[fld] = float(entry[fld])
                                except: pass
                    return _port
    except Exception:
        pass
    # Fallback to file
    try:
        with open(_PORTFOLIO_FILE, "r") as _f:
            data = json.load(_f)
        for entry in data:
            for fld in ("pv","pv_bp","delta","gamma","vega","theta","bpv","strike","forward","notional_mm"):
                if fld in entry:
                    try: entry[fld] = float(entry[fld])
                    except: pass
        return data
    except Exception:
        return []

def ccy_eod_utc(ccy: str, date_str: str) -> str:
    """Return ISO UTC timestamp for the canonical EOD close of a currency on a given date.
    date_str: 'YYYY-MM-DD'
    Returns: 'YYYY-MM-DD HH:MM:SS+00:00' (UTC)
    """
    from datetime import datetime
    tz = CCY_TZ.get(ccy, SYDNEY_TZ)
    h, m = CCY_EOD.get(ccy, (16, 30))
    d = datetime.strptime(date_str, "%Y-%m-%d")
    local_dt = d.replace(hour=h, minute=m, second=0, tzinfo=tz)
    utc_dt = local_dt.astimezone(__import__('datetime').timezone.utc)
    return utc_dt.strftime("%Y-%m-%d %H:%M:%S")


def ccy_eod_label(ccy: str, snap_utc) -> str:
    """Format a UTC snapshot_date as local close time string for display."""
    try:
        from datetime import timezone as _dtz
        tz = CCY_TZ.get(ccy, SYDNEY_TZ)
        if hasattr(snap_utc, 'replace'):
            if snap_utc.tzinfo is None:
                snap_utc = snap_utc.replace(tzinfo=_dtz.utc)
            local_dt = snap_utc.astimezone(tz)
            offset_h = local_dt.utcoffset().total_seconds() / 3600
            tz_labels = {
                "AUD": "AEDT" if offset_h == 11 else "AEST",
                "NZD": "NZDT" if offset_h == 13 else "NZST",
                "USD": "EDT"  if offset_h == -4 else "EST",
            }
            tz_lbl = tz_labels.get(ccy, "")
            return local_dt.strftime(f"%d-%b-%Y %H:%M {tz_lbl}")
        return str(snap_utc)[:16]
    except Exception:
        return str(snap_utc)[:16]


def is_admin() -> bool:
    """Admin — full access including SABR calibration, user management, all DB ops."""
    return st.session_state.get("user_role", "read_only") in ("admin", "super_admin")

def is_super_admin() -> bool:
    """Super admin — can change roles, unhide hidden tabs. Will only."""
    return st.session_state.get("user_role", "read_only") == "super_admin"

def is_user() -> bool:
    """User — can edit vol surfaces, upload EOD/intraday snapshots, edit CFS wedges."""
    return st.session_state.get("user_role", "read_only") in ("admin", "user")

def is_view_only() -> bool:
    """View-only — can load IRS/vols, price options, copy Quick Tix. No editing."""
    return st.session_state.get("user_role", "read_only") == "read_only"

def is_read_only() -> bool:
    """Read-only or trainee — cannot edit or save anything."""
    return st.session_state.get("user_role", "read_only") in ("read_only", "trainee")

def is_trainee() -> bool:
    """Trainee — can load IRS/vols and price locally only. No saving or Quick Tix."""
    return st.session_state.get("user_role", "read_only") == "trainee"

def can_upload_vol() -> bool:
    """Can upload/save vol snapshots to DB."""
    return st.session_state.get("user_role", "read_only") in ("super_admin", "admin", "user")

def can_edit_wedges() -> bool:
    """Can edit CFS wedge spreads."""
    return st.session_state.get("user_role", "read_only") in ("super_admin", "admin", "user")

def can_price() -> bool:
    """Can price options."""
    return st.session_state.get("user_role", "read_only") in ("super_admin", "admin", "user", "read_only")

def can_quick_tix() -> bool:
    """Can use Quick Tix copy function."""
    return st.session_state.get("user_role", "read_only") in ("super_admin", "admin", "user", "read_only")

# ── SABR Calibration Reference Data (31-Mar-2026) ──────────────────────────
# Source: Market vol cube calibration. Per-cell rho and nu from smile data.
# Rows = expiry, Cols = tenor. Values interpolated for missing cells.
_SABR_REF = {
    "AUD": {
        "expiries": ["1m","3m","6m","1y","2y","5y","10y","15y","20y","30y"],
        "tenors":   ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","15Y","20Y","30Y"],
        "rho": {
            "1m":  [-0.20,-0.17,-0.14,-0.17,-0.12,-0.06, 0.03, 0.06, 0.06, 0.06],
            "3m":  [-0.09,-0.01, 0.02, 0.04, 0.06, 0.04, 0.07, 0.08, 0.08, 0.07],
            "6m":  [-0.03, 0.03, 0.05, 0.04, 0.06, 0.07, 0.08, 0.09, 0.10, 0.09],
            "1y":  [ 0.03, 0.07, 0.09, 0.10, 0.09, 0.11, 0.11, 0.13, 0.14, 0.14],
            "2y":  [ 0.08, 0.13, 0.15, 0.15, 0.16, 0.18, 0.19, 0.19, 0.20, 0.20],
            "5y":  [ 0.16, 0.16, 0.19, 0.19, 0.20, 0.22, 0.22, 0.22, 0.23, 0.22],
            "10y": [ 0.19, 0.17, 0.18, 0.18, 0.25, 0.27, 0.28, 0.28, 0.29, 0.26],
            "15y": [ 0.22, 0.22, 0.22, 0.22, 0.29, 0.29, 0.33, 0.33, 0.34, 0.29],
            "20y": [ 0.25, 0.25, 0.25, 0.25, 0.35, 0.35, 0.40, 0.40, 0.38, 0.31],
            "30y": [ 0.18, 0.18, 0.18, 0.18, 0.31, 0.31, 0.40, 0.40, 0.40, 0.40],
        },
        "nu": {
            "1m":  [0.50, 0.44, 0.43, 0.47, 0.42, 0.46, 0.40, 0.41, 0.38, 0.39],
            "3m":  [0.48, 0.46, 0.44, 0.45, 0.41, 0.42, 0.36, 0.38, 0.35, 0.37],
            "6m":  [0.44, 0.43, 0.41, 0.41, 0.37, 0.38, 0.34, 0.36, 0.33, 0.36],
            "1y":  [0.42, 0.36, 0.34, 0.34, 0.32, 0.31, 0.28, 0.29, 0.27, 0.28],
            "2y":  [0.38, 0.31, 0.29, 0.29, 0.27, 0.24, 0.23, 0.23, 0.23, 0.23],
            "5y":  [0.31, 0.29, 0.26, 0.26, 0.24, 0.20, 0.19, 0.19, 0.19, 0.18],
            "10y": [0.25, 0.22, 0.20, 0.20, 0.18, 0.16, 0.14, 0.14, 0.14, 0.14],
            "15y": [0.21, 0.21, 0.18, 0.18, 0.15, 0.15, 0.13, 0.13, 0.13, 0.12],
            "20y": [0.19, 0.19, 0.16, 0.16, 0.13, 0.13, 0.12, 0.12, 0.12, 0.11],
            "30y": [0.19, 0.19, 0.16, 0.16, 0.13, 0.13, 0.12, 0.12, 0.12, 0.11],
        },
    },
    "USD": {
        "expiries": ["1m","3m","6m","9m","1y","2y","3y","5y","7y","10y","15y","20y","30y"],
        "tenors":   ["1Y","2Y","5Y","10Y","15Y","20Y","30Y"],
        "rho": {
            "1m":  [-0.211,-0.015, 0.164, 0.408, 0.261, 0.142, 0.125],
            "3m":  [-0.164,-0.057, 0.087, 0.149, 0.133, 0.120, 0.127],
            "6m":  [-0.055, 0.030, 0.037, 0.138, 0.106, 0.078, 0.109],
            "9m":  [-0.078, 0.012, 0.067, 0.162, 0.142, 0.126, 0.139],
            "1y":  [-0.093, 0.006, 0.104, 0.191, 0.184, 0.180, 0.173],
            "2y":  [ 0.093, 0.158, 0.155, 0.187, 0.168, 0.157, 0.172],
            "3y":  [ 0.273, 0.223, 0.238, 0.223, 0.223, 0.229, 0.208],
            "5y":  [ 0.387, 0.290, 0.311, 0.286, 0.278, 0.276, 0.270],
            "7y":  [ 0.360, 0.323, 0.301, 0.321, 0.315, 0.313, 0.322],
            "10y": [ 0.378, 0.341, 0.373, 0.467, 0.436, 0.415, 0.420],
            "15y": [ 0.336, 0.319, 0.375, 0.506, 0.470, 0.441, 0.396],
            "20y": [ 0.313, 0.286, 0.398, 0.558, 0.491, 0.434, 0.378],
            "30y": [ 0.260, 0.247, 0.358, 0.406, 0.373, 0.359, 0.488],
        },
        "nu": {
            "1m":  [0.923, 0.710, 0.609, 0.722, 0.618, 0.516, 0.492],
            "3m":  [0.832, 0.700, 0.514, 0.533, 0.496, 0.460, 0.481],
            "6m":  [0.674, 0.602, 0.493, 0.446, 0.440, 0.434, 0.431],
            "9m":  [0.576, 0.503, 0.428, 0.404, 0.390, 0.378, 0.386],
            "1y":  [0.468, 0.400, 0.360, 0.360, 0.340, 0.320, 0.338],
            "2y":  [0.255, 0.262, 0.264, 0.266, 0.266, 0.266, 0.266],
            "3y":  [0.210, 0.244, 0.223, 0.214, 0.213, 0.213, 0.216],
            "5y":  [0.180, 0.193, 0.183, 0.183, 0.190, 0.197, 0.192],
            "7y":  [0.173, 0.184, 0.180, 0.193, 0.195, 0.199, 0.187],
            "10y": [0.153, 0.170, 0.179, 0.151, 0.152, 0.154, 0.157],
            "15y": [0.144, 0.160, 0.151, 0.134, 0.136, 0.139, 0.147],
            "20y": [0.140, 0.156, 0.138, 0.126, 0.128, 0.130, 0.135],
            "30y": [0.138, 0.148, 0.132, 0.123, 0.120, 0.117, 0.110],
        },
    },
}

def _apply_sabr_calibration(ccy: str) -> int:
    """Apply per-cell rho/nu from calibration reference. Returns cells updated."""
    ref = _SABR_REF.get(ccy)
    if ref is None: return 0
    _, a, b, r, n = get_ccy_vol_data(ccy)
    atm = get_working_atm_surface(ccy)
    if atm is None: return 0
    
    exp_ref  = ref["expiries"]
    ten_ref  = ref["tenors"]
    exp_yrs  = [label_to_years(e) for e in exp_ref]
    ten_yrs  = [float(t[:-1]) for t in ten_ref]
    rho_grid = np.array([ref["rho"][e] for e in exp_ref], dtype=float)
    nu_grid  = np.array([ref["nu"][e]  for e in exp_ref], dtype=float)
    
    # Get ATM expiry/tenor labels from surface
    atm_exps = atm["Expiry"].astype(str).str.strip().str.lower().tolist()
    atm_tens = [c for c in atm.columns if c != "Expiry"]
    
    # Build new rho and nu dataframes via bilinear interpolation
    _df_rho = atm[["Expiry"]].copy()
    _df_nu  = atm[["Expiry"]].copy()
    
    updated = 0
    for _tc in atm_tens:
        _ty = label_to_years(str(_tc))
        _rho_col = []
        _nu_col  = []
        for _exp_lbl in atm_exps:
            _ey = label_to_years(_exp_lbl)
            _rho_interp = float(np.interp(_ey, exp_yrs,
                [float(np.interp(_ty, ten_yrs, rho_grid[i])) for i in range(len(exp_ref))]))
            _nu_interp  = float(np.interp(_ey, exp_yrs,
                [float(np.interp(_ty, ten_yrs, nu_grid[i]))  for i in range(len(exp_ref))]))
            _rho_col.append(round(np.clip(_rho_interp, -0.95, 0.95), 3))
            _nu_col.append(round(max(_nu_interp, 0.001), 3))
            updated += 1
        _df_rho[_tc] = _rho_col
        _df_nu[_tc]  = _nu_col
    
    # Preserve alpha and beta, update rho and nu
    if "vol_data" not in st.session_state: st.session_state["vol_data"] = {}
    if ccy not in st.session_state["vol_data"]: st.session_state["vol_data"][ccy] = {}
    _vd = st.session_state["vol_data"][ccy]
    _old_atm, _old_a, _old_b, _, _ = get_ccy_vol_data(ccy)
    set_ccy_vol_data(ccy, _old_atm, _old_a, _old_b, _df_rho, _df_nu)
    return updated


def is_trainee() -> bool:
    """Check if current user is trainee — view only, no pricing or editing."""
    return st.session_state.get("user_role", "read_only") == "trainee"

def require_admin(label="🔒 Admin only"):
    """Show a lock message if user is not admin. Returns True if admin."""
    if not is_admin():
        st.warning(f"{label} — contact wpo@rateedge.au to request edit access.")
        return False
    return True

def init_session():
    if "sw_pending_reload" not in st.session_state:
        st.session_state["sw_pending_reload"] = None
    if "swaption_portfolio" not in st.session_state:
        st.session_state["swaption_portfolio"] = _load_portfolio()
    if "portfolio" not in st.session_state:
        st.session_state["portfolio"] = list(st.session_state["swaption_portfolio"])
    if "vol_data" not in st.session_state:
        st.session_state["vol_data"] = {}
    if "curves" not in st.session_state:
        st.session_state["curves"] = {}  # {ccy: {"irs": df, "6v3": df, "3v1": df, "ois": df}}
    if "basis_curves" not in st.session_state:
        st.session_state["basis_curves"] = {}  # {ccy: {"6v3": df, "3v1": df}}
    if "vol_editor" not in st.session_state:
        st.session_state["vol_editor"] = {"working": {}, "base": {}, "history": {}, "future": {}, "redo_stack": {}}
    if "history_configs" not in st.session_state:
        st.session_state["history_configs"] = {}
    if "theme_name" not in st.session_state:
        st.session_state["theme_name"] = "Dealer Dark"
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
    if "username" not in st.session_state:
        st.session_state["username"] = None
    # Timestamps for data loads
    if "load_timestamps" not in st.session_state:
        st.session_state["load_timestamps"] = {
            "atm": {},    # {ccy: datetime}
            "sabr": {},   # {ccy: datetime}
            "curves": {}, # {ccy: datetime}
        }
    # Forward swap matrix cache
    if "fwd_matrix" not in st.session_state:
        st.session_state["fwd_matrix"] = {}
    # Spread defaults - load from file if exists, otherwise use hardcoded defaults
    _spread_defaults = {
        "cf_spr_3m1y": 10.0, "cf_spr_1y1y": 11.5, "cf_spr_2y1y": 13.0,
        "cf_spr_3y1y": 17.5, "cf_spr_4y1y": 20.0, "cf_spr_5y2y": 45.0,
        "cf_spr_7y3y": 50.0, "cf_spr_10y2y": 35.0, "cf_spr_12y3y": 100.0,
    }
    _spreads_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cfs_spreads.json")
    if not any(k in st.session_state for k in _spread_defaults):
        _loaded = {}
        # Try DB first
        if HAS_POSTGRES and st.session_state.get("authenticated") and st.session_state.get("username"):
            try:
                _user_id = st.session_state.get("username")
                _db_spreads = load_user_config(_user_id, "cf_spreads", "AUD")
                if _db_spreads:
                    _loaded = _db_spreads
            except Exception:
                pass
        # Fallback to file
        if not _loaded:
            try:
                with open(_spreads_file, "r") as _f:
                    _loaded = json.load(_f)
            except Exception:
                pass
        for k, v in _spread_defaults.items():
            st.session_state[k] = float(_loaded.get(k, v))
    # Track if we've auto-loaded from DB this session
    if "db_auto_loaded" not in st.session_state:
        st.session_state["db_auto_loaded"] = False
    if "sabr_panel_visible" not in st.session_state:
        st.session_state["sabr_panel_visible"] = True
    # Correlation matrix + CMS bumps   —   load from file if exists
    if not any(f"corr_{k}" in st.session_state for k in _CORR_DEFAULTS):
        _loaded_cfg = _load_exotics_config()
        _loaded_corr = _loaded_cfg.get("correlations", {})
        _loaded_bumps = _loaded_cfg.get("cms_bumps", {})
        for k, v in _CORR_DEFAULTS.items():
            st.session_state[f"corr_{k}"] = float(_loaded_corr.get(k, v))
        for t in _CORR_TENORS:
            st.session_state[f"cms_bump_{t}"] = float(_loaded_bumps.get(t, 0.0))

# Auth credentials
# Auth handled by email OTP


def get_timestamp_str(category: str, ccy: str) -> str:
    """Get formatted timestamp string for a category/currency (Sydney time, AEST/AEDT aware)."""
    from datetime import datetime, timezone, timedelta
    ts = st.session_state.get("load_timestamps", {}).get(category, {}).get(ccy)
    if ts is None:
        return "Not loaded"
    try:
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        ts_syd = ts.astimezone(SYDNEY_TZ)
        # Determine AEST (UTC+10) vs AEDT (UTC+11) from actual UTC offset
        utc_offset_hours = ts_syd.utcoffset().total_seconds() / 3600
        tz_label = "AEDT" if utc_offset_hours == 11 else "AEST"
        return ts_syd.strftime(f"%d-%b-%Y %H:%M:%S {tz_label}")
    except:
        return ts.strftime("%Y-%m-%d %H:%M:%S")


def set_timestamp(category: str, ccy: str):
    """Set current timestamp for a category/currency"""
    from datetime import datetime, timezone
    if "load_timestamps" not in st.session_state:
        st.session_state["load_timestamps"] = {"atm": {}, "sabr": {}, "curves": {}}
    st.session_state["load_timestamps"][category][ccy] = datetime.now(timezone.utc)


def get_basis_curve(ccy: str, basis_type: str = "6v3") -> Optional[pd.DataFrame]:
    """Get basis curve for currency"""
    return st.session_state.get("basis_curves", {}).get(ccy, {}).get(basis_type)


def set_basis_curve(ccy: str, basis_type: str, df: pd.DataFrame):
    """Set basis curve for currency"""
    if "basis_curves" not in st.session_state:
        st.session_state["basis_curves"] = {}
    if ccy not in st.session_state["basis_curves"]:
        st.session_state["basis_curves"][ccy] = {}
    st.session_state["basis_curves"][ccy][basis_type] = df
    if basis_type == "ois":
        _cids = st.session_state.setdefault("_curve_commit_ids", {})
        _cids[f"{ccy}_ois"] = _cids.get(f"{ccy}_ois", 0) + 1
        # Clear fwd_ann cache for this ccy
        _fc = st.session_state.get("_fwd_ann_cache", {})
        for k in [k for k in _fc if k[0] == ccy]:
            del _fc[k]


def parse_tenor_to_years(tenor_str: str) -> float:
    """Parse tenor string like '3m', '6m', '1y', '2Y', 'AUD IRS 5y SS' to years"""
    s = str(tenor_str).strip().lower()
    # Handle descriptive formats like "AUD IRS 5y SS"
    import re
    match = re.search(r'(\d+(?:\.\d+)?)\s*y', s)
    if match:
        return float(match.group(1))
    match = re.search(r'(\d+(?:\.\d+)?)\s*m', s)
    if match:
        return float(match.group(1)) / 12.0
    match = re.search(r'(\d+(?:\.\d+)?)\s*w', s)
    if match:
        return float(match.group(1)) / 52.0
    match = re.search(r'(\d+(?:\.\d+)?)\s*d', s)
    if match:
        return float(match.group(1)) / 365.0
    # Try direct float
    try:
        return float(s)
    except:
        return 0.0


def load_curve_flexible(df: pd.DataFrame, name: str) -> pd.DataFrame:
    """Load curve with flexible tenor parsing - handles both numeric years and tenor strings."""
    if {"Maturity (Years)", "Zero Rate (%)"}.issubset(df.columns):
        out = df[["Maturity (Years)", "Zero Rate (%)"]].copy()
        out.rename(columns={"Maturity (Years)": "MaturityY", "Zero Rate (%)": "ZeroRatePct"}, inplace=True)
        # MaturityY column may be numeric (0.25, 1.0) or string tenor ('1Y', '3M')
        def _to_years(v):
            try:
                f = float(v)
                return f if f > 0 else 0.0
            except (ValueError, TypeError):
                return parse_tenor_to_years(str(v))
        out["MaturityY"] = out["MaturityY"].apply(_to_years)
        out = out[out["MaturityY"] > 0].sort_values("MaturityY").reset_index(drop=True)
        return out
    raise ValueError(f"{name}: expected columns 'Maturity (Years)', 'Zero Rate (%)'")


def load_basis_curve_flexible(df: pd.DataFrame, name: str) -> pd.DataFrame:
    """Load basis curve with flexible tenor parsing - handles both numeric years and tenor strings."""
    if {"Tenor (Years)", "Basis (bp)"}.issubset(df.columns):
        out = df[["Tenor (Years)", "Basis (bp)"]].copy()
        out.rename(columns={"Tenor (Years)": "MaturityY", "Basis (bp)": "BasisBp"}, inplace=True)
        def _to_years(v):
            try:
                f = float(v)
                return f if f > 0 else 0.0
            except (ValueError, TypeError):
                return parse_tenor_to_years(str(v))
        out["MaturityY"] = out["MaturityY"].apply(_to_years)
        out = out[out["MaturityY"] > 0].sort_values("MaturityY").reset_index(drop=True)
        return out
    raise ValueError(f"{name}: expected columns 'Tenor (Years)', 'Basis (bp)'")


def get_ccy_vol_data(ccy: str):
    v = st.session_state["vol_data"].get(ccy, {})
    return (v.get("atm"), v.get("alpha"), v.get("beta"), v.get("rho"), v.get("nu"))


def set_ccy_vol_data(ccy: str, atm, a, b, r, n):
    # Never overwrite a DB-loaded surface with None or stale data
    if atm is None and st.session_state.get(f"_vol_loaded_{ccy}"):
        existing_atm, _, _, _, _ = (st.session_state.get("vol_data", {}).get(ccy, {}).get("atm"), None, None, None, None)
        atm = existing_atm  # preserve the DB-loaded surface
    st.session_state["vol_data"][ccy] = {"atm": atm, "alpha": a, "beta": b, "rho": r, "nu": n}
    if atm is not None:
        _h = st.session_state.get(f"_atm_hash_{ccy}", 0)
        st.session_state[f"_atm_hash_{ccy}"] = _h + 1
    ve = st.session_state["vol_editor"]
    if atm is not None:
        ve["base"][ccy] = atm.copy()
        ve["working"][ccy] = atm.copy()
        ve["history"][ccy] = []
        ve["future"][ccy] = []


def get_ccy_curve(ccy: str) -> Optional[pd.DataFrame]:
    return st.session_state["curves"].get(ccy)


def set_ccy_curve(ccy: str, curve_df: pd.DataFrame):
    st.session_state["curves"][ccy] = curve_df
    # Increment curve commit ID — used to invalidate fwd_ann cache cheaply
    _cids = st.session_state.setdefault("_curve_commit_ids", {})
    _cids[ccy] = _cids.get(ccy, 0) + 1
    # Clear fwd_ann cache for this ccy
    _fc = st.session_state.get("_fwd_ann_cache", {})
    keys_to_del = [k for k in _fc if k[0] == ccy]
    for k in keys_to_del:
        del _fc[k]


def bootstrap_aud_zeros_from_bbg_feed(xl: pd.ExcelFile) -> Optional[pd.DataFrame]:
    """
    Bootstrap AUD zero curve from BBG_Feed par swap rates.
    Q/Q (quarterly) for ≤3Y, S/S (semi-annual) for ≥4Y. AUD T+1 settlement.
    Returns DataFrame(MaturityY, ZeroRatePct) with 19 points, or None on failure.
    """
    import re as _re
    if "BBG_Feed" not in xl.sheet_names:
        return None
    try:
        raw = pd.read_excel(xl, sheet_name="BBG_Feed", header=None)

        QQ_MAP = {"6m QQ":0.5,"9m QQ":0.75,"1Y QQ":1.0,"18m QQ":1.5,"2Y QQ":2.0,"3Y QQ":3.0}
        SS_MAP = {"4Y SS":4.0,"5Y SS":5.0,"6Y SS":6.0,"7Y SS":7.0,"8Y SS":8.0,"9Y SS":9.0,
                  "10Y SS":10.0,"12Y SS":12.0,"15Y SS":15.0,"20Y SS":20.0,"25Y SS":25.0,"30Y SS":30.0,
                  "40Y SS":40.0,"50Y SS":50.0}
        # 40Y/50Y labeled as spread rows in BBG_Feed (absolute par rate in col E)
        SPREAD_40_50 = {"aud 30y v 40y spread": 40.0, "aud 30y v 50y spread": 50.0,
                        "30y v 40y": 40.0, "30y v 50y": 50.0}
        OIS_MAP = {"OIS 1W":1/52,"OIS 1M":1/12,"OIS 2M":2/12,"OIS 3M":3/12,
                   "OIS 4M":4/12,"OIS 5M":5/12,"OIS 6M":6/12,"OIS 9M":9/12,
                   "OIS 1Y":1.0,"OIS 2Y":2.0,"OIS 3Y":3.0}
        # Note: OIS 40Y/50Y are NOT added here — they would seed bootstrap DFs
        # at those maturities before IRS par bootstrap runs, contaminating the zero curve.
        # Long-end OIS is read from OIS_AUD sheet and stored in config_basis for charting only.

        par_qq: dict = {}
        par_ss: dict = {}
        ois_rates: dict = {}
        in_usd = False

        for _, row in raw.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            # Stop reading at USD section to prevent cross-contamination
            if "USD" in label and any(x in label for x in ["IRS","OIS","SOFR","Fed Funds"]):
                in_usd = True
            if in_usd:
                continue
            try:
                mid = float(row.iloc[4])
                if math.isnan(mid):
                    continue
            except (TypeError, ValueError):
                continue
            ll = label.lower()
            # Word-boundary regex prevents "5Y SS" matching "25Y SS" etc
            for k, v in QQ_MAP.items():
                if _re.search(r"(?<![0-9])" + _re.escape(k.lower()), ll):
                    par_qq[v] = mid
            for k, v in SS_MAP.items():
                if _re.search(r"(?<![0-9])" + _re.escape(k.lower()), ll):
                    par_ss[v] = mid
            for k, v in SPREAD_40_50.items():
                if v not in par_ss:
                    if k in ll:
                        par_ss[v] = mid
            for k, v in OIS_MAP.items():
                if v not in ois_rates:  # take FIRST occurrence only — prevents USD OIS rows overwriting AUD AONIA
                    if _re.search(r"(?<![0-9])" + _re.escape(k.lower()), ll):
                        ois_rates[v] = mid

        if len(par_qq) < 3 or len(par_ss) < 6:
            return None

        SPOT = 1.0 / 252.0

        # Read Curves_AUD: col 0 = maturity, col 1 = par rate (mislabeled "Zero Rate")
        try:
            if "Curves_AUD" in xl.sheet_names:
                _ca = pd.read_excel(xl, sheet_name="Curves_AUD", usecols=[0,1], header=0)
                _ca.columns = ["MaturityY","ParRate"]
                _ca["MaturityY"] = pd.to_numeric(_ca["MaturityY"], errors="coerce")
                _ca["ParRate"] = pd.to_numeric(_ca["ParRate"], errors="coerce")
                # Override OIS 3M with 0.25Y entry
                _r025 = _ca[(_ca["MaturityY"] - 0.25).abs() < 0.01]
                if len(_r025) > 0:
                    ois_rates[3/12] = float(_r025.iloc[0]["ParRate"])
                # Add 40Y/50Y SS par if not already in BBG_Feed
                for _ext_t in [40.0, 50.0]:
                    if _ext_t not in par_ss:
                        _re = _ca[(_ca["MaturityY"] - _ext_t).abs() < 0.1]
                        if len(_re) > 0:
                            par_ss[_ext_t] = float(_re.iloc[0]["ParRate"])
        except Exception:
            pass

        # Seed with OIS discount factors (short-end anchor)
        dfs: dict = {0.0: 1.0}
        for t, r in sorted(ois_rates.items()):
            if not math.isnan(r):
                dfs[t] = math.exp(-r / 100.0 * t)

        def _df(t: float) -> float:
            """Log-linear interpolation of current discount factor curve."""
            ts = sorted(dfs.keys())
            dfv = [dfs[x] for x in ts]
            if t <= ts[0]: return dfv[0]
            if t >= ts[-1]:
                z = -math.log(dfv[-1]) / ts[-1]
                return math.exp(-z * t)
            for i in range(len(ts) - 1):
                if ts[i] <= t <= ts[i+1]:
                    w = (t - ts[i]) / (ts[i+1] - ts[i])
                    return math.exp((1-w)*math.log(dfv[i]) + w*math.log(dfv[i+1]))
            return dfv[-1]

        # Bootstrap: solve for unknown terminal DF given known intermediate DFs
        bootstrapped: dict = {}  # tenor -> (swap_end_time, df)

        def _boot(par_rate: float, tenor: float, freq: float) -> bool:
            c = par_rate / 100.0
            swap_end = SPOT + tenor
            times = []
            t = SPOT + freq
            while t <= swap_end + 1e-9:
                times.append(round(min(t, swap_end), 8))
                t += freq
            if not times: return False
            ann = sum(_df(ti) * freq for ti in times[:-1])
            df_end = (_df(SPOT) - c * ann) / (1.0 + c * freq)
            if df_end <= 0 or math.isnan(df_end): return False
            dfs[swap_end] = df_end
            bootstrapped[tenor] = (swap_end, df_end)
            return True

        for tenor in sorted(par_qq.keys()):
            _boot(par_qq[tenor], tenor, 0.25)
        for tenor in sorted(par_ss.keys()):
            _boot(par_ss[tenor], tenor, 0.50)

        if len(bootstrapped) < 10:
            return None

        # Store par rates for display in curves tab
        try:
            import streamlit as _st
            _par_rows = []
            for _t, _r in sorted(par_qq.items()): _par_rows.append({"Tenor": f"{_t}Y", "Par Rate (%)": _r, "Conv": "Q/Q"})
            for _t, _r in sorted(par_ss.items()): _par_rows.append({"Tenor": f"{_t}Y", "Par Rate (%)": _r, "Conv": "S/S"})
            if "config_basis" not in _st.session_state: _st.session_state["config_basis"] = {}
            if "_irs_par_rates" not in _st.session_state: _st.session_state["_irs_par_rates"] = {}
            _st.session_state["_irs_par_rates"]["AUD"] = pd.DataFrame(_par_rows)
            # Store raw QQ/SS dicts
            _st.session_state["_aud_par_qq"] = par_qq
            _st.session_state["_aud_par_ss"] = par_ss

            # Build and store PURE QQ and SS zero curves for matrix generation.
            # Pure QQ: <=3Y direct 3M BBSW, >=4Y = (SS par - basis) bootstrapped at Q/Q freq.
            # Pure SS: >=4Y direct 6M BBSW, <=3Y = (QQ par + basis) bootstrapped at S/S freq.
            # Using separate curves avoids the blended-curve basis error on 4y+ Q/Q forwards.
            try:
                # Read basis directly from Excel — session state not yet populated at this point
                # since load_config_excel loads curves before basis.
                _bx = _by = None
                try:
                    if "Basis_AUD_6v3" in xl.sheet_names:
                        _raw_b6 = pd.read_excel(xl, sheet_name="Basis_AUD_6v3", usecols=[0,1], header=0)
                        _raw_b6.columns = ["Tenor (Years)", "Basis (bp)"]
                        _raw_b6 = _raw_b6.dropna()
                        _bx = pd.to_numeric(_raw_b6["Tenor (Years)"], errors="coerce").to_numpy()
                        _by = pd.to_numeric(_raw_b6["Basis (bp)"], errors="coerce").to_numpy()
                        _mask = ~(np.isnan(_bx) | np.isnan(_by))
                        _bx = _bx[_mask]; _by = _by[_mask]
                except Exception:
                    _bx = _by = None
                def _basis_at(t):
                    if _bx is None or len(_bx) == 0: return 0.0
                    return float(np.interp(t, _bx, _by))

                def _build_pure_zero(par_inputs, all_qq):
                    """Bootstrap a zero curve from par_inputs {tenor: rate_pct}, all at same freq."""
                    _dfs = {0.0: 1.0}
                    for _t, _r in sorted(ois_rates.items()):
                        _dfs[_t] = math.exp(-_r / 100.0 * _t)
                    def _dfi(t):
                        _ts = sorted(_dfs.keys()); _dfv = [_dfs[x] for x in _ts]
                        if t <= _ts[0]: return _dfv[0]
                        if t >= _ts[-1]:
                            _z = -math.log(_dfv[-1]) / _ts[-1]; return math.exp(-_z * t)
                        for _i in range(len(_ts) - 1):
                            if _ts[_i] <= t <= _ts[_i+1]:
                                _w = (t - _ts[_i]) / (_ts[_i+1] - _ts[_i])
                                return math.exp((1-_w)*math.log(_dfv[_i]) + _w*math.log(_dfv[_i+1]))
                        return _dfv[-1]
                    _freq = 0.25 if all_qq else 0.50
                    for _tenor in sorted(par_inputs.keys()):
                        _c = par_inputs[_tenor] / 100.0
                        _swap_end = SPOT + _tenor
                        _times = []; _t = SPOT + _freq
                        while _t <= _swap_end + 1e-9:
                            _times.append(round(min(_t, _swap_end), 8)); _t += _freq
                        if not _times: continue
                        _ann = sum(_dfi(_ti) * _freq for _ti in _times[:-1])
                        _df_end = (_dfi(SPOT) - _c * _ann) / (1.0 + _c * _freq)
                        if _df_end > 0 and not math.isnan(_df_end):
                            _dfs[_swap_end] = _df_end
                    _MATS = [0.25,0.5,0.75,1.0,1.5,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,10.0,12.0,15.0,20.0,25.0,30.0,40.0,50.0]
                    _zc = {}
                    for _m in _MATS:
                        _d = _dfi(_m)
                        if _d > 0 and not math.isnan(_d):
                            _zc[_m] = -math.log(_d) / _m * 100
                    return _zc

                # Pure QQ par inputs: <=3Y direct, >=4Y = SS - basis
                _par_qq_full = dict(par_qq)
                for _t, _r in par_ss.items(): _par_qq_full[_t] = _r - _basis_at(_t) / 100.0
                # Pure SS par inputs: >=4Y direct, <=3Y = QQ + basis
                _par_ss_full = dict(par_ss)
                for _t, _r in par_qq.items(): _par_ss_full[_t] = _r + _basis_at(_t) / 100.0

                _st.session_state["_aud_zc_qq"] = _build_pure_zero(_par_qq_full, all_qq=True)
                _st.session_state["_aud_zc_ss"] = _build_pure_zero(_par_ss_full, all_qq=False)
            except Exception as _zce:
                pass
            # Bootstrap sanity check: flag any par rate >100bp from its neighbour.
            # Catches BBG_Feed column-read errors before they corrupt the zero curve.
            _all_par = sorted(list(par_qq.items()) + list(par_ss.items()))
            _boot_warnings = []
            for _i in range(1, len(_all_par)):
                _t0, _r0 = _all_par[_i-1]
                _t1, _r1 = _all_par[_i]
                if abs(_r1 - _r0) > 1.0:  # >100bp jump between adjacent par rates
                    _boot_warnings.append(
                        f"Par rate spike: {_t0}Y={_r0:.4f}% -> {_t1}Y={_r1:.4f}% "
                        f"(delta={abs(_r1-_r0)*100:.1f}bp) — check column E in BBG_Feed"
                    )
            if _boot_warnings:
                _st.session_state["_bootstrap_warnings"] = _boot_warnings
            else:
                _st.session_state.pop("_bootstrap_warnings", None)
        except Exception:
            pass

        # Build output grid using full dfs (OIS seed + IRS nodes) — correct short-end zeros
        def _irs_df(t: float) -> float:
            ts = sorted(dfs.keys()); dfv = [dfs[x] for x in ts]
            if not ts: return 0.0
            if t <= ts[0]: return dfv[0]
            if t >= ts[-1]:
                z = -math.log(dfv[-1]) / ts[-1]; return math.exp(-z * t)
            for i in range(len(ts) - 1):
                if ts[i] <= t <= ts[i+1]:
                    w = (t - ts[i]) / (ts[i+1] - ts[i])
                    return math.exp((1-w)*math.log(dfv[i]) + w*math.log(dfv[i+1]))
            return dfv[-1]

        MATURITIES = [0.25, 0.50, 0.75, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0,
                      7.0, 8.0, 9.0, 10.0, 12.0, 15.0, 20.0, 25.0, 30.0, 40.0, 50.0]
        rows = []
        for m in MATURITIES:
            d = _irs_df(m)
            if d > 0 and not math.isnan(d):
                z = -math.log(d) / m * 100
                rows.append({"MaturityY": m, "ZeroRatePct": round(z, 6)})

        if len(rows) < 15:
            return None

        return pd.DataFrame(rows)

    except Exception:
        return None


def load_config_excel(upload, load_type: str = "all") -> dict:
    """
    Load config from Excel with selective loading.
    load_type: "atm", "sabr", "curves", or "all"
    Returns dict with counts of what was loaded.
    """
    xl = pd.ExcelFile(upload)
    loaded = {"atm": 0, "sabr": 0, "curves": 0, "basis": 0}
    
    for ccy in SUPPORTED_CURRENCIES:
        # Load ATM vols
        _load_atm = (load_type in ["atm","all"] or
                     (load_type=="atm_aud" and ccy=="AUD") or
                     (load_type=="atm_usd_nzd" and ccy in ["USD","NZD"]))
        if _load_atm:
            atm_name = f"ATM_Vols_{ccy}"
            if atm_name in xl.sheet_names:
                atm_raw = pd.read_excel(xl, sheet_name=atm_name)
                atm_df = load_atm_surface(atm_raw, atm_name)
                _, old_a, old_b, old_r, old_n = get_ccy_vol_data(ccy)
                set_ccy_vol_data(ccy, atm_df, old_a, old_b, old_r, old_n)
                set_timestamp("atm", ccy)
                loaded["atm"] += 1
                # Clear vol_editor so it reloads from fresh surface
                if "vol_editor" in st.session_state:
                    st.session_state["vol_editor"]["working"].pop(ccy, None)
                    st.session_state["vol_editor"]["base"].pop(ccy, None)

        # Load SABR grids
        if load_type in ["sabr", "all"]:
            sabr_a = sabr_b = sabr_r = sabr_n = None
            has_sabr = False
            for base in ["SABR_Alpha", "SABR_Beta", "SABR_Rho", "SABR_Nu"]:
                sname = f"{base}_{ccy}"
                if sname in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sname)
                    df = ensure_sabr_matrix(df, sname)
                    has_sabr = True
                else:
                    df = None
                if base == "SABR_Alpha":
                    sabr_a = df
                elif base == "SABR_Beta":
                    sabr_b = df
                elif base == "SABR_Rho":
                    sabr_r = df
                elif base == "SABR_Nu":
                    sabr_n = df
            
            if has_sabr:
                # Get existing ATM to preserve
                old_atm, _, _, _, _ = get_ccy_vol_data(ccy)
                set_ccy_vol_data(ccy, old_atm, sabr_a, sabr_b, sabr_r, sabr_n)
                set_timestamp("sabr", ccy)
                loaded["sabr"] += 1

        # Load curves and basis curves
        if load_type in ["curves", "all"]:
            curve_df = None

            # AUD: bootstrap from BBG_Feed par rates (live rates → correct zeros)
            if ccy == "AUD":
                bootstrapped = bootstrap_aud_zeros_from_bbg_feed(xl)
                if bootstrapped is not None and len(bootstrapped) >= 15:
                    curve_df = bootstrapped

            # Fallback for AUD (if bootstrap fails) and primary for NZD/USD:
            # always read directly from Curves_{CCY} sheet
            if curve_df is None:
                curve_name = f"Curves_{ccy}"
                if curve_name in xl.sheet_names:
                    raw_curve = pd.read_excel(xl, sheet_name=curve_name, usecols=[0, 1])
                    try:
                        curve_df = load_curve_flexible(raw_curve, curve_name)
                    except:
                        curve_df = load_curve(raw_curve, curve_name)

            if curve_df is not None and len(curve_df) > 0:
                set_ccy_curve(ccy, curve_df)
                if "config_curves" not in st.session_state:
                    st.session_state["config_curves"] = {}
                st.session_state["config_curves"][ccy] = curve_df
                set_timestamp("curves", ccy)
                loaded["curves"] += 1
            
            # 6v3 basis curve
            basis_6v3_name = f"Basis_{ccy}_6v3"
            if basis_6v3_name in xl.sheet_names:
                raw_basis = pd.read_excel(xl, sheet_name=basis_6v3_name, usecols=[0, 1])
                try:
                    basis_df = load_basis_curve_flexible(raw_basis, basis_6v3_name)
                    set_basis_curve(ccy, "6v3", basis_df)
                    if "config_basis" not in st.session_state: st.session_state["config_basis"] = {}
                    if ccy not in st.session_state["config_basis"]: st.session_state["config_basis"][ccy] = {}
                    st.session_state["config_basis"][ccy]["6v3"] = basis_df
                    loaded["basis"] += 1
                except Exception as e:
                    pass  # Basis loading is optional
            
            # 3v1 basis curve (if exists)
            basis_3v1_name = f"Basis_{ccy}_3v1"
            if basis_3v1_name in xl.sheet_names:
                raw_basis = pd.read_excel(xl, sheet_name=basis_3v1_name, usecols=[0, 1])
                try:
                    basis_df = load_basis_curve_flexible(raw_basis, basis_3v1_name)
                    set_basis_curve(ccy, "3v1", basis_df)
                    if "config_basis" not in st.session_state: st.session_state["config_basis"] = {}
                    if ccy not in st.session_state["config_basis"]: st.session_state["config_basis"][ccy] = {}
                    st.session_state["config_basis"][ccy]["3v1"] = basis_df
                    loaded["basis"] += 1
                except:
                    pass
            
            # OIS curve (if exists)
            ois_name = f"OIS_{ccy}"
            if ois_name in xl.sheet_names:
                raw_ois = pd.read_excel(xl, sheet_name=ois_name, usecols=[0, 1])
                try:
                    ois_df = load_curve_flexible(raw_ois, ois_name)
                    set_basis_curve(ccy, "ois", ois_df)
                    if "config_basis" not in st.session_state: st.session_state["config_basis"] = {}
                    if ccy not in st.session_state["config_basis"]: st.session_state["config_basis"][ccy] = {}
                    st.session_state["config_basis"][ccy]["ois"] = ois_df
                except:
                    pass
    
    return loaded


def get_working_atm_surface(ccy: str) -> Optional[pd.DataFrame]:
    """Returns working vol editor draft if user has made unpublished edits,
    otherwise returns the committed surface from vol_data."""
    ve = st.session_state.get("vol_editor", {})
    working = ve.get("working", {})
    base = ve.get("base", {})
    # Only use working copy if it differs from base (i.e. user has made edits)
    if ccy in working and isinstance(working[ccy], pd.DataFrame):
        if ccy in base and isinstance(base[ccy], pd.DataFrame):
            # If working == base, it's just a mirror — use committed vol_data
            if not working[ccy].equals(base[ccy]):
                return working[ccy]
        # No base to compare — use committed instead to avoid stale state
    atm, _, _, _, _ = get_ccy_vol_data(ccy)
    return atm


def get_published_atm_surface(ccy: str) -> Optional[pd.DataFrame]:
    """Returns the base (last published/loaded) surface — ignores unpublished vol editor edits."""
    ve = st.session_state.get("vol_editor", {})
    base = ve.get("base", {})
    if ccy in base and isinstance(base[ccy], pd.DataFrame):
        return base[ccy]
    atm, _, _, _, _ = get_ccy_vol_data(ccy)
    return atm


def push_vol_history(ccy: str):
    ve = st.session_state["vol_editor"]
    working = ve["working"].get(ccy)
    if working is not None:
        ve["history"].setdefault(ccy, []).append(working.copy())
        ve["future"][ccy] = []


def undo_vol(ccy: str):
    ve = st.session_state["vol_editor"]
    hist = ve["history"].get(ccy, [])
    if hist:
        current = ve["working"].get(ccy)
        ve["future"].setdefault(ccy, [])
        if current is not None:
            ve["future"][ccy].append(current.copy())
        ve["working"][ccy] = hist.pop()


def redo_vol(ccy: str):
    ve = st.session_state["vol_editor"]
    fut = ve["future"].get(ccy, [])
    if fut:
        current = ve["working"].get(ccy)
        ve["history"].setdefault(ccy, [])
        if current is not None:
            ve["history"][ccy].append(current.copy())
        ve["working"][ccy] = fut.pop()


def publish_vol(ccy: str):
    ve = st.session_state["vol_editor"]
    working = ve["working"].get(ccy)
    if working is not None:
        ve["base"][ccy] = working.copy()
        atm, a, b, r, n = get_ccy_vol_data(ccy)
        set_ccy_vol_data(ccy, working.copy(), a, b, r, n)


# ============================
# Tabs
# ============================

def vol_config_tab():
    st.subheader(" Vol / SABR Config & Upload")
    
    # Get theme colors
    is_dark = st.session_state.get("theme_name", "Dealer Dark") == "Dealer Dark"
    card_bg = "#1e293b" if is_dark else "#ffffff"
    border_color = "#334155" if is_dark else "#e2e8f0"
    text_color = "#f1f5f9" if is_dark else "#1e3a5f"
    muted_color = "#94a3b8" if is_dark else "#64748b"
    
    # Database load button - ALWAYS visible at top
    db_url = get_db_url()
    if HAS_POSTGRES and db_url:
        # Date picker row — default to latest available curve date in DB
        from datetime import date as _dt_date
        _date_col, _lbl_col = st.columns([2, 6])
        with _date_col:
            # Get available dates — cached in session state, refreshed every 5 min
            import time as _time_vc
            _avail_dates = st.session_state.get("_avail_curve_dates_cache", [])
            _avail_cache_age = st.session_state.get("_avail_curve_dates_ts", 0)
            if not _avail_dates or (_time_vc.time() - _avail_cache_age > 300):
                try:
                    _dc = get_db_connection()
                    if _dc:
                        _dcur = _dc.cursor()
                        _dcur.execute("SELECT DISTINCT date FROM swap_rates WHERE currency='AUD' ORDER BY date DESC LIMIT 60")
                        _avail_dates = [r[0] for r in _dcur.fetchall()]
                        _dc.close()
                        st.session_state["_avail_curve_dates_cache"] = _avail_dates
                        st.session_state["_avail_curve_dates_ts"] = _time_vc.time()
                except: pass
            _today = _dt_date.today()
            _latest_db = _avail_dates[0] if _avail_dates else _today
            _default_date = _today if _today >= _latest_db else _latest_db
            # Force widget to today if it's still showing an older date
            if st.session_state.get("db_load_date") and st.session_state["db_load_date"] < _default_date:
                st.session_state["db_load_date"] = _default_date
            _load_date = st.date_input(
                "Curve/Vol date to load",
                value=_default_date,
                key="db_load_date",
                format="DD/MM/YYYY",
                help="Pick a historical date to restore that day's curves and vols. Defaults to latest available."
            )
        with _lbl_col:
            if _avail_dates:
                st.caption(f"Curve dates in DB: {', '.join(str(d) for d in _avail_dates[:5])}{'...' if len(_avail_dates)>5 else ''}")

        col_db1, col_db2, col_db3 = st.columns([2, 2, 4])
        with col_db1:
            if st.button(" Load from Database", key="load_db_btn_top", type="primary"):
                user_id = st.session_state.get("username", "default")
                st.session_state.pop("_load_debug", None)
                st.session_state["_db_load_date"] = str(_load_date)
                loaded_count = load_all_session_data(user_id, load_date=str(_load_date))
                _load_dbg = st.session_state.pop("_load_debug", [])
                if loaded_count > 0:
                    for _msg in _load_dbg:
                        st.toast(f"📊 {_msg}", icon="✅")
                    st.toast(f"Loaded {loaded_count} configs ({_load_date})", icon="✅")
                    st.session_state["_post_load_msgs"] = _load_dbg
                    st.session_state["_post_load_count"] = loaded_count
                    st.session_state["_post_load_date"] = str(_load_date)
                    st.rerun()
                else:
                    _conn_test = get_db_connection()
                    if not _conn_test:
                        st.error("❌ Database connection failed — check network.")
                    else:
                        _conn_test.close()
                        # Check if swap_rates has any data (curves load separately)
                        _curve_loaded = any([
                            st.session_state.get("config_curves", {}).get(c) is not None
                            for c in ["AUD","NZD","USD"]
                        ])
                        if _curve_loaded:
                            st.warning(
                                f"⚠️ Curves loaded from DB but no saved configs found for **{user_id}**. "
                                "Upload your RateEdge_Config.xlsx and click Commit to save vols and curves, "
                                "then Save to Database."
                            )
                        else:
                            st.warning(
                                f"⚠️ No saved data found for **{user_id}**. "
                                "You need to upload RateEdge_Config.xlsx and Save to Database from a desktop session first."
                            )
        with col_db2:
            if st.button(" Save to Database", key="save_db_btn_top", type="secondary"):
                user_id = st.session_state.get("username", "default")
                # Test connection first
                conn_test = get_db_connection()
                if not conn_test:
                    st.error(f"Database connection failed. URL configured: {bool(get_db_url())}")
                else:
                    conn_test.close()
                    _vd = st.session_state.get("vol_data", {})
                    _cv = st.session_state.get("curves", {})
                    _dbg = " | ".join(f"{c}: atm={'ok' if _vd.get(c,{}).get('atm') is not None else 'MISSING'} curve={'ok' if _cv.get(c) is not None else 'MISSING'}" for c in ["AUD","NZD","USD"])
                    # Ensure tables exist before saving
                    init_database()
                    saved = save_all_session_data(user_id)
                    _last_err = st.session_state.pop("_save_last_error", None)
                    _save_debug = st.session_state.pop("_save_debug_msgs", [])
                    if saved > 0:
                        st.success(f"Saved {saved} configs: {' | '.join(_save_debug)}")
                    elif _last_err:
                        st.error(f"Save failed: {_last_err}")
                    else:
                        st.warning(f"Nothing saved. Session: {_dbg}")
        with col_db3:
            if st.button(" Clear Corrupted DB Data", key="clear_db_btn"):
                user_id = st.session_state.get("username", "default")
                conn = get_db_connection()
                if conn:
                    try:
                        cur = conn.cursor()
                        cur.execute("DELETE FROM user_configs WHERE user_id = %s", (user_id,))
                        conn.commit()
                        cur.close()
                        conn.close()
                        st.success(" Cleared all database configs. Now upload Excel and Save.")
                    except Exception as e:
                        st.error(f"Clear failed: {e}")
            st.caption(" Database connected")

        with col_db3:
            if st.button("🔄 Reload Vols from DB", key="reload_vols_btn", type="secondary"):
                # Clear cached vol data and reload from vol_history
                for _ccy in ["AUD","NZD","USD"]:
                    st.session_state.get("vol_data", {}).pop(_ccy, None)
                    st.session_state.pop(f"_sabr_init_{_ccy}", None)
                st.session_state["db_auto_loaded"] = False
                st.rerun()
    
    st.markdown("---")
    
    # File upload
    st.markdown("#### Upload Config File")
    upload = st.file_uploader(
        "Upload RateEdge_Config.xlsx",
        type=["xlsx"],
        key="cfg_upload",
        help="Excel file with sheets: ATM_Vols_[CCY], SABR_*_[CCY], Curves_[CCY]"
    )
    
    if upload is not None:
        st.markdown("#### Select what to commit:")
        
        load_type = st.radio(
            "Commit options",
            ["All", "SOD IRS", "AUD Vol", "USD & NZD Vol"],
            index=0,
            horizontal=True,
            key="load_type_radio"
        )
        
        # Map selection to load_type
        type_map = {
            "All": "all",
            "SOD IRS": "curves",
            "AUD Vol": "atm_aud",
            "USD & NZD Vol": "atm_usd_nzd"
        }
        
        if st.button(" Commit Selected Data", key="commit_btn", type="primary", disabled=not can_upload_vol()):
            selected_type = type_map[load_type]
            loaded = load_config_excel(upload, selected_type)
            
            # Show what was loaded
            msgs = []
            if loaded["atm"] > 0:
                msgs.append(f"ATM Vols: {loaded['atm']} currencies")
            if loaded["sabr"] > 0:
                msgs.append(f"SABR Grids: {loaded['sabr']} currencies")
            if loaded["curves"] > 0:
                msgs.append(f"Curves: {loaded['curves']} currencies")
            
            if msgs:
                _cdebug = []
                for _c in SUPPORTED_CURRENCIES:
                    _cv = st.session_state.get("config_curves", {}).get(_c)
                    if _cv is not None and len(_cv) > 0:
                        _z025 = _cv[_cv["MaturityY"].sub(0.25).abs() < 0.01]["ZeroRatePct"]
                        _z1 = _cv[_cv["MaturityY"].sub(1.0).abs() < 0.01]["ZeroRatePct"]
                        z025 = float(_z025.iloc[0]) if len(_z025) else 0
                        z1 = float(_z1.iloc[0]) if len(_z1) else 0
                        _cdebug.append(f"{_c}: {len(_cv)} pts | 0.25Y={z025:.4f}% | 1Y={z1:.4f}%")
                st.success(f" Loaded: {', '.join(msgs)}")
                if _cdebug:
                    for _cd in _cdebug:
                        st.caption(f"📊 {_cd}")
                # Show bootstrap warnings immediately after commit
                _bwarn = st.session_state.get("_bootstrap_warnings", [])
                for _bw in _bwarn:
                    st.error(f"🔴 BOOTSTRAP ERROR — {_bw}")
                # Show AUD par rates parsed from BBG_Feed for visual confirmation
                _aud_par = st.session_state.get("_irs_par_rates", {}).get("AUD")
                if _aud_par is not None and not _aud_par.empty:
                    st.caption("AUD par rates read from BBG_Feed (verify these are correct before using the forward matrix):")
                    _par_disp = _aud_par.copy()
                    _par_disp["Par Rate (%)"] = _par_disp["Par Rate (%)"].apply(lambda x: round(float(x), 4))
                    st.dataframe(_par_disp.set_index("Tenor").T.style.format("{:.4f}", subset=pd.IndexSlice["Par Rate (%)", :]), use_container_width=True)
                # Auto-save to DB so it persists across sessions
                if HAS_POSTGRES and is_admin():
                    try:
                        _uid = st.session_state.get("username", "default")
                        _saved = save_all_session_data(_uid)
                        if _saved > 0:
                            st.success(f"✅ Auto-saved {_saved} configs to database.")
                        # Clear load_user_config cache so next load gets fresh data
                        load_user_config.clear()
                    except Exception as _e:
                        st.warning(f"Auto-save failed: {_e}")

                # Save uploaded curves to swap_rates (AUD 6M BBSW/3M BBSW/AONIA, NZD 3M BKBM/NZONIA, USD SOFR)
                if HAS_POSTGRES and is_admin():
                    try:
                        import datetime as _dt
                        _today = str(_dt.date.today())
                        _conn = get_db_connection()
                        if _conn:
                            _cur = _conn.cursor()
                            _swap_rows_saved = 0
                            _curve_saves = [
                                ("AUD", "6M BBSW", None),
                                ("AUD", "3M BBSW", None),
                                ("AUD", "AONIA",   "ois"),
                                ("NZD", "3M BKBM", None),
                                ("NZD", "NZONIA",  "ois"),
                                ("USD", "SOFR",    None),
                            ]
                            _commit_sanity_warns = []
                            for _ccy, _fr, _basis_type in _curve_saves:
                                if _basis_type:
                                    _cdf = st.session_state.get("config_basis", {}).get(_ccy, {}).get(_basis_type)
                                else:
                                    _cdf = st.session_state.get("config_curves", {}).get(_ccy)
                                if _cdf is None or len(_cdf) == 0:
                                    continue
                                # Build rate dict for sanity check before INSERT
                                _pre_check = {}
                                for _, _row in _cdf.iterrows():
                                    _mat = float(_row.get("MaturityY", 0))
                                    _rate = float(_row.get("ZeroRatePct", 0))
                                    _months = round(_mat * 12)
                                    if _mat < 1.0:
                                        _tenor = f"{_months}M"
                                    else:
                                        _tenor = f"{int(round(_mat))}Y"
                                    _pre_check[_tenor] = _rate
                                _sw = check_swap_rates_sanity(_pre_check, _fr, _ccy)
                                _commit_sanity_warns.extend(_sw)
                                # INSERT to DB
                                for _tenor, _rate in _pre_check.items():
                                    _cur.execute("""
                                        INSERT INTO swap_rates (date, currency, tenor, floating_rate, rate)
                                        VALUES (%s, %s, %s, %s, %s)
                                        ON CONFLICT (date, currency, tenor, floating_rate) DO NOTHING
                                    """, (_today, _ccy, _tenor, _fr, _rate))
                                    _swap_rows_saved += 1
                            # Show commit sanity warnings BEFORE success message
                            for _csw in _commit_sanity_warns:
                                st.error(f"🔴 COMMIT SANITY — {_csw}")
                            _conn.commit()
                            _cur.close()
                            _conn.close()
                            _load_curve_from_db_latest.clear()
                            if _swap_rows_saved > 0:
                                st.success(f"✅ Saved {_swap_rows_saved} curve points to swap_rates ({_today})")
                    except Exception as _se:
                        st.warning(f"Curve save to swap_rates failed: {_se}")
            else:
                st.warning("No matching data found in file for selected option.")
    
    st.markdown("---")
    st.markdown("#### Currently Loaded Status")
    
    # Show auto-load result if present
    _auto_msg = st.session_state.pop("_auto_load_msg", None)
    if _auto_msg:
        st.info(_auto_msg)

    # Show post-load summary (persists after rerun from Load from Database)
    _post_msgs = st.session_state.pop("_post_load_msgs", None)
    _post_count = st.session_state.pop("_post_load_count", None)
    _post_date = st.session_state.pop("_post_load_date", None)
    if _post_count is not None:
        st.success(f"✅ Loaded {_post_count} configs from database ({_post_date})")
        for _pm in (_post_msgs or []):
            st.info(f"📊 {_pm}")

    # Show database status
    if HAS_POSTGRES and get_db_url():
        st.caption(" Database: Connected")
    else:
        st.caption(" Database: Not configured")

    # Pull latest vol snapshot label per currency from vol_history (cached 60s)
    _latest_snaps = st.session_state.get("_latest_vol_snaps_cache", {})
    _snaps_cache_age = st.session_state.get("_latest_vol_snaps_ts", 0)
    import time as _time
    if HAS_POSTGRES and get_db_url() and (_time.time() - _snaps_cache_age > 60):
        try:
            _sc = get_db_connection()
            if _sc:
                _scur = _sc.cursor()
                _scur.execute("""
                    SELECT DISTINCT ON (currency) currency, snapshot_date, label
                    FROM vol_history
                    ORDER BY currency, snapshot_date DESC
                """)
                for _row in _scur.fetchall():
                    try:
                        _snap_date_str = ccy_eod_label(_row[0], _row[1])
                    except Exception:
                        _snap_date_str = str(_row[1])[:16]
                    _latest_snaps[_row[0]] = {"date": _snap_date_str, "label": _row[2]}
                _scur.close()
                _sc.close()
                st.session_state["_latest_vol_snaps_cache"] = _latest_snaps
                st.session_state["_latest_vol_snaps_ts"] = _time.time()
        except Exception:
            pass

    for ccy in SUPPORTED_CURRENCIES:
        atm, a, b, r, n = get_ccy_vol_data(ccy)
        _cc = st.session_state.get("config_curves", {}).get(ccy)
        curve = _cc if _cc is not None else get_ccy_curve(ccy)

        # ATM Vol timestamps
        if atm is not None:
            _atm_rows = atm.shape[0] if hasattr(atm, 'shape') else "?"
            _atm_cols = atm.shape[1] if hasattr(atm, 'shape') else "?"
            atm_status = f"✅ {_atm_rows}×{_atm_cols}"
            _snap = _latest_snaps.get(ccy, {})
            atm_saved   = _snap.get('label', '—') if _snap else '—'
            atm_loaded  = get_timestamp_str("atm", ccy)
        else:
            atm_status = "❌ Not loaded"
            atm_saved = '—'; atm_loaded = '—'

        # IRS Curve timestamps
        if curve is not None and len(curve) > 0:
            _src_date = curve["_source_date"].iloc[0] if "_source_date" in curve.columns else "—"
            curve_status = f"✅ {len(curve)} pts"
            curve_saved  = str(_src_date) if _src_date else '—'
            curve_loaded = get_timestamp_str("curves", ccy)
        else:
            curve_status = "❌ Not loaded"
            curve_saved = '—'; curve_loaded = '—'

        st.markdown(
            f"""
            <div style="background:{card_bg};border:1px solid {border_color};border-radius:10px;padding:1rem;margin:0.5rem 0;">
                <div style="font-weight:600;font-size:1.1rem;color:{text_color};margin-bottom:0.5rem;">{ccy}</div>
                <table style="width:100%;color:{text_color};font-size:0.85rem;border-collapse:collapse;">
                    <tr style="color:{muted_color};font-size:0.72rem;border-bottom:1px solid {border_color};">
                        <td style="padding:0.2rem 0.5rem 0.2rem 0;width:110px;"></td>
                        <td style="padding:0.2rem 0.5rem;">Status</td>
                        <td style="padding:0.2rem 0.5rem;">Last Saved</td>
                        <td style="padding:0.2rem 0.5rem;">Loaded</td>
                        <td style="padding:0.2rem 0.5rem;">Live</td>
                    </tr>
                    <tr>
                        <td style="padding:0.3rem 0.5rem 0.3rem 0;color:{muted_color};font-size:0.8rem;">ATM Vol</td>
                        <td style="padding:0.3rem 0.5rem;">{atm_status}</td>
                        <td style="padding:0.3rem 0.5rem;color:{muted_color};font-size:0.78rem;">{atm_saved}</td>
                        <td style="padding:0.3rem 0.5rem;color:{muted_color};font-size:0.78rem;">{atm_loaded}</td>
                        <td style="padding:0.3rem 0.5rem;color:#64748b;font-size:0.78rem;">—</td>
                    </tr>
                    <tr>
                        <td style="padding:0.3rem 0.5rem 0.3rem 0;color:{muted_color};font-size:0.8rem;">IRS Curve</td>
                        <td style="padding:0.3rem 0.5rem;">{curve_status}</td>
                        <td style="padding:0.3rem 0.5rem;color:{muted_color};font-size:0.78rem;">{curve_saved}</td>
                        <td style="padding:0.3rem 0.5rem;color:{muted_color};font-size:0.78rem;">{curve_loaded}</td>
                        <td style="padding:0.3rem 0.5rem;color:#64748b;font-size:0.78rem;">—</td>
                    </tr>
                </table>
            </div>
            """,
            unsafe_allow_html=True,
        )
    
    # Vol History Management Section
    if HAS_POSTGRES and get_db_url():
        st.markdown("---")
        st.markdown("### 📸 Vol History Snapshots")
        st.caption("Save and manage historical volatility snapshots")
        
        tab_save, tab_manage = st.tabs(["💾 Save Snapshot", "📅 Manage Snapshots"])
        
        with tab_save:
            st.markdown("#### Save Current Vol Surface")
            
            col1, col2 = st.columns(2)
            with col1:
                snap_ccy = st.selectbox("Currency", SUPPORTED_CURRENCIES, key="snap_ccy")
            with col2:
                from datetime import datetime as _dt_sl, timezone as _tz_sl, timedelta as _td_sl
                _utc_now = _dt_sl.now(_tz_sl.utc)
                # Sydney: AEDT=UTC+11 (Oct-Apr), AEST=UTC+10 (Apr-Oct)
                _month = _utc_now.month
                _syd_off = 11 if (_month >= 10 or _month <= 4) else 10
                _syd_tz2 = _tz_sl(timedelta(hours=_syd_off))
                _now_syd = _utc_now.astimezone(_syd_tz2)
                _tz_lbl2 = "AEDT" if _syd_off == 11 else "AEST"
                _auto_label = f"EOD {_now_syd.strftime('%d-%b-%Y')} {_tz_lbl2}" if _now_syd.hour >= 16 else f"Intraday {_now_syd.strftime('%d-%b-%Y %H:%M')} {_tz_lbl2}"
                snap_label = st.text_input("Label", value=_auto_label, placeholder="e.g. EOD 01-Apr-2026 AEDT", key="snap_label")
            
            snap_notes = st.text_area("Notes (optional)", placeholder="Additional context about this snapshot...", key="snap_notes", height=100)
            
            if st.button("💾 Save Snapshot", key="save_snapshot_btn", type="primary", disabled=not can_upload_vol()):
                if not snap_label.strip():
                    st.error("Please provide a label for this snapshot")
                else:
                    user_id = st.session_state.get("username", "default")
                    snapshot_id = save_vol_snapshot(user_id, snap_ccy, snap_label.strip(), snap_notes.strip())
                    if snapshot_id:
                        st.success(f"✅ Snapshot saved! ID: {snapshot_id}")
                    else:
                        st.error("Failed to save snapshot. Make sure vol data is loaded.")
        
        with tab_manage:
            st.markdown("#### Saved Snapshots")
            manage_ccy = st.selectbox("Filter by Currency", ["All"] + SUPPORTED_CURRENCIES, key="manage_snap_ccy")
            user_id = st.session_state.get("username", "default")
            filter_ccy = None if manage_ccy == "All" else manage_ccy
            if st.button("🔄 Load Snapshots", key="load_snaps_btn"):
                st.session_state["_snap_list_cache"] = list_vol_snapshots(user_id, filter_ccy)
                st.session_state["_snap_list_ccy"] = filter_ccy
            # Use cached list, refresh if currency changed
            if st.session_state.get("_snap_list_ccy") != filter_ccy:
                st.session_state.pop("_snap_list_cache", None)
            snapshots = st.session_state.get("_snap_list_cache", None)
            if snapshots is None:
                st.info("Click Load Snapshots to view saved vol surfaces.")
                snapshots = []
            
            if not snapshots:
                st.info("No snapshots saved yet. Create your first snapshot above!")
            else:
                st.caption(f"Found {len(snapshots)} snapshot(s)")
                
                for snap in snapshots:
                    with st.expander(f"📸 {snap['currency']} - {snap['label']} ({snap['snapshot_date'].strftime('%Y-%m-%d %H:%M')})", expanded=False):
                        col1, col2, col3 = st.columns([3, 1, 1])
                        
                        with col1:
                            st.markdown(f"**Currency:** {snap['currency']}")
                            st.markdown(f"**Date:** {snap['snapshot_date'].strftime('%Y-%m-%d %H:%M:%S')}")
                            if snap['notes']:
                                st.markdown(f"**Notes:** {snap['notes']}")
                            st.caption(f"Created: {snap['created_at'].strftime('%Y-%m-%d %H:%M:%S')}")
                        
                        with col2:
                            if st.button("📂 Load", key=f"load_snap_{snap['id']}", use_container_width=True):
                                loaded_snap = load_vol_snapshot(snap['id'])
                                if loaded_snap:
                                    # Load into session state
                                    ccy = loaded_snap['currency']
                                    if "vol_data" not in st.session_state:
                                        st.session_state["vol_data"] = {}
                                    if ccy not in st.session_state["vol_data"]:
                                        st.session_state["vol_data"][ccy] = {}
                                    
                                    st.session_state["vol_data"][ccy]["atm"] = loaded_snap["atm"]
                                    st.session_state["vol_data"][ccy]["alpha"] = loaded_snap["alpha"]
                                    st.session_state["vol_data"][ccy]["beta"] = loaded_snap["beta"]
                                    st.session_state["vol_data"][ccy]["rho"] = loaded_snap["rho"]
                                    st.session_state["vol_data"][ccy]["nu"] = loaded_snap["nu"]
                                    # Clear vol_editor so it reloads from fresh surface
                                    if "vol_editor" in st.session_state:
                                        st.session_state["vol_editor"]["working"].pop(ccy, None)
                                        st.session_state["vol_editor"]["base"].pop(ccy, None)
                                    _h = st.session_state.get(f"_atm_hash_{ccy}", 0)
                                    st.session_state[f"_atm_hash_{ccy}"] = _h + 1
                                    
                                    # Update timestamps
                                    if "timestamps" not in st.session_state:
                                        st.session_state["timestamps"] = {}
                                    st.session_state["timestamps"][f"atm_{ccy}"] = loaded_snap['snapshot_date'].strftime('%Y-%m-%d %H:%M:%S')
                                    st.session_state["timestamps"][f"sabr_{ccy}"] = loaded_snap['snapshot_date'].strftime('%Y-%m-%d %H:%M:%S')
                                    
                                    st.success(f"✅ Loaded snapshot: {snap['label']}")
                                    st.rerun()
                                else:
                                    st.error("Failed to load snapshot")
                        
                        with col3:
                            if st.button("🗑️ Delete", key=f"del_snap_{snap['id']}", use_container_width=True):
                                if delete_vol_snapshot(snap['id']):
                                    st.success("✅ Deleted")
                                    st.rerun()
                                else:
                                    st.error("Failed to delete")


@st.cache_data(ttl=300, show_spinner=False)
def _load_curve_from_db_latest(floating_rate: str, ccy: str = "AUD", load_date: str = None) -> pd.DataFrame:
    """Load swap rates from swap_rates table. load_date: specific date or None for latest."""
    try:
        conn = get_db_connection()
        if conn is None:
            return None
        cur = conn.cursor()
        # Use specified date or find the most recent available
        if load_date:
            # Try exact date first, fall back to most recent on or before that date
            cur.execute(
                "SELECT MAX(date) FROM swap_rates WHERE currency=%s AND floating_rate=%s AND date <= %s",
                (ccy, floating_rate, load_date)
            )
        else:
            cur.execute(
                "SELECT MAX(date) FROM swap_rates WHERE currency=%s AND floating_rate=%s",
                (ccy, floating_rate)
            )
        row = cur.fetchone()
        if not row or row[0] is None:
            conn.close()
            return None
        latest_date = row[0]
        # Get all tenors for that date
        cur.execute(
            "SELECT tenor, rate FROM swap_rates WHERE currency=%s AND floating_rate=%s AND date=%s ORDER BY tenor",
            (ccy, floating_rate, latest_date)
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return None
        # Convert to MaturityY / ZeroRatePct format
        records = []
        for tenor, rate in rows:
            t = str(tenor).strip().upper()
            import re as _re
            m = _re.match(r"(\d+(?:\.\d+)?)(Y|M)", t)
            if not m: continue
            v, u = float(m.group(1)), m.group(2)
            mat_y = v if u == "Y" else v / 12
            records.append({"MaturityY": mat_y, "ZeroRatePct": float(rate)})  # Supabase stores as % (e.g. 4.78); ZeroRatePct in percent
        if not records:
            return None
        df = pd.DataFrame(records).sort_values("MaturityY").reset_index(drop=True)
        df["_source_date"] = str(latest_date)
        return df
    except Exception:
        return None


@st.cache_data(ttl=300, show_spinner=False)
def _load_ois_from_db_latest(ccy: str = "AUD") -> pd.DataFrame:
    """Load latest AONIA OIS swap rates from swap_rates table."""
    try:
        conn = get_db_connection()
        if conn is None:
            return None
        cur = conn.cursor()
        cur.execute(
            "SELECT MAX(date) FROM swap_rates WHERE currency=%s AND floating_rate=%s",
            (ccy, "AONIA")
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            conn.close()
            return None
        latest_date = row[0]
        cur.execute(
            "SELECT tenor, rate FROM swap_rates WHERE currency=%s AND floating_rate=%s AND date=%s ORDER BY tenor",
            (ccy, "AONIA", latest_date)
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return None
        records = []
        for tenor, rate in rows:
            t = str(tenor).strip().upper()
            import re as _re
            m = _re.match(r"(\d+(?:\.\d+)?)(Y|M)", t)
            if not m: continue
            v, u = float(m.group(1)), m.group(2)
            mat_y = v if u == "Y" else v / 12
            records.append({"MaturityY": mat_y, "ZeroRatePct": float(rate)})  # Supabase stores as % (e.g. 4.78); ZeroRatePct in percent
        if not records:
            return None
        df = pd.DataFrame(records).sort_values("MaturityY").reset_index(drop=True)
        df["_source_date"] = str(latest_date)
        return df
    except Exception:
        return None


def generate_forward_matrix(ccy: str, curve: pd.DataFrame, basis_6v3: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """Public wrapper   —   converts DataFrames → tuples and calls the cached matrix generator."""
    if curve is None or curve.empty:
        return pd.DataFrame()
    curve_tuple = (
        tuple(curve["MaturityY"].to_numpy().astype(float).tolist()),
        tuple(curve["ZeroRatePct"].to_numpy().astype(float).tolist()),
    )
    basis_tuple = None
    if basis_6v3 is not None and not basis_6v3.empty and "MaturityY" in basis_6v3.columns and "BasisBp" in basis_6v3.columns:
        basis_tuple = (
            tuple(basis_6v3["MaturityY"].to_numpy().astype(float).tolist()),
            tuple(basis_6v3["BasisBp"].to_numpy().astype(float).tolist()),
        )
    return _generate_forward_matrix_cached(ccy, curve_tuple, basis_tuple, convention="market", ois_tuple=None)


def generate_forward_matrix_convention(ccy: str, curve: pd.DataFrame, basis_6v3: Optional[pd.DataFrame] = None,
                                        convention: str = "market") -> pd.DataFrame:
    """Public wrapper   —   like generate_forward_matrix but with explicit convention override."""
    if curve is None or curve.empty:
        return pd.DataFrame()
    curve_tuple = (
        tuple(curve["MaturityY"].to_numpy().astype(float).tolist()),
        tuple(curve["ZeroRatePct"].to_numpy().astype(float).tolist()),
    )
    basis_tuple = None
    if basis_6v3 is not None and not basis_6v3.empty and "MaturityY" in basis_6v3.columns and "BasisBp" in basis_6v3.columns:
        basis_tuple = (
            tuple(basis_6v3["MaturityY"].to_numpy().astype(float).tolist()),
            tuple(basis_6v3["BasisBp"].to_numpy().astype(float).tolist()),
        )
    return _generate_forward_matrix_cached(ccy, curve_tuple, basis_tuple, convention=convention, ois_tuple=None)


def curves_tab():
    import plotly.graph_objects as go
    st.subheader("📐 IRS Curves & Forward Matrix")

    ccy = st.selectbox("Currency", SUPPORTED_CURRENCIES, key="curve_ccy")

    curve     = st.session_state.get("config_curves", {}).get(ccy)
    basis_6v3 = st.session_state.get("config_basis", {}).get(ccy, {}).get("6v3")
    basis_3v1 = st.session_state.get("config_basis", {}).get(ccy, {}).get("3v1")
    ois_curve = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    par_rates = st.session_state.get("_irs_par_rates", {}).get(ccy)

    if curve is None:
        st.warning("⬆️ Upload RateEdge_Config.xlsx in Vol/Upload tab → Commit All")
        return

    def _clean(df):
        if df is None: return None
        return df.drop(columns=["_source_date"], errors="ignore")

    curve_c = _clean(curve)
    b6c     = _clean(basis_6v3)
    b3c     = _clean(basis_3v1)
    oisc    = _clean(ois_curve)

    # ── Chart toggles ─────────────────────────────────────────────────────────
    _ck = st.columns(5)
    with _ck[0]: _show_par = st.checkbox("IRS Par", value=True, key="chart_par")
    with _ck[1]: _show_irs = st.checkbox("IRS Zero", value=True, key="chart_irs")
    with _ck[2]: _show_ois = st.checkbox("OIS", value=True, key="chart_ois")
    with _ck[3]: _show_b6  = st.checkbox("6v3 Basis", value=True, key="chart_b6")
    with _ck[4]: _show_b3  = st.checkbox("3v1 Basis", value=True, key="chart_b3")

    try:
        fig = go.Figure()
        if _show_par:
            if par_rates is not None and not par_rates.empty:
                try:
                    _par_x = list(par_rates["Tenor"].apply(
                        lambda x: float(x[:-1]) if str(x).endswith("Y") else float(str(x)[:-1])/12))
                    _par_y = list(par_rates["Par Rate (%)"])
                    # Supplement with 40Y/50Y from _aud_par_ss (uploaded SS par rates)
                    _par_ss_ext = st.session_state.get("_aud_par_ss", {})
                    for _et in [40.0, 50.0]:
                        if _et in _par_ss_ext and _et not in _par_x:
                            _par_x.append(_et); _par_y.append(_par_ss_ext[_et])
                    # Also check curve_c for 40Y/50Y par (col "ZeroRatePct" = par on Curves_AUD)
                    for _, _cr in curve_c.iterrows():
                        _m = float(_cr["MaturityY"])
                        if _m in [40.0, 50.0] and _m not in _par_x:
                            _par_x.append(_m); _par_y.append(float(_cr["ZeroRatePct"]))
                    _ext_pairs = sorted(zip(_par_x, _par_y))
                    fig.add_trace(go.Scatter(
                        x=[p[0] for p in _ext_pairs],
                        y=[p[1] for p in _ext_pairs],
                        mode="lines+markers", name="IRS Par", line=dict(color="#22c55e", width=2)))
                except Exception:
                    pass
            else:
                fig.add_trace(go.Scatter(x=curve_c["MaturityY"], y=curve_c["ZeroRatePct"],
                    mode="lines+markers", name="IRS Par", line=dict(color="#22c55e", width=2)))
        if _show_irs:
            # Use blended curve (correct QQ/SS zero rates), extend to 40Y/50Y from _aud_zc_ss
            _zx = list(curve_c["MaturityY"].astype(float))
            _zy = list(curve_c["ZeroRatePct"].astype(float))
            if ccy == "AUD":
                _zc_ss = st.session_state.get("_aud_zc_ss") or {}
                for _et in [40.0, 50.0]:
                    if _et in _zc_ss and _et not in _zx:
                        _zx.append(_et); _zy.append(_zc_ss[_et])
            _pairs = sorted(zip(_zx, _zy))
            fig.add_trace(go.Scatter(
                x=[p[0] for p in _pairs], y=[p[1] for p in _pairs],
                mode="lines+markers", name="IRS Zero", line=dict(color="#3b82f6", width=2)))
        if _show_ois and oisc is not None and not oisc.empty:
            fig.add_trace(go.Scatter(x=oisc["MaturityY"], y=oisc["ZeroRatePct"],
                mode="lines+markers", name="OIS", line=dict(color="#f59e0b", width=2)))
        if _show_b6 and b6c is not None and not b6c.empty:
            fig.add_trace(go.Scatter(x=b6c["MaturityY"], y=b6c["BasisBp"],
                mode="lines+markers", name="6v3 Basis (bp)",
                yaxis="y2", line=dict(color="#ef4444", width=2, dash="dot")))
        if _show_b3 and b3c is not None and not b3c.empty:
            fig.add_trace(go.Scatter(x=b3c["MaturityY"], y=b3c["BasisBp"],
                mode="lines+markers", name="3v1 Basis (bp)",
                yaxis="y2", line=dict(color="#a855f7", width=2, dash="dot")))
        fig.update_layout(height=380, template="plotly_dark",
            xaxis=dict(title="Maturity (Years)"),
            yaxis=dict(title="Rate (%)"),
            yaxis2=dict(title="Basis (bp)", overlaying="y", side="right", showgrid=False),
            legend=dict(orientation="h", y=1.1),
            margin=dict(l=40, r=40, t=20, b=40))
        st.plotly_chart(fig, use_container_width=True)
    except Exception as _e:
        st.warning(f"Chart: {_e}")

    with st.expander("IRS Par Rates & Curve Data", expanded=False):
        _cols_to_show = []
        if _show_par:
            if par_rates is not None and not par_rates.empty:
                # Add 40Y/50Y to par table if available
                _par_table = par_rates.copy()
                _par_ss_full = st.session_state.get("_aud_par_ss", {})
                _existing_tenors = set(_par_table["Tenor"].astype(str).str.upper())
                _extra_rows = []
                for _et in [40.0, 50.0]:
                    _tk = f"{int(_et)}Y"
                    if _tk not in _existing_tenors and _et in _par_ss_full:
                        _extra_rows.append({"Tenor": _tk, "Par Rate (%)": round(_par_ss_full[_et], 4), "Conv": "S/S"})
                if _extra_rows:
                    import pandas as _pd2
                    _par_table = _pd2.concat([_par_table, _pd2.DataFrame(_extra_rows)], ignore_index=True)
            else:
                _par_table = curve_c
            _cols_to_show.append(("IRS Par Rates (%)", _par_table))
        if _show_irs:
            _cols_to_show.append(("IRS Zero Curve (%)", curve_c))
        if _show_ois and oisc is not None and not oisc.empty:
            _cols_to_show.append(("OIS Curve (%)", oisc))
        if _show_b6 and b6c is not None and not b6c.empty:
            _cols_to_show.append(("6v3 Basis (bp)", b6c))
        if _show_b3 and b3c is not None and not b3c.empty:
            _cols_to_show.append(("3v1 Basis (bp)", b3c))
        if not _cols_to_show:
            _par_table = par_rates if (par_rates is not None and not par_rates.empty) else curve_c
            _cols_to_show.append(("IRS Par Rates (%)", _par_table))
        _tcols = st.columns(len(_cols_to_show))
        for _i, (_label, _df) in enumerate(_cols_to_show):
            with _tcols[_i]:
                st.caption(_label)
                st.dataframe(_df, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── IRS Forward Matrix ─────────────────────────────────────────────────────
    if "fwd_matrix"   not in st.session_state: st.session_state["fwd_matrix"]   = {}
    if "basis_matrix" not in st.session_state: st.session_state["basis_matrix"] = {}
    if "fwd_section_open" not in st.session_state: st.session_state["fwd_section_open"] = True

    has_fwd = ccy in st.session_state.get("fwd_matrix", {}) and \
              not st.session_state["fwd_matrix"][ccy].empty

    _fl = "▼ Hide Forward Swap Rates" if st.session_state["fwd_section_open"] else "▶ Show Forward Swap Rates"
    if st.button(_fl, key="fwd_toggle"):
        st.session_state["fwd_section_open"] = not st.session_state["fwd_section_open"]
        st.rerun()

    if st.session_state["fwd_section_open"]:
        _r1, _r2, _r3, _r4 = st.columns([3, 1, 3, 3])
        with _r1:
            leg_conv = st.radio("Leg Convention", ["Market","Q/Q","S/S"],
                                horizontal=True, key="fwd_leg_convention")
        with _r2:
            show_hm = st.checkbox("Heatmap", value=False, key="show_heatmap")
        with _r3:
            gen_fwd = st.button("▶ Generate Forward Matrix", key="gen_fwd_matrix",
                                type="primary", use_container_width=True,
                                help="Force regenerate. Convention changes auto-update below.")
        with _r4:
            st.download_button("⬇ Download",
                               st.session_state["fwd_matrix"][ccy].to_csv() if has_fwd else "",
                               f"{ccy}_fwd_matrix.csv", key="dl_fwd",
                               use_container_width=True, type="primary",
                               disabled=not has_fwd)

        _mc = st.session_state.get("config_curves", {}).get(ccy)
        _mb = st.session_state.get("config_basis", {}).get(ccy, {}).get("6v3")
        _conv_now = {"Market":"market","Q/Q":"qq","S/S":"ss"}.get(leg_conv,"market")
        _prev_conv = st.session_state.get("fwd_convention", "market")
        _prev_ccy  = st.session_state.get("fwd_ccy", ccy)

        # Auto-regen: convention changed, or currency changed, or explicit button press
        _need_regen = gen_fwd or _conv_now != _prev_conv or _prev_ccy != ccy

        if _need_regen and _mc is not None:
            with st.spinner("Calculating..."):
                fm = generate_forward_matrix_convention(ccy, _mc, _mb, _conv_now)
                st.session_state["fwd_matrix"][ccy] = fm
                st.session_state["fwd_convention"]  = _conv_now
                st.session_state["fwd_ccy"]         = ccy
                if _mb is not None:
                    st.session_state["basis_matrix"][ccy] = generate_basis_matrix(ccy, _mb)
            has_fwd = True
        elif _need_regen and _mc is None:
            st.error("No curve — upload config first")

        # Also generate 3v1 basis matrix if basis available
        _mb3v1 = st.session_state.get("config_basis", {}).get(ccy, {}).get("3v1")
        if _need_regen and _mc is not None and _mb3v1 is not None:
            if "basis_matrix_3v1" not in st.session_state:
                st.session_state["basis_matrix_3v1"] = {}
            st.session_state["basis_matrix_3v1"][ccy] = generate_basis_matrix(ccy, _mb3v1)

        if has_fwd:
            has_basis   = ccy in st.session_state.get("basis_matrix", {})
            has_basis3v1 = ccy in st.session_state.get("basis_matrix_3v1", {})
            _vo = ["IRS Fwd"]
            if has_basis:    _vo.append("6v3 Basis")
            if has_basis3v1: _vo.append("3v1 Basis")
            _rv = st.radio("View", _vo, horizontal=True, key="rate_view_toggle")
            if _rv == "IRS Fwd":
                _disp = st.session_state["fwd_matrix"][ccy]
            elif _rv == "6v3 Basis":
                _disp = st.session_state["basis_matrix"].get(ccy)
            else:
                _disp = st.session_state.get("basis_matrix_3v1", {}).get(ccy)
            if _disp is not None:
                _nc  = [c for c in _disp.columns if c != "Expiry"]
                _fmt = {c: "{:.4f}" for c in _nc}
                if show_hm:
                    _cm = "RdYlGn_r" if _rv == "IRS Fwd" else "RdYlGn"
                    st.dataframe(_disp.style.format(_fmt).background_gradient(_cm, axis=None, subset=_nc),
                                 use_container_width=True, height=820)
                else:
                    st.dataframe(_disp.style.format(_fmt), use_container_width=True, height=820)
        else:
            st.info("Click **▶ Generate Forward Matrix**")

    # ── 📡 Publish to Options Whiteboard ──────────────────────────────────────────────────
    st.markdown("---")
    _pb1, _pb2 = st.columns([2, 4])
    with _pb1:
        if st.button("📡 Publish ATM Vols / Prems / FWDs to Options Whiteboard", disabled=is_trainee(),
                     key="curves_publish_blotter", type="primary",
                     use_container_width=True):
            if not HAS_POSTGRES:
                st.error("Database not connected.")
            elif not is_admin():
                st.warning("Admin only.")
            else:
                _pub_ccy = ccy
                _all_mids = {}
                # ATM Vol mids
                _atm_s = get_working_atm_surface(_pub_ccy)
                if _atm_s is not None:
                    _av = _atm_s.copy()
                    if "Expiry" in _av.columns: _av = _av.set_index("Expiry")
                    for _el in _av.index:
                        for _tc in _av.columns:
                            try:
                                _v = float(_av.loc[_el, _tc])
                                if not math.isnan(_v) and _v > 0:
                                    _all_mids[f"vol_{str(_el)}_{str(_tc)}"] = {
                                        "value": round(_v, 4),
                                        "label": f"ATM vol {_el} {_tc} bp"}
                            except: pass
                # BP Premium mids — from pricer's own Generate ATM Matrix
                _prem_df = st.session_state.get("atm_prem_matrix", {}).get(_pub_ccy, {}).get("prem")
                if _prem_df is not None:
                    _pv = _prem_df.copy()
                    if "Expiry" in _pv.columns: _pv = _pv.set_index("Expiry")
                    for _el in _pv.index:
                        for _tc in _pv.columns:
                            try:
                                _v = float(_pv.loc[_el, _tc])
                                if not math.isnan(_v) and _v > 0:
                                    _all_mids[f"prem_{str(_el)}_{str(_tc)}"] = {
                                        "value": round(_v, 4),
                                        "label": f"Straddle prem {_el} {_tc} bp"}
                            except: pass
                # FWD ATM IRS mids
                _fwd_m = st.session_state.get("fwd_matrix", {}).get(_pub_ccy)
                if _fwd_m is not None and not _fwd_m.empty:
                    for _fexp in _fwd_m.index:
                        for _ften in _fwd_m.columns:
                            try:
                                _fv = float(_fwd_m.loc[_fexp, _ften])
                                if not math.isnan(_fv) and _fv > 0:
                                    _all_mids[f"fwd_{str(_fexp)}_{str(_ften)}"] = {
                                        "value": round(_fv, 4),
                                        "label": f"FWD ATM {_fexp} {_ften} %"}
                            except: pass
                if _all_mids:
                    _n = publish_blotter_mids(_pub_ccy, _all_mids)
                    # Save snapshot fast — direct INSERT using already-built matrices
                    try:
                        from psycopg2.extras import Json as _Json
                        _snap_conn = get_db_connection()
                        if _snap_conn:
                            _snap_uid = st.session_state.get("username", "wpo@rateedge.au")
                            if _snap_uid in {"wpo70@icloud.com"}: _snap_uid = "wpo@rateedge.au"
                            # Build atm_vols JSON from working surface
                            _snap_atm = get_working_atm_surface(_pub_ccy)
                            _snap_prem_df = st.session_state.get("atm_prem_matrix", {}).get(_pub_ccy, {}).get("prem")
                            if _snap_atm is not None:
                                _atm_j = _snap_atm.copy()
                                if _atm_j.index.name == "Expiry": _atm_j = _atm_j.reset_index()
                                _atm_json = _Json({"values": _atm_j.to_dict(orient="records")})
                                _prem_json = None
                                if _snap_prem_df is not None:
                                    _pv2 = _snap_prem_df.copy()
                                    if _pv2.index.name == "Expiry": _pv2 = _pv2.reset_index()
                                    _prem_json = _Json({"values": _pv2.to_dict(orient="records")})
                                from datetime import datetime as _dtnow2
                                from datetime import datetime as _dtnow2, timezone as _tz_pub, timedelta as _td_pub
                                _utc_now2 = _dtnow2.now(_tz_pub.utc)
                                _syd_off2 = 11 if (_utc_now2.month >= 10 or _utc_now2.month <= 4) else 10
                                _now_local = _utc_now2.astimezone(_tz_pub(timedelta(hours=_syd_off2)))
                                _tz_lbl = "AEDT" if _syd_off2 == 11 else "AEST"
                                _slbl = f"{_pub_ccy} {_now_local.strftime('%d-%b-%Y %H:%M')} {_tz_lbl}"
                                _sc2 = _snap_conn.cursor()
                                _sc2.execute("""
                                    INSERT INTO vol_history
                                    (user_id, currency, snapshot_date, label, atm_vols, atm_prems, notes)
                                    VALUES (%s,%s,NOW(),%s,%s,%s,%s)
                                """, (_snap_uid, _pub_ccy, _slbl, _atm_json, _prem_json, "Published from Curves tab"))
                                _snap_conn.commit()
                                _sc2.close()
                            _snap_conn.close()
                    except Exception as _se:
                        pass  # snapshot failure doesn't block publish
                    st.success(f"✅ Published {_n} mids + snapshot saved for {_pub_ccy}")
                else:
                    st.warning("No data — generate forward matrix and ATM matrix first.")
    with _pb2:
        st.caption("Publishes ATM vols, bp premiums and FWD ATM rates to the blotter, and saves a vol snapshot. Generate ATM Matrix and Forward Matrix first.")

    st.markdown("---")

    # ── Swap Rates Validator ───────────────────────────────────────────────────
    if st.session_state.get("_swap_load_warnings"):
        for _slw in st.session_state["_swap_load_warnings"]:
            st.error(_slw)
        if st.button("Clear swap rate warnings", key="clear_swap_warns"):
            st.session_state.pop("_swap_load_warnings", None)
            st.rerun()


    st.markdown("---")

    # ── ATM Vol / Forward Premium / Vega ──────────────────────────────────────
    if "atm_prem_matrix" not in st.session_state: st.session_state["atm_prem_matrix"] = {}
    if "atm_section_open" not in st.session_state: st.session_state["atm_section_open"] = True

    _al = "▼ Hide ATM Vol / Premium / Vega" if st.session_state["atm_section_open"] else "▶ Show ATM Vol / Premium / Vega"
    if st.button(_al, key="atm_toggle"):
        st.session_state["atm_section_open"] = not st.session_state["atm_section_open"]
        st.rerun()

    if st.session_state["atm_section_open"]:
        atm_vols, _, _, _, _ = get_ccy_vol_data(ccy)
        if atm_vols is None:
            st.info("No ATM vols — upload config first")
        else:
            has_atm = ccy in st.session_state.get("atm_prem_matrix", {})
            _av = "ATM Vol (bp)"

            # View radio | Heatmap | Generate | Download  (same layout as FWD matrix)
            _aa1, _aa2, _aa3, _aa4 = st.columns([3, 1, 3, 3])
            with _aa1:
                if has_atm:
                    _av = st.radio("View", ["ATM Vol (bp)", "Forward Premium (bp)", "Vega ($/1bp 100mm)"],
                                   horizontal=False, key="atm_view_toggle")
            with _aa2:
                show_atm_hm = st.checkbox("Heatmap", value=False, key="show_atm_heatmap")
            with _aa3:
                gen_atm = st.button("▶ Generate ATM Matrix", key="gen_atm_matrix",
                                    type="primary", use_container_width=True)
            with _aa4:
                if has_atm:
                    _ad = st.session_state["atm_prem_matrix"][ccy]
                    _adf0 = {"ATM Vol (bp)": _ad["vol"],
                             "Forward Premium (bp)": _ad["prem"],
                             "Vega ($/1bp 100mm)": _ad["vega"]}.get(_av, _ad["vol"])
                st.download_button("⬇ Download",
                                   _adf0.to_csv() if has_atm else "",
                                   f"{ccy}_atm_matrix.csv", key="dl_atm",
                                   use_container_width=True, type="primary",
                                   disabled=not has_atm)

            if gen_atm:
                _mc = st.session_state.get("config_curves", {}).get(ccy)
                _mb = st.session_state.get("config_basis", {}).get(ccy, {}).get("6v3")
                if _mc is None:
                    st.error("No curve — upload config first")
                else:
                    with st.spinner("Calculating..."):
                        pm, vm = calculate_atm_premium_matrix(ccy, _mc, atm_vols, _mb)
                        st.session_state["atm_prem_matrix"][ccy] = {"vol": atm_vols, "prem": pm, "vega": vm}
                    st.rerun()

            if has_atm:
                _ad = st.session_state["atm_prem_matrix"][ccy]
                _adf = {"ATM Vol (bp)": _ad["vol"],
                        "Forward Premium (bp)": _ad["prem"],
                        "Vega ($/1bp 100mm)": _ad["vega"]}.get(_av, _ad["vol"])
                _anc  = [c for c in _adf.columns if c != "Expiry"]
                _afmt = {c: "{:.2f}" for c in _anc}
                if show_atm_hm:
                    st.dataframe(_adf.style.format(_afmt).background_gradient("RdYlGn_r", axis=None, subset=_anc),
                                 use_container_width=True, height=820)
                else:
                    st.dataframe(_adf.style.format(_afmt), use_container_width=True, height=820)
            else:
                st.info("Click **▶ Generate ATM Matrix**")




def check_swap_rates_sanity(rates_dict: dict, floating_rate: str, ccy: str) -> list:
    """
    Sanity check a swap rate curve (tenor→rate dict) for anomalies.
    Flags any adjacent tenor jump > 100bp.
    Returns list of warning strings (empty list = clean).
    """
    warns = []
    if not rates_dict:
        return warns
    try:
        # Sort by numeric maturity
        def _t2y(t):
            t = str(t).strip().upper()
            if t.endswith('M'): return float(t[:-1]) / 12.0
            if t.endswith('Y'): return float(t[:-1])
            try: return float(t)
            except: return 0.0
        sorted_items = sorted(rates_dict.items(), key=lambda x: _t2y(x[0]))
        for i in range(len(sorted_items) - 1):
            t0, r0 = sorted_items[i]
            t1, r1 = sorted_items[i+1]
            delta_bp = abs(float(r1) - float(r0)) * 100.0  # rates in %, delta in bp
            if delta_bp > 100.0:
                warns.append(
                    f"🔴 {ccy} {floating_rate}: {t0}→{t1} jump = {delta_bp:.0f}bp "
                    f"({float(r0):.4f}% → {float(r1):.4f}%)"
                )
    except Exception as e:
        warns.append(f"⚠️ {ccy} {floating_rate} sanity check error: {str(e)}")
    return warns

@st.cache_data(ttl=3600, show_spinner=False)
def _load_swap_rates_from_db(floating_rate: str, load_date: str = None) -> pd.DataFrame:
    """Load swap rates from Supabase swap_rates table for a given date (default: today)."""
    try:
        conn = get_db_connection()
        if conn is None: return pd.DataFrame()
        cur = conn.cursor()
        import datetime as _dt
        if load_date is None:
            load_date = str(_dt.date.today())
        else:
            load_date = str(load_date)
        # Try exact date first; fall back to most recent available date
        cur.execute(
            "SELECT tenor, rate FROM swap_rates WHERE currency=%s AND floating_rate=%s AND date=%s ORDER BY tenor",
            ("AUD", floating_rate, load_date))
        rows = cur.fetchall()
        if not rows:
            # Fallback: most recent date
            cur.execute(
                "SELECT tenor, rate FROM swap_rates WHERE currency=%s AND floating_rate=%s ORDER BY date DESC, tenor LIMIT 30",
                ("AUD", floating_rate))
            rows = cur.fetchall()
        conn.close()
        if not rows: return pd.DataFrame()
        df = pd.DataFrame(rows, columns=["tenor","rate"])
        df["rate"] = df["rate"].astype(float)
        # Run sanity check and store any warnings
        import streamlit as _st
        _rd = dict(zip(df["tenor"], df["rate"]))
        _sw = check_swap_rates_sanity(_rd, floating_rate, "AUD")
        if _sw:
            existing = _st.session_state.get("_swap_load_warnings", [])
            _st.session_state["_swap_load_warnings"] = existing + _sw
        return df.set_index("tenor")
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def _load_swap_rates_history(floating_rate: str, years_back: int = 20) -> pd.DataFrame:
    """Load historical swap rates (date × tenor). Default last 2 years."""
    try:
        conn = get_db_connection()
        if conn is None: return pd.DataFrame()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '15s'")
        cur.execute(
            """SELECT date, tenor, rate FROM swap_rates
               WHERE currency=%s AND floating_rate=%s
                 AND date >= CURRENT_DATE - INTERVAL %s
               ORDER BY date""",
            ("AUD", floating_rate, f"{years_back} years"))
        rows = cur.fetchall()
        conn.close()
        if not rows: return pd.DataFrame()
        df = pd.DataFrame(rows, columns=["date","tenor","rate"])
        df["date"] = pd.to_datetime(df["date"])
        df["rate"] = df["rate"].astype(float)
        return df.pivot_table(index="date", columns="tenor", values="rate", aggfunc="last").sort_index()
    except Exception:
        return pd.DataFrame()


def fwd_analysis_tab():
    """FWD Swap & Basis Historical Analysis tab"""
    st.subheader("📈 FWD IRS Analysis")

    # Load history only when explicitly requested — prevents hang on every render
    if st.button("🔄 Load Swap Rate History", key="fwd_load_history", type="secondary"):
        _load_swap_rates_history.clear()
        st.session_state["_fwd_w3"] = _load_swap_rates_history("3M BBSW")
        st.session_state["_fwd_w6"] = _load_swap_rates_history("6M BBSW")

    _w3 = st.session_state.get("_fwd_w3", pd.DataFrame())
    _w6 = st.session_state.get("_fwd_w6", pd.DataFrame())

    if _w3.empty and _w6.empty:
        st.info("Click **🔄 Load Swap Rate History** to populate charts. Spreads can be saved without loading.")

    def _fwd(wide, start_y, tenor_y):
        end_y = start_y + tenor_y
        def _lbl(y):
            if y == int(y): return f"{int(y)}Y"
            return f"{round(y*12)}M"
        s, e = _lbl(start_y), _lbl(end_y)
        if s not in wide.columns or e not in wide.columns: return None
        return (wide[e] * end_y - wide[s] * start_y) / tenor_y

    _sp_colors = ["#3b82f6","#ef4444","#22c55e","#f59e0b","#a855f7","#06b6d4","#f43f5e","#84cc16"]

    # Convention helper   —   market default: <=3Y = 3M BBSW, >=4Y = 6M BBSW
    def _conv_rate(tenor_y, conv="Market"):
        """Return rate series for a given tenor using selected convention."""
        def _lbl(y): return f"{int(y)}Y" if y == int(y) else f"{round(y*12)}M"
        t = _lbl(tenor_y)
        if conv == "Q/Q (3M BBSW)":
            return _w3[t] if t in _w3.columns else None
        elif conv == "S/S (6M BBSW)":
            return _w6[t] if t in _w6.columns else None
        else:  # Market convention
            if tenor_y <= 3:
                return _w3[t] if t in _w3.columns else None
            else:
                return _w6[t] if t in _w6.columns else None

    def _fwd_conv(start_y, tenor_y, conv="Market"):
        """Fwd-fwd rate using convention-appropriate curve."""
        end_y = start_y + tenor_y
        r_s = _conv_rate(start_y, conv)
        r_e = _conv_rate(end_y, conv)
        if r_s is None or r_e is None: return None
        return (r_e * end_y - r_s * start_y) / tenor_y

    # Convention selector — only relevant when history is loaded
    if not _w3.empty or not _w6.empty:
        _conv = st.radio("Rate Convention", ["Market (≤3Y Q/Q, ≥4Y S/S)", "Q/Q (3M BBSW)", "S/S (6M BBSW)"],
                         horizontal=True, key="fwd_conv")
        _conv_key = "Market" if "Market" in _conv else ("Q/Q (3M BBSW)" if "Q/Q" in _conv else "S/S (6M BBSW)")
        st.caption("3M BBSW: 1Y-3Y | 6M BBSW: 4Y-50Y")
    else:
        _conv_key = st.session_state.get("fwd_conv", "Market (≤3Y Q/Q, ≥4Y S/S)")
        _conv_key = "Market" if "Market" in _conv_key else ("Q/Q (3M BBSW)" if "Q/Q" in _conv_key else "S/S (6M BBSW)")

    _an_tabs = st.tabs(["IRS Spreads", "IRS Butterflies", "Fwd-Fwd Rates (3M)", "6v3 Outright", "6v3 Fwd-Fwd", "6v3 Spreads", "6v3 Butterflies"])

    def _autosave_fwd_prefs():
        """Persist FWD analysis series lists to DB. Debounced — skips if saved in last 5s."""
        if not HAS_POSTGRES or not get_db_url():
            return
        import time as _t
        if _t.time() - st.session_state.get("_fwd_prefs_last_save", 0) < 5:
            return
        _uid = st.session_state.get("username", "default")
        _prefs = {
            "irs_sp_list": [list(x) for x in st.session_state.get("irs_sp_list", [])],
            "irs_fl_list": [list(x) for x in st.session_state.get("irs_fl_list", [])],
            "fvfv_list":   [list(x) for x in st.session_state.get("fvfv_list",   [])],
            "b6_list":     list(st.session_state.get("b6_list", [])),
            "fv6_list":    [list(x) for x in st.session_state.get("fv6_list",    [])],
            "bsp_list":    [list(x) for x in st.session_state.get("bsp_list",    [])],
        }
        save_user_config(_uid, "fwd_analysis_prefs", "GLB", _prefs)
        st.session_state["_fwd_prefs_last_save"] = _t.time()

    def _chart_tools(fig, series_dict: dict, key: str, ylab: str = "bp"):
        """📂 Download + date-range picker + Hi/Lo/Mean/Std/Current stats box."""
        import plotly.io as _pio

        # ── Download button ──────────────────────────────────────
        try:
            _img = _pio.to_image(fig, format="png", width=1400, height=520)
            st.download_button("📂 Copy Chart", _img, f"RateEdge_{key}.png", "image/png",
                               key=f"dl_{key}", use_container_width=False)
        except Exception:
            _html = fig.to_html(include_plotlyjs="cdn", full_html=True)
            st.download_button("📂 Copy Chart", _html.encode(), f"RateEdge_{key}.html", "text/html",
                               key=f"dl_{key}", use_container_width=False)

        if not series_dict:
            return

        # ── Date range for stats window ──────────────────────────
        _all_idx = pd.DatetimeIndex([])
        for _s in series_dict.values():
            if hasattr(_s, "index") and not _s.empty:
                _all_idx = _all_idx.union(_s.index)
        if _all_idx.empty:
            return
        _min_d = _all_idx.min().date()
        _max_d = _all_idx.max().date()

        st.markdown("**📊 Range Stats**")
        _rc1, _rc2, _rc3 = st.columns([2, 2, 2])
        with _rc1:
            _pt1 = st.date_input("From", value=_min_d, min_value=_min_d, max_value=_max_d,
                                 key=f"pt1_{key}", format="DD/MM/YYYY")
        with _rc2:
            _pt2 = st.date_input("To", value=_max_d, min_value=_min_d, max_value=_max_d,
                                 key=f"pt2_{key}", format="DD/MM/YYYY")
        with _rc3:
            st.markdown(f"<div style='padding-top:30px;color:#94a3b8;font-size:0.8rem'>"
                        f"{_pt1} → {_pt2}</div>", unsafe_allow_html=True)

        _t1 = pd.Timestamp(_pt1)
        _t2 = pd.Timestamp(_pt2) + pd.Timedelta(days=1)
        _rows = []
        for _name, _ser in series_dict.items():
            _w = _ser[(_ser.index >= _t1) & (_ser.index < _t2)].dropna()
            if _w.empty:
                continue
            _cur = _ser.dropna()
            _rows.append({
                "Series": _name,
                f"Hi": round(_w.max(), 4),
                f"Lo": round(_w.min(), 4),
                f"Mean": round(_w.mean(), 4),
                f"Std": round(_w.std(), 4),
                f"Current": round(_cur.iloc[-1], 4) if not _cur.empty else None,
                f"vs Mean": round(_cur.iloc[-1] - _w.mean(), 4) if not _cur.empty else None,
            })
        if _rows:
            _sdf = pd.DataFrame(_rows).set_index("Series")
            st.dataframe(_sdf.style.format("{:.4f}", na_rep="  —  "),
                         use_container_width=True, height=min(38 + 38*len(_rows), 280))

    # Always use standard tenor set — never depends on DB data being loaded
    _STANDARD_TENORS = [1,2,3,4,5,6,7,8,9,10,12,15,20,25,30,40,50]
    _yr_tenors = sorted(list(set(
        _STANDARD_TENORS +
        [int(c[:-1]) for c in _w3.columns if c.endswith("Y") and c[:-1].isdigit()] +
        [int(c[:-1]) for c in _w6.columns if c.endswith("Y") and c[:-1].isdigit()]
    )))
    _tn_opts = [f"{y}Y" for y in _yr_tenors]

    # Fwd start/tenor options
    _fwd_starts = [1,2,3,4,5,7,10,12,15,20]
    _fwd_tenors = [1,2,3,5,7,10]

    def _fig_layout(fig, cut, ylab):
        fig.update_layout(
            height=460, margin=dict(l=50,r=20,t=40,b=40),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(15,23,42,0.8)",
            legend=dict(orientation="h", y=1.06, font=dict(color="#e2e8f0", size=12)),
            yaxis_title=ylab,
            xaxis=dict(gridcolor="#334155", color="#94a3b8", range=[cut, pd.Timestamp.now()]),
            yaxis=dict(gridcolor="#334155", color="#94a3b8"),
            font=dict(color="#94a3b8"),
        )

    def _add_series(fig, label, series, color, bands=False):
        fig.add_trace(go.Scatter(x=series.index, y=series.values, mode="lines",
            name=label, line=dict(color=color, width=1.8)))
        if bands:
            mu, sd = series.mean(), series.std()
            fig.add_hline(y=mu, line=dict(color=color, dash="dash", width=1), opacity=0.5)
            fig.add_hrect(y0=mu-sd, y1=mu+sd, fillcolor=color, opacity=0.06, line_width=0)
        else:
            fig.add_hline(y=series.mean(), line=dict(color=color, dash="dot", width=1), opacity=0.4)

    # ── TAB 1: IRS SPREADS ──────────────────────────────────────
    with _an_tabs[0]:
        st.markdown("#### IRS Curve Spreads")
        # Init active spreads list
        if "irs_sp_list" not in st.session_state:
            st.session_state["irs_sp_list"] = []  # DB populates; empty = no defaults

        # Builder row
        bc1, bc2, bc3, bc4 = st.columns([1.2, 1.2, 0.8, 1.8])
        with bc1:
            _sp_l1 = st.selectbox("Leg 1 (short)", _tn_opts, index=_tn_opts.index("2Y") if "2Y" in _tn_opts else 0, key="sp_l1")
        with bc2:
            _sp_l2_default = _tn_opts.index("10Y") if "10Y" in _tn_opts else min(4, len(_tn_opts)-1)
            _sp_l2 = st.selectbox("Leg 2 (long)", _tn_opts, index=_sp_l2_default, key="sp_l2")
        with bc3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _sp_add_clicked = st.button("➕ Add", key="sp_add", use_container_width=True)
        with bc4:
            rc1, rc2 = st.columns([3,1])
            with rc1:
                _sp_remove = st.selectbox("Remove spread", ["  —  "] + [f"{a} → {b}" for a,b in st.session_state["irs_sp_list"]], key="sp_rm")
            with rc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("➖", key="sp_rm_btn", use_container_width=True) and _sp_remove != "  —  ":
                    _rm_parts = _sp_remove.split(" → ")
                    if len(_rm_parts)==2 and (_rm_parts[0],_rm_parts[1]) in st.session_state["irs_sp_list"]:
                        st.session_state["irs_sp_list"].remove((_rm_parts[0],_rm_parts[1]))
                        _autosave_fwd_prefs()

        if _sp_add_clicked:
            _l1 = st.session_state.get("sp_l1", _sp_l1)
            _l2 = st.session_state.get("sp_l2", _sp_l2)
            if _l1 == _l2:
                st.warning("Leg 1 and Leg 2 must be different tenors.")
            elif (_l1, _l2) in st.session_state["irs_sp_list"]:
                st.warning(f"{_l1} → {_l2} is already in the list.")
            else:
                st.session_state["irs_sp_list"].append((_l1, _l2))
                _autosave_fwd_prefs()
                st.success(f"✅ Added {_l1} → {_l2}")
                # No rerun — avoids retriggering the history load

        c1, c2, c3 = st.columns(3)
        with c1: _sp_yr = st.slider("History (years)", 1, 8, 5, key="sp_yr")
        with c2: _sp_bands = st.checkbox("Mean ± 1σ bands", True, key="sp_bands")

        _cut = pd.Timestamp.now() - pd.DateOffset(years=_sp_yr)
        _fig = go.Figure()
        _sp_series = {}
        _no_data_spreads = []
        for _a, _b in st.session_state["irs_sp_list"]:
            try:
                _ay = int(_a[:-1]); _by = int(_b[:-1])
            except (ValueError, IndexError):
                _no_data_spreads.append(f"{_a} → {_b}")
                continue
            _ra = _conv_rate(_ay, _conv_key); _rb = _conv_rate(_by, _conv_key)
            if _ra is None or _rb is None:
                _no_data_spreads.append(f"{_a} → {_b}")
                continue
            _sr = (_rb - _ra).dropna()
            _sr = _sr[_sr.index >= _cut] * 100
            if not _sr.empty:
                _sp_series[f"{_a} → {_b}"] = _sr
            else:
                _no_data_spreads.append(f"{_a} → {_b}")

        _sp_keys = list(_sp_series.keys())
        with c3:
            _sp_as_spread = st.checkbox("Show as spread", False, key="sp_as_spread")
        if _sp_as_spread and len(_sp_keys) >= 2:
            _sc1, _sc2 = st.columns(2)
            with _sc1:
                _sp_s1 = st.selectbox("Series A", _sp_keys, index=0, key="sp_s1")
            with _sc2:
                _sp_s2_opts = [k for k in _sp_keys if k != _sp_s1]
                _sp_s2 = st.selectbox("Series B (subtract)", _sp_s2_opts, index=0, key="sp_s2") if _sp_s2_opts else None
            if _sp_s2 and _sp_s1 in _sp_series and _sp_s2 in _sp_series:
                _cmb = (_sp_series[_sp_s1] - _sp_series[_sp_s2]).dropna()
                _fig.add_trace(go.Scatter(x=_cmb.index, y=_cmb.values, mode="lines",
                    name=f"{_sp_s1}  →  {_sp_s2}", line=dict(color=_sp_colors[0], width=1.8)))
                _fig.add_hline(y=_cmb.mean(), line=dict(color="#94a3b8", dash="dash", width=1))
                _sp_active = {f"{_sp_s1}  →  {_sp_s2}": _cmb}
            else:
                _sp_as_spread = False
                _sp_active = _sp_series
        else:
            _sp_active = _sp_series

        if not _sp_as_spread:
            for _i, (_lbl, _sr) in enumerate(_sp_series.items()):
                _add_series(_fig, _lbl, _sr, _sp_colors[_i % len(_sp_colors)], _sp_bands)

        if _sp_series:
            _fig_layout(_fig, _cut, "Spread (bp)")
            st.plotly_chart(_fig, use_container_width=True)
            _chart_tools(_fig, _sp_active, "sp", "bp")
        if _no_data_spreads:
            st.info(f"✅ Saved — no historical data yet to chart: {', '.join(_no_data_spreads)}")
        if not _sp_series and not _no_data_spreads and st.session_state["irs_sp_list"]:
            st.info("✅ Spreads saved. No historical swap rate data in DB yet — load via FWD Analysis or swap_rates table.")

    # ── TAB 2: IRS BUTTERFLIES ──────────────────────────────────
    with _an_tabs[1]:
        st.markdown("#### IRS Rate Butterflies")
        if "irs_fl_list" not in st.session_state:
            st.session_state["irs_fl_list"] = []  # DB populates

        bc1,bc2,bc3,bc4,bc5 = st.columns([1,1,1,0.7,1.5])
        with bc1: _fl_w = st.selectbox("Wing 1", _tn_opts, index=_tn_opts.index("2Y") if "2Y" in _tn_opts else 0, key="fl_w")
        with bc2: _fl_m = st.selectbox("Body", _tn_opts, index=_tn_opts.index("5Y") if "5Y" in _tn_opts else 2, key="fl_m")
        with bc3: _fl_e = st.selectbox("Wing 2", _tn_opts, index=_tn_opts.index("10Y") if "10Y" in _tn_opts else 4, key="fl_e")
        with bc4:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _fl_add_clicked = st.button("➕ Add", key="fl_add", use_container_width=True)
        with bc5:
            rc1, rc2 = st.columns([3,1])
            with rc1:
                _fl_rm = st.selectbox("Remove", ["  —  "]+[f"{w}/{m}/{e}" for w,m,e in st.session_state["irs_fl_list"]], key="fl_rm")
            with rc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("➖", key="fl_rm_btn", use_container_width=True) and _fl_rm != "  —  ":
                    _rp = _fl_rm.split("/")
                    if len(_rp)==3 and tuple(_rp) in st.session_state["irs_fl_list"]:
                        st.session_state["irs_fl_list"].remove(tuple(_rp))
                        _autosave_fwd_prefs()

        if _fl_add_clicked:
            _fw = st.session_state.get("fl_w", _fl_w)
            _fm = st.session_state.get("fl_m", _fl_m)
            _fe = st.session_state.get("fl_e", _fl_e)
            if len({_fw,_fm,_fe}) < 3:
                st.warning("Wing 1, Body and Wing 2 must all be different tenors.")
            elif (_fw,_fm,_fe) in st.session_state["irs_fl_list"]:
                st.warning(f"{_fw}/{_fm}/{_fe} is already in the list.")
            else:
                st.session_state["irs_fl_list"].append((_fw,_fm,_fe))
                _autosave_fwd_prefs()

        c1,c2,c3 = st.columns(3)
        with c1: _fl_yr = st.slider("History (years)",1,8,5,key="fl_yr")

        _cut_fl = pd.Timestamp.now() - pd.DateOffset(years=_fl_yr)
        _fig_fl = go.Figure()
        _fl_series = {}
        for _fw,_fm,_fe in st.session_state["irs_fl_list"]:
            _wy=int(_fw[:-1]); _my=int(_fm[:-1]); _ey=int(_fe[:-1])
            _rw=_conv_rate(_wy,_conv_key); _rm=_conv_rate(_my,_conv_key); _re=_conv_rate(_ey,_conv_key)
            if _rw is None or _rm is None or _re is None: continue
            _fly = (_rm - 0.5*(_rw+_re)).dropna()
            _fly = _fly[_fly.index>=_cut_fl]*100
            _fl_series[f"{_fw}/{_fm}/{_fe}"] = _fly

        _fl_keys = list(_fl_series.keys())
        with c2:
            _fl_as_spread = st.checkbox("Show as spread", False, key="fl_as_spread")
        if _fl_as_spread and len(_fl_keys) >= 2:
            _fc1, _fc2 = st.columns(2)
            with _fc1:
                _fl_s1 = st.selectbox("Series A", _fl_keys, index=0, key="fl_s1")
            with _fc2:
                _fl_s2_opts = [k for k in _fl_keys if k != _fl_s1]
                _fl_s2 = st.selectbox("Series B (subtract)", _fl_s2_opts, index=0, key="fl_s2") if _fl_s2_opts else None
            if _fl_s2 and _fl_s1 in _fl_series and _fl_s2 in _fl_series:
                _cmb = (_fl_series[_fl_s1] - _fl_series[_fl_s2]).dropna()
                _fig_fl.add_trace(go.Scatter(x=_cmb.index,y=_cmb.values,mode="lines",
                    name=f"{_fl_s1}  →  {_fl_s2}",line=dict(color=_sp_colors[0],width=1.8)))
                _fig_fl.add_hline(y=_cmb.mean(),line=dict(color="#94a3b8",dash="dash",width=1))
                _fl_active = {f"{_fl_s1}  →  {_fl_s2}": _cmb}
            else:
                _fl_as_spread = False
                _fl_active = _fl_series
        else:
            _fl_active = _fl_series

        if not _fl_as_spread:
            for _i,(_lbl,_fly) in enumerate(_fl_series.items()):
                _add_series(_fig_fl, _lbl, _fly, _sp_colors[_i%len(_sp_colors)])

        _fig_fl.add_hline(y=0, line=dict(color="#64748b",width=1))
        _fig_layout(_fig_fl, _cut_fl, "Fly (bp)")
        st.plotly_chart(_fig_fl, use_container_width=True)
        _chart_tools(_fig_fl, _fl_active, "fl", "bp")

    # ── TAB 3: FWD-FWD RATES ────────────────────────────────────
    with _an_tabs[2]:
        st.markdown("#### Forward-Forward Swap Rates")
        if "fvfv_list" not in st.session_state:
            st.session_state["fvfv_list"] = []  # DB populates

        bc1,bc2,bc3,bc4 = st.columns([1,1,0.7,1.5])
        with bc1: _fv_st = st.selectbox("Start (years)", _fwd_starts, index=_fwd_starts.index(2) if 2 in _fwd_starts else 0, key="fv_st")
        with bc2: _fv_tn = st.selectbox("Tenor (years)", _fwd_tenors, index=_fwd_tenors.index(2) if 2 in _fwd_tenors else 0, key="fv_tn")
        with bc3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _fv_add_clicked = st.button("➕ Add", key="fv_add", use_container_width=True)
        with bc4:
            rc1, rc2 = st.columns([3,1])
            with rc1:
                _fv_rm = st.selectbox("Remove", ["  —  "]+[f"{s}y{t}y" for s,t in st.session_state["fvfv_list"]], key="fv_rm")
            with rc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("➖", key="fv_rm_btn", use_container_width=True) and _fv_rm != "  —  ":
                    _rp = _fv_rm[:-1].split("y")
                    if len(_rp)==2:
                        try:
                            _rs,_rt = int(_rp[0]),int(_rp[1])
                            if (_rs,_rt) in st.session_state["fvfv_list"]:
                                st.session_state["fvfv_list"].remove((_rs,_rt))
                                _autosave_fwd_prefs()
                        except: pass

        if _fv_add_clicked:
            _fvs = st.session_state.get("fv_st", _fv_st)
            _fvt = st.session_state.get("fv_tn", _fv_tn)
            if (_fvs, _fvt) in st.session_state["fvfv_list"]:
                st.warning(f"{_fvs}y{_fvt}y is already in the list.")
            else:
                st.session_state["fvfv_list"].append((_fvs, _fvt))
                _autosave_fwd_prefs()

        c1,c2,c3 = st.columns(3)
        with c1: _fv_yr = st.slider("History (years)",1,8,5,key="fv_yr")

        _cut_fv = pd.Timestamp.now() - pd.DateOffset(years=_fv_yr)
        _fig_fv = go.Figure()
        _fv_series = {}
        for _s,_t in st.session_state["fvfv_list"]:
            _r = _fwd_conv(_s,_t,_conv_key)
            if _r is not None:
                _fv_series[f"{_s}y{_t}y"] = _r[_r.index>=_cut_fv].dropna()

        _fv_keys = list(_fv_series.keys())
        with c2:
            _fv_as_spread = st.checkbox("Show as spread", False, key="fv_sprd")
        if _fv_as_spread and len(_fv_keys) >= 2:
            _vc1, _vc2 = st.columns(2)
            with _vc1:
                _fv_s1 = st.selectbox("Series A", _fv_keys, index=0, key="fv_s1")
            with _vc2:
                _fv_s2_opts = [k for k in _fv_keys if k != _fv_s1]
                _fv_s2 = st.selectbox("Series B (subtract)", _fv_s2_opts, index=0, key="fv_s2") if _fv_s2_opts else None
            if _fv_s2 and _fv_s1 in _fv_series and _fv_s2 in _fv_series:
                _cmb=(_fv_series[_fv_s1]-_fv_series[_fv_s2]).dropna()*100
                _fig_fv.add_trace(go.Scatter(x=_cmb.index,y=_cmb.values,mode="lines",
                    name=f"{_fv_s1}  →  {_fv_s2}",line=dict(color=_sp_colors[0],width=1.8)))
                _fig_fv.add_hline(y=_cmb.mean(),line=dict(color="#94a3b8",dash="dash",width=1))
                _fig_layout(_fig_fv, _cut_fv, "Spread (bp)")
                _fv_active = {f"{_fv_s1}  →  {_fv_s2}": _cmb}
            else:
                _fv_as_spread = False
                _fv_active = _fv_series
        else:
            _fv_active = _fv_series

        if not _fv_as_spread:
            for _i,(_l,_s) in enumerate(_fv_series.items()):
                _add_series(_fig_fv, _l, _s, _sp_colors[_i%len(_sp_colors)])
            _fig_layout(_fig_fv, _cut_fv, "Rate (%)")
        st.plotly_chart(_fig_fv, use_container_width=True)
        _chart_tools(_fig_fv, _fv_active, "fv", "%")
    with _an_tabs[3]:
        st.markdown("#### 6v3 Basis   —   Outright (6M BBSW  →  3M BBSW)")
        _com6v3 = sorted([c for c in _w6.columns if c in _w3.columns and c.endswith("Y")],
                          key=lambda x: int(x[:-1]))
        if not _com6v3:
            st.info("No overlapping tenors between 3M and 6M BBSW.")
        else:
            if "b6_list" not in st.session_state:
                st.session_state["b6_list"] = []  # DB populates
            bc1,bc2,bc3 = st.columns([1.5,0.7,1.5])
            with bc1: _b6_add_tn = st.selectbox("Add tenor", [t for t in _com6v3 if t not in st.session_state["b6_list"]] or _com6v3, key="b6_add_tn")
            with bc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                _b6_add_clicked = st.button("➕ Add", key="b6_add", use_container_width=True)
            with bc3:
                rc1, rc2 = st.columns([3,1])
                with rc1:
                    _b6_rm = st.selectbox("Remove", ["  —  "]+st.session_state["b6_list"], key="b6_rm")
                with rc2:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    if st.button("➖", key="b6_rm_btn", use_container_width=True) and _b6_rm != "  —  " and _b6_rm in st.session_state["b6_list"]:
                        st.session_state["b6_list"].remove(_b6_rm)
                        _autosave_fwd_prefs()

            if _b6_add_clicked:
                _btn = st.session_state.get("b6_add_tn", _b6_add_tn)
                if _btn not in st.session_state["b6_list"]:
                    st.session_state["b6_list"].append(_btn)
                    _autosave_fwd_prefs()

            c1,c2 = st.columns(2)
            with c1: _b6_yr = st.slider("History (years)",1,8,5,key="b6_yr")
            _cut_b6 = pd.Timestamp.now() - pd.DateOffset(years=_b6_yr)
            _fig_b6 = go.Figure()
            _b6_series = {}
            for _i,_tn in enumerate(st.session_state["b6_list"]):
                if _tn not in _w6.columns or _tn not in _w3.columns: continue
                _b6 = (_w6[_tn]-_w3[_tn]).dropna()
                _b6 = _b6[_b6.index>=_cut_b6]*100
                _b6_series[f"{_tn} 6v3"] = _b6
                _add_series(_fig_b6, f"{_tn} 6v3", _b6, _sp_colors[_i%len(_sp_colors)])
            _fig_b6.add_hline(y=0,line=dict(color="#64748b",width=1))
            _fig_layout(_fig_b6, _cut_b6, "6v3 Basis (bp)")
            st.plotly_chart(_fig_b6, use_container_width=True)
            _chart_tools(_fig_b6, _b6_series, "b6", "bp")

    # ── TAB 5: 6v3 FWD-FWD ─────────────────────────────────────
    with _an_tabs[4]:
        st.markdown("#### 6v3 Forward-Forward Basis")
        st.caption("Fwd-fwd 6M BBSW  →  fwd-fwd 3M BBSW for same start/tenor")
        if "fv6_list" not in st.session_state:
            st.session_state["fv6_list"] = []  # DB populates

        bc1,bc2,bc3,bc4 = st.columns([1,1,0.7,1.5])
        with bc1: _fv6_st = st.selectbox("Start (years)", _fwd_starts, index=_fwd_starts.index(2) if 2 in _fwd_starts else 0, key="fv6_st")
        with bc2: _fv6_tn = st.selectbox("Tenor (years)", _fwd_tenors, index=_fwd_tenors.index(2) if 2 in _fwd_tenors else 0, key="fv6_tn")
        with bc3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _fv6_add_clicked = st.button("➕ Add", key="fv6_add", use_container_width=True)
        with bc4:
            rc1, rc2 = st.columns([3,1])
            with rc1:
                _fv6_rm = st.selectbox("Remove", ["  —  "]+[f"{s}y{t}y" for s,t in st.session_state["fv6_list"]], key="fv6_rm")
            with rc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("➖", key="fv6_rm_btn", use_container_width=True) and _fv6_rm != "  —  ":
                    _rp = _fv6_rm[:-1].split("y")
                    if len(_rp)==2:
                        try:
                            _rs,_rt=int(_rp[0]),int(_rp[1])
                            if (_rs,_rt) in st.session_state["fv6_list"]:
                                st.session_state["fv6_list"].remove((_rs,_rt))
                                _autosave_fwd_prefs()
                        except: pass

        if _fv6_add_clicked:
            _fv6s = st.session_state.get("fv6_st", _fv6_st)
            _fv6t = st.session_state.get("fv6_tn", _fv6_tn)
            if (_fv6s, _fv6t) in st.session_state["fv6_list"]:
                st.warning(f"{_fv6s}y{_fv6t}y is already in the list.")
            else:
                st.session_state["fv6_list"].append((_fv6s, _fv6t))
                _autosave_fwd_prefs()

        c1,c2 = st.columns(2)
        with c1: _fv6_yr = st.slider("History (years)",1,8,5,key="fv6_yr")

        _cut_fv6 = pd.Timestamp.now() - pd.DateOffset(years=_fv6_yr)
        _fig_fv6 = go.Figure()
        _fv6_series = {}
        for _s,_t in st.session_state["fv6_list"]:
            _r6 = _fwd(_w6,_s,_t); _r3 = _fwd(_w3,_s,_t)
            if _r6 is not None and _r3 is not None:
                _b = (_r6-_r3).dropna()
                _fv6_series[f"{_s}y{_t}y 6v3"] = _b[_b.index>=_cut_fv6]*100

        _fv6_keys = list(_fv6_series.keys())
        with c2:
            _fv6_as_spread = st.checkbox("Show as spread", False, key="fv6_as_spread")
        if _fv6_as_spread and len(_fv6_keys) >= 2:
            _v6c1, _v6c2 = st.columns(2)
            with _v6c1:
                _fv6_s1 = st.selectbox("Series A", _fv6_keys, index=0, key="fv6_s1")
            with _v6c2:
                _fv6_s2_opts = [k for k in _fv6_keys if k != _fv6_s1]
                _fv6_s2 = st.selectbox("Series B (subtract)", _fv6_s2_opts, index=0, key="fv6_s2") if _fv6_s2_opts else None
            if _fv6_s2 and _fv6_s1 in _fv6_series and _fv6_s2 in _fv6_series:
                _cmb=(_fv6_series[_fv6_s1]-_fv6_series[_fv6_s2]).dropna()
                _fig_fv6.add_trace(go.Scatter(x=_cmb.index,y=_cmb.values,mode="lines",
                    name=f"{_fv6_s1}  →  {_fv6_s2}",line=dict(color=_sp_colors[0],width=1.8)))
                _fig_fv6.add_hline(y=_cmb.mean(),line=dict(color="#94a3b8",dash="dash",width=1))
                _fv6_active = {f"{_fv6_s1}  →  {_fv6_s2}": _cmb}
            else:
                _fv6_as_spread = False
                _fv6_active = _fv6_series
        else:
            _fv6_active = _fv6_series

        if not _fv6_as_spread:
            for _i,(_l,_b) in enumerate(_fv6_series.items()):
                _add_series(_fig_fv6, _l, _b, _sp_colors[_i%len(_sp_colors)])
        _fig_fv6.add_hline(y=0,line=dict(color="#64748b",width=1))
        _fig_layout(_fig_fv6, _cut_fv6, "6v3 Fwd-Fwd Basis (bp)")
        st.plotly_chart(_fig_fv6, use_container_width=True)
        _chart_tools(_fig_fv6, _fv6_active, "fv6", "bp")
    with _an_tabs[5]:
        st.markdown("#### 6v3 Basis Spreads")
        _com6v3_sp = sorted([c for c in _w6.columns if c in _w3.columns and c.endswith("Y")],
                              key=lambda x: int(x[:-1]))
        if len(_com6v3_sp) < 2:
            st.info("Need at least 2 overlapping tenors.")
        else:
            if "bsp_list" not in st.session_state:
                st.session_state["bsp_list"] = []  # DB populates

            bc1,bc2,bc3,bc4 = st.columns([1.2,1.2,0.7,1.5])
            with bc1: _bsp_l1 = st.selectbox("Leg 1 (6v3 tenor)", _com6v3_sp, index=0, key="bsp_l1")
            with bc2: _bsp_l2 = st.selectbox("Leg 2 (6v3 tenor)", _com6v3_sp, index=min(2,len(_com6v3_sp)-1), key="bsp_l2")
            with bc3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                _bsp_add_clicked = st.button("➕ Add", key="bsp_add", use_container_width=True)
            with bc4:
                rc1, rc2 = st.columns([3,1])
                with rc1:
                    _bsp_rm = st.selectbox("Remove", ["  —  "]+[f"{a} → {b}" for a,b in st.session_state["bsp_list"]], key="bsp_rm")
                with rc2:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    if st.button("➖", key="bsp_rm_btn", use_container_width=True) and _bsp_rm != "  —  ":
                        _rp=_bsp_rm.split(" → ")
                        if len(_rp)==2 and tuple(_rp) in st.session_state["bsp_list"]:
                            st.session_state["bsp_list"].remove(tuple(_rp))
                            _autosave_fwd_prefs()

            if _bsp_add_clicked:
                _bl1 = st.session_state.get("bsp_l1", _bsp_l1)
                _bl2 = st.session_state.get("bsp_l2", _bsp_l2)
                if _bl1 == _bl2:
                    st.warning("Leg 1 and Leg 2 must be different tenors.")
                elif (_bl1, _bl2) in st.session_state["bsp_list"]:
                    st.warning(f"{_bl1} → {_bl2} is already in the list.")
                else:
                    st.session_state["bsp_list"].append((_bl1, _bl2))
                    _autosave_fwd_prefs()

            c1,c2,c3 = st.columns(3)
            with c1: _bsp_yr = st.slider("History (years)",1,8,5,key="bsp_yr")
            with c2: _bsp_as_spread = st.checkbox("Show as spread", False, key="bsp_as_spread")

            _cut_bsp = pd.Timestamp.now() - pd.DateOffset(years=_bsp_yr)
            _fig_bsp = go.Figure()
            _bsp_series = {}
            for _a,_b in st.session_state["bsp_list"]:
                if _a not in _w6.columns or _a not in _w3.columns: continue
                if _b not in _w6.columns or _b not in _w3.columns: continue
                _ba=(_w6[_a]-_w3[_a]).dropna()*100
                _bb=(_w6[_b]-_w3[_b]).dropna()*100
                _bsprd=(_ba-_bb).dropna()
                _bsp_series[f"{_a} → {_b} 6v3 sprd"] = _bsprd[_bsprd.index>=_cut_bsp]

            _bsp_keys = list(_bsp_series.keys())
            if _bsp_as_spread and len(_bsp_keys) >= 2:
                _bc1, _bc2 = st.columns(2)
                with _bc1:
                    _bsp_s1 = st.selectbox("Series A", _bsp_keys, index=0, key="bsp_s1")
                with _bc2:
                    _bsp_s2_opts = [k for k in _bsp_keys if k != _bsp_s1]
                    _bsp_s2 = st.selectbox("Series B (subtract)", _bsp_s2_opts, index=0, key="bsp_s2") if _bsp_s2_opts else None
                if _bsp_s2 and _bsp_s1 in _bsp_series and _bsp_s2 in _bsp_series:
                    _cmb=(_bsp_series[_bsp_s1]-_bsp_series[_bsp_s2]).dropna()
                    _fig_bsp.add_trace(go.Scatter(x=_cmb.index,y=_cmb.values,mode="lines",
                        name=f"{_bsp_s1}  →  {_bsp_s2}",line=dict(color=_sp_colors[0],width=1.8)))
                    _fig_bsp.add_hline(y=_cmb.mean(),line=dict(color="#94a3b8",dash="dash",width=1))
                    _bsp_active = {f"{_bsp_s1}  →  {_bsp_s2}": _cmb}
                else:
                    _bsp_as_spread = False
                    _bsp_active = _bsp_series
            else:
                _bsp_active = _bsp_series

            if not _bsp_as_spread:
                for _i,(_lbl,_bsprd) in enumerate(_bsp_series.items()):
                    _add_series(_fig_bsp, _lbl, _bsprd, _sp_colors[_i%len(_sp_colors)])
            _fig_bsp.add_hline(y=0,line=dict(color="#64748b",width=1))
            _fig_layout(_fig_bsp, _cut_bsp, "6v3 Spread (bp)")
            st.plotly_chart(_fig_bsp, use_container_width=True)
            _chart_tools(_fig_bsp, _bsp_active, "bsp", "bp")

    # ── TAB 7: 6v3 BUTTERFLIES ──────────────────────────────────
    with _an_tabs[6]:
        st.markdown("#### 6v3 Basis Butterflies")
        st.caption("Fly = 6v3(body)  →  0.5≈[6v3(wing1) + 6v3(wing2)]")
        _com6v3_bfly = sorted([c for c in _w6.columns if c in _w3.columns and c.endswith("Y")],
                               key=lambda x: int(x[:-1]))
        if len(_com6v3_bfly) < 3:
            st.info("Need at least 3 overlapping tenors.")
        else:
            if "b6bfly_list" not in st.session_state:
                st.session_state["b6bfly_list"] = []

            bc1,bc2,bc3,bc4,bc5 = st.columns([1,1,1,0.7,1.5])
            with bc1: _b6bfly_w1 = st.selectbox("Wing 1", _com6v3_bfly, index=0, key="b6bfly_w1")
            with bc2: _b6bfly_bd = st.selectbox("Body",   _com6v3_bfly, index=min(2,len(_com6v3_bfly)-1), key="b6bfly_bd")
            with bc3: _b6bfly_w2 = st.selectbox("Wing 2", _com6v3_bfly, index=min(4,len(_com6v3_bfly)-1), key="b6bfly_w2")
            with bc4:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                _b6bfly_add = st.button("➕ Add", key="b6bfly_add", use_container_width=True)
            with bc5:
                rc1,rc2 = st.columns([3,1])
                with rc1:
                    _b6bfly_rm = st.selectbox("Remove", ["  —  "]+[f"{w}/{m}/{e}" for w,m,e in st.session_state["b6bfly_list"]], key="b6bfly_rm")
                with rc2:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    if st.button("➖", key="b6bfly_rm_btn", use_container_width=True) and _b6bfly_rm != "  —  ":
                        _rp = _b6bfly_rm.split("/")
                        if len(_rp)==3 and tuple(_rp) in st.session_state["b6bfly_list"]:
                            st.session_state["b6bfly_list"].remove(tuple(_rp))
                            st.rerun()

            if _b6bfly_add:
                _bw1 = st.session_state.get("b6bfly_w1", _b6bfly_w1)
                _bbd = st.session_state.get("b6bfly_bd", _b6bfly_bd)
                _bw2 = st.session_state.get("b6bfly_w2", _b6bfly_w2)
                if len({_bw1,_bbd,_bw2}) < 3:
                    st.warning("Wing 1, Body and Wing 2 must all be different tenors.")
                elif (_bw1,_bbd,_bw2) in st.session_state["b6bfly_list"]:
                    st.warning(f"{_bw1}/{_bbd}/{_bw2} already in list.")
                else:
                    st.session_state["b6bfly_list"].append((_bw1,_bbd,_bw2))
                    st.rerun()

            c1,c2,c3 = st.columns(3)
            with c1: _b6bfly_yr = st.slider("History (years)",1,8,5,key="b6bfly_yr")
            with c2: _b6bfly_as_spread = st.checkbox("Show as spread", False, key="b6bfly_as_spread")

            _cut_b6bfly = pd.Timestamp.now() - pd.DateOffset(years=_b6bfly_yr)
            _fig_b6bfly = go.Figure()
            _b6bfly_series = {}

            for _bw1,_bbd,_bw2 in st.session_state["b6bfly_list"]:
                for _tn in [_bw1,_bbd,_bw2]:
                    if _tn not in _w6.columns or _tn not in _w3.columns:
                        break
                else:
                    _b6_w1 = (_w6[_bw1]-_w3[_bw1]).dropna()*100
                    _b6_bd = (_w6[_bbd]-_w3[_bbd]).dropna()*100
                    _b6_w2 = (_w6[_bw2]-_w3[_bw2]).dropna()*100
                    _bfly = (_b6_bd - 0.5*(_b6_w1+_b6_w2)).dropna()
                    _b6bfly_series[f"{_bw1}/{_bbd}/{_bw2}"] = _bfly[_bfly.index>=_cut_b6bfly]

            _b6bfly_keys = list(_b6bfly_series.keys())
            if _b6bfly_as_spread and len(_b6bfly_keys) >= 2:
                _bbc1,_bbc2 = st.columns(2)
                with _bbc1:
                    _bbs1 = st.selectbox("Series A", _b6bfly_keys, index=0, key="b6bfly_s1")
                with _bbc2:
                    _bbs2_opts = [k for k in _b6bfly_keys if k != _bbs1]
                    _bbs2 = st.selectbox("Series B (subtract)", _bbs2_opts, index=0, key="b6bfly_s2") if _bbs2_opts else None
                if _bbs2 and _bbs1 in _b6bfly_series and _bbs2 in _b6bfly_series:
                    _cmb = (_b6bfly_series[_bbs1]-_b6bfly_series[_bbs2]).dropna()
                    _fig_b6bfly.add_trace(go.Scatter(x=_cmb.index,y=_cmb.values,mode="lines",
                        name=f"{_bbs1}  →  {_bbs2}",line=dict(color=_sp_colors[0],width=1.8)))
                    _fig_b6bfly.add_hline(y=_cmb.mean(),line=dict(color="#94a3b8",dash="dash",width=1))
                    _b6bfly_active = {f"{_bbs1}  →  {_bbs2}": _cmb}
                else:
                    _b6bfly_as_spread = False
                    _b6bfly_active = _b6bfly_series
            else:
                _b6bfly_active = _b6bfly_series

            if not _b6bfly_as_spread:
                for _i,(_lbl,_bfly) in enumerate(_b6bfly_series.items()):
                    _add_series(_fig_b6bfly, _lbl, _bfly, _sp_colors[_i%len(_sp_colors)])

            _fig_b6bfly.add_hline(y=0,line=dict(color="#64748b",width=1))
            _fig_layout(_fig_b6bfly, _cut_b6bfly, "6v3 Fly (bp)")
            st.plotly_chart(_fig_b6bfly, use_container_width=True)
            _chart_tools(_fig_b6bfly, _b6bfly_active, "b6bfly", "bp")

def _build_aud_par_splines():
    """
    Build AUD blended QQ and SS par-rate cubic splines from session state.
    QQ curve: <=3Y direct from 3M BBSW, >3Y = SS par - 6v3 basis
    SS curve: >=4Y direct from 6M BBSW, <4Y = QQ par + 6v3 basis
    Returns (cs_qq, cs_ss) or (None, None) if data unavailable.
    """
    from scipy.interpolate import CubicSpline as _CS
    par_qq = st.session_state.get("_aud_par_qq", {})
    par_ss = st.session_state.get("_aud_par_ss", {})
    basis_6v3 = st.session_state.get("config_basis", {}).get("AUD", {}).get("6v3")
    if not par_qq or not par_ss:
        return None, None
    bx = by = None
    if basis_6v3 is not None and not basis_6v3.empty:
        bx = basis_6v3["MaturityY"].to_numpy().astype(float)
        by = basis_6v3["BasisBp"].to_numpy().astype(float)
    def _b(t):
        if bx is None: return 0.0
        return float(np.interp(t, bx, by))
    # QQ full: <=3Y direct, >3Y = SS - basis
    qq = dict(par_qq)
    for t, r in par_ss.items():
        qq[t] = r - _b(t) / 100.0
    # SS full: >=4Y direct, <4Y = QQ + basis
    ss = dict(par_ss)
    for t, r in par_qq.items():
        ss[t] = r + _b(t) / 100.0
    qx = np.array(sorted(qq)); qy = np.array([qq[k] for k in qx])
    sx = np.array(sorted(ss)); sy = np.array([ss[k] for k in sx])
    return _CS(qx, qy), _CS(sx, sy)


def _par_fwd(cs, exp_y: float, tenor_y: float) -> float:
    """
    Forward par rate from cubic spline of par rates.
    fwd(exp, tenor) = (par(exp+tenor)*(exp+tenor) - par(exp)*exp) / tenor
    For exp=0 (spot), uses par(tenor) directly.
    """
    t1 = exp_y
    t2 = exp_y + tenor_y
    p2 = float(cs(t2)) * t2
    p1 = float(cs(t1)) * t1 if t1 > 0 else 0.0
    return (p2 - p1) / tenor_y


def _generate_forward_matrix_cached(ccy: str, curve_tuple: tuple, basis_tuple: Optional[tuple] = None,
                                     freq_override: Optional[float] = None, convention: str = "market",
                                     ois_tuple: Optional[tuple] = None) -> pd.DataFrame:
    """Generate forward swap rate matrix.
    AUD: uses par formula directly (fwd = (par(t2)*t2 - par(t1)*t1)/tenor)
         on blended QQ/SS cubic spline — never uses zero/bootstrapped rates.
    NZD/USD: zero curve with IRS-only discounting.
    """

    expiries = ["1w", "1m", "2m", "3m", "6m", "9m", "1y", "18m", "2y", "3y", "4y", "5y", "6y", "7y", "8y", "9y", "10y", "12y", "15y", "20y", "25y", "30y"]
    tenors = ["1Y", "2Y", "3Y", "4Y", "5Y", "7Y", "10Y", "12Y", "15Y", "20Y", "25Y", "30Y"]

    curve_x = np.array(curve_tuple[0])
    curve_y = np.array(curve_tuple[1]) / 100.0

    basis_x = basis_y = None
    if basis_tuple is not None:
        basis_x = np.array(basis_tuple[0])
        basis_y = np.array(basis_tuple[1])

    ois_x = ois_y = None
    if ois_tuple is not None:
        ois_x = np.array(ois_tuple[0])
        ois_y = np.array(ois_tuple[1]) / 100.0

    # AUD: load pure QQ and SS zero curves built during bootstrap
    aud_zc_qq = aud_zc_ss = None
    if ccy == "AUD":
        aud_zc_qq = st.session_state.get("_aud_zc_qq")
        aud_zc_ss = st.session_state.get("_aud_zc_ss")

    SPOT_M = 1.0 / 252.0

    def _fwd_from_zc(zc, exp, tenor, freq):
        """Forward swap rate from a zero curve dict {maturity: zero_rate_pct}."""
        xs = np.array(sorted(zc.keys()))
        ys = np.array([zc[k] / 100.0 for k in xs])
        t_s = exp + SPOT_M; t_e = t_s + tenor
        times = []; t = t_s + freq
        while t <= t_e + 1e-9:
            times.append(min(t, t_e)); t += freq
        if not times: return 0.0
        prev = t_s; ann = 0.0
        for ti in times:
            z = float(np.interp(ti, xs, ys))
            ann += math.exp(-z * ti) * (ti - prev); prev = ti
        if ann <= 0: return 0.0
        zs = float(np.interp(t_s, xs, ys)); ze = float(np.interp(t_e, xs, ys))
        df_s = math.exp(-zs * t_s); df_e = math.exp(-ze * t_e)
        return (df_s - df_e) / ann * 100.0

    matrix = []

    for exp in expiries:
        exp_y = label_to_years(exp)
        row = {"Expiry": exp}

        for tenor in tenors:
            tenor_y = float(tenor[:-1])
            try:
                if ccy == "AUD" and aud_zc_qq is not None and aud_zc_ss is not None:
                    # ── AUD: pure separate zero curves, no post-hoc basis adjustment ──
                    # Market: tenor <=3Y → QQ zero curve @ Q/Q freq
                    #         tenor  >3Y → SS zero curve @ S/S freq
                    if convention == "market":
                        if tenor_y <= 3.0:
                            fwd = _fwd_from_zc(aud_zc_qq, exp_y, tenor_y, 0.25)
                        else:
                            fwd = _fwd_from_zc(aud_zc_ss, exp_y, tenor_y, 0.50)
                    elif convention == "qq":
                        fwd = _fwd_from_zc(aud_zc_qq, exp_y, tenor_y, 0.25)
                    elif convention == "ss":
                        fwd = _fwd_from_zc(aud_zc_ss, exp_y, tenor_y, 0.50)
                    else:
                        if tenor_y <= 3.0:
                            fwd = _fwd_from_zc(aud_zc_qq, exp_y, tenor_y, 0.25)
                        else:
                            fwd = _fwd_from_zc(aud_zc_ss, exp_y, tenor_y, 0.50)
                    row[tenor] = fwd
                else:
                    # ── NZD/USD: zero curve IRS-only discounting ──────────────
                    mkt_rate = fast_forward_rate(curve_x, curve_y, exp_y, tenor_y, ccy,
                                                 freq_override=None,
                                                 ois_x=ois_x, ois_y=ois_y,
                                                 basis6v3_x=basis_x, basis6v3_y=basis_y)
                    basis_bp = 0.0
                    if basis_x is not None:
                        mid_t = exp_y + tenor_y / 2
                        basis_bp = float(np.interp(mid_t, basis_x, basis_y))
                    if convention == "qq" and tenor_y > 3.0:
                        fwd = mkt_rate - basis_bp / 10000.0
                    elif convention == "ss" and tenor_y <= 3.0:
                        fwd = mkt_rate + basis_bp / 10000.0
                    else:
                        fwd = mkt_rate
                    row[tenor] = fwd * 100

            except:
                row[tenor] = None

        matrix.append(row)

    df = pd.DataFrame(matrix)
    df = df.set_index("Expiry")
    return df


def fast_forward_rate(curve_x: np.ndarray, curve_y: np.ndarray, expiry: float, tenor: float, ccy: str,
                      freq_override: Optional[float] = None,
                      ois_x: Optional[np.ndarray] = None, ois_y: Optional[np.ndarray] = None,
                      basis6v3_x: Optional[np.ndarray] = None, basis6v3_y: Optional[np.ndarray] = None) -> float:
    """Year-fraction forward swap rate. curve_y in decimal."""
    SPOT = 1.0 / 252.0
    if freq_override is not None:
        freq = freq_override
    elif ccy == "AUD":
        freq = 0.25 if tenor <= 3 else 0.5
    elif ccy == "NZD":
        freq = 0.25 if tenor <= 2 else 0.5
    else:
        freq = 0.5

    t_start = expiry + SPOT
    t_end   = t_start + tenor
    times = []
    t = t_start + freq
    while t <= t_end + 1e-9:
        times.append(min(t, t_end))
        t += freq
    if not times:
        return 0.0

    disc_x = ois_x if ois_x is not None else curve_x
    disc_y = ois_y if ois_y is not None else curve_y

    prev = t_start; ann = 0.0
    for ti in times:
        ann += math.exp(-float(np.interp(ti, disc_x, disc_y)) * ti) * (ti - prev)
        prev = ti
    if ann <= 0:
        return 0.0

    # Plain zero curve for start/end DFs — no basis on endpoints
    df_s = math.exp(-float(np.interp(t_start, curve_x, curve_y)) * t_start)
    df_e = math.exp(-float(np.interp(t_end,   curve_x, curve_y)) * t_end)
    return (df_s - df_e) / ann


def interpolate_basis(basis_df: pd.DataFrame, t: float) -> float:
    """Interpolate basis curve at time t"""
    if basis_df is None or basis_df.empty:
        return 0.0
    xs = basis_df["MaturityY"].to_numpy().astype(float)
    ys = basis_df["BasisBp"].to_numpy().astype(float)
    if t <= xs[0]:
        return float(ys[0])
    if t >= xs[-1]:
        return float(ys[-1])
    return float(np.interp(t, xs, ys))


def swaptions_tab(vol_mode: str):
    st.subheader(" Swaptions")

    # Apply any pending reload from blotter/portfolio BEFORE widgets render
    _pending = st.session_state.get("sw_pending_reload")
    if _pending:
        EXPIRY_PRESETS = ["1w","2w","1m","2m","3m","6m","9m","1y","18m","2y","3y","5y","7y","10y","12y","15y","20y","📅 Custom Date..."]
        tenor_options = ["1Y","2Y","3Y","4Y","5Y","6Y","7Y","8Y","9Y","10Y","12Y","15Y","20Y","25Y","30Y"]
        structures = ["Payer","Receiver","ATM Straddle","Strangle","Risk Reversal","Payer Ladder","Receiver Ladder"]
        if _pending.get("expiry") in EXPIRY_PRESETS:
            st.session_state["sw_expiry"] = _pending["expiry"]
        if _pending.get("tenor") in tenor_options:
            st.session_state["sw_tenor"] = _pending["tenor"]
        if _pending.get("structure") in structures:
            st.session_state["sw_structure"] = _pending["structure"]
        if _pending.get("notional_mm") is not None:
            st.session_state["sw_not"] = _pending["notional_mm"]
        if _pending.get("strike") is not None:
            st.session_state["sw_strike_1"] = _pending["strike"]
        st.session_state["sw_pending_reload"] = None  # clear after applying

    # Currency selector
    col_ccy, col_spacer = st.columns([1, 3])
    with col_ccy:
        ccy_select = st.selectbox("📎 Currency", ALL_CURRENCIES, key="sw_ccy")
    
    # Extract actual currency code (remove PENDING)
    ccy = ccy_select.split(" ")[0]
    
    # Check if pending currency selected
    if "PENDING" in ccy_select:
        st.warning(f"├ö├àÔöé {ccy} pricing coming soon. Currently supported: AUD, NZD, USD")
        return
    
    # Get curves and data
    fwd_matrix = st.session_state.get("fwd_matrix", {}).get(ccy)
    _cb6 = st.session_state.get("config_basis", {}).get(ccy, {}).get("6v3")
    basis_6v3 = _cb6 if _cb6 is not None else get_basis_curve(ccy, "6v3")
    _cbo = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _cbo if _cbo is not None else get_basis_curve(ccy, "ois")
    _cc = st.session_state.get("config_curves", {}).get(ccy)
    curve = _cc if _cc is not None else get_ccy_curve(ccy)

    # ── SABR Smile Mode & Alpha Monitor ──────────────────────────────
    _sabr_visible = st.session_state.get("sabr_panel_visible", True)
    _toggle_label = "🔄 Hide SABR Panel" if _sabr_visible else "📋 Show SABR Smile Mode & Alpha Monitor"
    if st.button(_toggle_label, key="sabr_panel_toggle"):
        st.session_state["sabr_panel_visible"] = not _sabr_visible
        st.rerun()

    if _sabr_visible:
      with st.container():
        _sm_col, _info_col = st.columns([2, 4])
        with _sm_col:
            _smile_mode = st.radio(
                "Smile Mode",
                ["Sticky-ATM (alpha-sticky)", "Sticky-Delta", "Sticky-Strike"],
                index=0,
                key="sabr_smile_mode",
                help=(
                    "Sticky-ATM: recalibrate alpha to match ATM vol daily, keep ρ,ν/×/~ fixed. "
                    "Sticky-Delta: params fixed, smile moves with forward. "
                    "Sticky-Strike: OTM options repriced at original absolute strikes."
                )
            )
        with _info_col:
            if "Sticky-ATM" in _smile_mode:
                st.caption("⚙️ ~, ρ,ν, × locked from config. ~ recalibrated to ATM surface each session. Stale ~ cells flagged below.")
            elif "Sticky-Delta" in _smile_mode:
                st.caption("⚙️ All SABR params fixed. Smile moves with forward. May understate repricing on large rate moves.")
            else:
                st.caption("⚙️ OTM options priced at original absolute strikes. Most conservative   —   can overstate downside moves.")

        # Alpha comparison table
        _, _a, _b, _r, _n = get_ccy_vol_data(ccy)
        _atm_surf = get_published_atm_surface(ccy)  # Use published not working draft

        # Auto-init default SABR if ATM loaded but SABR missing — once per session
        if _a is None and _atm_surf is not None and not st.session_state.get(f"_sabr_init_{ccy}"):
            try:
                _ar = _atm_surf.copy()
                _tc = [c for c in _ar.columns if c != "Expiry"]
                _vd = st.session_state.setdefault("vol_data", {}).setdefault(ccy, {})
                for _pp, _dv in [("beta", 0.5), ("rho", 0.20), ("nu", 0.30)]:
                    _dp = _ar[["Expiry"]].copy()
                    for _t in _tc: _dp[_t] = _dv
                    _vd[_pp] = _dp
                _da = _ar.copy()
                for _t in _tc: _da[_t] = _da[_t] / 10000.0
                _vd["alpha"] = _da
                st.session_state[f"_sabr_init_{ccy}"] = True
                _, _a, _b, _r, _n = get_ccy_vol_data(ccy)
            except Exception:
                pass

        if _a is not None and _atm_surf is not None and curve is not None:
            _EXPIRIES = ["1m","3m","6m","1y","2y","3y","5y","7y","10y","15y","20y"]
            _TENORS   = ["1Y","2Y","3Y","5Y","7Y","10Y","15Y","20Y","25Y","30Y"]

            # Gate: only run expensive alpha grid on explicit button click — never on render
            if st.button("🔍 Check α consistency", key="show_alpha_grid", type="secondary"):
                st.session_state["_alpha_check_result"] = None  # force recompute
            _alpha_result = st.session_state.get("_alpha_check_result")
            if _alpha_result is None and st.session_state.get("_run_alpha_check"):
                pass  # computed below
            if st.button("▶ Run α Check", key="run_alpha_check_btn", type="secondary"):
                _rows = []
                _any_stale = False
                # Use committed ATM only — not the vol editor working draft
                _committed_atm = st.session_state.get("vol_data", {}).get(ccy, {}).get("atm")
                _check_surf = _committed_atm if _committed_atm is not None else _atm_surf
                for _exp in _EXPIRIES:
                    _row = {"Expiry": _exp}
                    _exp_y = label_to_years(_exp)
                    for _ten in _TENORS:
                        _ten_y = label_to_years(_ten)
                        _atm_bp = get_matrix_value(_check_surf, _exp, _ten_y)
                        _s = get_sabr_params_from_matrices(_a, _b, _r, _n, _exp, _ten_y)
                        if _atm_bp is None or _s is None or _exp_y <= 0:
                            _row[_ten] = "  —  "; continue
                        try:
                            _F, _, _ = forward_and_annuity_from_curve(curve, ccy, _exp_y, _ten_y, ois_curve)
                        except Exception:
                            _F = 0.05
                        if _F <= 0: _F = 0.05
                        _atm_dec = _atm_bp / 10000.0
                        _impl_alpha = sabr_implied_alpha_from_atm(_atm_dec, _F, _exp_y, _s["beta"], _s["rho"], _s["nu"])
                        _stored_alpha = _s["alpha"]
                        if _stored_alpha > 0:
                            _pct_diff = (_impl_alpha - _stored_alpha) / _stored_alpha * 100.0
                            _stale = abs(_pct_diff) > 10.0
                            if _stale: _any_stale = True
                            _row[_ten] = f"{'🔴' if abs(_pct_diff)>20 else '🟡' if _stale else '🟢'} {_pct_diff:+.1f}%"
                        else:
                            _row[_ten] = "  —  "
                    _rows.append(_row)
                st.session_state["_alpha_check_result"] = {"rows": _rows, "stale": _any_stale}

            _alpha_result = st.session_state.get("_alpha_check_result")
            if _alpha_result:
                _rows = _alpha_result["rows"]; _any_stale = _alpha_result["stale"]
                if _any_stale:
                    st.warning("⚙️ Stale α — 🟡 >10%, 🔴 >20%. Consider recalibrating.")
                else:
                    st.success("✅ α consistent with committed ATM surface (within 10%)")
                if _rows:
                    _alpha_df = pd.DataFrame(_rows).set_index("Expiry")
                    st.dataframe(_alpha_df, use_container_width=True)
                    st.caption("Divergence = (implied α from ATM vol − stored α) / stored α × 100%. β, ρ, ν held fixed.")

            _rc1, _rc2 = st.columns([2, 4])
            with _rc1:
                if st.button("🔄 Recalibrate Alpha (Sticky-ATM)", key="recal_alpha_btn", type="primary"):
                    _, _a2, _b2, _r2, _n2 = get_ccy_vol_data(ccy)
                    _atm2 = get_working_atm_surface(ccy)
                    if _a2 is not None and _atm2 is not None:
                        _new_alpha = _a2.copy()
                        _exp_col = "Expiry" if "Expiry" in _new_alpha.columns else _new_alpha.columns[0]
                        _tenor_cols = [c for c in _new_alpha.columns if c != _exp_col]
                        _updated = 0
                        for _i, _erow in _new_alpha.iterrows():
                            _exp_lbl = str(_erow[_exp_col]).strip()
                            _exp_y2 = label_to_years(_exp_lbl)
                            if _exp_y2 <= 0:
                                continue
                            for _tc in _tenor_cols:
                                _ten_y2 = label_to_years(str(_tc))
                                _atm_bp2 = get_matrix_value(_atm2, _exp_lbl, _ten_y2)
                                _s2 = get_sabr_params_from_matrices(_a2, _b2, _r2, _n2, _exp_lbl, _ten_y2)
                                if _atm_bp2 is None or _s2 is None:
                                    continue
                                try:
                                    _F2, _, _ = forward_and_annuity_from_curve(curve, ccy, _exp_y2, _ten_y2, ois_curve)
                                except Exception:
                                    _F2 = 0.05
                                if _F2 <= 0:
                                    _F2 = 0.05
                                _new_a = sabr_implied_alpha_from_atm(_atm_bp2 / 10000.0, _F2, _exp_y2, _s2["beta"], _s2["rho"], _s2["nu"])
                                if _new_a > 0:
                                    _new_alpha.at[_i, _tc] = _new_a
                                    _updated += 1
                        _old_atm, _, _b2, _r2, _n2 = get_ccy_vol_data(ccy)
                        set_ccy_vol_data(ccy, _old_atm, _new_alpha, _b2, _r2, _n2)
                        # Auto-save to DB
                        if HAS_POSTGRES:
                            try:
                                _uid = st.session_state.get("username", "default")
                                save_all_session_data(_uid)
                            except Exception:
                                pass
                        st.success(f"✅ Alpha recalibrated   —   {_updated} cells updated. ~, ρ,ν, × unchanged.")
                        st.rerun()
            with _rc2:
                st.caption("Updates ~ to match current ATM surface. ~, ρ,ν, × remain locked. Run daily at session start in Sticky-ATM mode.")

    # Row 1: Structure Type and Model
    col_struct, col_model, col_prem = st.columns([2, 1, 1])
    with col_struct:
        structure = st.selectbox(
            "Structure",
            ["Payer", "Receiver", "ATM Straddle", "Strangle", "Risk Reversal", "Payer Ladder", "Receiver Ladder"],
            index=2,
            key="sw_structure"
        )
    with col_model:
        model_choice = st.selectbox("Model", ["Normal", "Black"], index=0, key="sw_model")
    with col_prem:
        premium_type = st.selectbox("Premium", ["Fwd", "Spot"], index=0, key="sw_prem_type")
    
    # For backwards compatibility
    side = structure

    # Row 2: Notional, Expiry, Expiry Date, Tenor, Leg Convention
    col_not, col_exp, col_expdt, col_delay, col_tenor, col_conv = st.columns([2, 2, 2, 2, 2, 2])
    with col_not:
        notional = st.number_input("Notional (mm)", min_value=1.0, max_value=10000.0, value=100.0, step=10.0, key="sw_not")
    with col_exp:
        EXPIRY_PRESETS = ["1w","2w","1m","2m","3m","6m","9m","1y","18m","2y","3y","5y","7y","10y","12y","15y","20y"]
        expiry_sel = st.selectbox("Expiry", EXPIRY_PRESETS, index=4, key="sw_expiry")
        expiry = expiry_sel
        expiry_y = label_to_years(expiry)
        expiry_display = expiry
    with col_expdt:
        from datetime import date as _sw_date
        _calc_dt = modified_following(_sw_date.today() + __import__('datetime').timedelta(days=int(expiry_y * 365.25)))
        _default_dt_str = _calc_dt.strftime("%d/%m/%Y")
        _custom_dt_str = st.text_input("Expiry Date (DD/MM/YY)", value=_default_dt_str, key="sw_expiry_date_override")
        try:
            from datetime import datetime as _swdt
            _formats = ["%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"]
            _parsed = None
            for _fmt in _formats:
                try: _parsed = _swdt.strptime(_custom_dt_str.strip(), _fmt).date(); break
                except: pass
            if _parsed and _parsed != _calc_dt:
                expiry_y = max((_parsed - _sw_date.today()).days / 365.0, 1/365.0)
                expiry_display = _parsed.strftime("%d/%m/%Y")
        except: pass
    with col_delay:
        DELAY_PRESETS = ["None","1m","2m","3m","6m","9m","1y","18m","2y"]
        delay_sel = st.selectbox("Delay Start (Midcurve)", DELAY_PRESETS, index=0, key="sw_delay_start")
        delay_y = 0.0 if delay_sel == "None" else label_to_years(delay_sel)
        is_midcurve = delay_y > 0
    with col_tenor:
        tenor_options = ["1Y","2Y","3Y","4Y","5Y","6Y","7Y","8Y","9Y","10Y","12Y","15Y","20Y","25Y","30Y"]
        swap_tenor = st.selectbox("Swap Tenor", tenor_options, index=4, key="sw_tenor")
        tenor_y = float(swap_tenor[:-1])
    with col_conv:
        leg_conv = st.radio("Leg Convention", ["Market", "Q/Q", "S/S"], horizontal=True, key="sw_leg_conv")
        freq_override = None if leg_conv == "Market" else (0.25 if leg_conv == "Q/Q" else 0.5)

    # Forward rate calculation
    # Vanilla:   fwd = forward_and_annuity(expiry_y, tenor_y)
    # Midcurve:  ATM = [Ann(expiry+delay, tenor) * R(expiry+delay, tenor)
    #                 - Ann(expiry, delay) * R(expiry, delay)]
    #                 / Ann(expiry, tenor)   — swap triangle
    if curve is not None:
        if is_midcurve:
            # Swap triangle for midcurve: Xm→YmTY
            # Long:  start=expiry, tenor=delay+tenor  e.g. 6m×2Y for 6m→1y1Y
            # Short: start=expiry, tenor=delay         e.g. 6m×1Y for 6m→1y1Y
            # Mid annuity: start=expiry+delay, tenor=tenor  e.g. 1.5y×1Y
            # R(mid) = [Ann(long)*R(long) - Ann(short)*R(short)] / Ann(mid)
            fwd_long, ann_long, _   = forward_and_annuity_from_curve(curve, ccy, expiry_y, delay_y + tenor_y, ois_curve, freq_override=freq_override)
            fwd_short, ann_short, _ = forward_and_annuity_from_curve(curve, ccy, expiry_y, delay_y, ois_curve)
            fwd_mid, ann, _         = forward_and_annuity_from_curve(curve, ccy, expiry_y + delay_y, tenor_y, ois_curve, freq_override=freq_override)
            if ann > 0 and ann_long > 0 and ann_short > 0:
                fwd = (ann_long * fwd_long - ann_short * fwd_short) / ann
            else:
                fwd = fwd_mid
            fwd_source = f"midcurve ({expiry}→{delay_sel}{swap_tenor})"
        else:
            fwd, ann, _ = forward_and_annuity_from_curve(curve, ccy, expiry_y, tenor_y, ois_curve, freq_override=freq_override)
            fwd_source = "curve"
        if basis_6v3 is not None and ccy == "AUD":
            basis_bp = interpolate_basis(basis_6v3, expiry_y + delay_y + tenor_y / 2)
            if leg_conv == "Q/Q" and tenor_y > 3.0:
                fwd = fwd - basis_bp / 10000.0
            elif leg_conv == "S/S" and tenor_y <= 3.0:
                fwd = fwd + basis_bp / 10000.0
    else:
        fwd = 0.04
        ann = tenor_y
        fwd_source = "default"
        is_midcurve = False

    fwd_pct = fwd * 100
    # Midcurve label suffix
    if is_midcurve:
        expiry_display = f"{expiry_display}→{delay_sel}"
    # Safety defaults in case col blocks don't execute
    eff_disc_rate = 0.035
    disc_source = "Flat (default)"

    col_fwd, col_disc = st.columns(2)
    
    with col_fwd:
        st.metric("Forward (%)", f"{fwd_pct:.4f}", delta=fwd_source)
    
    with col_disc:
        has_ois = ois_curve is not None
        if has_ois:
            disc_method = st.radio("Discount", ["OIS", "Flat"], horizontal=True, key="sw_disc_method")
        else:
            disc_method = "Flat"
            st.caption(" No OIS curve")
        
        if disc_method == "Flat" or ois_curve is None:
            flat_rate = st.number_input("Rate (%)", min_value=0.0, max_value=20.0, value=4.0, key="sw_disc_flat")
            eff_disc_rate = flat_rate / 100.0
            disc_source = "Flat"
        else:
            try:
                _ois_col_m = next((c for c in ois_curve.columns if "matur" in str(c).lower()), None)
                _ois_col_z = next((c for c in ois_curve.columns if "zero" in str(c).lower() or "rate" in str(c).lower()), None)
                if _ois_col_m is None or _ois_col_z is None:
                    raise ValueError("OIS curve missing MaturityY/ZeroRatePct columns")
                ois_xs = ois_curve[_ois_col_m].to_numpy().astype(float)
                ois_ys = ois_curve[_ois_col_z].to_numpy().astype(float) / 100.0
                _disc_t = expiry_y + delay_y if is_midcurve else expiry_y
                eff_disc_rate = float(np.interp(_disc_t, ois_xs, ois_ys))
                disc_source = "OIS"
                st.caption(f"OIS rate: {eff_disc_rate*100:.2f}% @ {_disc_t:.2f}Y")
            except Exception:
                eff_disc_rate = 0.035
                disc_source = "Flat (fallback)"

    # Strike inputs - varies by structure
    st.markdown("---")
    st.markdown("##### Strikes")
    
    # Initialize strike variables
    strike_pct = fwd_pct
    strike_pct_2 = fwd_pct
    strike_pct_3 = fwd_pct
    
    if structure == "ATM Straddle":
        strike_pct = fwd_pct
        st.info(f"ATM Strike: **{fwd_pct:.4f}%**")
        
    elif structure in ["Payer", "Receiver"]:
        strike_mode = st.radio("Strike Mode", ["ATM", "10 bp", "25 bp", "50 bp", "100 bp", "Manual"], 
                               horizontal=True, key="sw_strike_mode")
        offset_map = {"ATM": 0, "10 bp": 10, "25 bp": 25, "50 bp": 50, "100 bp": 100, "Manual": None}
        offset = offset_map[strike_mode]
        if strike_mode == "Manual":
            strike_pct = st.number_input("Strike (%)", min_value=0.0, max_value=20.0, 
                                         value=round(fwd_pct, 4), format="%.4f", key="sw_strike_pr")
        else:
            strike_pct = fwd_pct + (offset/100.0 if structure == "Payer" else -offset/100.0)
            st.info(f"Strike: **{strike_pct:.4f}%** ({strike_mode} {'OTM' if offset > 0 else ''})")
        moneyness_bp = (strike_pct - fwd_pct) * 100
        st.caption(f"Moneyness: **{moneyness_bp:+.1f} bp**" if abs(moneyness_bp) >= 0.5 else "Moneyness: **ATM**")
        
    elif structure == "Strangle":
        st.caption("Buy OTM Payer + Buy OTM Receiver (long vol)")
        col_k1, col_k2 = st.columns(2)
        with col_k1:
            strike_pct = st.number_input("Payer Strike (%)", min_value=0.0, max_value=20.0,
                                         value=round(fwd_pct + 0.25, 4), format="%.4f", key="sw_strike_stg1",
                                         help="Higher strike - OTM payer")
        with col_k2:
            strike_pct_2 = st.number_input("Receiver Strike (%)", min_value=0.0, max_value=20.0,
                                           value=round(fwd_pct - 0.25, 4), format="%.4f", key="sw_strike_stg2",
                                           help="Lower strike - OTM receiver")
        st.caption(f"Width: **{(strike_pct - strike_pct_2)*100:.0f} bp**")
        
    elif structure == "Risk Reversal":
        st.caption("Buy Payer + Sell Receiver (or vice versa) - rate protection")
        rr_strike_mode = st.radio("Strike Mode", 
                                   ["Symmetric (25bp)", "Symmetric (50bp)", "Symmetric (100bp)", "Manual (independent)"], 
                                   horizontal=True, key="sw_rr_mode")
        col_k1, col_k2, col_dir = st.columns([1, 1, 1])
        if "Symmetric" in rr_strike_mode:
            offset = {"25bp": 0.25, "50bp": 0.50, "100bp": 1.00}.get(
                next((x for x in ["25bp","50bp","100bp"] if x in rr_strike_mode), "25bp"), 0.25)
            strike_pct = fwd_pct + offset
            strike_pct_2 = fwd_pct - offset
            with col_k1: st.metric("Payer Strike (%)", f"{strike_pct:.4f}")
            with col_k2: st.metric("Receiver Strike (%)", f"{strike_pct_2:.4f}")
        else:
            with col_k1:
                strike_pct = st.number_input("Payer Strike (%)", min_value=0.0, max_value=20.0,
                                             value=round(fwd_pct + 0.25, 4), format="%.4f", key="sw_strike_rr1")
            with col_k2:
                strike_pct_2 = st.number_input("Receiver Strike (%)", min_value=0.0, max_value=20.0,
                                               value=round(fwd_pct - 0.25, 4), format="%.4f", key="sw_strike_rr2")
        with col_dir:
            collar_dir = st.radio("Direction", ["Long Payer/Short Rec", "Short Payer/Long Rec"], 
                                  key="sw_collar_dir")
        width_bp = (strike_pct - strike_pct_2) * 100
        st.caption(f"Width: **{width_bp:.0f} bp** | Payer +{(strike_pct-fwd_pct)*100:.0f}bp | Receiver {(strike_pct_2-fwd_pct)*100:.0f}bp")
        
    elif structure == "Payer Ladder":
        st.caption("Buy 1x K1 Payer + Sell 1x K2 + Sell 1x K3 — limited cost, capped upside")
        col_k1, col_k2mode, col_k3mode = st.columns(3)
        with col_k1:
            strike_pct = st.number_input("K1 Strike (%)", min_value=0.0, max_value=20.0,
                                         value=round(fwd_pct + 0.25, 4), format="%.4f", key="sw_strike_pl1")
        with col_k2mode:
            _k2_off_sel = st.radio("K2 offset from K1", ["10bp","25bp","50bp","Manual"], horizontal=True, key="sw_k2_off_pl")
            if _k2_off_sel == "Manual":
                strike_pct_2 = st.number_input("K2 (%)", min_value=0.0, max_value=20.0,
                                               value=round(strike_pct + 0.25, 4), format="%.4f", key="sw_strike_pl2")
            else:
                strike_pct_2 = round(strike_pct + {"10bp":0.10,"25bp":0.25,"50bp":0.50}[_k2_off_sel], 4)
                st.metric("K2", f"{strike_pct_2:.4f}%")
        with col_k3mode:
            _k3_off_sel = st.radio("K3 offset from K2", ["10bp","25bp","50bp","Manual"], horizontal=True, key="sw_k3_off_pl")
            if _k3_off_sel == "Manual":
                strike_pct_3 = st.number_input("K3 (%)", min_value=0.0, max_value=20.0,
                                               value=round(strike_pct_2 + 0.25, 4), format="%.4f", key="sw_strike_pl3")
            else:
                strike_pct_3 = round(strike_pct_2 + {"10bp":0.10,"25bp":0.25,"50bp":0.50}[_k3_off_sel], 4)
                st.metric("K3", f"{strike_pct_3:.4f}%")
        st.caption(f"K1={strike_pct:.4f}% (+{(strike_pct-fwd_pct)*100:.0f}bp) | K2={strike_pct_2:.4f}% | K3={strike_pct_3:.4f}%")
        
    elif structure == "Receiver Ladder":
        st.caption("Buy 1x K1 Receiver + Sell 1x K2 + Sell 1x K3 — limited cost, capped downside")
        col_k1, col_k2mode, col_k3mode = st.columns(3)
        with col_k1:
            strike_pct = st.number_input("K1 Strike (%)", min_value=0.0, max_value=20.0,
                                         value=round(fwd_pct - 0.25, 4), format="%.4f", key="sw_strike_rl1")
        with col_k2mode:
            _k2_off_sel = st.radio("K2 offset from K1", ["10bp","25bp","50bp","Manual"], horizontal=True, key="sw_k2_off_rl")
            if _k2_off_sel == "Manual":
                strike_pct_2 = st.number_input("K2 (%)", min_value=0.0, max_value=20.0,
                                               value=round(strike_pct - 0.25, 4), format="%.4f", key="sw_strike_rl2")
            else:
                strike_pct_2 = round(strike_pct - {"10bp":0.10,"25bp":0.25,"50bp":0.50}[_k2_off_sel], 4)
                st.metric("K2", f"{strike_pct_2:.4f}%")
        with col_k3mode:
            _k3_off_sel = st.radio("K3 offset from K2", ["10bp","25bp","50bp","Manual"], horizontal=True, key="sw_k3_off_rl")
            if _k3_off_sel == "Manual":
                strike_pct_3 = st.number_input("K3 (%)", min_value=0.0, max_value=20.0,
                                               value=round(strike_pct_2 - 0.25, 4), format="%.4f", key="sw_strike_rl3")
            else:
                strike_pct_3 = round(strike_pct_2 - {"10bp":0.10,"25bp":0.25,"50bp":0.50}[_k3_off_sel], 4)
                st.metric("K3", f"{strike_pct_3:.4f}%")
        st.caption(f"K1={strike_pct:.4f}% (-{abs((strike_pct-fwd_pct)*100):.0f}bp) | K2={strike_pct_2:.4f}% | K3={strike_pct_3:.4f}%")

    # Vol source
    st.markdown("---")
    vol_src = st.radio("Vol", ["Surface", "Manual"], horizontal=True, key="sw_volsrc")

    # 3D Vol Surface   —   collapsible, below Vol source buttons
    with st.expander("📊 ATM Vol Surface (3D)", expanded=False):
        atm_3d = get_working_atm_surface(ccy)
        if atm_3d is not None:
            try:
                import plotly.graph_objects as go
                import numpy as np
                surf_mode_sw = st.radio("Display", ["Vol (bp)", "Fwd Premium (bp)"], horizontal=True, key="sw_surf_mode_top")
                # ATM surface has Expiry as a column, not index
                surf = atm_3d.copy()
                if "Expiry" in surf.columns:
                    surf = surf.set_index("Expiry")
                def _yrs(lbl):
                    lbl = str(lbl).strip().lower()
                    if lbl.endswith("w"): return float(lbl[:-1])/52
                    if lbl.endswith("m"): return float(lbl[:-1])/12
                    if lbl.endswith("y"): return float(lbl[:-1])
                    try: return float(lbl)
                    except: return 0
                expiry_labels = [str(e) for e in surf.index]
                tenor_labels  = [str(c) for c in surf.columns]
                sorted_exp = sorted(expiry_labels, key=_yrs)
                sorted_ten = sorted(tenor_labels,  key=_yrs)
                exp_yrs = [_yrs(e) for e in sorted_exp]
                ten_yrs = [_yrs(t) for t in sorted_ten]
                # For Fwd Premium use real premium matrix if available
                prem_3d = None
                if surf_mode_sw == "Fwd Premium (bp)":
                    prem_store = st.session_state.get("prem_matrix", {})
                    if ccy in prem_store:
                        p = prem_store[ccy].copy()
                        if "Expiry" in p.columns:
                            p = p.set_index("Expiry")
                        prem_3d = p

                z_vals = []
                for exp in sorted_exp:
                    row = []
                    for ten in sorted_ten:
                        try:
                            if surf_mode_sw == "Fwd Premium (bp)" and prem_3d is not None:
                                v = float(prem_3d.loc[exp, ten]) if exp in prem_3d.index and ten in prem_3d.columns else np.nan
                            else:
                                v = float(surf.loc[exp, ten])
                            if pd.isna(v): v = np.nan
                        except:
                            v = np.nan
                        row.append(v)
                    z_vals.append(row)
                zlabel = "Vol (bp)" if surf_mode_sw == "Vol (bp)" else "Fwd Premium (bp)"
                z_arr = np.array(z_vals, dtype=float)

                # Surface trace   —   matches Vol Editor style
                surf_trace = go.Surface(
                    x=ten_yrs, y=exp_yrs, z=z_arr,
                    colorscale=[
                        [0.0,  "#0ea5e9"],
                        [0.25, "#22d3ee"],
                        [0.5,  "#4ade80"],
                        [0.75, "#facc15"],
                        [1.0,  "#ef4444"],
                    ],
                    opacity=0.92,
                    colorbar=dict(title=dict(text=zlabel, font=dict(color="#94a3b8", size=11)),
                                  tickfont=dict(color="#94a3b8"), thickness=10, len=0.6, x=1.02),
                    hovertemplate=f"Tenor: %{{x:.1f}}y<br>Expiry: %{{y:.2f}}y<br>{zlabel}: %{{z:.1f}}<extra></extra>",
                    lighting=dict(ambient=0.7, diffuse=0.8, specular=0.3, roughness=0.5),
                    lightposition=dict(x=1000, y=1000, z=2000),
                )
                # Scatter dots on grid points   —   matches Vol Editor nodes
                dot_x, dot_y, dot_z, dot_text = [], [], [], []
                for i, exp in enumerate(sorted_exp):
                    for j, ten in enumerate(sorted_ten):
                        v = z_arr[i][j] if i < len(z_arr) and j < len(z_arr[i]) else np.nan
                        if not np.isnan(v):
                            dot_x.append(ten_yrs[j])
                            dot_y.append(exp_yrs[i])
                            dot_z.append(v)
                            dot_text.append(f"{ten} x {exp}: {v:.1f}")
                dots_trace = go.Scatter3d(
                    x=dot_x, y=dot_y, z=dot_z,
                    mode="markers",
                    marker=dict(
                        size=4,
                        color="rgba(255,255,255,0.85)",
                        line=dict(width=0),
                    ),
                    text=dot_text,
                    hoverinfo="text",
                    showlegend=False
                )
                fig3d = go.Figure(data=[surf_trace, dots_trace])
                fig3d.update_layout(
                    scene=dict(
                        # X = Tenor (left→right, short→long) matching vol editor
                        xaxis=dict(title="Tenor", tickmode="array",
                                   tickvals=ten_yrs, ticktext=sorted_ten,
                                   tickangle=0, nticks=len(sorted_ten),
                                   gridcolor="#1e3a5f", color="#94a3b8",
                                   backgroundcolor="rgba(2,6,23,0.8)", showbackground=True),
                        # Y = Expiry (front→back) matching vol editor
                        yaxis=dict(title="Expiry", tickmode="array",
                                   tickvals=exp_yrs, ticktext=sorted_exp,
                                   nticks=len(sorted_exp),
                                   gridcolor="#1e3a5f", color="#94a3b8",
                                   backgroundcolor="rgba(2,6,23,0.8)", showbackground=True),
                        zaxis=dict(title=zlabel,
                                   gridcolor="#1e3a5f", color="#94a3b8",
                                   backgroundcolor="rgba(2,6,23,0.6)", showbackground=True),
                        bgcolor="rgba(2,6,23,0.0)",
                        # Camera matches vol editor angle   —   tenor left-right, expiry front-back
                        camera=dict(eye=dict(x=-1.8, y=-1.8, z=1.0)),
                        aspectmode="manual",
                        aspectratio=dict(x=2.0, y=1.5, z=0.8),
                    ),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    margin=dict(l=0, r=0, t=10, b=0),
                    height=520,
                    font=dict(color="#94a3b8", family="Inter, Arial", size=10),
                )
                st.plotly_chart(fig3d, use_container_width=True)
            except Exception as e:
                st.warning(f"3D surface error: {e}")
        else:
            st.info("Load a vol surface in Vol/SABR tab first.")
    
    if vol_src == "Manual":
        vol_input = st.number_input("Vol (bp normal or % Black)", value=80.0, key="sw_volinput")
        vol_used_display = vol_input
        vol = vol_input / 10000.0 if vol_mode.startswith("Normal") else vol_input / 100.0
    else:
        atm = get_working_atm_surface(ccy)
        _, a, b, r, n = get_ccy_vol_data(ccy)
        atm_val = get_matrix_value(atm, expiry, tenor_y) if atm is not None else None
        if atm_val is None:
            st.warning("No ATM vol - using 80bp")
            atm_val = 80.0
        if vol_mode.startswith("Normal"):
            vol = atm_val / 10000.0
            vol_used_display = atm_val
        else:
            # Only compute SABR vol on render if params already cached — avoid hang
            sabr = get_sabr_params_from_matrices(a, b, r, n, expiry, tenor_y)
            if sabr and sabr.get("alpha", 0) > 0:
                try:
                    vol = sabr_implied_vol_black(fwd_pct/100.0, strike_pct/100.0, expiry_y,
                                                 sabr["alpha"], sabr["beta"], sabr["rho"], sabr["nu"])
                    vol_used_display = vol * 100.0
                except Exception:
                    vol = atm_val / 100.0
                    vol_used_display = atm_val
            else:
                vol = atm_val / 100.0
                vol_used_display = vol * 100.0
        st.caption(f"Vol: {vol_used_display:.1f} {'bp' if vol_mode.startswith('Normal') else '%'}")

    # Dates
    from datetime import datetime, timedelta
    today = datetime.now()
    expiry_date = today + timedelta(days=int(expiry_y * 365))
    swap_start = expiry_date + timedelta(days=1)
    swap_end = swap_start + timedelta(days=int(tenor_y * 365))
    
    # Roll convention
    if ccy == "AUD":
        roll = "Q/Q" if tenor_y <= 3 else "S/S"
    elif ccy == "NZD":
        roll = "Q/Q" if tenor_y <= 2 else "S/S"
    else:
        roll = "S/S"

    # Helper function to get vol for a specific strike
    def get_vol_for_strike(k_pct):
        if vol_src == "Manual":
            return vol_input / 10000.0 if vol_mode.startswith("Normal") else vol_input / 100.0
        else:
            _smile = st.session_state.get("sabr_smile_mode", "Sticky-ATM (alpha-sticky)")
            sabr = get_sabr_params_from_matrices(a, b, r, n, expiry, tenor_y)
            if vol_mode.startswith("Normal"):
                if sabr and sabr.get("alpha", 0) > 0:
                    # Apply Normal SABR smile for OTM strikes
                    return sabr_normal_vol_smile(
                        fwd_pct/100.0, k_pct/100.0, expiry_y,
                        sabr["alpha"], sabr["beta"], sabr["rho"], sabr["nu"])
                else:
                    return atm_val / 10000.0
            else:
                if sabr and sabr.get("alpha", 0) > 0:
                    return sabr_implied_vol_black(fwd_pct/100.0, k_pct/100.0, expiry_y,
                                                   sabr["alpha"], sabr["beta"], sabr["rho"], sabr["nu"])
                else:
                    return atm_val / 100.0

    # Price button
    st.markdown("---")
    
    if st.button(" Price Swaption", key="sw_price", type="primary"):
        try:
            model_type = "Normal" if vol_mode.startswith("Normal") else "Black"
            legs = []  # Store individual leg results
            
            if structure == "Payer":
                vol_k1 = get_vol_for_strike(strike_pct)
                ticket = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Payer {expiry_display}x{swap_tenor}", use_curve=curve is not None)
                res = price_swaption(ticket)
                legs.append(("Payer", strike_pct, 1, res))
                label = f"Payer {expiry_display}x{swap_tenor} K={strike_pct:.2f}%"
                
            elif structure == "Receiver":
                vol_k1 = get_vol_for_strike(strike_pct)
                ticket = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Receiver {expiry_display}x{swap_tenor}", use_curve=curve is not None)
                res = price_swaption(ticket)
                legs.append(("Receiver", strike_pct, 1, res))
                label = f"Receiver {expiry_display}x{swap_tenor} K={strike_pct:.2f}%"
                
            elif structure == "ATM Straddle":
                vol_atm = get_vol_for_strike(fwd_pct)
                ticket_p = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=fwd_pct/100.0, vol=vol_atm, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Payer ATM", use_curve=curve is not None)
                ticket_r = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=fwd_pct/100.0, vol=vol_atm, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Receiver ATM", use_curve=curve is not None)
                res_p = price_swaption(ticket_p)
                res_r = price_swaption(ticket_r)
                legs.append(("Payer", fwd_pct, 1, res_p))
                legs.append(("Receiver", fwd_pct, 1, res_r))
                res = {k: res_p.get(k,0) + res_r.get(k,0) for k in res_p}
                res["bpv"] = res_p["bpv"]
                label = f"ATM Straddle {expiry_display}x{swap_tenor}"
                
            elif structure == "Strangle":
                vol_k1 = get_vol_for_strike(strike_pct)  # Payer strike (higher)
                vol_k2 = get_vol_for_strike(strike_pct_2)  # Receiver strike (lower)
                ticket_p = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Payer K={strike_pct:.2f}%", use_curve=curve is not None)
                ticket_r = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_2/100.0, vol=vol_k2, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Receiver K={strike_pct_2:.2f}%", use_curve=curve is not None)
                res_p = price_swaption(ticket_p)
                res_r = price_swaption(ticket_r)
                legs.append(("Long Payer", strike_pct, 1, res_p))
                legs.append(("Long Receiver", strike_pct_2, 1, res_r))
                res = {k: res_p.get(k,0) + res_r.get(k,0) for k in res_p}
                res["bpv"] = res_p["bpv"]
                label = f"Strangle {expiry_display}x{swap_tenor} ({strike_pct_2:.2f}/{strike_pct:.2f})"
                
            elif structure == "Risk Reversal":
                vol_k1 = get_vol_for_strike(strike_pct)  # Payer strike
                vol_k2 = get_vol_for_strike(strike_pct_2)  # Receiver strike
                is_long_payer = collar_dir.startswith("Long Payer")
                ticket_p = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Payer K={strike_pct:.2f}%", use_curve=curve is not None)
                ticket_r = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_2/100.0, vol=vol_k2, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Receiver K={strike_pct_2:.2f}%", use_curve=curve is not None)
                res_p = price_swaption(ticket_p)
                res_r = price_swaption(ticket_r)
                if is_long_payer:
                    legs.append(("Long Payer", strike_pct, 1, res_p))
                    legs.append(("Short Receiver", strike_pct_2, -1, res_r))
                    res = {k: res_p.get(k,0) - res_r.get(k,0) for k in res_p}
                else:
                    legs.append(("Short Payer", strike_pct, -1, res_p))
                    legs.append(("Long Receiver", strike_pct_2, 1, res_r))
                    res = {k: res_r.get(k,0) - res_p.get(k,0) for k in res_p}
                res["bpv"] = res_p["bpv"]
                label = f"Risk Reversal {expiry_display}x{swap_tenor} ({strike_pct_2:.2f}/{strike_pct:.2f})"
                
            elif structure == "Payer Ladder":
                vol_k1 = get_vol_for_strike(strike_pct)
                vol_k2 = get_vol_for_strike(strike_pct_2)
                vol_k3 = get_vol_for_strike(strike_pct_3)
                ticket_1 = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Long 1x Payer K1={strike_pct:.2f}%", use_curve=curve is not None)
                ticket_2 = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_2/100.0, vol=vol_k2, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Short 1x Payer K2={strike_pct_2:.2f}%", use_curve=curve is not None)
                ticket_3 = SwaptionTicket(
                    side="Payer", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_3/100.0, vol=vol_k3, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Short 1x Payer K3={strike_pct_3:.2f}%", use_curve=curve is not None)
                res_1 = price_swaption(ticket_1)
                res_2 = price_swaption(ticket_2)
                res_3 = price_swaption(ticket_3)
                legs.append(("Long 1x Payer", strike_pct, 1, res_1))
                legs.append(("Short 1x Payer", strike_pct_2, -1, res_2))
                legs.append(("Short 1x Payer", strike_pct_3, -1, res_3))
                res = {k: res_1.get(k,0) - res_2.get(k,0) - res_3.get(k,0) for k in res_1}
                res["bpv"] = res_1["bpv"]
                label = f"Payer Ladder {expiry_display}x{swap_tenor} ({strike_pct:.2f}/{strike_pct_2:.2f}/{strike_pct_3:.2f})"
                
            elif structure == "Receiver Ladder":
                vol_k1 = get_vol_for_strike(strike_pct)
                vol_k2 = get_vol_for_strike(strike_pct_2)
                vol_k3 = get_vol_for_strike(strike_pct_3)
                ticket_1 = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct/100.0, vol=vol_k1, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Long 1x Receiver K1={strike_pct:.2f}%", use_curve=curve is not None)
                ticket_2 = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_2/100.0, vol=vol_k2, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Short 1x Receiver K2={strike_pct_2:.2f}%", use_curve=curve is not None)
                ticket_3 = SwaptionTicket(
                    side="Receiver", payoff_type="vanilla", notional=notional*1e6, currency=ccy,
                    expiry_years=expiry_y, swap_tenor_years=tenor_y, forward=fwd_pct/100.0,
                    strike=strike_pct_3/100.0, vol=vol_k3, discount_rate=eff_disc_rate, annuity=ann,
                    model=model_type, label=f"Short 1x Receiver K3={strike_pct_3:.2f}%", use_curve=curve is not None)
                res_1 = price_swaption(ticket_1)
                res_2 = price_swaption(ticket_2)
                res_3 = price_swaption(ticket_3)
                legs.append(("Long 1x Receiver", strike_pct, 1, res_1))
                legs.append(("Short 1x Receiver", strike_pct_2, -1, res_2))
                legs.append(("Short 1x Receiver", strike_pct_3, -1, res_3))
                res = {k: res_1.get(k,0) - res_2.get(k,0) - res_3.get(k,0) for k in res_1}
                res["bpv"] = res_1["bpv"]
                label = f"Receiver Ladder {expiry_display}x{swap_tenor} ({strike_pct:.2f}/{strike_pct_2:.2f}/{strike_pct_3:.2f})"
            
            st.success(f" Priced: **{label}** | PV = ${res['pv']:,.0f} ({res.get('pv_bp_fwd', res['pv_bp']):.2f} bp fwd)")
        
            # Store results in session state
            moneyness_bp = (strike_pct - fwd_pct) * 100 if structure in ["Payer", "Receiver"] else 0
            if premium_type == "Fwd":
                display_prem_bp = res.get("pv_bp_fwd", res["pv_bp"])
            else:
                display_prem_bp = res.get("pv_bp_spot", res["pv_bp"])
            st.session_state["sw_last_result"] = {
                "res": res, "label": label, "structure": structure, "legs": legs,
                "params": {
                    "Structure": structure, "Expiry": expiry_display,
                    "Tenor": swap_tenor, "Forward (%)": f"{fwd_pct:.4f}",
                    "Annuity (PV01)": f"{ann:.4f}", "Discount": f"{eff_disc_rate*100:.3f}% ({disc_source})",
                    "Notional": f"{notional:,.0f}mm"
                },
                "notional": notional,
                "premium_type": premium_type,
                "display_prem_bp": display_prem_bp,
            }
            
            # Add to Trade Blotter
            if premium_type == "Fwd":
                display_prem_bp = res.get("pv_bp_fwd", res["pv_bp"])
            else:
                display_prem_bp = res.get("pv_bp_spot", res["pv_bp"])
            entry = dict(instrument_type="Swaption", currency=ccy, structure=structure,
                         expiry=expiry, tenor=swap_tenor, model=vol_mode,
                         notional_mm=notional, strike=strike_pct, forward=fwd_pct, pv=res["pv"],
                         pv_bp=display_prem_bp, premium_type=premium_type,
                         delta=res["delta"], gamma=res["gamma"], vega=res["vega"],
                         theta=res["theta"], bpv=res["bpv"], label=label,
                         legs=[{"name": l[0], "strike": l[1], "qty": l[2],
                                "pv": l[3].get("pv",0), "pv_bp": l[3].get("pv_bp",0),
                                "delta": l[3].get("delta",0)} for l in legs],
                         expiry_date=expiry_date.strftime('%d-%b-%Y') if 'expiry_date' in dir() else "",
                         swap_start=expiry_date.strftime('%d-%b-%Y') if 'expiry_date' in dir() else "",
                         )
            st.session_state["swaption_portfolio"].append(entry)
            st.session_state["portfolio"].append(entry)
            _save_portfolio()
            
        except Exception as e:
            st.error(f" Pricing error: {e}")
            import traceback
            st.code(traceback.format_exc())
    
    # Display last result (persists after pricing)
    if "sw_last_result" in st.session_state:
        r = st.session_state["sw_last_result"]
        res = r["res"]
        stored_notional = r.get("notional", 100)
        legs = r.get("legs", [])
        
        st.markdown("###  Results")
        col_params, col_greeks = st.columns(2)
        
        with col_params:
            st.markdown("##### Parameters")
            params_df = pd.DataFrame(list(r["params"].items()), columns=["Parameter", "Value"])
            st.dataframe(params_df, use_container_width=True, hide_index=True)
            
            # Show leg breakdown for multi-leg structures
            if len(legs) > 1:
                st.markdown("##### Leg Breakdown")
                leg_data = []
                for leg_name, leg_strike, leg_mult, leg_res in legs:
                    leg_data.append({
                        "Leg": leg_name,
                        "Strike (%)": f"{leg_strike:.4f}",
                        "Premium (bp)": f"{leg_res.get('pv_bp_fwd', leg_res['pv_bp']) * leg_mult:.2f}",
                        "PV": f"${leg_res['pv'] * leg_mult:,.0f}",
                        "Delta": f"{leg_res['delta'] * leg_mult:,.0f}"
                    })
                st.dataframe(pd.DataFrame(leg_data), use_container_width=True, hide_index=True)
        
        with col_greeks:
            st.markdown("##### Valuation")
            stored_prem_type = r.get("premium_type", "Fwd")
            display_prem_bp = r.get("display_prem_bp", res.get("pv_bp", 0))

            if stored_prem_type == "Fwd":
                st.metric("Premium (bp)   —   Fwd", f"{display_prem_bp:.2f}",
                          help="Forward premium in bp: df cancels in PV/BPV = market convention")
            else:
                st.metric("Premium (bp)   —   Spot", f"{display_prem_bp:.2f}",
                          help="Spot premium: PV/BPV ≈ df_expiry")
            st.metric("Total PV", f"${res['pv']:,.0f}")
            
            st.markdown("##### Greeks (Net)")
            d_ratio = res.get("delta_ratio", res["delta"] / stored_notional if stored_notional else 0)
            d_dv01  = res.get("delta_dv01", res["delta"] * 0.0001)
            greeks_df = pd.DataFrame({
                "Greek": ["Delta (swap hedge)", "Delta % notional", "Delta DV01 ($/bp)", "Gamma ($/bp)", "Vega ($/bp vol)", "Theta ($/day)", "BPV ($/bp)"],
                "Value": [
                    f"${res['delta']:,.0f}",
                    f"{d_ratio*100:.1f}%",
                    f"{d_dv01:,.1f}",
                    f"{res['gamma']:,.2f}",
                    f"{res['vega']:,.1f}",
                    f"{res['theta']:,.1f}",
                    f"{res['bpv']:,.1f}"
                ],
                "Per 1mm notional": [
                    f"${d_ratio*1e6:,.0f}",
                    f"{d_ratio*100:.1f}%",
                    f"{d_dv01/stored_notional:,.1f}",
                    f"{res['gamma']/stored_notional:,.3f}",
                    f"{res['vega']/stored_notional:,.1f}",
                    f"{res['theta']/stored_notional:,.1f}",
                    f"{res['bpv']/stored_notional:,.1f}"
                ]
            })
            st.dataframe(greeks_df, use_container_width=True, hide_index=True)

    # Display portfolio
    if st.session_state["swaption_portfolio"]:
        ph1, ph2 = st.columns([3, 1])
        with ph1:
            st.markdown("### Swaption Options Whiteboard")
        with ph2:
            if st.button("🗑️ Clear All", key="sw_clear_portfolio"):
                st.session_state["swaption_portfolio"] = []
                st.session_state["portfolio"] = []
                for _ck in [k for k in st.session_state if k.startswith("_sw_price_") or k.startswith("_sw_calc_")]:
                    del st.session_state[_ck]
                _save_portfolio()
                st.rerun()

        df = pd.DataFrame(st.session_state["swaption_portfolio"])
        df["_expiry_sort"] = df["expiry"].apply(lambda e: label_to_years(str(e)))
        df = df.sort_values("_expiry_sort").reset_index(drop=True)

        # ── Per-row display with legs breakdown ──────────────────────
        for idx, row in df.iterrows():
            _struct = row.get("structure", "")
            _label  = row.get("label", f"{row.get('expiry','')}x{row.get('tenor','')}")
            _pv     = float(row.get("pv", 0))
            _pv_bp  = float(row.get("pv_bp", 0))
            _delta  = float(row.get("delta", 0))
            _notl   = float(row.get("notional_mm", 100))
            _ccy    = row.get("currency", "AUD")
            _strike = float(row.get("strike", 0))
            _fwd    = float(row.get("forward", 0))
            _expiry = row.get("expiry", "")
            _tenor  = row.get("tenor", "")
            _legs   = row.get("legs", [])
            _exp_date = row.get("expiry_date", "")

            with st.container():
                c1, c2, c3, c4, c5, c6 = st.columns([3, 1.2, 1.2, 1.2, 1.2, 0.8])
                with c1: st.markdown(f"**{idx+1}. {_label}**")
                with c2: st.metric("PV (bp)", f"{_pv_bp:.2f}")
                with c3: st.metric("PV ($k)", f"{_pv/1000:,.1f}")
                with c4: st.metric("Delta ($k)", f"{_delta/1000:,.1f}")
                with c5: st.metric("Notional", f"{_notl:.0f}mm")
                with c6:
                    if st.button("🗑️", key=f"sw_del_{idx}_{_label[:8]}", help="Remove"):
                        st.session_state["swaption_portfolio"].pop(idx)
                        st.session_state["portfolio"] = [p for p in st.session_state["portfolio"]
                                                          if p.get("label") != _label or
                                                          p.get("expiry") != _expiry]
                        _save_portfolio()
                        st.rerun()

                # Legs breakdown for R/R and Ladders
                if _struct in ["Risk Reversal", "Payer Ladder", "Receiver Ladder"] and _legs:
                    _leg_rows = []
                    for _lg in _legs:
                        _qty = int(_lg.get("qty", 1))
                        _side = "Long" if _qty > 0 else "Short"
                        _leg_rows.append({
                            "Leg": _lg.get("name", ""),
                            "B/S": _side,
                            "Strike (%)": f"{float(_lg.get('strike',0)):.4f}",
                            "PV ($k)": f"{float(_lg.get('pv',0))/1000:,.1f}",
                            "PV (bp)": f"{float(_lg.get('pv_bp',0)):.2f}",
                            "Delta ($k)": f"{float(_lg.get('delta',0))/1000:,.1f}",
                        })
                    st.dataframe(pd.DataFrame(_leg_rows), use_container_width=True, hide_index=True)

                # Quick Tix
                if can_quick_tix():
                 with st.expander(f"📋 Quick Tix — {_label[:40]}", expanded=False):
                    # Build dates
                    from datetime import date as _date
                    from dateutil.relativedelta import relativedelta as _rdelta
                    _today = _date.today()
                    try:
                        _exp_y  = label_to_years(str(_expiry))
                        _ten_y  = float(str(_tenor).replace("Y","").replace("y",""))
                        _exp_dt = _today + _rdelta(days=int(_exp_y*365.25))
                        _start_dt = _exp_dt + _rdelta(days=2)  # T+2 start
                        _end_dt = _start_dt + _rdelta(months=int(_ten_y*12))
                        # Quarterly roll dates (first 4)
                        _rolls = []
                        _r = _start_dt + _rdelta(months=3)
                        for _ in range(4):
                            _rolls.append(_r.strftime('%d-%b-%Y'))
                            _r += _rdelta(months=3)
                        _rolls_str = ", ".join(_rolls[:4]) + ("..." if _ten_y > 1 else "")

                        # Net delta direction
                        _net_delta = _delta
                        _delta_dir = "Pay Fixed (hedge = pay fixed IRS)" if _net_delta > 0 else "Rec Fixed (hedge = receive fixed IRS)"
                        _delta_mm = abs(_net_delta) / 10000  # DV01 → notional approx

                        _tix = f"""=== {_ccy} {_struct.upper()} ===
Expiry:      {_exp_dt.strftime('%d-%b-%Y')} ({_expiry})
Swap Start:  {_start_dt.strftime('%d-%b-%Y')}
Swap End:    {_end_dt.strftime('%d-%b-%Y')}
Tenor:       {_tenor}
Rolls:       {_rolls_str}
Fwd Rate:    {_fwd:.4f}%
"""
                        if _legs and _struct in ["Risk Reversal", "Payer Ladder", "Receiver Ladder"]:
                            for _lg in _legs:
                                _s = "BUY" if int(_lg.get("qty",1)) > 0 else "SELL"
                                _tix += f"{_s} {_lg.get('name',''):20s} K={float(_lg.get('strike',0)):.4f}%\n"
                        else:
                            _tix += f"Strike:      {_strike:.4f}%\n"

                        _tix += f"""Premium:     {_pv_bp:.2f} bp ({'+' if _pv>=0 else ''}{_pv/1000:,.1f}k)
Net Delta:   {_net_delta/1000:,.1f}k  →  {_delta_dir}
Notional:    {_notl:.0f}mm {_ccy}"""

                        st.code(_tix, language=None)
                        st.caption("Copy ↑ and paste into chat/IB/email")
                    except Exception as _te:
                        st.caption(f"Could not build tix: {_te}")

                st.markdown("---")

        # Reload into Pricer
        st.markdown("##### Reload into Pricer")
        row_labels = [
            f"{i}: {r.get('label', str(r.get('expiry','?')) + 'x' + str(r.get('tenor','?')))}"
            for i, r in df.iterrows()
        ]
        reload_sel = st.selectbox("Select ticket", ["  —  "] + row_labels, key="sw_reload_sel")
        if st.button("🔄 Reload & Reprice", key="sw_reload_btn") and reload_sel != "  —  ":
            row_idx = int(reload_sel.split(":")[0])
            row = df.loc[row_idx]
            EXPIRY_PRESETS = ["1w","2w","1m","2m","3m","6m","9m","1y","18m","2y","3y","5y","7y","10y","12y","15y","20y","📅 Custom Date..."]
            tenor_options = ["1Y","2Y","3Y","4Y","5Y","6Y","7Y","8Y","9Y","10Y","12Y","15Y","20Y","25Y","30Y"]
            exp_val = str(row.get("expiry", "5y"))
            tenor_val = str(row.get("tenor", "5Y"))
            st.session_state["sw_pending_reload"] = {
                "expiry": exp_val if exp_val in EXPIRY_PRESETS else "5y",
                "tenor": tenor_val if tenor_val in tenor_options else "5Y",
                "structure": str(row.get("structure", "ATM Straddle")),
                "notional_mm": float(row.get("notional_mm", 100)),
                "strike": float(row["strike"]) if "strike" in row and not pd.isna(row.get("strike", float("nan"))) else None,
            }
            st.rerun()


def caps_floors_tab(vol_mode: str):
    st.subheader("Caps & Floors")
    
    # Currency selector
    col_ccy, col_spacer = st.columns([1, 3])
    with col_ccy:
        ccy_select = st.selectbox("📎 Currency", ALL_CURRENCIES, key="cf_ccy")
    
    # Extract actual currency code (remove PENDING)
    ccy = ccy_select.split(" ")[0]
    
    # Check if pending currency selected
    if "PENDING" in ccy_select:
        st.warning(f"├ö├àÔöé {ccy} pricing coming soon. Currently supported: AUD, NZD, USD")
        return

    col_type, col_model = st.columns(2)
    with col_type:
        cf_type = st.selectbox("Instrument", ["Cap", "Floor", "Straddle", "Collar", "Strangle", "Digital Cap", "Digital Floor"], index=2, key="cf_type")
    with col_model:
        model = st.selectbox("Model", ["Normal", "Black"], index=0, key="cf_model")

    col_not, col_first, col_tenor = st.columns(3)
    with col_not:
        notional = st.number_input("Notional (mm)", min_value=1.0, max_value=10000.0, value=100.0, step=10.0, key="cf_not")
    with col_first:
        first_fixing = st.selectbox(
            "First fixing",
            ["1m","2m","3m","6m","9m","1y","18m","2y","3y","4y","5y","7y","10y"],
            index=2,
            key="cf_first",
        )
        first_fixing_y = label_to_years(first_fixing)
    with col_tenor:
        tenor = st.selectbox(
            "Final maturity",
            ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y"],
            index=0,
            key="cf_tenor",
        )
        tenor_y = float(tenor[:-1])  # Final maturity from today

    _cc = st.session_state.get("config_curves", {}).get(ccy)
    curve = _cc if _cc is not None else get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    if curve is not None:
        # Forward swap rate: first_fixing start, full tenor length
        # e.g. 3m x 5Y = fwd starting in 3m for 5Y (NOT 4.75Y)
        fwd, _, _ = forward_and_annuity_from_curve(curve, ccy, first_fixing_y, tenor_y, ois_curve)
        
        # Build QUARTERLY cap schedule   —   MUST use same 1/252 base as bootstrap
        # so pricer T values exactly match the bootstrapped vol curve anchor points.
        # Skip caplets where T_fix <= first_fixing_y (those fixings are "known").
        base = 1.0 / 252.0
        cap_start = base
        cap_end   = tenor_y + base
        sched = []
        t = cap_start
        while t < cap_end - 1e-8:
            t_next = min(t + 0.25, cap_end)
            accrual = t_next - t
            sched.append((t_next, accrual))
            t = t_next
    else:
        sched = [(i * 0.25, 0.25) for i in range(int(tenor_y / 0.25))]
        fwd = 0.04

    fwd_pct = fwd * 100
    
    # Show structure dates
    from datetime import date, timedelta
    from dateutil.relativedelta import relativedelta
    today = date.today()
    # Use proper calendar months for first fixing
    first_fixing_date = today + relativedelta(months=int(first_fixing_y * 12))
    final_maturity = today + relativedelta(months=int(tenor_y * 12))
    num_caplets = sum(1 for T_i, _ in sched if T_i > first_fixing_y + 1.0/252.0)
    
    st.markdown(f"""
    <div style="background: rgba(30,41,59,0.5); border-radius: 8px; padding: 12px; margin: 10px 0;">
        <div style="display: flex; justify-content: space-between; flex-wrap: wrap; gap: 20px;">
            <div><span style="color: #94a3b8;">Forward:</span> <strong>{fwd_pct:.4f}%</strong></div>
            <div><span style="color: #94a3b8;">First Fixing:</span> <strong>{first_fixing_date.strftime('%d-%b-%Y')}</strong></div>
            <div><span style="color: #94a3b8;">Final Maturity:</span> <strong>{final_maturity.strftime('%d-%b-%Y')}</strong></div>
            <div><span style="color: #94a3b8;">Caplets:</span> <strong>{num_caplets}</strong></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Strike inputs based on structure
    st.markdown("##### Strikes")
    strike_pct_2 = fwd_pct  # Initialize
    
    if cf_type in ["Cap", "Floor"]:
        # Strike mode selector
        strike_mode = st.radio("Strike Mode", ["ATM", "10 bp", "25 bp", "50 bp", "100 bp", "Manual"], 
                               horizontal=True, key="cf_strike_mode")
        
        offset_map = {"ATM": 0, "10 bp": 10, "25 bp": 25, "50 bp": 50, "100 bp": 100, "Manual": None}
        offset = offset_map[strike_mode]
        
        if strike_mode == "Manual":
            strike = st.number_input("Strike (%)", value=round(fwd_pct, 4), format="%.4f", key="cf_strike") / 100.0
        else:
            # For Cap, add offset (OTM); for Floor, subtract offset (OTM)
            if cf_type == "Cap":
                strike = fwd + offset/10000.0
            else:
                strike = fwd - offset/10000.0
            st.info(f"Strike: **{strike*100:.4f}%** ({strike_mode} {'OTM' if offset > 0 else ''})")
            
    elif cf_type == "Straddle":
        strike = fwd  # ATM
        st.info(f"ATM Strike: **{fwd_pct:.4f}%**")
        
    elif cf_type == "Collar":
        st.caption("Long Cap + Short Floor - rate protection")
        
        collar_mode = st.radio("Strike Mode", 
                               ["Symmetric (25bp)", "Symmetric (50bp)", "Symmetric (100bp)", "Manual"], 
                               horizontal=True, key="cf_collar_mode")
        
        if "Symmetric" in collar_mode:
            if "25bp" in collar_mode:
                offset = 0.0025
            elif "50bp" in collar_mode:
                offset = 0.0050
            else:
                offset = 0.0100
            
            strike = fwd + offset  # Cap strike
            strike_pct_2 = fwd - offset  # Floor strike
            
            col_k1, col_k2 = st.columns(2)
            with col_k1:
                st.metric("Cap Strike (%)", f"{strike*100:.4f}")
            with col_k2:
                st.metric("Floor Strike (%)", f"{strike_pct_2*100:.4f}")
        else:
            col_k1, col_k2 = st.columns(2)
            with col_k1:
                strike = st.number_input("Cap Strike (%)", value=round(fwd_pct + 0.50, 4), 
                                         format="%.4f", key="cf_strike") / 100.0
            with col_k2:
                strike_pct_2 = st.number_input("Floor Strike (%)", value=round(fwd_pct - 0.50, 4), 
                                               format="%.4f", key="cf_strike_2") / 100.0
        
        width_bp = (strike - strike_pct_2) * 10000
        st.caption(f"Width: **{width_bp:.0f} bp** | Cap +{(strike-fwd)*10000:.0f}bp | Floor {(strike_pct_2-fwd)*10000:.0f}bp")
        
    elif cf_type == "Strangle":
        st.caption("Long OTM Cap + Long OTM Floor (long vol)")
        
        strangle_mode = st.radio("Strike Mode", 
                                 ["Symmetric (25bp)", "Symmetric (50bp)", "Symmetric (100bp)", "Manual"], 
                                 horizontal=True, key="cf_strangle_mode")
        
        if "Symmetric" in strangle_mode:
            if "25bp" in strangle_mode:
                offset = 0.0025
            elif "50bp" in strangle_mode:
                offset = 0.0050
            else:
                offset = 0.0100
            
            strike = fwd + offset  # Cap strike
            strike_pct_2 = fwd - offset  # Floor strike
            
            col_k1, col_k2 = st.columns(2)
            with col_k1:
                st.metric("Cap Strike (%)", f"{strike*100:.4f}")
            with col_k2:
                st.metric("Floor Strike (%)", f"{strike_pct_2*100:.4f}")
        else:
            col_k1, col_k2 = st.columns(2)
            with col_k1:
                strike = st.number_input("Cap Strike (%)", value=round(fwd_pct + 0.50, 4), 
                                         format="%.4f", key="cf_strike") / 100.0
            with col_k2:
                strike_pct_2 = st.number_input("Floor Strike (%)", value=round(fwd_pct - 0.50, 4), 
                                               format="%.4f", key="cf_strike_2") / 100.0
        
        width_bp = (strike - strike_pct_2) * 10000
        st.caption(f"Width: **{width_bp:.0f} bp**")

    elif cf_type in ["Digital Cap", "Digital Floor"]:
        st.caption(f"{'Binary Cap' if cf_type=='Digital Cap' else 'Binary Floor'} — pays **100bp × notional × accrual** on each reset date if BBSW fixes {'at or above' if cf_type=='Digital Cap' else 'at or below'} strike. Priced as tight call spread.")
        strike_mode = st.radio("Strike Mode", ["ATM", "10 bp", "25 bp", "50 bp", "100 bp", "Manual"],
                               horizontal=True, key="cf_dig_strike_mode")
        offset_map = {"ATM": 0, "10 bp": 10, "25 bp": 25, "50 bp": 50, "100 bp": 100, "Manual": None}
        offset = offset_map[strike_mode]
        if strike_mode == "Manual":
            strike = st.number_input("Strike (%)", value=round(fwd_pct, 4), format="%.4f", key="cf_dig_strike_manual") / 100.0
        else:
            strike = fwd + (offset/10000.0 if cf_type == "Digital Cap" else -offset/10000.0)
            st.info(f"Strike: **{strike*100:.4f}%** ({strike_mode} {'OTM' if offset > 0 else 'ATM'})")

    vol_src = st.radio(
        "Vol source",
        ["WEDGES (Edit Spreads)", "Surface (Auto)", "Manual Flat", "Manual Term Structure"],
        horizontal=True,
        index=0,
        key="cf_volsrc",
    )

    # Initialise spread defaults so caplet curve build never hits NameError
    spread_3m1y  = st.session_state.get("cf_spr_3m1y",  10.0)
    spread_1y1y  = st.session_state.get("cf_spr_1y1y",  11.5)
    spread_2y1y  = st.session_state.get("cf_spr_2y1y",  13.0)
    spread_3y1y  = st.session_state.get("cf_spr_3y1y",  17.5)
    spread_4y1y  = st.session_state.get("cf_spr_4y1y",  20.0)
    spread_5y2y  = st.session_state.get("cf_spr_5y2y",  45.0)
    spread_7y3y  = st.session_state.get("cf_spr_7y3y",  50.0)
    spread_10y2y = st.session_state.get("cf_spr_10y2y", 35.0)
    spread_12y3y = st.session_state.get("cf_spr_12y3y", 75.0)
    caplet_vol_curve = None  # Initialize
    
    if vol_src == "Manual Flat":
        vol_input = st.number_input(
            "Vol (normal bp or Black %)",
            value=35.0,
            key="cf_vol",
            help="Normal bp or Black %, depending on selected model",
        )
        caplet_vol_curve = {t: vol_input for t in [0.25, 0.5, 0.75, 1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]}
        st.caption(f"Using flat vol: **{vol_input:.2f}** ({'bp normal' if model=='Normal' else '% Black'})")
        
    elif vol_src == "Manual Term Structure":
        st.caption("Enter caplet vols at anchor maturities (bp)   —   cubic spline interpolated between points")

        MTS_ANCHORS = [
            ("cf_mts_025", "0.25Y"), ("cf_mts_050", "0.50Y"), ("cf_mts_075", "0.75Y"),
            ("cf_mts_1y",  "1Y"),    ("cf_mts_2y",  "2Y"),    ("cf_mts_3y",  "3Y"),
            ("cf_mts_4y",  "4Y"),    ("cf_mts_5y",  "5Y"),    ("cf_mts_7y",  "7Y"),
            ("cf_mts_10y", "10Y"),   ("cf_mts_12y", "12Y"),   ("cf_mts_15y", "15Y"),
        ]
        MTS_MATS = [0.25, 0.5, 0.75, 1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0, 12.0, 15.0]
        MTS_DEFAULTS = [72.0, 73.0, 74.0, 75.0, 78.0, 80.0, 82.0, 84.0, 88.0, 90.0, 88.0, 85.0]

        # Header
        mh1, mh2 = st.columns([1.2, 2.0])
        mh1.markdown("<div style='font-size:0.85rem;font-weight:600;color:#64748b'>Maturity</div>", unsafe_allow_html=True)
        mh2.markdown("<div style='font-size:0.85rem;font-weight:600;color:#64748b'>Vol (bp)</div>", unsafe_allow_html=True)
        st.markdown("<hr style='margin:2px 0 4px 0;border-color:#334155'>", unsafe_allow_html=True)

        anchor_vols = {}
        for (key, lbl), mat, default in zip(MTS_ANCHORS, MTS_MATS, MTS_DEFAULTS):
            mc1, mc2 = st.columns([1.2, 2.0])
            mc1.markdown(f"<div style='font-size:0.95rem;padding-top:8px'>{lbl}</div>", unsafe_allow_html=True)
            val = mc2.number_input("", value=st.session_state.get(key, default), key=key,
                                   format="%.2f", step=0.5, label_visibility="collapsed")
            anchor_vols[mat] = val

        # Cubic spline interpolation
        from scipy.interpolate import CubicSpline
        mats_arr = np.array(list(anchor_vols.keys()))
        vols_arr = np.array(list(anchor_vols.values()))
        cs_mts = CubicSpline(mats_arr, vols_arr)
        caplet_vol_curve = {}
        t = 0.25
        while t <= 15.01:
            caplet_vol_curve[round(t, 2)] = max(float(cs_mts(t)), 1.0)
            t += 0.25


    elif vol_src == "WEDGES (Edit Spreads)":
        st.caption("**Calendar Spread Adjustments** (bp premium: CFS straddle - Swaption straddle)")
        
        # Initialize defaults in session state
        # Committed spreads
        spread_3m1y  = st.session_state["cf_spr_3m1y"]
        spread_1y1y  = st.session_state["cf_spr_1y1y"]
        spread_2y1y  = st.session_state["cf_spr_2y1y"]
        spread_3y1y  = st.session_state["cf_spr_3y1y"]
        spread_4y1y  = st.session_state["cf_spr_4y1y"]
        spread_5y2y  = st.session_state["cf_spr_5y2y"]
        spread_7y3y  = st.session_state["cf_spr_7y3y"]
        spread_10y2y = st.session_state["cf_spr_10y2y"]
        spread_12y3y = st.session_state["cf_spr_12y3y"]

        # Toggle
        if "wedges_expanded" not in st.session_state:
            st.session_state["wedges_expanded"] = True
        icon = "▼ Hide Spreads & SABRs" if st.session_state["wedges_expanded"] else "▶ Show Spreads & SABRs"
        if st.button(icon, key="wedges_toggle"):
            st.session_state["wedges_expanded"] = not st.session_state["wedges_expanded"]
            st.rerun()

        if st.session_state["wedges_expanded"]:
            if "cfs_table_data" not in st.session_state:
                st.session_state["cfs_table_data"] = {}

            # SABR defaults per expiry
            # 11 SABR rows: 3m above spreads, 1y-15y align with 9 spread rows, 20y below
            _SABR_ROWS = [
                ("3m",  0.5, 0.10, 0.60, 0.01),
                ("1y",  0.5, 0.35, 0.50, 0.01),
                ("2y",  0.5, 0.35, 0.45, 0.01),
                ("3y",  0.5, 0.30, 0.40, 0.01),
                ("4y",  0.5, 0.28, 0.38, 0.01),
                ("5y",  0.5, 0.25, 0.35, 0.01),
                ("7y",  0.5, 0.20, 0.30, 0.01),
                ("10y", 0.5, 0.20, 0.30, 0.01),
                ("12y", 0.5, 0.20, 0.30, 0.01),
                ("15y", 0.5, 0.20, 0.28, 0.01),
                ("20y", 0.5, 0.20, 0.28, 0.01),
            ]
            for _exp, _b, _r, _n, _sh in _SABR_ROWS:
                for _p, _d in [("beta",_b),("rho",_r),("nu",_n),("shift",_sh)]:
                    _k = f"cf_sabr_{_exp}_{_p}"
                    if _k not in st.session_state:
                        st.session_state[_k] = _d

            ROW_DATA = [
                ("cf_spr_3m1y",  "3m1y→1Y",     "3m1y",  "3mx1",  "1Y CFS",  spread_3m1y),
                ("cf_spr_1y1y",  "1y1y vs 1x2",    "1y1y",  "1x2",   "2Y CFS",  spread_1y1y),
                ("cf_spr_2y1y",  "2y1y vs 2x3",    "2y1y",  "2x3",   "3Y CFS",  spread_2y1y),
                ("cf_spr_3y1y",  "3y1y vs 3x4",    "3y1y",  "3x4",   "4Y CFS",  spread_3y1y),
                ("cf_spr_4y1y",  "4y1y vs 4x5",    "4y1y",  "4x5",   "5Y CFS",  spread_4y1y),
                ("cf_spr_5y2y",  "5y2y vs 5x7",    "5y2y",  "5x7",   "7Y CFS",  spread_5y2y),
                ("cf_spr_7y3y",  "7y3y vs 7x10",   "7y3y",  "7x10",  "10Y CFS", spread_7y3y),
                ("cf_spr_10y2y", "10y2y vs 10x12", "10y2y", "10x12", "12Y CFS", spread_10y2y),
                ("cf_spr_12y3y", "12y3y vs 12x15", "12y3y", "12x15", "15Y CFS", spread_12y3y),
            ]

            # Two-column layout: spreads left, SABR right
            col_spr, col_sabr = st.columns([2.2, 1.0])

            with col_spr:
                CW = [1.3, 0.55, 1.0, 0.55, 0.05, 0.55, 0.9, 0.55, 0.55, 0.9, 0.65]

                def _h(txt, align="left"):
                    return f"<div style='font-size:0.75rem;font-weight:600;color:#64748b;text-align:{align}'>{txt}</div>"

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                hc = st.columns(CW)
                hc[0].markdown(_h("Wedge"),               unsafe_allow_html=True)
                hc[1].markdown(_h("Last","right"),        unsafe_allow_html=True)
                hc[2].markdown(_h("Current(bp)","center"), unsafe_allow_html=True)
                hc[3].markdown(_h("Chg","right"),         unsafe_allow_html=True)
                hc[5].markdown(_h("Label","right"),       unsafe_allow_html=True)
                hc[6].markdown(_h("Swptn","right"),       unsafe_allow_html=True)
                hc[7].markdown(_h("Wdg","right"),         unsafe_allow_html=True)
                hc[8].markdown(_h("Sprd","right"),        unsafe_allow_html=True)
                hc[9].markdown(_h("FWD CFS","right"),     unsafe_allow_html=True)
                hc[10].markdown(_h("Target","right"),     unsafe_allow_html=True)
                st.markdown("<hr style='margin:2px 0 0 0;border-color:#334155'>", unsafe_allow_html=True)

                new_spread_values = {}
                for spr_key, wedge_lbl, tbl_lbl, tbl_wedge, cfs_lbl, spread in ROW_DATA:
                    last_val = st.session_state[spr_key]
                    cur_val  = st.session_state.get(f"{spr_key}_temp", last_val)
                    tdata  = st.session_state["cfs_table_data"].get(tbl_lbl, {})
                    swpt   = tdata.get("swaption", None)
                    new_val = cur_val  # will be overwritten by number_input below
                    rc = st.columns(CW)
                    fs = "font-size:0.80rem;padding-top:6px"
                    rc[0].markdown(f"<div style='{fs}'>{wedge_lbl}</div>", unsafe_allow_html=True)
                    rc[1].markdown(f"<div style='{fs};text-align:right;color:#94a3b8'>{last_val:.1f}</div>", unsafe_allow_html=True)
                    new_val = rc[2].number_input("", value=cur_val, key=f"{spr_key}_new", format="%.1f", step=0.5, label_visibility="collapsed")
                    delta = new_val - last_val
                    dc = "#22c55e" if delta > 0 else "#ef4444" if delta < 0 else "#94a3b8"
                    rc[3].markdown(f"<div style='{fs};text-align:right;color:{dc}'>{delta:+.1f}</div>", unsafe_allow_html=True)
                    # Use new_val (pending spread) for live CFS display
                    if swpt is not None:
                        cfs = swpt + new_val
                        st.session_state["cfs_table_data"].setdefault(tbl_lbl, {})["cfs_straddle"] = cfs
                        swpt_str = f"{swpt:.4f}"
                        cfs_str  = f"{cfs:.4f}"
                    else:
                        cfs = None
                        swpt_str = "  —  "
                        cfs_str  = "  —  "
                    st.session_state["cfs_table_data"].setdefault(tbl_lbl, {})["cfs_label"] = cfs_lbl
                    rc[5].markdown(f"<div style='{fs};text-align:right;color:#94a3b8'>{tbl_lbl}</div>", unsafe_allow_html=True)
                    rc[6].markdown(f"<div style='{fs};text-align:right;color:#cbd5e1'>{swpt_str}</div>", unsafe_allow_html=True)
                    rc[7].markdown(f"<div style='{fs};text-align:right;color:#94a3b8'>{tbl_wedge}</div>", unsafe_allow_html=True)
                    rc[8].markdown(f"<div style='{fs};text-align:right;color:#94a3b8'>{new_val:.1f}</div>", unsafe_allow_html=True)
                    rc[9].markdown(f"<div style='{fs};text-align:right;color:#38bdf8;font-weight:600'>{cfs_str}</div>", unsafe_allow_html=True)
                    rc[10].markdown(f"<div style='{fs};text-align:right;color:#64748b'>{cfs_lbl}</div>", unsafe_allow_html=True)
                    new_spread_values[spr_key] = new_val
                    st.session_state[f"{spr_key}_temp"] = new_val

            with col_sabr:
                with st.expander("⚙️ SABR Skew Params", expanded=False):
                 st.markdown("<div style='font-size:0.75rem;font-weight:600;color:#64748b;margin-bottom:2px'>SABR Parameters (Caplet Skew)</div>", unsafe_allow_html=True)
                _sh_cols = st.columns([0.5, 0.9, 0.9, 0.9, 0.9])
                for _lbl, _c in zip(["Tenor","B","r,v","x","Shift"], _sh_cols):
                    _c.markdown(f"<div style='font-size:0.75rem;font-weight:600;color:#64748b;text-align:center'>{_lbl}</div>", unsafe_allow_html=True)
                st.markdown("<hr style='margin:2px 0 0 0;border-color:#334155'>", unsafe_allow_html=True)
                for _exp, _b_def, _r_def, _n_def, _sh_def in _SABR_ROWS:
                    sr = st.columns([0.5, 0.9, 0.9, 0.9, 0.9])
                    sr[0].markdown(f"<div style='font-size:0.80rem;padding-top:6px;color:#94a3b8;text-align:center'>{_exp}</div>", unsafe_allow_html=True)
                    st.session_state[f"cf_sabr_{_exp}_beta"]  = sr[1].number_input("", value=st.session_state[f"cf_sabr_{_exp}_beta"],  key=f"s_b_{_exp}", min_value=0.0, max_value=1.0,  step=0.05, format="%.2f", label_visibility="collapsed")
                    st.session_state[f"cf_sabr_{_exp}_rho"]   = sr[2].number_input("", value=st.session_state[f"cf_sabr_{_exp}_rho"],   key=f"s_r_{_exp}", min_value=-1.0, max_value=1.0, step=0.05, format="%.2f", label_visibility="collapsed")
                    st.session_state[f"cf_sabr_{_exp}_nu"]    = sr[3].number_input("", value=st.session_state[f"cf_sabr_{_exp}_nu"],    key=f"s_n_{_exp}", min_value=0.0, max_value=2.0,  step=0.05, format="%.2f", label_visibility="collapsed")
                    st.session_state[f"cf_sabr_{_exp}_shift"] = sr[4].number_input("", value=st.session_state[f"cf_sabr_{_exp}_shift"], key=f"s_sh_{_exp}", min_value=0.0, max_value=0.05, step=0.005, format="%.3f", label_visibility="collapsed")

            new_spread_3m1y  = new_spread_values["cf_spr_3m1y"]
            new_spread_1y1y  = new_spread_values["cf_spr_1y1y"]
            new_spread_2y1y  = new_spread_values["cf_spr_2y1y"]
            new_spread_3y1y  = new_spread_values["cf_spr_3y1y"]
            new_spread_4y1y  = new_spread_values["cf_spr_4y1y"]
            new_spread_5y2y  = new_spread_values["cf_spr_5y2y"]
            new_spread_7y3y  = new_spread_values["cf_spr_7y3y"]
            new_spread_10y2y = new_spread_values["cf_spr_10y2y"]
            new_spread_12y3y = new_spread_values["cf_spr_12y3y"]

            st.markdown("<hr style='margin:4px 0;border-color:#334155'>", unsafe_allow_html=True)

            bl, _, br = st.columns([2, 0.2, 2])
            if bl.button("✅ Calculate CFS from Spreads", key="apply_spreads", type="primary") and require_admin("Edit Spreads"):
                for spr_key, *_ in ROW_DATA:
                    st.session_state[spr_key] = new_spread_values[spr_key]
                # Persist to disk
                try:
                    _spreads_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cfs_spreads.json")
                    with open(_spreads_file, "w") as _f:
                        json.dump({k: st.session_state[k] for k, *_ in ROW_DATA}, _f)
                except Exception:
                    pass
                # Also persist to DB
                if HAS_POSTGRES:
                    try:
                        _user_id = st.session_state.get("username", "default")
                        _spread_data = {k: st.session_state[k] for k, *_ in ROW_DATA}
                        save_user_config(_user_id, "cf_spreads", ccy, _spread_data)
                    except Exception:
                        pass
                # Bust caplet and CFS cache so curve re-builds with new spreads
                st.session_state.pop("_caplet_curve_key", None)
                st.session_state.pop("_atm_cfs_cache_key", None)
                st.session_state.pop("_atm_cfs_rows_cache", None)
                st.rerun()
            if br.button("🔔 Generate Swaption Premiums", key="gen_swpt_prem", type="primary"):
                curve     = get_ccy_curve(ccy)
                atm       = get_working_atm_surface(ccy)
                _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
                ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
                if curve is not None and atm is not None:
                    for lbl, exp, tenor, cfs_lbl in [
                        ("3m1y","3m",1.0,"1Y CFS"),("1y1y","1y",1.0,"2Y CFS"),
                        ("2y1y","2y",1.0,"3Y CFS"),("3y1y","3y",1.0,"4Y CFS"),
                        ("4y1y","4y",1.0,"5Y CFS"),("5y2y","5y",2.0,"7Y CFS"),
                        ("7y3y","7y",3.0,"10Y CFS"),("10y2y","10y",2.0,"12Y CFS"),
                        ("12y3y","12y",3.0,"15Y CFS"),
                    ]:
                        try:
                            vol_bp = get_matrix_value(atm, exp, tenor)
                            if vol_bp is None: continue
                            exp_y = label_to_years(exp)
                            _, ann, _ = forward_and_annuity_from_curve(curve, ccy, exp_y, tenor, ois_curve)
                            sigma_n = vol_bp / 10000.0
                            sqrt_t  = math.sqrt(max(exp_y, 0.001))
                            premium_bp = 2 * 0.3989 * sigma_n * sqrt_t * ann * 10000
                            st.session_state["cfs_table_data"][lbl] = {
                                "swaption": round(premium_bp, 4),
                                "cfs_label": cfs_lbl,
                                "cfs_straddle": round(premium_bp, 4)
                            }
                        except:
                            pass
                st.rerun()

            # ── Publish Wedge Mids to Blotter ─────────────────────────
            st.markdown("<hr style='margin:6px 0;border-color:#1e3050'>", unsafe_allow_html=True)
            _pub_col1, _pub_col2 = st.columns([2, 3])
            with _pub_col1:
                if st.button("📌 Publish Wedge Mids to Options Whiteboard", key="publish_wedge_mids", use_container_width=True) and require_admin("Publish Mids"):
                    if not HAS_POSTGRES:
                        st.error("Database not connected.")
                    else:
                        _mids_to_pub = {}
                        for spr_key, wedge_lbl, tbl_lbl, tbl_wedge, cfs_lbl, spread in ROW_DATA:
                            tdata = st.session_state.get("cfs_table_data", {}).get(tbl_lbl, {})
                            swpt_mid = tdata.get("swaption")
                            if swpt_mid is not None:
                                _mids_to_pub[tbl_lbl] = {
                                    "value": round(float(swpt_mid), 4),
                                    "label": f"{wedge_lbl} swptn mid"
                                }
                            cfs_mid = tdata.get("cfs_straddle")
                            if cfs_mid is not None:
                                _mids_to_pub[f"cfs_{tbl_lbl}"] = {
                                    "value": round(float(cfs_mid), 4),
                                    "label": f"{cfs_lbl} straddle mid"
                                }
                            # Wedge spread
                            _mids_to_pub[f"wedge_{spr_key}"] = {
                                "value": round(float(st.session_state.get(spr_key, 0)), 4),
                                "label": f"{wedge_lbl} wedge bp"
                            }
                        # Add ATM CFS straddle data
                        _atm_cfs_pub = st.session_state.get("atm_cfs_data", {})
                        _mids_to_pub.update(_atm_cfs_pub)
                        _n = publish_blotter_mids(ccy, _mids_to_pub)
                        if _n > 0:
                            st.success(f"✅ Published {_n} mid values to blotter.")
                        else:
                            st.error("No mids published — generate premiums first.")
            with _pub_col2:
                st.caption("Publishes swaption premiums, CFS straddle mids and wedge spreads to Supabase for the live blotter to consume via 🔄 Load Fresh Mids.")

            # ── Publish ALL Blotter Mids (Vol + Premium + FWD ATM) ─────────
            st.markdown("<hr style='margin:6px 0;border-color:#1e3050'>", unsafe_allow_html=True)
            _pb2, _pb3 = st.columns([2, 3])
            with _pb2:
                if st.button("📡 Publish All Mids to Options Whiteboard", key="publish_all_mids",
                             type="primary", use_container_width=True) and require_admin("Publish All Mids"):
                    if not HAS_POSTGRES:
                        st.error("Database not connected.")
                    else:
                        _all_mids = {}
                        _pub_ccy = ccy

                        # ── 1. ATM BP Vol mids from working vol surface ──────
                        _atm_s = get_working_atm_surface(_pub_ccy)
                        if _atm_s is not None:
                            _atm_w = _atm_s.copy()
                            if "Expiry" in _atm_w.columns:
                                _atm_w = _atm_w.set_index("Expiry")
                            for _exp_lbl in _atm_w.index:
                                for _ten_col in _atm_w.columns:
                                    try:
                                        _v = float(_atm_w.loc[_exp_lbl, _ten_col])
                                        if not math.isnan(_v) and _v > 0:
                                            _k = f"vol_{str(_exp_lbl).replace(' ','_')}_{str(_ten_col)}"
                                            _all_mids[_k] = {"value": round(_v, 4),
                                                             "label": f"ATM vol {_exp_lbl} {_ten_col} bp"}
                                    except Exception:
                                        pass

                        # ── 2. BP Premium mids — Bachelier ATM straddle ──────
                        # prem = 2 * N(0) * sigma_n * sqrt(T) * annuity * 10000 bp
                        _curve_pub = get_ccy_curve(_pub_ccy)
                        _ois_pub   = st.session_state.get("config_basis", {}).get(_pub_ccy, {}).get("ois")
                        if _atm_s is not None and _curve_pub is not None:
                            _atm_p = _atm_s.copy()
                            if "Expiry" in _atm_p.columns:
                                _atm_p = _atm_p.set_index("Expiry")
                            for _exp_lbl in _atm_p.index:
                                _exp_y = label_to_years(str(_exp_lbl))
                                if _exp_y <= 0:
                                    continue
                                for _ten_col in _atm_p.columns:
                                    try:
                                        _vbp = float(_atm_p.loc[_exp_lbl, _ten_col])
                                        if math.isnan(_vbp) or _vbp <= 0:
                                            continue
                                        _ten_y = label_to_years(str(_ten_col))
                                        _, _ann, _ = forward_and_annuity_from_curve(
                                            _curve_pub, _pub_ccy, _exp_y, _ten_y, _ois_pub)
                                        _sig_n = _vbp / 10000.0
                                        _prem_bp = 2 * 0.3989422804 * _sig_n * math.sqrt(max(_exp_y, 1e-6)) * _ann * 10000
                                        _k2 = f"prem_{str(_exp_lbl).replace(' ','_')}_{str(_ten_col)}"
                                        _all_mids[_k2] = {"value": round(_prem_bp, 4),
                                                          "label": f"Straddle prem {_exp_lbl} {_ten_col} bp"}
                                    except Exception:
                                        pass

                        # ── 3. FWD ATM IRS mids from forward matrix ──────────
                        _fwd_m = st.session_state.get("fwd_matrix", {}).get(_pub_ccy)
                        if _fwd_m is not None and not _fwd_m.empty:
                            for _fexp in _fwd_m.index:
                                for _ften in _fwd_m.columns:
                                    try:
                                        _fv = float(_fwd_m.loc[_fexp, _ften])
                                        if not math.isnan(_fv) and _fv > 0:
                                            _k3 = f"fwd_{str(_fexp).replace(' ','_')}_{str(_ften)}"
                                            _all_mids[_k3] = {"value": round(_fv, 4),
                                                              "label": f"FWD ATM {_fexp} {_ften} %"}
                                    except Exception:
                                        pass

                        # ── 4. CFS + wedge mids (existing) ───────────────────
                        for _sk, _wl, _tl, _tw, _cl, _sp in ROW_DATA:
                            _td = st.session_state.get("cfs_table_data", {}).get(_tl, {})
                            if _td.get("swaption") is not None:
                                _all_mids[_tl] = {"value": round(float(_td["swaption"]), 4),
                                                  "label": f"{_wl} swptn mid"}
                            if _td.get("cfs_straddle") is not None:
                                _all_mids[f"cfs_{_tl}"] = {"value": round(float(_td["cfs_straddle"]), 4),
                                                            "label": f"{_cl} straddle mid"}
                            _all_mids[f"wedge_{_sk}"] = {"value": round(float(st.session_state.get(_sk, 0)), 4),
                                                          "label": f"{_wl} wedge bp"}
                        _all_mids.update(st.session_state.get("atm_cfs_data", {}))

                        _n_pub = publish_blotter_mids(_pub_ccy, _all_mids)
                        if _n_pub > 0:
                            st.success(f"✅ Published {_n_pub} mids to blotter — vol, premium, FWD ATM + CFS.")
                        else:
                            st.error("Nothing published — load curves and generate forward matrix first.")
            with _pb3:
                st.caption("Publishes ATM BP vols, Bachelier straddle premiums, FWD ATM IRS rates and CFS mids. Generate forward matrix first.")


        # ── ATM CFS Straddle Table ──────────────────────────────────
        st.markdown("<hr style='margin:6px 0;border-color:#1e3050'>", unsafe_allow_html=True)
        if "atm_cfs_expanded" not in st.session_state:
            st.session_state["atm_cfs_expanded"] = True
        _atm_icon = "▼ Hide ATM CFS Straddles" if st.session_state["atm_cfs_expanded"] else "▶ Show ATM CFS Straddles"
        if st.button(_atm_icon, key="atm_cfs_toggle"):
            st.session_state["atm_cfs_expanded"] = not st.session_state["atm_cfs_expanded"]
            st.rerun()

        if st.session_state["atm_cfs_expanded"]:
            _CFS_MAP = [
                (1, "3m1y"), (2, "1y1y"), (3, "2y1y"), (4, "3y1y"), (5, "4y1y"),
                (7, "5y2y"), (10, "7y3y"), (12, "10y2y"), (15, "12y3y"),
            ]
            _cfs_tdata = st.session_state.get("cfs_table_data", {})
            _caplet_vc = st.session_state.get("caplet_vol_curve_aud")
            _curve_local = get_ccy_curve(ccy)
            _ois_tmp_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
            _ois_tmp = _ois_tmp_cb if _ois_tmp_cb is not None else get_basis_curve(ccy, "ois")
            _ois_local = _ois_tmp if (_ois_tmp is not None and not isinstance(_ois_tmp, bool)) else _curve_local

            # Cache: only recalculate when curve id or tdata length changes
            _cfs_id = (ccy, id(_curve_local) if _curve_local is not None else 0, len(_cfs_tdata))
            _cfs_cached = st.session_state.get("_atm_cfs_cache_key") == _cfs_id
            if _cfs_cached and st.session_state.get("_atm_cfs_rows_cache"):
                st.dataframe(pd.DataFrame(st.session_state["_atm_cfs_rows_cache"]),
                             use_container_width=True, hide_index=True)
            elif _cfs_tdata and _curve_local is not None:
                st.session_state["_atm_cfs_cache_key"] = _cfs_id
                from datetime import date
                from dateutil.relativedelta import relativedelta
                _today = date.today()
                _start_dt = _today + relativedelta(months=3, days=1)
                _spot_dt  = _today + relativedelta(days=1)
                _fwd_start_y = 0.25
                _cum_prem = 0.0
                _atm_cfs_rows = []
                _atm_cfs_data = {}

                for _t, _key in _CFS_MAP:
                    try:
                        _end_dt = _spot_dt + relativedelta(years=_t)
                        # ATM fwd swap rate: cap tenor _t starts 3m fwd, so underlying swap = _t - 0.25
                        # e.g. 1Y cap = 3m start, 9m swap; 2Y cap = 3m start, 21m swap
                        _cap_swap_tenor = float(_t) - 0.25
                        _fwd_rate = None
                        if _curve_local is not None:
                            import numpy as _np
                            _cx = _curve_local["MaturityY"].to_numpy().astype(float)
                            _cy = _curve_local["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            _ox = _oy = None
                            if _ois_local is not None:
                                _oc2 = _ois_local.drop(columns=["_source_date"], errors="ignore") if hasattr(_ois_local, "drop") else _ois_local
                                if hasattr(_oc2, "columns") and "MaturityY" in _oc2.columns:
                                    _ox = _oc2["MaturityY"].to_numpy().astype(float)
                                    _oy = _oc2["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            _b6v3 = get_basis_curve(ccy, "6v3")
                            _bx = _b6v3["MaturityY"].to_numpy().astype(float) if _b6v3 is not None else None
                            _by = _b6v3["BasisBp"].to_numpy().astype(float) if _b6v3 is not None else None
                            # Use pure QQ zero curve for AUD cap strikes (caps are Q/Q)
                            _zc_qq_local = st.session_state.get("_aud_zc_qq")
                            if ccy == "AUD" and _zc_qq_local:
                                import numpy as _np2
                                _zqx = _np2.array(sorted(_zc_qq_local.keys()))
                                _zqy = _np2.array([_zc_qq_local[k] / 100.0 for k in _zqx])
                                _fwd_rate = fast_forward_rate(_zqx, _zqy, _fwd_start_y, _cap_swap_tenor, ccy, freq_override=0.25, ois_x=_ox, ois_y=_oy)
                            else:
                                _fwd_rate = fast_forward_rate(_cx, _cy, _fwd_start_y, _cap_swap_tenor, ccy, freq_override=None, ois_x=_ox, ois_y=_oy, basis6v3_x=_bx, basis6v3_y=_by)
                        # Cumulative CFS straddle
                        _wedge_straddle = _cfs_tdata.get(_key, {}).get("cfs_straddle")
                        if _wedge_straddle is not None:
                            _cum_prem += float(_wedge_straddle)
                        _straddle_prem = round(_cum_prem, 4)
                        # Flat vol from caplet curve
                        _flat_vol = None
                        if _caplet_vc:
                            _mats = sorted(_caplet_vc.keys())
                            if float(_t) in _caplet_vc:
                                _flat_vol = _caplet_vc[float(_t)]
                            elif _t <= _mats[0]:
                                _flat_vol = _caplet_vc[_mats[0]]
                            elif _t >= _mats[-1]:
                                _flat_vol = _caplet_vc[_mats[-1]]
                            else:
                                for _j in range(len(_mats)-1):
                                    if _mats[_j] <= _t <= _mats[_j+1]:
                                        _a = (_t - _mats[_j]) / (_mats[_j+1] - _mats[_j])
                                        _flat_vol = _caplet_vc[_mats[_j]] + _a * (_caplet_vc[_mats[_j+1]] - _caplet_vc[_mats[_j]])
                                        break
                        _atm_cfs_rows.append({
                            "Tenor": f"{_t}Y",
                            "Start": _start_dt.strftime("%d %b %y"),
                            "End":   _end_dt.strftime("%d %b %y"),
                            "ATM Fwd %": f"{_fwd_rate*100:.3f}" if _fwd_rate else "—",
                            "Straddle bp": f"{_straddle_prem:.4f}",
                            "Flat Vol bp": f"{_flat_vol:.1f}" if _flat_vol else "—",
                        })
                        _atm_cfs_data[f"cf_straddle_{_t}y"] = {"value": _straddle_prem, "label": f"{_t}Y ATM CFS straddle"}
                        if _flat_vol: _atm_cfs_data[f"cf_vol_{_t}y"] = {"value": round(float(_flat_vol), 2), "label": f"{_t}Y ATM CFS flat vol"}
                        if _fwd_rate: _atm_cfs_data[f"cf_strike_{_t}y"] = {"value": round(float(_fwd_rate*100), 4), "label": f"{_t}Y ATM fwd strike"}
                    except Exception:
                        _atm_cfs_rows.append({"Tenor": f"{_t}Y", "Start": "—", "End": "—", "ATM Fwd %": "—", "Straddle bp": "—", "Flat Vol bp": "—"})

                st.session_state["atm_cfs_data"] = _atm_cfs_data
                st.session_state["_atm_cfs_rows_cache"] = _atm_cfs_rows
                if _atm_cfs_rows:
                    st.dataframe(pd.DataFrame(_atm_cfs_rows), use_container_width=True, hide_index=True)
            else:
                st.info("Generate swaption premiums first to compute ATM CFS straddles.")

                # CALCULATE CFS VALUES BEFORE BUILDING VOL CURVE
        if "cfs_table_data" in st.session_state:
            spreads_map = {
                "3m1y": spread_3m1y, "1y1y": spread_1y1y, "2y1y": spread_2y1y,
                "3y1y": spread_3y1y, "4y1y": spread_4y1y, "5y2y": spread_5y2y,
                "7y3y": spread_7y3y, "10y2y": spread_10y2y, "12y3y": spread_12y3y
            }
            for label in ["3m1y", "1y1y", "2y1y", "3y1y", "4y1y", "5y2y", "7y3y", "10y2y", "12y3y"]:
                if label in st.session_state["cfs_table_data"]:
                    swpt = st.session_state["cfs_table_data"][label].get("swaption", "")
                    spread = spreads_map.get(label, 0)
                    if swpt != "":
                        cfs_straddle = swpt + spread
                        st.session_state["cfs_table_data"][label]["cfs_straddle"] = cfs_straddle
        
        # Build caplet curve — only rebuild when spreads or ATM surface change
        atm = get_working_atm_surface(ccy)
        # Use stable ATM hash — id() changes every render if object is recreated
        _atm_hash = st.session_state.get(f"_atm_hash_{ccy}", 0)
        _caplet_key = (spread_3m1y, spread_1y1y, spread_2y1y, spread_3y1y,
                       spread_4y1y, spread_5y2y, spread_7y3y, spread_10y2y, spread_12y3y,
                       _atm_hash)
        _cached_key = st.session_state.get("_caplet_curve_key")
        if _cached_key != _caplet_key or st.session_state.get("caplet_vol_curve_aud") is None:
            caplet_vol_curve = build_caplet_vol_curve(
                ccy, atm, None,
                spread_3m1y=spread_3m1y,
                spread_1y1y=spread_1y1y,
                spread_2y1y=spread_2y1y,
                spread_3y1y=spread_3y1y,
                spread_4y1y=spread_4y1y,
                spread_5y2y=spread_5y2y,
                spread_7y3y=spread_7y3y,
                spread_10y2y=spread_10y2y,
                spread_12y3y=spread_12y3y,
            )
            st.session_state["_caplet_curve_key"] = _caplet_key
        else:
            caplet_vol_curve = st.session_state.get("caplet_vol_curve_aud")
        
        if caplet_vol_curve and len(caplet_vol_curve) > 0:
            st.session_state["caplet_vol_curve_aud"] = caplet_vol_curve
        if caplet_vol_curve is None or len(caplet_vol_curve) == 0:
            caplet_vol_curve = st.session_state.get("caplet_vol_curve_aud")
        if caplet_vol_curve is None or len(caplet_vol_curve) == 0:
            st.warning("No ATM surface found. Falling back to 35bp flat.")
            caplet_vol_curve = {t: 35.0 for t in [0.25, 0.5, 0.75, 1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]}
        
        # Show the curve
        if caplet_vol_curve:
            with st.expander("📊 Resulting Caplet Vol Curve", expanded=False):
                # Show exact bootstrapped vols in table
                curve_data = []
                for t in sorted(caplet_vol_curve.keys()):
                    curve_data.append({"Maturity (Y)": f"{t:.2f}", "Vol (bp)": f"{caplet_vol_curve[t]:.2f}"})
                st.dataframe(pd.DataFrame(curve_data), use_container_width=True, hide_index=True)
                
                from scipy.interpolate import CubicSpline
                import plotly.graph_objects as _pgo

                maturities = np.array(sorted(caplet_vol_curve.keys()))
                vols       = np.array([caplet_vol_curve[t] for t in maturities])

                fig = _pgo.Figure()

                # Smooth spline line
                if len(maturities) >= 4:
                    cs       = CubicSpline(maturities, vols)
                    mat_fine = np.linspace(maturities[0], maturities[-1], 300)
                    vol_fine = cs(mat_fine)
                    fig.add_trace(_pgo.Scatter(
                        x=mat_fine, y=vol_fine, mode="lines",
                        line=dict(color="#2563eb", width=2),
                        hoverinfo="skip", name="Spline"
                    ))

                # Dots with hover
                hover_text = [f"T = {t:.2f}Y<br>Vol = {v:.2f} bp" for t, v in zip(maturities, vols)]
                fig.add_trace(_pgo.Scatter(
                    x=maturities, y=vols, mode="markers",
                    marker=dict(color="#2563eb", size=5, line=dict(color="white", width=1)),
                    text=hover_text, hovertemplate="%{text}<extra></extra>",
                    name="Bootstrapped"
                ))

                max_yr = int(np.ceil(maturities[-1]))
                fig.update_layout(
                    xaxis=dict(
                        title="Maturity (Years)",
                        tickmode="array",
                        tickvals=list(range(0, max_yr + 1)),
                        ticktext=[str(y) for y in range(0, max_yr + 1)],
                        gridcolor="#e2e8f0",
                    ),
                    yaxis=dict(title="Vol (bp)", gridcolor="#e2e8f0"),
                    plot_bgcolor="#f8fafc", paper_bgcolor="white",
                    margin=dict(l=50, r=20, t=30, b=40),
                    height=320,
                    showlegend=False,
                    hovermode="closest",
                )
                st.plotly_chart(fig, use_container_width=True)
        
    else:  # Surface (Auto)
        atm = get_working_atm_surface(ccy)
        _surf_key = id(atm) if atm is not None else 0
        _cached_surf_key = st.session_state.get(f"_surf_caplet_key_{ccy}")
        if _cached_surf_key != _surf_key or st.session_state.get(f"_surf_caplet_curve_{ccy}") is None:
            caplet_vol_curve = build_caplet_vol_curve_from_surface(ccy, atm)
            st.session_state[f"_surf_caplet_key_{ccy}"] = _surf_key
            st.session_state[f"_surf_caplet_curve_{ccy}"] = caplet_vol_curve
        else:
            caplet_vol_curve = st.session_state[f"_surf_caplet_curve_{ccy}"]
        
        if caplet_vol_curve is None or len(caplet_vol_curve) == 0:
            st.warning("Unable to build caplet vol curve from surface. Falling back to 35bp flat.")
            caplet_vol_curve = {t: 35.0 for t in [0.25, 0.5, 0.75, 1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]}
        else:
            with st.expander("📊 Caplet Vol Term Structure (Auto)", expanded=False):
                curve_data = []
                for t in sorted(caplet_vol_curve.keys()):
                    curve_data.append({"Maturity (Y)": f"{t:.2f}", "Vol (bp)": f"{caplet_vol_curve[t]:.2f}"})
                st.dataframe(pd.DataFrame(curve_data), use_container_width=True, hide_index=True)

                import plotly.graph_objects as _pgo

                maturities = np.array(sorted(caplet_vol_curve.keys()))
                vols       = np.array([caplet_vol_curve[t] for t in maturities])

                fig2 = _pgo.Figure()
                fig2.add_trace(_pgo.Scatter(
                    x=maturities, y=vols, mode="lines",
                    line=dict(color="#2563eb", width=2),
                    hoverinfo="skip", name="Curve"
                ))
                hover_text2 = [f"T = {t:.2f}Y<br>Vol = {v:.2f} bp" for t, v in zip(maturities, vols)]
                fig2.add_trace(_pgo.Scatter(
                    x=maturities, y=vols, mode="markers",
                    marker=dict(color="#2563eb", size=5, line=dict(color="white", width=1)),
                    text=hover_text2, hovertemplate="%{text}<extra></extra>",
                    name="Points"
                ))
                max_yr2 = int(np.ceil(maturities[-1]))
                fig2.update_layout(
                    xaxis=dict(
                        title="Maturity (Years)",
                        tickmode="array",
                        tickvals=list(range(0, max_yr2 + 1)),
                        ticktext=[str(y) for y in range(0, max_yr2 + 1)],
                        gridcolor="#e2e8f0",
                    ),
                    yaxis=dict(title="Vol (bp)", gridcolor="#e2e8f0"),
                    plot_bgcolor="#f8fafc", paper_bgcolor="white",
                    margin=dict(l=50, r=20, t=30, b=40),
                    height=320,
                    showlegend=False,
                    hovermode="closest",
                )
                st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    
    if st.button(" Price Cap/Floor", key="cf_price", type="primary", disabled=is_trainee()):
        try:
            pv_total = 0.0
            delta_total = 0.0
            vega_total = 0.0
            gamma_total = 0.0
            
            # Store individual caplet/floorlet results
            caplet_details = []
            
            # Helper to price a strip and collect caplet details
            def price_strip(strike_val, is_cap_flag, leg_name=""):
                pv = delta = vega = gamma = 0.0
                caplets = []
                
                for i, (T_i, accrual) in enumerate(sched):
                    # Skip caplets in the "known" period (at or before first fixing)
                    if T_i <= first_fixing_y + 1.0 / 252.0:
                        continue
                    
                    # Get caplet-specific vol from term structure
                    caplet_vol_bp = get_caplet_vol_for_fixing(caplet_vol_curve, T_i)
                    if caplet_vol_bp is None:
                        caplet_vol_bp = 35.0  # Fallback
                    
                    # Convert to sigma based on model
                    if model == "Normal":
                        sigma = caplet_vol_bp / 10000.0
                    else:
                        sigma = caplet_vol_bp / 100.0
                    
                    # Get discount rate from OIS curve
                    disc_rate = interpolate_zero(ois_curve, T_i)
                    
                    # Calculate individual forward for THIS caplet period (3m)
                    period_start = max(T_i - 0.25, 0.001)
                    period_tenor = 0.25
                    F_i, _, _ = forward_and_annuity_from_curve(curve, ccy, period_start, period_tenor, ois_curve)
                    
                    # For ATM (strike_val == fwd), use F_i as strike to ensure F=K for each caplet
                    if abs(strike_val - fwd) < 0.0001:  # ATM straddle
                        caplet_strike = F_i
                    else:
                        caplet_strike = strike_val
                    
                    if model == "Black":
                        res = black_caplet(notional * 1e6, accrual, F_i, caplet_strike, sigma, T_i, disc_rate, is_cap=is_cap_flag)
                    else:
                        res = bachelier_caplet(notional * 1e6, accrual, F_i, caplet_strike, sigma, T_i, disc_rate, is_cap=is_cap_flag)
                    pv += res["pv"]
                    delta += res["delta"]
                    vega += res["vega"]
                    gamma += res["gamma"]
                    
                    # Fixing date derived from T_i (years from today)
                    from dateutil.relativedelta import relativedelta
                    fixing_months = round((T_i - base) * 12)
                    caplet_date = today + relativedelta(months=fixing_months)
                    
                    caplets.append({
                        "Leg": leg_name,
                        "#": i,
                        "Fixing": caplet_date.strftime('%d-%b-%Y'),
                        "T (yrs)": f"{T_i:.4f}",
                        "Fwd (%)": f"{F_i*100:.4f}",
                        "Vol (bp)": f"{caplet_vol_bp:.2f}",
                        "Accrual": f"{accrual:.4f}",
                        "PV": f"${res['pv']:,.0f}",
                        "Delta": f"{res['delta']:,.0f}",
                    })
                return {"pv": pv, "delta": delta, "vega": vega, "gamma": gamma, "caplets": caplets}

            def price_digital_strip(strike_val, is_cap_flag, leg_name=""):
                """Digital cap/floor — 100bp × notional × accrual payout per period.
                Priced as tight call spread: [V(K-0.5bp) - V(K+0.5bp)] / 1bp × 100bp"""
                eps = 0.0001  # 1bp spread
                lo = price_strip(strike_val - eps/2, is_cap_flag, leg_name)
                hi = price_strip(strike_val + eps/2, is_cap_flag, leg_name)
                scale = 0.01 / eps  # 100bp payout / 1bp spread width = 100
                pv    = (lo["pv"]    - hi["pv"])    * scale
                delta = (lo["delta"] - hi["delta"]) * scale
                vega  = (lo["vega"]  - hi["vega"])  * scale
                gamma = (lo["gamma"] - hi["gamma"]) * scale
                return {"pv": pv, "delta": delta, "vega": vega, "gamma": gamma, "caplets": lo["caplets"]}
            
            legs = []
            
            if cf_type == "Cap":
                res = price_strip(strike, True, "Cap")
                pv_total, delta_total, vega_total, gamma_total = res["pv"], res["delta"], res["vega"], res["gamma"]
                caplet_details = res["caplets"]
                legs.append(("Cap", strike*100, 1, res))
                label = f"Cap {first_fixing}-{tenor} K={strike*100:.2f}%"
                
            elif cf_type == "Floor":
                res = price_strip(strike, False, "Floor")
                pv_total, delta_total, vega_total, gamma_total = res["pv"], res["delta"], res["vega"], res["gamma"]
                caplet_details = res["caplets"]
                legs.append(("Floor", strike*100, 1, res))
                label = f"Floor {first_fixing}-{tenor} K={strike*100:.2f}%"
                
            elif cf_type == "Straddle":
                res_cap = price_strip(fwd, True, "Cap")
                res_floor = price_strip(fwd, False, "Floor")
                pv_total = res_cap["pv"] + res_floor["pv"]
                delta_total = res_cap["delta"] + res_floor["delta"]
                vega_total = res_cap["vega"] + res_floor["vega"]
                gamma_total = res_cap["gamma"] + res_floor["gamma"]
                caplet_details = res_cap["caplets"] + res_floor["caplets"]
                legs.append(("Cap", fwd*100, 1, res_cap))
                legs.append(("Floor", fwd*100, 1, res_floor))
                label = f"Straddle {first_fixing}-{tenor} ATM"
                
            elif cf_type == "Collar":
                res_cap = price_strip(strike, True, "Long Cap")
                res_floor = price_strip(strike_pct_2, False, "Short Floor")
                # Long cap, short floor (protection against rising rates)
                pv_total = res_cap["pv"] - res_floor["pv"]
                delta_total = res_cap["delta"] - res_floor["delta"]
                vega_total = res_cap["vega"] - res_floor["vega"]
                gamma_total = res_cap["gamma"] - res_floor["gamma"]
                caplet_details = res_cap["caplets"] + res_floor["caplets"]
                legs.append(("Long Cap", strike*100, 1, res_cap))
                legs.append(("Short Floor", strike_pct_2*100, -1, res_floor))
                label = f"Collar {first_fixing}-{tenor} ({strike_pct_2*100:.2f}/{strike*100:.2f})"
                
            elif cf_type == "Strangle":
                res_cap = price_strip(strike, True, "OTM Cap")
                res_floor = price_strip(strike_pct_2, False, "OTM Floor")
                pv_total = res_cap["pv"] + res_floor["pv"]
                delta_total = res_cap["delta"] + res_floor["delta"]
                vega_total = res_cap["vega"] + res_floor["vega"]
                gamma_total = res_cap["gamma"] + res_floor["gamma"]
                caplet_details = res_cap["caplets"] + res_floor["caplets"]
                legs.append(("OTM Cap", strike*100, 1, res_cap))
                legs.append(("OTM Floor", strike_pct_2*100, 1, res_floor))
                label = f"Strangle {first_fixing}-{tenor} ({strike_pct_2*100:.2f}/{strike*100:.2f})"

            elif cf_type == "Digital Cap":
                res = price_digital_strip(strike, True, "Digital Cap")
                pv_total, delta_total, vega_total, gamma_total = res["pv"], res["delta"], res["vega"], res["gamma"]
                caplet_details = res["caplets"]
                legs.append(("Digital Cap", strike*100, 1, res))
                label = f"Digital Cap {first_fixing}-{tenor} K={strike*100:.2f}% | 100bp payout per fixing"

            elif cf_type == "Digital Floor":
                res = price_digital_strip(strike, False, "Digital Floor")
                pv_total, delta_total, vega_total, gamma_total = res["pv"], res["delta"], res["vega"], res["gamma"]
                caplet_details = res["caplets"]
                legs.append(("Digital Floor", strike*100, 1, res))
                label = f"Digital Floor {first_fixing}-{tenor} K={strike*100:.2f}% | 100bp payout per fixing"

            # Calculate premium in bp: (PV / Notional) * 10000
            pv_bp = (pv_total / (notional * 1e6)) * 10000.0 if notional > 0 else 0.0
            
            # one_bp = sum of caplet annuities = sum(notional * accrual_i * df(Ti) * 0.0001)
            # Matches how bachelier_caplet computes delta so delta_ratio = 50% ATM exactly
            one_bp_annuity = 0.0
            for _Ti, _acc in sched:
                if _Ti <= first_fixing_y + 1.0/252.0:
                    continue
                _df_i = math.exp(-interpolate_zero(ois_curve, _Ti) * _Ti)
                one_bp_annuity += notional * 1e6 * _acc * _df_i * 0.0001

            st.success(f"✅ Priced: **{label}** | PV = ${pv_total:,.0f} ({pv_bp:.4f} bp)")
            
            # Store for display
            st.session_state["cf_last_result"] = {
                "legs": legs,
                "caplet_details": caplet_details,
                "pv_total": pv_total,
                "pv_bp": pv_bp,
                "delta_total": delta_total,
                "gamma_total": gamma_total,
                "vega_total": vega_total,
                "one_bp": one_bp_annuity,
                "label": label,
                "notional": notional,
            }

            st.session_state["portfolio"].append(
                dict(
                    instrument_type="Cap/Floor",
                    currency=ccy,
                    structure=cf_type,
                    expiry=first_fixing,
                    tenor=tenor,
                    model=model,
                    vol_input="Caplet Term Structure",
                    notional_mm=notional,
                    strike=strike * 100.0,
                    forward=fwd * 100.0,
                    pv=pv_total,
                    pv_bp=pv_bp,
                    delta=delta_total,
                    gamma=gamma_total,
                    vega=vega_total,
                    theta=0.0,
                    bpv=one_bp_annuity,
                    label=label,
                )
            )
        except Exception as e:
            st.error(f" Pricing error: {e}")
            import traceback
            st.code(traceback.format_exc())
    
    # Display results if available
    if "cf_last_result" in st.session_state:
        r = st.session_state["cf_last_result"]
        
        st.markdown("###  Results")
        col_params, col_greeks = st.columns(2)
        
        with col_params:
            # Show leg breakdown
            if len(r["legs"]) > 0:
                st.markdown("##### Leg Breakdown")
                leg_data = []
                notional_val = r.get("notional", 100.0)
                for leg_name, leg_strike, leg_mult, leg_res in r["legs"]:
                    leg_pv_bp = ((leg_res['pv'] * leg_mult) / (notional_val * 1e6)) * 10000.0 if notional_val > 0 else 0
                    leg_data.append({
                        "Leg": leg_name,
                        "Strike (%)": f"{leg_strike:.4f}",
                        "Premium (bp)": f"{leg_pv_bp:.4f}",
                        "PV": f"${leg_res['pv'] * leg_mult:,.0f}",
                        "Delta": f"{leg_res['delta'] * leg_mult:,.0f}"
                    })
                st.dataframe(pd.DataFrame(leg_data), use_container_width=True, hide_index=True)
        
        with col_greeks:
            st.markdown("##### Valuation")
            st.metric("Premium (bp)", f"{r['pv_bp']:.4f}")
            st.metric("Total PV", f"${r['pv_total']:,.0f}")
            
            st.markdown("##### Greeks (Net)")
            # Cap/Floor delta   —   same convention as swaptions
            # delta_total = sum of caplet DV01s
            # delta_ratio = delta_total / one_bp = hedge % of notional (50% ATM)
            # delta_swap  = delta_ratio * notional_$ = notional-equivalent swap hedge
            _notional_d = r['notional'] * 1e6
            _one_bp = r['one_bp'] if r['one_bp'] > 0 else 1e-8
            delta_ratio = r['delta_total'] / _one_bp          # e.g. 0.50 ATM
            delta_swap  = delta_ratio * _notional_d            # e.g. 50mm for 100mm notional
            delta_dv01  = r['delta_total']                     # already in $/bp
            greeks_df = pd.DataFrame({
                "Greek": ["Delta (swap hedge)", "Delta % notional", "Delta DV01 ($/bp)", "Gamma ($/bp)", "Vega ($/bp vol)", "BPV ($/bp)"],
                "Value": [
                    f"${delta_swap:,.0f}",
                    f"{delta_ratio*100:.1f}%",
                    f"{delta_dv01:,.1f}",
                    f"{r['gamma_total']:,.2f}",
                    f"{r['vega_total']:,.1f}",
                    f"{r['one_bp']:,.1f}"
                ],
                "Per 1mm notional": [
                    f"${delta_swap/r['notional']:,.0f}",
                    f"{delta_ratio*100:.1f}%",
                    f"{delta_dv01/r['notional']:,.1f}",
                    f"{r['gamma_total']/r['notional']:,.3f}",
                    f"{r['vega_total']/r['notional']:,.1f}",
                    f"{r['one_bp']/r['notional']:,.1f}"
                ]
            })
            st.dataframe(greeks_df, use_container_width=True, hide_index=True)
        
        # Caplet/Floorlet breakdown in expander
        if r["caplet_details"]:
            with st.expander(" Caplet/Floorlet Breakdown", expanded=False):
                st.dataframe(pd.DataFrame(r["caplet_details"]), use_container_width=True, hide_index=True)

    # ── Cap/Floor Portfolio Blotter ───────────────────────────────────────────
    _cf_port = [t for t in st.session_state.get("portfolio", []) if t.get("instrument_type") == "Cap/Floor"]
    if _cf_port:
        st.markdown("---")
        _ph1, _ph2 = st.columns([3, 1])
        with _ph1:
            st.markdown("### Cap/Floor Blotter")
        with _ph2:
            if st.button("🗑️ Clear All", key="cf_clear_portfolio"):
                st.session_state["portfolio"] = [t for t in st.session_state.get("portfolio", []) if t.get("instrument_type") != "Cap/Floor"]
                st.rerun()

        _df = pd.DataFrame(_cf_port)
        _df["_sort"] = _df["expiry"].apply(lambda e: label_to_years(str(e)))
        _df = _df.sort_values("_sort").reset_index(drop=True)
        _df_disp = _df.drop(columns=["_sort"]).copy()
        for _col in ["pv", "delta", "vega", "gamma"]:
            if _col in _df_disp.columns:
                _df_disp[_col] = (_df_disp[_col] / 1e3).round(1)
        if "pv_bp" in _df_disp.columns:
            _df_disp["pv_bp"] = _df_disp["pv_bp"].round(4)
        _df_disp.rename(columns={
            "pv": "PV (k)", "pv_bp": "PV (bp)",
            "delta": "Delta (k)", "vega": "Vega (k)",
            "gamma": "Gamma (k)", "instrument_type": "Type",
        }, inplace=True)
        _show_cols = [c for c in ["Type","currency","structure","expiry","tenor","strike","forward","notional_mm","PV (bp)","PV (k)","Delta (k)","Vega (k)","Gamma (k)","model"] if c in _df_disp.columns]
        st.dataframe(_df_disp[_show_cols], use_container_width=True, hide_index=True)

        # Net greeks
        _net_pv   = sum(t.get("pv", 0) for t in _cf_port)
        _net_delt = sum(t.get("delta", 0) for t in _cf_port)
        _net_vega = sum(t.get("vega", 0) for t in _cf_port)
        _net_bp   = sum(t.get("pv_bp", 0) for t in _cf_port)
        _nc1, _nc2, _nc3, _nc4 = st.columns(4)
        _nc1.metric("Net PV (bp)", f"{_net_bp:.2f}")
        _nc2.metric("Net PV ($)", f"${_net_pv:,.0f}")
        _nc3.metric("Net Delta ($)", f"${_net_delt:,.0f}")
        _nc4.metric("Net Vega ($)", f"${_net_vega:,.0f}")


def exotics_tab(vol_mode: str):
    st.subheader("Exotics / Structured Rates")

    # Currency selector
    col_ccy, col_spacer = st.columns([1, 3])
    with col_ccy:
        ccy_select = st.selectbox("📎 Currency", ALL_CURRENCIES, key="ex_ccy")
    ccy = ccy_select.split(" ")[0]
    if "PENDING" in ccy_select:
        st.warning(f"├ö├àÔöé {ccy} pricing coming soon. Currently supported: AUD, NZD, USD")
        return

    curve     = get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    basis_6v3 = get_basis_curve(ccy, "6v3")
    atm       = get_working_atm_surface(ccy)
    _, a_m, b_m, r_m, n_m = get_ccy_vol_data(ccy)

    # ── Correlation & CMS Config (always visible   —   check before pricing) ──────
    with st.expander("⚙️ Correlation Matrix & CMS Convexity Config", expanded=False):
        st.markdown("**Swap Rate Correlation Matrix**   —   used by Spread Options and CMS Spread products")
        st.caption("Edit cells then click Save. These persist across sessions.")

        cfg_changed = False

        # Render editable upper-triangular matrix as a grid
        # Row = shorter tenor, Col = longer tenor
        header_cols = st.columns([1] + [1]*len(_CORR_TENORS))
        header_cols[0].markdown("**ρ,ν**")
        for ci, t in enumerate(_CORR_TENORS):
            header_cols[ci+1].markdown(f"**{t}**")

        for ri, t_row in enumerate(_CORR_TENORS):
            row_cols = st.columns([1] + [1]*len(_CORR_TENORS))
            row_cols[0].markdown(f"**{t_row}**")
            for ci, t_col in enumerate(_CORR_TENORS):
                if ci < ri:
                    row_cols[ci+1].markdown("")          # below diagonal   —   blank
                elif ci == ri:
                    row_cols[ci+1].markdown("1.000")     # diagonal
                else:
                    key_corr = f"{t_row}_{t_col}"
                    alt_key  = f"{t_col}_{t_row}"
                    dict_key = key_corr if key_corr in _CORR_DEFAULTS else alt_key
                    ss_key   = f"corr_{dict_key}"
                    current  = float(st.session_state.get(ss_key, _CORR_DEFAULTS.get(dict_key, 0.90)))
                    new_val  = row_cols[ci+1].number_input(
                        "", min_value=0.0, max_value=1.0, value=current,
                        step=0.001, format="%.3f", key=f"cfg_{ss_key}",
                        label_visibility="collapsed")
                    if abs(new_val - current) > 1e-6:
                        st.session_state[ss_key] = new_val
                        cfg_changed = True

        st.markdown("---")
        st.markdown("**CMS Convexity Adjustment Bumps (bp)**   —   added to Hagan formula output")
        st.caption("Positive = increase convexity adj above model. Zero = use Hagan formula as-is.")
        bump_cols = st.columns(len(_CORR_TENORS))
        for ci, t in enumerate(_CORR_TENORS):
            ss_key_b = f"cms_bump_{t}"
            current_b = float(st.session_state.get(ss_key_b, 0.0))
            new_b = bump_cols[ci].number_input(
                t, min_value=-20.0, max_value=20.0, value=current_b,
                step=0.1, format="%.1f", key=f"cfg_{ss_key_b}")
            if abs(new_b - current_b) > 1e-6:
                st.session_state[ss_key_b] = new_b
                cfg_changed = True

        save_col, reset_col, _ = st.columns([1, 1, 4])
        with save_col:
            if st.button("💾 Save Config", key="ex_save_cfg"):
                _save_exotics_config()
                st.success("Saved to exotics_config.json")
        with reset_col:
            if st.button("🔄 Reset Defaults", key="ex_reset_cfg"):
                for k, v in _CORR_DEFAULTS.items():
                    st.session_state[f"corr_{k}"] = v
                for t in _CORR_TENORS:
                    st.session_state[f"cms_bump_{t}"] = 0.0
                _save_exotics_config()
                st.rerun()

    sub = st.tabs(["📋 Spread / Curve Options", "📊 CMS Products", "🔒 Bermudan / Callable", "🔢 Digital Ladder", "🔄 Zero Coupon Swap"])

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 1   —   SPREAD / CURVE OPTIONS
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with sub[0]:
        st.markdown("### Curve Spread Options")
        st.caption("Option on the spread between two swap rates   —   steepener/flattener. Bachelier (Normal) model.")

        # ── Leg inputs ──────────────────────────────────────────────
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("**Long leg (higher tenor)**")
            long_tenor_sel = st.selectbox("Long Tenor", ["5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"],
                                           index=4, key="so_long_tenor")
            long_y = float(long_tenor_sel[:-1])
        with col2:
            st.markdown("**Short leg (lower tenor)**")
            short_tenor_sel = st.selectbox("Short Tenor", ["1Y","2Y","3Y","4Y","5Y","7Y","10Y"],
                                            index=2, key="so_short_tenor")
            short_y = float(short_tenor_sel[:-1])
        with col3:
            st.markdown("**Option**")
            so_expiry_sel = st.selectbox("Expiry",
                ["1m","2m","3m","6m","9m","1y","2y","3y","5y"], index=3, key="so_expiry")
            so_T = label_to_years(so_expiry_sel)
            so_type = st.selectbox("Type", [
                "Steepener Payer (spread > K)",
                "Flattener Receiver (spread < K)",
                "Straddle (both legs)"],
                index=2, key="so_type")
            so_notional = st.number_input("Notional (mm)", 1.0, 10000.0, 100.0, step=10.0, key="so_notl")

        # ── Forward rates from curve ────────────────────────────────
        if curve is not None:
            fwd_long,  ann_long,  _ = forward_and_annuity_from_curve(
                curve, ccy, so_T, long_y,  ois_curve)
            fwd_short, ann_short, _ = forward_and_annuity_from_curve(
                curve, ccy, so_T, short_y, ois_curve)
        else:
            fwd_long, fwd_short = 0.045, 0.040
            ann_long, ann_short = long_y, short_y

        fwd_spread_bp = (fwd_long - fwd_short) * 10000.0

        # ── Vol inputs ──────────────────────────────────────────────
        st.markdown("---")
        st.markdown("**Volatility Inputs**")
        vc1, vc2, vc3, vc4 = st.columns(4)
        # Reset vol + strike session state when tenors/expiry change
        _so_tenor_key = f"{long_tenor_sel}_{short_tenor_sel}_{so_expiry_sel}"
        if st.session_state.get("_so_last_tenor_key") != _so_tenor_key:
            st.session_state["so_strike"] = float(round(fwd_spread_bp, 1))
            st.session_state["_so_last_tenor_key"] = _so_tenor_key
            # Pre-compute vols so we can seed the keys before widgets render
            _so_expiry_order_pre = ["1m","2m","3m","6m","9m","1y","18m","2y","3y","4y","5y","7y","10y","12y","15y","20y"]
            def _pre_vol(tenor_y_):
                if atm is None: return 35.0
                v_ = get_matrix_value(atm, so_expiry_sel, tenor_y_)
                if v_: return v_
                vals_ = []
                for lbl_ in _so_expiry_order_pre:
                    vv_ = get_matrix_value(atm, lbl_, tenor_y_)
                    if vv_: vals_.append((label_to_years(lbl_), vv_))
                if not vals_: return 35.0
                vals_.sort()
                for ii in range(len(vals_)-1):
                    t0_, v0_ = vals_[ii]; t1_, v1_ = vals_[ii+1]
                    if t0_ <= so_T <= t1_:
                        return v0_ + (so_T-t0_)/(t1_-t0_) * (v1_-v0_)
                return vals_[0][1] if so_T < vals_[0][0] else vals_[-1][1]
            st.session_state["so_vl2"] = float(round(_pre_vol(long_y), 1))
            st.session_state["so_vs2"] = float(round(_pre_vol(short_y), 1))

        with vc1:
            # Pull vol from ATM surface   —   try exact expiry then interpolate between neighbours
            _so_expiry_order = ["1m","2m","3m","6m","9m","1y","18m","2y","3y","4y","5y","7y","10y","12y","15y","20y"]
            def _get_atm_vol_interp(tenor_y):
                if atm is None:
                    return 35.0
                # Try exact match first
                v = get_matrix_value(atm, so_expiry_sel, tenor_y)
                if v:
                    return v
                # Interpolate between available neighbours in expiry dimension
                vals = []
                for lbl in _so_expiry_order:
                    vv = get_matrix_value(atm, lbl, tenor_y)
                    if vv:
                        vals.append((label_to_years(lbl), vv))
                if not vals:
                    return 35.0
                if len(vals) == 1:
                    return vals[0][1]
                # Linear interpolation at so_T
                vals.sort()
                for i in range(len(vals)-1):
                    t0, v0 = vals[i]; t1, v1 = vals[i+1]
                    if t0 <= so_T <= t1:
                        alpha_ = (so_T - t0) / (t1 - t0) if t1 > t0 else 0
                        return v0 + alpha_ * (v1 - v0)
                # Extrapolate from nearest
                return vals[0][1] if so_T < vals[0][0] else vals[-1][1]

            long_vol_default  = _get_atm_vol_interp(long_y)
            short_vol_default = _get_atm_vol_interp(short_y)
            vol_long_bp = st.number_input(f"Vol {long_tenor_sel} (bp pa)",
                                           0.1, 2000.0, float(min(round(long_vol_default, 1), 2000.0)),
                                           step=0.5, key="so_vl2")
        with vc2:
            vol_short_bp = st.number_input(f"Vol {short_tenor_sel} (bp pa)",
                                            0.1, 2000.0, float(min(round(short_vol_default, 1), 2000.0)),
                                            step=0.5, key="so_vs2")
        with vc3:
            # Pull from config matrix, fall back to distance formula
            def _nearest_tenor(y_):
                mapping = {1:"1Y",2:"2Y",3:"3Y",4:"4Y",5:"5Y",7:"7Y",10:"10Y",
                           12:"12Y",15:"15Y",20:"20Y",25:"25Y",30:"30Y"}
                closest = min(mapping.keys(), key=lambda k: abs(k - y_))
                return mapping[closest]
            t1_key = _nearest_tenor(long_y)
            t2_key = _nearest_tenor(short_y)
            rho_cfg = get_correlation(t1_key, t2_key)
            rho = st.number_input("Correlation ρ,ν", 0.0, 1.0,
                                   round(rho_cfg, 3), step=0.005,
                                   format="%.3f", key="so_rho")
            st.caption(f"From config: {t1_key}/{t2_key}")
        with vc4:
            # Strike: default ATM (fwd spread)
            K_spread_bp = st.number_input("Strike Spread (bp)",
                                           -500.0, 1000.0,
                                           float(round(fwd_spread_bp, 1)),
                                           step=1.0, key="so_strike")

        # ── OIS df ─────────────────────────────────────────────────
        if ois_curve is not None:
            xs = ois_curve["MaturityY"].to_numpy().astype(float)
            ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
            df_T_so = math.exp(-float(np.interp(so_T, xs, ys)) * so_T)
        else:
            df_T_so = math.exp(-0.043 * so_T)

        # ── Bachelier spread option formula ────────────────────────
        # ┬ñ├ó_spread~ = ┬ñ├ó1~ + ┬ñ├ó2~ - 2ρ,ν┬ñ├ó1┬ñ├ó2
        sigma_long  = vol_long_bp  / 10000.0
        sigma_short = vol_short_bp / 10000.0
        sigma_spread = math.sqrt(max(
            sigma_long**2 + sigma_short**2 - 2*rho*sigma_long*sigma_short, 1e-12))
        vol_spread_bp = sigma_spread * 10000.0

        X_fwd = (fwd_long - fwd_short) * 10000.0   # bp
        K     = K_spread_bp
        sqrt_T = math.sqrt(max(so_T, 1e-8))
        from statistics import NormalDist as _Nd
        _nd = _Nd()

        is_straddle    = "Straddle" in so_type
        is_payer_spread = "Steepener" in so_type  # payer = benefits from steepening (spread widens)

        d_so = (X_fwd - K) / (vol_spread_bp * sqrt_T) if vol_spread_bp * sqrt_T > 0 else 0.0

        def _payer_prem(Xf, Ks, sv, sqT, df):
            d_ = (Xf - Ks) / (sv * sqT) if sv * sqT > 0 else 0.0
            return df * ((Xf - Ks) * _nd.cdf(d_) + sv * sqT * _nd.pdf(d_))

        def _recv_prem(Xf, Ks, sv, sqT, df):
            d_ = (Xf - Ks) / (sv * sqT) if sv * sqT > 0 else 0.0
            return df * ((Ks - Xf) * _nd.cdf(-d_) + sv * sqT * _nd.pdf(d_))

        if is_straddle:
            prem_payer = _payer_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so)
            prem_recv  = _recv_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so)
            prem_bp    = prem_payer + prem_recv
            delta_long  = df_T_so * (_nd.cdf(d_so) - _nd.cdf(-d_so))
            delta_short = -delta_long
        elif is_payer_spread:
            prem_bp     = _payer_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so)
            delta_long  =  df_T_so * _nd.cdf(d_so)
            delta_short = -df_T_so * _nd.cdf(d_so)
        else:
            prem_bp     = _recv_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so)
            delta_long  = -df_T_so * _nd.cdf(-d_so)
            delta_short =  df_T_so * _nd.cdf(-d_so)

        prem_dollar = prem_bp / 10000.0 * so_notional * 1e6
        vega_dollar = df_T_so * sqrt_T * _nd.pdf(d_so) * so_notional * 1e6 / 10000.0
        if is_straddle:
            vega_dollar *= 2.0

        # ── Results ─────────────────────────────────────────────────
        st.markdown("---")
        m1,m2,m3,m4,m5,m6 = st.columns(6)
        m1.metric(f"Fwd {long_tenor_sel}", f"{fwd_long*100:.4f}%")
        m2.metric(f"Fwd {short_tenor_sel}", f"{fwd_short*100:.4f}%")
        m3.metric("Fwd Spread", f"{X_fwd:.2f}bp")
        m4.metric("Spread Vol (bp)", f"{vol_spread_bp:.2f}bp")
        m5.metric("Premium (bp notl)", f"{prem_bp:.3f}bp")
        m6.metric("Premium (AUD)", f"${prem_dollar:,.0f}")

        if is_straddle:
            rs1, rs2, rs3, rs4, rs5 = st.columns(5)
            rs1.metric("Payer leg (bp)", f"{_payer_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so):.3f}")
            rs2.metric("Receiver leg (bp)", f"{_recv_prem(X_fwd, K, vol_spread_bp, sqrt_T, df_T_so):.3f}")
            rs3.metric(f"Delta {long_tenor_sel}", f"{delta_long:.4f}")
            rs4.metric("Vega ($/bp vol)", f"${vega_dollar:,.0f}")
            rs5.metric("df(T)", f"{df_T_so:.6f}")
        else:
            r1,r2,r3,r4 = st.columns(4)
            r1.metric(f"Delta {long_tenor_sel}", f"{delta_long:.4f}")
            r2.metric(f"Delta {short_tenor_sel}", f"{delta_short:.4f}")
            r3.metric("Vega ($/bp vol)", f"${vega_dollar:,.0f}")
            r4.metric("df(T)", f"{df_T_so:.6f}")
        st.caption(f"┬ñ├ó_long={vol_long_bp:.1f}bp  ┬ñ├ó_short={vol_short_bp:.1f}bp  "
                   f"ρ,ν={rho:.3f}  ┬ñ├ó_spread={vol_spread_bp:.2f}bp  d={d_so:.4f}  T={so_T:.4f}y")

        # ── Payoff chart ─────────────────────────────────────────────
        spread_range = np.linspace(X_fwd - 150, X_fwd + 150, 100)
        prems_payer, prems_recv, prems_strad = [], [], []
        for sp in spread_range:
            c_ = max(_payer_prem(sp, K, vol_spread_bp, sqrt_T, df_T_so), 0.0)
            p_ = max(_recv_prem(sp, K, vol_spread_bp, sqrt_T, df_T_so), 0.0)
            prems_payer.append(c_)
            prems_recv.append(p_)
            prems_strad.append(c_ + p_)

        fig_so = go.Figure()
        fig_so.add_trace(go.Scatter(x=list(spread_range), y=prems_payer,
                                    name="Steepener Payer", line=dict(color="#22c55e", width=2)))
        fig_so.add_trace(go.Scatter(x=list(spread_range), y=prems_recv,
                                    name="Flattener Receiver", line=dict(color="#ef4444", width=2)))
        fig_so.add_trace(go.Scatter(x=list(spread_range), y=prems_strad,
                                    name="Straddle", line=dict(color="#a78bfa", width=2, dash="dot")))
        fig_so.add_vline(x=X_fwd, line_dash="dash", line_color="#94a3b8",
                         annotation_text=f"Fwd {X_fwd:.1f}bp")
        fig_so.add_vline(x=K, line_dash="dot", line_color="#f59e0b",
                         annotation_text=f"Strike {K:.1f}bp")
        fig_so.update_layout(
            title=f"{long_tenor_sel}-{short_tenor_sel} Spread Option   —   Premium vs Forward Spread",
            xaxis_title="Forward Spread (bp)", yaxis_title="Premium (bp notional)",
            template="plotly_dark", height=370,
            legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig_so, use_container_width=True)

        # ── Correlation sensitivity ──────────────────────────────────
        with st.expander("📊 Correlation Sensitivity"):
            rho_range = np.linspace(0.60, 0.99, 40)
            prem_vs_rho = []
            for rho_ in rho_range:
                sv_ = math.sqrt(max(sigma_long**2 + sigma_short**2 - 2*rho_*sigma_long*sigma_short, 1e-12)) * 10000
                if is_straddle:
                    p_ = _payer_prem(X_fwd, K, sv_, sqrt_T, df_T_so) + _recv_prem(X_fwd, K, sv_, sqrt_T, df_T_so)
                elif is_payer_spread:
                    p_ = _payer_prem(X_fwd, K, sv_, sqrt_T, df_T_so)
                else:
                    p_ = _recv_prem(X_fwd, K, sv_, sqrt_T, df_T_so)
                prem_vs_rho.append(max(p_, 0.0))
            fig_rho = go.Figure()
            fig_rho.add_trace(go.Scatter(x=list(rho_range), y=prem_vs_rho,
                                          line=dict(color="#3b82f6", width=2)))
            fig_rho.add_vline(x=rho, line_dash="dash", line_color="#f59e0b",
                               annotation_text=f"ρ,ν={rho:.3f}")
            fig_rho.update_layout(title="Premium vs Correlation ρ,ν",
                                   xaxis_title="ρ,ν", yaxis_title="Premium (bp)",
                                   template="plotly_dark", height=300)
            st.plotly_chart(fig_rho, use_container_width=True)
            st.caption("Higher correlation → lower spread vol → cheaper spread option. "
                       "ρ,ν is the key uncertainty in spread option pricing.")

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 2   —   CMS PRODUCTS
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with sub[1]:
        st.markdown("### CMS Products")
        st.caption("CMS caplets/floorlets/swaplets   —   floating rate is the prevailing CMS swap rate, not BBSW. "
                   "Hagan linear swap rate (LSR) convexity adjustment + Bachelier pricing.")

        cms_product = st.radio("Product",
            ["CMS Caplet / Floorlet", "CMS Cap / Floor (series)", "CMS Swap", "CMS Spread Cap/Floor"],
            horizontal=True, key="cms_product")

        st.markdown("---")

        # ── Common CMS inputs ────────────────────────────────────────
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            cms_rate_tenor = st.selectbox("CMS Rate Tenor",
                ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","30Y"],
                index=4, key="cms_tenor")
            cms_n = float(cms_rate_tenor[:-1])
        with c2:
            cms_pay_freq = st.selectbox("Payment Frequency",
                ["Monthly","Quarterly","Semi-Annual","Annual"],
                index=1, key="cms_freq")
            cms_delta = {"Monthly": 1/12, "Quarterly": 0.25,
                         "Semi-Annual": 0.5, "Annual": 1.0}[cms_pay_freq]
        with c3:
            cms_notional = st.number_input("Notional (mm)", 1.0, 10000.0, 100.0,
                                            step=10.0, key="cms_notl")
        with c4:
            cms_opt_type = st.selectbox("Option Type", ["Cap (call on CMS rate)",
                                                         "Floor (put on CMS rate)"],
                                         key="cms_opt_type")
            cms_is_cap = "Cap" in cms_opt_type

        # ── Single caplet or series ──────────────────────────────────
        if cms_product in ["CMS Caplet / Floorlet", "CMS Swap"]:
            fix_col, _ = st.columns([2, 2])
            with fix_col:
                cms_fix_label = st.selectbox("Fixing Date",
                    ["3m","6m","9m","1y","18m","2y","3y","5y"],
                    index=3, key="cms_fix")
            cms_fix_dates = [label_to_years(cms_fix_label)]
            cms_settle_label = f"{cms_fix_label} + +"
        else:
            # Series: define start + tenor of the structure
            sc1, sc2, sc3 = st.columns(3)
            with sc1:
                series_start = st.selectbox("Structure Start",
                    ["0","3m","6m","1y"], index=0, key="cms_ser_start")
                t_start = 0.0 if series_start == "0" else label_to_years(series_start)
            with sc2:
                series_tenor = st.selectbox("Structure Tenor",
                    ["1Y","2Y","3Y","4Y","5Y","7Y","10Y"],
                    index=2, key="cms_ser_tenor")
                t_end = t_start + float(series_tenor[:-1])
            with sc3:
                pass
            # Build fixing schedule
            t = t_start + cms_delta
            cms_fix_dates = []
            while t <= t_end + 1e-8:
                cms_fix_dates.append(round(t, 6))
                t += cms_delta
            st.caption(f"{len(cms_fix_dates)} fixing dates from {t_start:.2f}y to {t_end:.2f}y "
                       f"({cms_pay_freq})")

        # ── Strike ───────────────────────────────────────────────────
        # CMS ATM ├ö├½├¬ forward CMS rate (incl convexity adj) at first fixing
        T0 = cms_fix_dates[0]
        if curve is not None:
            fwd_cms, ann_cms, _ = forward_and_annuity_from_curve(
                curve, ccy, T0, cms_n, ois_curve)
        else:
            fwd_cms, ann_cms = 0.043, cms_n
        fwd_cms_pct = fwd_cms * 100.0

        sk1, sk2 = st.columns([2, 2])
        with sk1:
            K_cms = st.number_input("Strike (% pa)", 0.01, 20.0,
                                     round(fwd_cms_pct, 3), step=0.01,
                                     format="%.4f", key="cms_strike") / 100.0
        with sk2:
            # Vol: pull ATM from surface at first fixing
            def _cms_vol(T_fix):
                if atm is None: return 35.0
                exp_lbl = None
                for lbl in ["3m","6m","9m","1y","18m","2y","3y","5y","7y","10y"]:
                    if abs(label_to_years(lbl) - T_fix) < 0.15:
                        exp_lbl = lbl
                        break
                if exp_lbl is None: return 35.0
                v = get_matrix_value(atm, exp_lbl, cms_n)
                return v if v else 35.0
            vol_cms_default = _cms_vol(T0)
            vol_cms_bp = st.number_input("Vol (bp pa)", 1.0, 2000.0,
                                          float(min(round(vol_cms_default, 1), 2000.0)),
                                          step=0.5, key="cms_vol")

        # ── Hagan LSR Convexity Adjustment ──────────────────────────
        # adj ├ö├½├¬ F~ ≈ ┬ñ├ó~ ≈ T ≈ G
        # G = (dA/dF)/A evaluated at F, using bond-math annuity derivative
        # For S/S AUD: A(F) = > + ≈ (1+F+)^(-k)  [simplified flat curve approx]
        # dA/dF = -> k≈+~ ≈ (1+F+)^(-(k+1))
        # G = (1/A) ≈ dA/dF

        def cms_convexity_adj(F: float, sigma_bp: float, T_fix: float,
                               tenor_y: float, delta_cpn: float = 0.5) -> float:
            """
            Hagan linear swap rate convexity adjustment (bp).
            Uses bond-math approximation for annuity derivative.
            AUD: semi-annual coupon for >3Y (delta=0.5), quarterly for <=3Y (delta=0.25).
            """
            if F <= 0 or sigma_bp <= 0 or T_fix <= 0:
                return 0.0
            sigma = sigma_bp / 10000.0
            n_cpn = int(round(tenor_y / delta_cpn))
            if n_cpn == 0:
                return 0.0
            # Annuity and its derivative w.r.t. F
            v_factor = 1.0 / (1.0 + F * delta_cpn)
            A = delta_cpn * sum(v_factor**k for k in range(1, n_cpn + 1))
            # dA/dF = -delta~ > k ≈ (1+F+)^(-(k+1))
            dA_dF = -delta_cpn**2 * sum(k * v_factor**(k + 1) for k in range(1, n_cpn + 1))
            G = dA_dF / A if A > 0 else 0.0
            # adj in rate units: F~ ≈ ┬ñ├ó~ ≈ T ≈ G (sign: positive for cap-style)
            adj = -F**2 * sigma**2 * T_fix * G   # G is negative → adj positive
            return adj * 10000.0   # return in bp

        # CMS coupon frequency for AUD
        cms_cpn_freq = 0.25 if cms_n <= 3.0 else 0.5

        # ── Price each caplet in schedule ────────────────────────────
        def price_cms_caplet(T_fix: float, K: float, vol_bp: float,
                              is_cap: bool) -> dict:
            """Price one CMS caplet/floorlet with convexity adjustment."""
            if curve is not None:
                F, ann, _ = forward_and_annuity_from_curve(
                    curve, ccy, T_fix, cms_n, ois_curve)
            else:
                F, ann = fwd_cms, ann_cms

            # Convexity adjustment
            adj_bp = cms_convexity_adj(F, vol_bp, T_fix, cms_n, cms_cpn_freq)
            # Add manual CMS bump from config (user override)
            _cms_tenor_key = min(_CORR_TENORS,
                key=lambda t: abs(float(t[:-1]) - cms_n))
            adj_bp += get_cms_bump(_cms_tenor_key)
            F_adj = F + adj_bp / 10000.0   # adjusted CMS forward rate

            # OIS df to payment date (T_fix + cms_delta for standard lag)
            T_pay = T_fix + cms_delta
            if ois_curve is not None:
                xs = ois_curve["MaturityY"].to_numpy().astype(float)
                ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                df_pay = math.exp(-float(np.interp(T_pay, xs, ys)) * T_pay)
            else:
                df_pay = math.exp(-0.043 * T_pay)

            # Bachelier caplet/floorlet on adjusted CMS rate
            sigma = vol_bp / 10000.0
            sqrt_T = math.sqrt(max(T_fix, 1e-8))
            from statistics import NormalDist as _nd_
            _nd_i = _nd_()
            d = (F_adj - K) / (sigma * sqrt_T) if sigma * sqrt_T > 0 else 0.0
            if is_cap:
                prem_rate = (F_adj - K) * _nd_i.cdf(d) + sigma * sqrt_T * _nd_i.pdf(d)
                delta = _nd_i.cdf(d)
            else:
                prem_rate = (K - F_adj) * _nd_i.cdf(-d) + sigma * sqrt_T * _nd_i.pdf(d)
                delta = -_nd_i.cdf(-d)
            # PV = premium ≈ delta_period ≈ notional ≈ df_pay
            prem_bp_period = prem_rate * 10000.0
            pv = prem_rate * cms_delta * cms_notional * 1e6 * df_pay
            vega = df_pay * sqrt_T * _nd_i.pdf(d) * cms_delta * cms_notional * 1e6 / 10000.0
            return {
                "T_fix": T_fix, "F": F, "adj_bp": adj_bp, "F_adj": F_adj,
                "df_pay": df_pay, "d": d, "prem_bp": prem_bp_period,
                "pv": pv, "delta": delta, "vega": vega
            }

        if cms_product != "CMS Spread Cap/Floor":
            results_cms = [price_cms_caplet(t, K_cms, vol_cms_bp, cms_is_cap)
                           for t in cms_fix_dates]

            total_pv      = sum(r["pv"] for r in results_cms)
            total_pv_bp   = total_pv / (cms_notional * 1e6) * 10000.0
            total_vega    = sum(r["vega"] for r in results_cms)

            # For CMS swap: also price the floating BBSW leg minus CMS leg
            if cms_product == "CMS Swap":
                # CMS swap: pay CMS, receive BBSW (or vice versa)
                # Fair CMS rate = PV(BBSW leg) / annuity_CMS
                # Simplified: just show the fair CMS rate incl convexity
                fair_cms_rate = sum(r["F_adj"] * r["df_pay"] * cms_delta
                                    for r in results_cms) / \
                                sum(r["df_pay"] * cms_delta for r in results_cms) \
                                if results_cms else fwd_cms
                cms_spread_to_bbsw = (fair_cms_rate - fwd_cms) * 10000.0

                st.markdown("---")
                s1, s2, s3, s4 = st.columns(4)
                s1.metric("Fair CMS Rate", f"{fair_cms_rate*100:.4f}%")
                s2.metric("Convexity Adj (avg bp)", f"{sum(r['adj_bp'] for r in results_cms)/len(results_cms):.2f}bp")
                s3.metric("CMS vs BBSW Spread", f"{cms_spread_to_bbsw:.2f}bp")
                s4.metric("# Payments", str(len(results_cms)))
            else:
                st.markdown("---")
                mr1, mr2, mr3, mr4, mr5 = st.columns(5)
                mr1.metric("Total PV (AUD)", f"${total_pv:,.0f}")
                mr2.metric("PV (bp notional)", f"{total_pv_bp:.3f}bp")
                mr3.metric("Vega (AUD/bp)", f"${total_vega:,.0f}")
                mr4.metric("Avg CMS Fwd", f"{sum(r['F'] for r in results_cms)/len(results_cms)*100:.4f}%")
                mr5.metric("Avg Conv Adj", f"{sum(r['adj_bp'] for r in results_cms)/len(results_cms):.2f}bp")

            # Schedule table
            with st.expander("📅 Caplet Schedule"):
                df_sched = pd.DataFrame([{
                    "Fix (y)": f"{r['T_fix']:.3f}",
                    "CMS Fwd (%)": f"{r['F']*100:.4f}",
                    "Conv Adj (bp)": f"{r['adj_bp']:.3f}",
                    "Adj Fwd (%)": f"{r['F_adj']*100:.4f}",
                    "df(pay)": f"{r['df_pay']:.6f}",
                    "d": f"{r['d']:.4f}",
                    "Prem (bp period)": f"{r['prem_bp']:.4f}",
                    "PV (AUD)": f"${r['pv']:,.0f}",
                } for r in results_cms])
                st.dataframe(df_sched, use_container_width=True, hide_index=True)

            # Premium vs strike chart
            k_range = np.linspace(max(K_cms - 0.015, 0.001), K_cms + 0.015, 60)
            prem_vs_k = []
            for k_ in k_range:
                total_ = sum(price_cms_caplet(t, k_, vol_cms_bp, cms_is_cap)["pv"]
                             for t in cms_fix_dates)
                prem_vs_k.append(total_ / (cms_notional * 1e6) * 10000.0)
            fig_cms = go.Figure()
            fig_cms.add_trace(go.Scatter(x=list(k_range * 100), y=prem_vs_k,
                                          line=dict(color="#3b82f6", width=2),
                                          name="Total PV (bp notional)"))
            fig_cms.add_vline(x=K_cms * 100, line_dash="dot", line_color="#f59e0b",
                               annotation_text=f"Strike {K_cms*100:.3f}%")
            fig_cms.update_layout(
                title=f"CMS {cms_rate_tenor} {'Cap' if cms_is_cap else 'Floor'}   —   PV vs Strike",
                xaxis_title="Strike (%)", yaxis_title="PV (bp notional)",
                template="plotly_dark", height=360)
            st.plotly_chart(fig_cms, use_container_width=True)

            # Convexity adjustment explainer
            with st.expander("📐 Convexity Adjustment Detail"):
                st.markdown(f"""
**Why CMS rates need a convexity adjustment:**

A CMS coupon pays the {cms_rate_tenor} swap rate at fixing, but discounted to the payment date using a shorter-maturity discount factor. The CMS rate is a martingale under the *swap annuity measure*, not under the *payment date forward measure*   —   so a measure-change correction is required.

**Hagan LSR formula used here:**

`adj ├ö├½├¬ F~ ≈ ┬ñ├ó~ ≈ T_fix ≈ |dA/dF| / A`

Where:
- `F` = forward {cms_rate_tenor} CMS rate
- `┬ñ├ó` = normal vol of the {cms_rate_tenor} rate
- `T_fix` = time to fixing
- `A(F)` = swap annuity (PV of $1 paid every {cms_cpn_freq*12:.0f} months for {cms_rate_tenor})
- `dA/dF` = sensitivity of annuity to rate   —   always negative (rates up → annuity down)

The adjustment is always **positive** (CMS forward rate > standard forward rate) because higher rates reduce the annuity, creating a natural bias toward higher CMS fixings under the payment measure.

**AUD convention:** {cms_rate_tenor} CMS uses {'quarterly (Q/Q)' if cms_cpn_freq == 0.25 else 'semi-annual (S/S)'} coupon frequency for the underlying swap annuity calculation.
                """)

        else:
            # ── CMS SPREAD CAP/FLOOR ──────────────────────────────────
            st.markdown("**CMS Spread Cap/Floor**   —   option on spread between two CMS rates")
            ss1, ss2, ss3 = st.columns(3)
            with ss1:
                cms2_tenor = st.selectbox("Second CMS Tenor",
                    ["1Y","2Y","3Y","5Y","10Y"], index=1, key="cms2_tenor")
                cms2_n = float(cms2_tenor[:-1])
            with ss2:
                vol_cms2_bp = st.number_input("Vol CMS2 (bp pa)", 1.0, 2000.0,
                    float(min(round(_cms_vol(T0), 1), 2000.0)), step=0.5, key="cms2_vol")
            with ss3:
                _t1c = min(_CORR_TENORS, key=lambda t: abs(float(t[:-1]) - cms_n))
                _t2c = min(_CORR_TENORS, key=lambda t: abs(float(t[:-1]) - cms2_n))
                rho_cms_default = get_correlation(_t1c, _t2c)
                rho_cms = st.number_input("CMS-CMS ρ,ν", 0.0, 1.0,
                    round(rho_cms_default, 3),
                    step=0.005, format="%.3f", key="cms_rho")
                st.caption(f"Config: {_t1c}/{_t2c}")

            results_sp = []
            for T_fix in cms_fix_dates:
                r1 = price_cms_caplet(T_fix, 0.0, vol_cms_bp, True)
                # CMS2 forward
                if curve is not None:
                    F2, _, _ = forward_and_annuity_from_curve(curve, ccy, T_fix, cms2_n, ois_curve)
                else:
                    F2 = 0.038
                adj2 = cms_convexity_adj(F2, vol_cms2_bp, T_fix, cms2_n, 0.25 if cms2_n <= 3 else 0.5)
                _cms2_tenor_key = min(_CORR_TENORS, key=lambda t: abs(float(t[:-1]) - cms2_n))
                adj2 += get_cms_bump(_cms2_tenor_key)
                F2_adj = F2 + adj2 / 10000.0

                # Spread forward
                X_fwd_cms = (r1["F_adj"] - F2_adj) * 10000.0
                K_sp = K_cms * 10000.0   # K_cms already in rate units here, reuse as bp strike
                sigma1 = vol_cms_bp / 10000.0
                sigma2 = vol_cms2_bp / 10000.0
                sigma_sp = math.sqrt(max(sigma1**2 + sigma2**2 - 2*rho_cms*sigma1*sigma2, 1e-12))
                sqrt_T = math.sqrt(max(T_fix, 1e-8))
                d_sp = (X_fwd_cms - K_cms*10000) / (sigma_sp*10000*sqrt_T) if sigma_sp*sqrt_T > 0 else 0.0
                from statistics import NormalDist as _nd_
                _nd_i = _nd_()
                if cms_is_cap:
                    prem_rate_sp = (X_fwd_cms - K_cms*10000) * _nd_i.cdf(d_sp) + sigma_sp*10000*sqrt_T*_nd_i.pdf(d_sp)
                else:
                    prem_rate_sp = (K_cms*10000 - X_fwd_cms) * _nd_i.cdf(-d_sp) + sigma_sp*10000*sqrt_T*_nd_i.pdf(d_sp)
                pv_sp = prem_rate_sp / 10000.0 * cms_delta * cms_notional * 1e6 * r1["df_pay"]
                results_sp.append({"T_fix": T_fix, "F1_adj": r1["F_adj"], "F2_adj": F2_adj,
                                   "spread_fwd": X_fwd_cms, "pv": pv_sp})

            total_pv_sp = sum(r["pv"] for r in results_sp)
            st.metric("Total PV (AUD)", f"${total_pv_sp:,.0f}")
            st.metric("PV (bp notional)", f"{total_pv_sp/(cms_notional*1e6)*10000:.3f}bp")

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 3   —   BERMUDAN / CALLABLE (Hull-White 1F trinomial tree)
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with sub[2]:
        st.markdown("### Bermudan Swaption / Callable Swap")
        st.caption("Hull-White 1F trinomial tree, calibrated to co-terminal swaptions from ATM surface.")

        berm_type = st.radio("Structure",
            ["Bermudan Payer", "Bermudan Receiver", "Callable Swap (pay-fixed)"],
            horizontal=True, key="berm_type")

        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            berm_final_sel = st.selectbox("Final Maturity",
                ["2y","3y","4y","5y","7y","10y","12y","15y","20y"],
                index=4, key="berm_final")
            T_final = label_to_years(berm_final_sel)
        with bc2:
            berm_swap_tenor = st.selectbox("Underlying Swap Tenor",
                ["1Y","2Y","3Y","4Y","5Y","7Y","10Y"],
                index=4, key="berm_tenor")
            T_swap = float(berm_swap_tenor[:-1])
        with bc3:
            berm_first_call = st.selectbox("First Call",
                ["1y","2y","3y","4y","5y"], index=0, key="berm_first")
            T_first = label_to_years(berm_first_call)
        bc4, bc5, bc6 = st.columns(3)
        with bc4:
            berm_freq_sel = st.selectbox("Exercise Freq",
                ["Annual","Semi-Annual","Quarterly"], index=0, key="berm_freq")
            berm_freq = {"Annual": 1.0, "Semi-Annual": 0.5, "Quarterly": 0.25}[berm_freq_sel]
        with bc5:
            berm_notional = st.number_input("Notional (mm)", 1.0, 10000.0, 100.0,
                                              step=10.0, key="berm_notl")
        with bc6:
            hw_mr = st.number_input("H-W Mean Reversion a (%)", 0.1, 20.0, 3.0,
                                     step=0.1, format="%.2f", key="hw_mr") / 100.0

        # Build exercise schedule
        ex_dates = []
        t = T_first
        while t <= T_final - T_swap + 1e-6:
            ex_dates.append(round(t, 4))
            t += berm_freq
        if not ex_dates:
            st.warning("No valid exercise dates   —   check First Call ≤ Final Maturity  →  Swap Tenor.")
            st.stop()

        st.caption(f"Exercise dates: {', '.join(f'{t}y' for t in ex_dates)} "
                   f"({len(ex_dates)} exercise opportunities)")

        # Forward and strike
        if curve is not None:
            fwd_b, ann_b, _ = forward_and_annuity_from_curve(
                curve, ccy, ex_dates[0], T_swap, ois_curve)
        else:
            fwd_b, ann_b = 0.043, T_swap
        K_b = st.number_input("Strike (% pa)", 0.01, 20.0,
                               round(fwd_b * 100, 3), step=0.01,
                               format="%.4f", key="berm_strike") / 100.0

        # ── Hull-White 1F trinomial tree ─────────────────────────────
        if st.button("⚙️ Price Bermudan (HW1F Tree)", key="berm_price"):
            with st.spinner("Calibrating HW1F to co-terminal swaptions and running trinomial tree├ö├ç┬¬"):
                try:
                    # Calibrate ┬ñ├ó to co-terminal swaptions
                    # Co-terminal: for each ex date T_i, calibrate to swaption expiring T_i into (T_final - T_i) swap
                    def hw_swaption_price(sigma_hw, a, T_exp, T_sw, curve_df, ois_df):
                        """Analytic HW1F swaption price (Jamshidian decomposition)."""
                        if curve_df is None:
                            return 0.0
                        # Bond prices P(0,T)
                        def P(t):
                            if ois_df is not None:
                                xs = ois_df["MaturityY"].to_numpy().astype(float)
                                ys = ois_df["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            else:
                                xs = curve_df["MaturityY"].to_numpy().astype(float)
                                ys = curve_df["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            r = float(np.interp(t, xs, ys))
                            return math.exp(-r * t)
                        # HW B function
                        def B(s, t):
                            return (1 - math.exp(-a * (t - s))) / a if a > 1e-8 else (t - s)
                        # HW sigma of ln P(T_exp, T_pay)
                        def sigma_P(T_exp_, T_pay_):
                            return sigma_hw * B(T_exp_, T_pay_) * math.sqrt(
                                (1 - math.exp(-2*a*T_exp_)) / (2*a)) if a > 1e-8 else \
                                sigma_hw * (T_pay_ - T_exp_) * math.sqrt(T_exp_)
                        # Coupon swap schedule
                        freq_sw = 0.25 if T_sw <= 3 else 0.5
                        pay_times = []
                        tt = T_exp + freq_sw
                        while tt <= T_exp + T_sw + 1e-6:
                            pay_times.append(round(tt, 6))
                            tt += freq_sw
                        if not pay_times:
                            return 0.0
                        # Coupon = forward swap rate
                        fwd_sw, ann_sw, _ = forward_and_annuity_from_curve(
                            curve_df, ccy, T_exp, T_sw, ois_df)
                        c = fwd_sw  # strike = ATM for calibration
                        # Jamshidian: find r* such that > c_i P(T_exp, T_i; r*) = 1
                        def swap_val_at_r(r_star):
                            total = 0.0
                            for i, T_i in enumerate(pay_times):
                                accrual = T_i - (pay_times[i-1] if i > 0 else T_exp)
                                coupon = K_b * accrual + (1.0 if i == len(pay_times)-1 else 0.0)
                                # P(T_exp, T_i | r*) in HW
                                Pb = P(T_i) / P(T_exp) * math.exp(
                                    -B(T_exp, T_i) * r_star -
                                    0.5 * sigma_hw**2 * B(T_exp, T_i)**2 *
                                    (1 - math.exp(-2*a*T_exp)) / (2*a))
                                total += coupon * Pb
                            return total - 1.0
                        try:
                            r_star = scipy.optimize.brentq(swap_val_at_r, -0.5, 0.5, xtol=1e-8)
                        except Exception:
                            r_star = 0.0
                        # Sum of bond puts
                        from statistics import NormalDist as _nd_j
                        _ndj = _nd_j()
                        total_pv_hw = 0.0
                        for i, T_i in enumerate(pay_times):
                            accrual = T_i - (pay_times[i-1] if i > 0 else T_exp)
                            X_i = (K_b * accrual + (1.0 if i == len(pay_times)-1 else 0.0))
                            # K* for this bond
                            Pb_star = P(T_i) / P(T_exp) * math.exp(
                                -B(T_exp, T_i) * r_star -
                                0.5 * sigma_hw**2 * B(T_exp, T_i)**2 *
                                (1 - math.exp(-2*a*T_exp)) / (2*a))
                            K_star_i = X_i * Pb_star
                            sig_i = sigma_P(T_exp, T_i)
                            if sig_i < 1e-10:
                                continue
                            d1_j = math.log(P(T_i) / (K_star_i * P(T_exp))) / sig_i + sig_i / 2
                            d2_j = d1_j - sig_i
                            total_pv_hw += X_i * (
                                P(T_exp) * K_star_i * _ndj.cdf(-d2_j) - P(T_i) * _ndj.cdf(-d1_j))
                        return total_pv_hw

                    # Get market ATM vols for co-terminal swaptions
                    def _co_vol(T_exp_, T_sw_):
                        if atm is None: return 35.0
                        exp_lbl_ = None
                        for lbl_ in ["3m","6m","9m","1y","18m","2y","3y","4y","5y","7y","10y","12y","15y","20y"]:
                            if abs(label_to_years(lbl_) - T_exp_) < 0.2:
                                exp_lbl_ = lbl_
                                break
                        if exp_lbl_ is None: return 35.0
                        v_ = get_matrix_value(atm, exp_lbl_, T_sw_)
                        return v_ if v_ else 35.0

                    # Calibrate one ┬ñ├ó_HW per exercise date (piecewise constant)
                    sigmas_hw = []
                    for T_ex in ex_dates:
                        T_remaining = T_final - T_ex
                        mkt_vol_bp = _co_vol(T_ex, T_remaining)
                        # Bachelier market price for this co-terminal swaption
                        fwd_co, ann_co, _ = forward_and_annuity_from_curve(
                            curve, ccy, T_ex, T_remaining, ois_curve) if curve is not None \
                            else (0.043, T_remaining, None)
                        if ois_curve is not None:
                            xs = ois_curve["MaturityY"].to_numpy().astype(float)
                            ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            df_co = math.exp(-float(np.interp(T_ex, xs, ys)) * T_ex)
                        else:
                            df_co = math.exp(-0.043 * T_ex)
                        sigma_n = mkt_vol_bp / 10000.0
                        mkt_prem = df_co * (0.3989 * sigma_n * math.sqrt(T_ex) * ann_co)

                        # Find ┬ñ├ó_HW that matches market
                        def obj_hw(sig_):
                            hw_p = hw_swaption_price(sig_[0], hw_mr, T_ex, T_remaining, curve, ois_curve)
                            return [(hw_p * berm_notional * 1e6 - mkt_prem * berm_notional * 1e6)]
                        try:
                            sol = scipy.optimize.fsolve(obj_hw, [0.01], full_output=True)
                            sig_cal = abs(sol[0][0])
                        except Exception:
                            sig_cal = 0.01
                        sigmas_hw.append(max(sig_cal, 0.0001))

                    # ── Trinomial tree backward induction ──────────────
                    # Build tree on [0, T_final] with steps every berm_freq/4
                    dt = min(berm_freq / 4.0, 0.25 / 4.0)
                    N_steps = max(int(T_final / dt) + 1, 20)
                    dt = T_final / N_steps
                    # Grid width: ~M nodes, M = 3 (covers ~3 std devs)
                    M = 3
                    dr = math.sqrt(3 * sigmas_hw[-1]**2 * dt) if sigmas_hw else 0.001

                    # Get OIS zero curve for discount
                    def ois_zero(t_):
                        if ois_curve is not None:
                            xs = ois_curve["MaturityY"].to_numpy().astype(float)
                            ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            return float(np.interp(t_, xs, ys))
                        elif curve is not None:
                            xs = curve["MaturityY"].to_numpy().astype(float)
                            ys = curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                            return float(np.interp(t_, xs, ys))
                        return 0.043

                    # Theta for HW: ensures fit to initial discount curve
                    # ±(t) ├ö├½├¬ ├ö├¬├®f/├ö├¬├®t + a≈f(t) + ┬ñ├ó~/(2a)(1-exp(-2at))
                    def hw_theta(t_):
                        dt_s = 0.001
                        f0 = ois_zero(t_)
                        f1 = ois_zero(t_ + dt_s)
                        dfdt = (f1 - f0) / dt_s
                        return dfdt + hw_mr * f0 + sigmas_hw[-1]**2 / (2*hw_mr) * \
                               (1 - math.exp(-2*hw_mr*t_)) if hw_mr > 1e-8 else dfdt

                    # Node short rates: r(i,j) = f(0,t_i) + x(j) + alpha(t_i)
                    # x(j) = j * dr, alpha calibrated to fit discount curve
                    # For simplicity: use r(i,j) = ois_zero(t_i) + j*dr
                    n_nodes = 2 * M + 1
                    # Initialise: terminal swap values
                    t_grid = [i * dt for i in range(N_steps + 1)]

                    # Option value at each node   —   backward induction
                    # V[j] = value at node j at current time step
                    # j ranges from -M to +M
                    import numpy as _np2

                    def swap_value_at_node(t_ex_, r_node_):
                        """PV of underlying swap at an exercise date node."""
                        if curve is None:
                            return (fwd_b - K_b) * ann_b * berm_notional * 1e6
                        # Forward swap rate from node (approx: shift OIS zero by r_node)
                        # Simplified: use flat rate shift
                        T_rem_ = T_final - t_ex_
                        fwd_, ann_, _ = forward_and_annuity_from_curve(
                            curve, ccy, t_ex_, T_rem_, ois_curve)
                        # Rate shift: fwd moves roughly 1:1 with short rate shock
                        fwd_shifted = fwd_ + r_node_
                        if "Payer" in berm_type or "pay-fixed" in berm_type:
                            return (fwd_shifted - K_b) * ann_ * berm_notional * 1e6
                        else:
                            return (K_b - fwd_shifted) * ann_ * berm_notional * 1e6

                    # Build value array at final maturity (swap has expired)
                    V = _np2.zeros(2 * M + 1)

                    # Backward from T_final to t=0
                    for i in range(N_steps - 1, -1, -1):
                        t_i = t_grid[i]
                        t_next = t_grid[i + 1]
                        sigma_i = sigmas_hw[min(
                            sum(1 for te in ex_dates if te <= t_i), len(sigmas_hw)-1)]
                        dr_i = math.sqrt(3 * sigma_i**2 * dt)
                        V_new = _np2.zeros(2 * M + 1)
                        for j in range(-M, M + 1):
                            r_j = ois_zero(t_i) + j * dr_i
                            df_step = math.exp(-r_j * dt)
                            # Branching probabilities (standard HW trinomial)
                            eta = hw_mr * (ois_zero(t_i) - r_j) * dt / dr_i \
                                  if dr_i > 1e-10 else 0.0
                            eta = max(min(eta, M - 1), -(M - 1))
                            k = round(eta)
                            p_u = (1/6 + (eta - k)**2/2 + (eta - k)/2)
                            p_m = (2/3 - (eta - k)**2)
                            p_d = (1/6 + (eta - k)**2/2 - (eta - k)/2)
                            p_u = max(p_u, 0); p_m = max(p_m, 0); p_d = max(p_d, 0)
                            s = p_u + p_m + p_d
                            if s > 0: p_u, p_m, p_d = p_u/s, p_m/s, p_d/s
                            # Continuation value
                            j_u = min(j + k + 1, M); j_m = min(j + k, M); j_d = max(j + k - 1, -M)
                            j_u = max(j_u, -M); j_d = min(j_d, M)
                            cont = df_step * (p_u * V[j_u + M] + p_m * V[j_m + M] + p_d * V[j_d + M])
                            # Exercise check
                            if any(abs(t_i - te) < dt / 2 for te in ex_dates):
                                exercise = max(swap_value_at_node(t_i, j * dr_i), 0.0)
                                V_new[j + M] = max(cont, exercise)
                            else:
                                V_new[j + M] = cont
                        V = V_new

                    # Price = V[M] (j=0 node at t=0)
                    berm_pv = V[M]
                    berm_pv_bp = berm_pv / (berm_notional * 1e6) * 10000.0

                    # European swaption for comparison (first exercise date)
                    if curve is not None:
                        fwd_e, ann_e, _ = forward_and_annuity_from_curve(
                            curve, ccy, ex_dates[0], T_swap, ois_curve)
                    else:
                        fwd_e, ann_e = fwd_b, ann_b
                    mkt_vol_e = _co_vol(ex_dates[0], T_swap) / 10000.0
                    if ois_curve is not None:
                        xs = ois_curve["MaturityY"].to_numpy().astype(float)
                        ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                        df_e = math.exp(-float(np.interp(ex_dates[0], xs, ys)) * ex_dates[0])
                    else:
                        df_e = math.exp(-0.043 * ex_dates[0])
                    from statistics import NormalDist as _nd_b
                    _ndb = _nd_b()
                    sqrt_e = math.sqrt(ex_dates[0])
                    if "Payer" in berm_type or "pay-fixed" in berm_type:
                        euro_pv = df_e * berm_notional * 1e6 * (
                            (fwd_e - K_b) * _ndb.cdf((fwd_e - K_b)/(mkt_vol_e*sqrt_e) if mkt_vol_e*sqrt_e > 0 else 0) +
                            mkt_vol_e * sqrt_e * _ndb.pdf((fwd_e - K_b)/(mkt_vol_e*sqrt_e) if mkt_vol_e*sqrt_e > 0 else 0)
                        ) * ann_e
                    else:
                        euro_pv = df_e * berm_notional * 1e6 * (
                            (K_b - fwd_e) * _ndb.cdf((K_b - fwd_e)/(mkt_vol_e*sqrt_e) if mkt_vol_e*sqrt_e > 0 else 0) +
                            mkt_vol_e * sqrt_e * _ndb.pdf((fwd_e - K_b)/(mkt_vol_e*sqrt_e) if mkt_vol_e*sqrt_e > 0 else 0)
                        ) * ann_e
                    euro_bp = euro_pv / (berm_notional * 1e6) * 10000.0

                    st.markdown("---")
                    br1, br2, br3, br4 = st.columns(4)
                    br1.metric("Bermudan PV (AUD)", f"${berm_pv:,.0f}")
                    br2.metric("Bermudan PV (bp)", f"{berm_pv_bp:.2f}bp")
                    br3.metric("European PV (first call)", f"${euro_pv:,.0f}")
                    br4.metric("Bermudan Premium", f"${berm_pv - euro_pv:,.0f}",
                               delta=f"{berm_pv_bp - euro_bp:.2f}bp over European")

                    with st.expander("📅 HW1F Calibration   —   Co-terminal ┬ñ├ó"):
                        df_cal = pd.DataFrame([{
                            "Exercise Date": f"{te:.2f}y",
                            "Co-terminal Swap": f"{T_final-te:.1f}Y",
                            "Mkt Vol (bp)": f"{_co_vol(te, T_final-te):.1f}",
                            "HW ┬ñ├ó (%)": f"{sigmas_hw[i]*100:.4f}",
                        } for i, te in enumerate(ex_dates)])
                        st.dataframe(df_cal, use_container_width=True, hide_index=True)
                        st.caption(f"Mean reversion a = {hw_mr*100:.2f}%  |  "
                                   f"Tree: {N_steps} steps, dt={dt:.4f}y, dr={dr_i:.6f}, M=~{M}")

                except Exception as e:
                    st.error(f"Pricing error: {e}")
                    import traceback
                    st.code(traceback.format_exc())

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 4   —   DIGITAL LADDER (existing)
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with sub[3]:
        st.markdown("### Digital Ladder")

        col_not, col_exp, col_tenor = st.columns(3)
        with col_not:
            notional = st.number_input("Notional (mm)", 0.0, 1e6, 100.0, step=10.0, key="ex_not")
        with col_exp:
            final_expiry = st.selectbox("Final expiry",
                ["1y","2y","3y","4y","5y","7y","10y","15y","20y"], index=2, key="ex_final_exp")
            final_expiry_y = label_to_years(final_expiry)
        with col_tenor:
            swap_tenor = st.selectbox("Underlying swap tenor",
                ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","15Y","20Y","30Y"],
                index=4, key="ex_tenor")
            tenor_y = float(swap_tenor[:-1])

        if curve is not None:
            fwd, ann, _ = forward_and_annuity_from_curve(curve, ccy, final_expiry_y, tenor_y, ois_curve)
        else:
            fwd, ann = 0.04, tenor_y

        strike_pct = st.number_input("ATM strike (%)", value=round(fwd * 100, 4), key="ex_strike") / 100.0

        atm_val_dl = get_matrix_value(atm, final_expiry, tenor_y) if atm is not None else None
        base_sigma = (atm_val_dl / 10000.0) if atm_val_dl else 0.0035

        # OIS df
        if ois_curve is not None:
            xs = ois_curve["MaturityY"].to_numpy().astype(float)
            ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
            disc_ois = float(np.interp(final_expiry_y, xs, ys))
        else:
            disc_ois = 0.04

        if st.button("Price Digital Ladder", key="ex_price_dl"):
            ladder_strikes = [strike_pct + k * 0.0025 for k in range(-2, 3)]
            rows = []
            for k in ladder_strikes:
                ticket = SwaptionTicket(
                    side="Payer", payoff_type="digital",
                    notional=notional * 1e6, currency=ccy,
                    expiry_years=final_expiry_y, swap_tenor_years=tenor_y,
                    forward=fwd, strike=k, vol=base_sigma,
                    discount_rate=disc_ois, annuity=ann,
                    model="Normal",
                    payout_bp=25.0, label=f"Digital @ {k*100:.2f}%",
                    use_curve=curve is not None,
                )
                res = price_swaption(ticket)
                rows.append(dict(strike_pct=k*100, pv=res["pv"], pv_bp=res["pv_bp"]))
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True)


    with sub[4]:
        st.markdown("### Zero Coupon Swap")
        st.caption(
            "**Notional = Ending (Maturity) Amount**   —   market convention. "
            "Starting notional is back-calculated from the fixed rate. "
            "Fixed leg: N_end  →  N_start (single payment). "
            "Floating leg: N_start ≈ [├ö├¬├à(1+BBSW_i≈+_i)  →  1] (single payment). "
            "AUD: annual compounding, Act/365, T+1 settlement."
        )

        # ── Inputs ───────────────────────────────────────────────────
        zi1, zi2, zi3, zi4 = st.columns(4)
        with zi1:
            zcs_notional = st.number_input("Ending Notional / Maturity Amount (AUD mm)", 1.0, 5000.0,
                                            100.0, step=10.0, key="zcs_notl")
            zcs_end_notional = zcs_notional * 1e6   # N_end   —   the maturity amount
        with zi2:
            zcs_tenor_sel = st.selectbox("Tenor",
                ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"],
                index=4, key="zcs_tenor")
            zcs_n = float(zcs_tenor_sel[:-1])
        with zi3:
            zcs_side = st.selectbox("Direction",
                ["Pay Fixed / Receive Floating", "Receive Fixed / Pay Floating"],
                key="zcs_side")
            zcs_pay_fixed = "Pay Fixed" in zcs_side
        with zi4:
            zcs_bbsw_freq = st.selectbox("BBSW Reset Freq",
                ["Quarterly (3m)", "Semi-Annual (6m)"],
                index=0, key="zcs_freq")
            zcs_delta = 0.25 if "Quarterly" in zcs_bbsw_freq else 0.5

        # Trade date / effective / maturity
        zd1, zd2, zd3 = st.columns(3)
        with zd1:
            from datetime import date as _d, timedelta as _td
            zcs_trade_date = st.date_input("Trade Date",
                value=_d.today(), format="DD/MM/YYYY", key="zcs_tdate")
        with zd2:
            # AUD: T+1 BD (AFMA)   —   uses full holiday calendar
            zcs_eff_date = au_spot_date(zcs_trade_date, ccy)
            st.metric("Effective Date", zcs_eff_date.strftime("%d %b %Y"))
        with zd3:
            # Modified Following maturity   —   AFMA convention
            zcs_mat_date = au_end_date(zcs_eff_date, zcs_n)
            st.metric("Maturity Date", zcs_mat_date.strftime("%d %b %Y"))

        # Actual year fraction (Act/365)
        zcs_T = (zcs_mat_date - zcs_eff_date).days / 365.0

        # ── Pricing ──────────────────────────────────────────────────
        # OIS df(T) for discounting
        if ois_curve is not None:
            xs_o = ois_curve["MaturityY"].to_numpy().astype(float)
            ys_o = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
            df_T_zcs = math.exp(-float(np.interp(zcs_T, xs_o, ys_o)) * zcs_T)
        else:
            df_T_zcs = math.exp(-0.043 * zcs_T)

        # Par ZCS fixed rate: r = df(T)^(-1/T) - 1  (annual compounding)
        zcs_par_rate = df_T_zcs ** (-1.0 / zcs_T) - 1.0 if zcs_T > 0 else 0.0

        zp1, zp2 = st.columns([2, 2])
        with zp1:
            zcs_fixed_rate = st.number_input(
                "Fixed Rate (% pa, annual compounding)",
                min_value=0.001, max_value=20.0,
                value=round(zcs_par_rate * 100, 4),
                step=0.001, format="%.4f", key="zcs_fixed_rate") / 100.0
        with zp2:
            st.caption(" ")
            st.markdown(f"**Par ZCS rate:** `{zcs_par_rate*100:.4f}%`  "
                        f"{'✅ ATM' if abs(zcs_fixed_rate - zcs_par_rate) < 0.00005 else f'> {(zcs_fixed_rate-zcs_par_rate)*10000:+.1f}bp'}")

        # ── Notional derivation ───────────────────────────────────────
        # N_start = N_end / (1 + r)^T     —   back-calculated from ending notional + fixed rate
        zcs_compound_factor_fixed = (1.0 + zcs_fixed_rate) ** zcs_T
        zcs_start_notional = zcs_end_notional / zcs_compound_factor_fixed
        # For legacy compatibility keep zcs_face = start notional (used in float schedule)
        zcs_face = zcs_start_notional

        # ── Fixed leg ─────────────────────────────────────────────────
        # Fixed payment = N_end - N_start  (the compounded interest only)
        zcs_fixed_payment = zcs_end_notional - zcs_start_notional
        zcs_fixed_pv = zcs_fixed_payment * df_T_zcs

        # ── Floating leg ─────────────────────────────────────────────
        # Build BBSW reset schedule using AFMA calendar (Modified Following)
        def _fwd_bbsw(t_start, accrual_frac_):
            """Forward BBSW rate for a period starting at t_start with given accrual."""
            if curve is None:
                return 0.043
            xs_c = curve["MaturityY"].to_numpy().astype(float)
            ys_c = curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
            r_s = float(np.interp(t_start, xs_c, ys_c))
            r_e = float(np.interp(t_start + accrual_frac_, xs_c, ys_c))
            df_s = math.exp(-r_s * t_start)
            df_e = math.exp(-r_e * (t_start + accrual_frac_))
            return (df_s / df_e - 1.0) / accrual_frac_ if accrual_frac_ > 0 else 0.0

        freq_months = 3 if zcs_delta == 0.25 else 6
        _raw_periods = build_au_reset_schedule(zcs_eff_date, zcs_mat_date, freq_months)
        reset_dates = []
        for i, p in enumerate(_raw_periods):
            t_start = (p["start_date"] - zcs_eff_date).days / 365.0
            fwd_rate = _fwd_bbsw(t_start, p["accrual_frac"])
            reset_dates.append({
                "period": i + 1,
                "reset_date": p["reset_date"],
                "start_date": p["start_date"],
                "end_date": p["end_date"],
                "accrual_days": p["accrual_days"],
                "accrual_frac": p["accrual_frac"],
                "t_start": t_start,
                "fwd_bbsw": fwd_rate,
            })

        # Compound floating leg: product of (1 + BBSW_i ≈ +_i) over all periods
        compounding_factor = 1.0
        running_balance = zcs_face
        for p in reset_dates:
            factor_i = 1.0 + p["fwd_bbsw"] * p["accrual_frac"]
            p["factor_i"] = factor_i
            compounding_factor *= factor_i
            running_balance = zcs_face * compounding_factor
            p["running_balance"] = running_balance

        zcs_float_payment = zcs_face * (compounding_factor - 1.0)
        zcs_float_pv = zcs_float_payment * df_T_zcs  # ├ö├½├¬ N ≈ (1 - df(T))

        # Net
        if zcs_pay_fixed:
            zcs_net_pv = zcs_float_pv - zcs_fixed_pv
            zcs_net_payment = zcs_float_payment - zcs_fixed_payment
        else:
            zcs_net_pv = zcs_fixed_pv - zcs_float_pv
            zcs_net_payment = zcs_fixed_payment - zcs_float_payment

        # DV01   —   bump fixed rate 1bp
        # DV01: bump fixed rate +1bp, recompute start notional and fixed payment
        _rate_up = zcs_fixed_rate + 0.0001
        _start_up = zcs_end_notional / ((1.0 + _rate_up) ** zcs_T)
        fixed_pv_up = (zcs_end_notional - _start_up) * df_T_zcs
        zcs_dv01 = abs(fixed_pv_up - zcs_fixed_pv)

        # ── Metrics ──────────────────────────────────────────────────
        st.markdown("---")
        mm1, mm2, mm3, mm4, mm5, mm6 = st.columns(6)
        mm1.metric("Fixed Payment (AUD)", f"${zcs_fixed_payment:,.0f}")
        mm2.metric("Float Payment (est.)", f"${zcs_float_payment:,.0f}")
        mm3.metric("Net at Maturity", f"${zcs_net_payment:,.0f}",
                    delta=f"{'Pay' if zcs_net_payment < 0 else 'Receive'}")
        mm4.metric("Fixed PV", f"${zcs_fixed_pv:,.0f}")
        mm5.metric("Float PV (est.)", f"${zcs_float_pv:,.0f}")
        mm6.metric("MTM (Net PV)", f"${zcs_net_pv:,.0f}")

        mn1, mn2, mn3, mn4, mn5, mn6 = st.columns(6)
        mn1.metric("Par ZCS Rate", f"{zcs_par_rate*100:.4f}%")
        mn2.metric("Fixed Rate", f"{zcs_fixed_rate*100:.4f}%")
        mn3.metric("Starting Notional", f"${zcs_start_notional/1e6:,.4f}mm")
        mn4.metric("Ending Notional", f"${zcs_end_notional/1e6:,.2f}mm")
        mn5.metric("DV01 (AUD/bp)", f"${zcs_dv01:,.0f}")
        mn6.metric("df(T)", f"{df_T_zcs:.6f}")
        st.caption(f"T = {zcs_T:.4f}y  |  {len(reset_dates)} BBSW resets  |  "
                   f"Float compounding factor = {compounding_factor:.6f}  |  "
                   f"Fixed compound = {zcs_compound_factor_fixed:.6f}  |  "
                   f"Eff: {zcs_eff_date.strftime('%d %b %Y')}  |  "
                   f"Mat: {zcs_mat_date.strftime('%d %b %Y')}")

        # ── Reset schedule table ─────────────────────────────────────
        st.markdown("#### Floating Leg Reset Schedule (Estimated)")
        st.caption(
            "Compounding: each period's interest accrues on the PREVIOUS period's balance (interest on interest). "
            "No payments until maturity. Forward BBSW rates from curve   —   actual fixings will differ."
        )
        _prev_bal = zcs_start_notional
        _rows_zcs = []
        for _pi, p in enumerate(reset_dates):
            opening_bal  = _prev_bal
            period_int   = opening_bal * p["fwd_bbsw"] * p["accrual_frac"]
            # Last period: closing balance = agreed N_end ($100mm)
            is_last      = (_pi == len(reset_dates) - 1)
            closing_bal  = zcs_end_notional if is_last else (opening_bal + period_int)
            if is_last:
                period_int = closing_bal - opening_bal   # back-solve final period interest
            accrued_tot  = closing_bal - zcs_start_notional
            _rows_zcs.append({
                "#":                        p["period"],
                "Reset Date":               p["reset_date"].strftime("%d %b %Y"),
                "Period Start":             p["start_date"].strftime("%d %b %Y"),
                "Period End":               p["end_date"].strftime("%d %b %Y"),
                "Days":                     p["accrual_days"],
                "Fwd BBSW (%)":             f"{p['fwd_bbsw']*100:.4f}",
                "Opening Balance (AUD)":    f"${opening_bal:,.2f}",
                "Period Interest (AUD)":    f"${period_int:,.2f}",
                "Closing Balance (AUD)":    f"${closing_bal:,.2f}",
                "Accrued Interest (AUD)":   f"${accrued_tot:,.2f}",
            })
            _prev_bal = closing_bal
        df_sched_zcs = pd.DataFrame(_rows_zcs)
        st.dataframe(df_sched_zcs, use_container_width=True, hide_index=True)
        st.info(
            f"**Agreed Ending Notional (both legs): AUD {zcs_end_notional:,.2f}**  |  "
            f"Est. Float Total Interest: AUD {zcs_float_payment:,.2f}  |  "
            f"Fixed Total Interest: AUD {zcs_fixed_payment:,.2f}  |  "
            f"Closing Balance at Maturity: AUD {_prev_bal:,.2f}"
        )

        st.markdown(f"""
| | **Fixed Leg** | **Floating Leg** |
|---|---|---|
| **Payment Type** | Single at maturity | Single at maturity |
| **Starting Notional** | AUD {zcs_start_notional:,.2f} | AUD {zcs_start_notional:,.2f} |
| **Ending Notional** | AUD {zcs_end_notional:,.2f} | AUD {zcs_end_notional:,.2f} |
| **Amount (est.)** | AUD {zcs_fixed_payment:,.2f} | AUD {zcs_float_payment:,.2f} |
| **Compounding** | Annual: N_end  →  N_start | N_start ≈ [├ö├¬├à(1+BBSW≈+)  →  1] |
| **Payer** | {'You' if zcs_pay_fixed else 'Counterparty'} | {'Counterparty' if zcs_pay_fixed else 'You'} |
""")

        # ── Excel schedule generator ─────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📄 Client Payment Schedule")

        st.caption(
            "LEI = Legal Entity Identifier (20-char)   —   required for ASIC trade reporting by both swap counterparties. "
            "RateEdge as arranger/broker is not a principal and does not require an LEI. "
            "BIC = SWIFT code identifying the party's LCH clearing member. "
            "AUD ZCS: LCH SwapClear eligible (BBSW, up to 31Y), CFTC-mandated cleared 28d-30Y."
        )

        zc1, zc2, zc3 = st.columns(3)
        with zc1:
            st.markdown("**Fixed Rate Payer**")
            client_name   = st.text_input("Name", value="[Counterparty Name]", key="zcs_client")
            client_lei    = st.text_input("LEI", value="[20-char LEI]", key="zcs_client_lei")
            client_bic    = st.text_input("BIC / Clearing Member", value="[SWIFT BIC]", key="zcs_client_bic")
        with zc2:
            st.markdown("**Floating Rate Payer**")
            dealer_name   = st.text_input("Name", value="RateEdge", key="zcs_dealer")
            dealer_lei    = st.text_input("LEI (counterparty's)", value="[20-char LEI]", key="zcs_dealer_lei")
            dealer_bic    = st.text_input("BIC / Clearing Member", value="[SWIFT BIC]", key="zcs_dealer_bic")
        with zc3:
            st.markdown("**Arranging Broker**")
            broker_co     = st.text_input("Firm", value="RateEdge", key="zcs_broker_co")
            broker_name   = st.text_input("Broker", value="WPO", key="zcs_broker_name")
            broker_email  = st.text_input("Email", value="wpo@rateedge.au", key="zcs_broker_email")
            # Legacy alias so rest of Excel code still works
            client_entity = client_lei
            dealer_entity = dealer_lei

        if st.button("📂 Generate Client Schedule (Excel)", key="zcs_gen_excel"):
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, numbers)
            from openpyxl.utils import get_column_letter
            import io

            wb = openpyxl.Workbook()

            # Party labels   —   always by role, not "client/dealer"
            party_fixed_payer    = dealer_name if zcs_pay_fixed else client_name
            party_fixed_receiver = client_name if zcs_pay_fixed else dealer_name
            lei_fixed_payer      = dealer_lei  if zcs_pay_fixed else client_lei
            bic_fixed_payer      = dealer_bic  if zcs_pay_fixed else client_bic
            lei_fixed_receiver   = client_lei  if zcs_pay_fixed else dealer_lei
            bic_fixed_receiver   = client_bic  if zcs_pay_fixed else dealer_bic

            # ── Styles ───────────────────────────────────────────────
            navy_fill   = PatternFill("solid", start_color="1E3A5F")
            red_fill    = PatternFill("solid", start_color="DC2626")
            light_fill  = PatternFill("solid", start_color="EFF6FF")
            mid_fill    = PatternFill("solid", start_color="DBEAFE")
            white_fill  = PatternFill("solid", start_color="FFFFFF")
            hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            sub_font    = Font(name="Arial", bold=True, color="1E3A5F", size=10)
            body_font   = Font(name="Arial", size=10)
            bold_font   = Font(name="Arial", bold=True, size=10)
            right_al    = Alignment(horizontal="right")
            left_al     = Alignment(horizontal="left")
            centre_al   = Alignment(horizontal="center")
            thin        = Side(style="thin", color="CBD5E1")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            btm_border  = Border(bottom=Side(style="medium", color="1E3A5F"))

            def _hdr(ws, row, col, val, fill=navy_fill, font=hdr_font, al=centre_al):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill; c.font = font; c.alignment = al
                c.border = border

            def _cell(ws, row, col, val, font=body_font, al=right_al, fill=white_fill, fmt=None):
                c = ws.cell(row=row, column=col, value=val)
                c.font = font; c.alignment = al; c.fill = fill; c.border = border
                if fmt: c.number_format = fmt

            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            # SHEET 1   —   TRADE CONFIRMATION SUMMARY
            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            ws1 = wb.active
            ws1.title = "Trade Summary"
            ws1.column_dimensions["A"].width = 28
            ws1.column_dimensions["B"].width = 26
            ws1.column_dimensions["C"].width = 8
            ws1.column_dimensions["D"].width = 28
            ws1.column_dimensions["E"].width = 26

            # Title banner
            ws1.merge_cells("A1:E1")
            c = ws1["A1"]
            c.value = "ZERO COUPON SWAP   —   INDICATIVE TRADE CONFIRMATION"
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
            c.fill = navy_fill; c.alignment = centre_al

            ws1.merge_cells("A2:E2")
            c = ws1["A2"]
            c.value = f"RateEdge Options Platform  |  {zcs_trade_date.strftime('%d %B %Y')}  |  Arranged by: {broker_co}   —   {broker_name}  |  {broker_email}"
            c.font = Font(name="Arial", color="FFFFFF", size=9, italic=True)
            c.fill = red_fill; c.alignment = centre_al

            # Parties
            ws1.merge_cells("A3:B3")
            ws1["A3"].value = "PARTIES"; ws1["A3"].font = hdr_font
            ws1["A3"].fill = PatternFill("solid", start_color="334155")
            ws1["A3"].alignment = centre_al

            party_data = [
                ("Payer Fixed (Fixed Leg)",
                 f"{party_fixed_payer}  |  LEI: {lei_fixed_payer}  |  BIC: {bic_fixed_payer}"),
                ("Payer Floating (Floating Leg)",
                 f"{party_fixed_receiver}  |  LEI: {lei_fixed_receiver}  |  BIC: {bic_fixed_receiver}"),
            ]
            for i, (lbl, val) in enumerate(party_data, start=4):
                ws1.cell(row=i, column=1, value=lbl).font = bold_font
                ws1.cell(row=i, column=1).fill = light_fill
                ws1.cell(row=i, column=2, value=val).font = body_font
                ws1.cell(row=i, column=2).fill = white_fill

            # Trade economics
            row = 7
            ws1.merge_cells(f"A{row}:B{row}")
            ws1[f"A{row}"].value = "TRADE ECONOMICS"
            ws1[f"A{row}"].font = hdr_font
            ws1[f"A{row}"].fill = PatternFill("solid", start_color="334155")
            ws1[f"A{row}"].alignment = centre_al

            econ = [
                ("Ending Notional (Maturity Amount)", f"AUD {zcs_end_notional:,.2f}"),
                ("Starting Notional (PV of Ending)", f"AUD {zcs_start_notional:,.2f}"),
                ("Trade Date", zcs_trade_date.strftime("%d %B %Y")),
                ("Effective Date", zcs_eff_date.strftime("%d %B %Y")),
                ("Maturity Date", zcs_mat_date.strftime("%d %B %Y")),
                ("Tenor", f"{zcs_n:.0f} Years"),
                ("Day Count", "Actual / 365"),
                ("Business Day Convention", "Following"),
                ("Settlement", "T+1 AUD Business Day (AFMA)"),
                ("BBSW Reset Frequency", "Quarterly (3 Month BBSW)" if zcs_delta == 0.25 else "Semi-Annual (6 Month BBSW)"),
            ]
            for i, (lbl, val) in enumerate(econ, start=row+1):
                ws1.cell(row=i, column=1, value=lbl).font = sub_font
                ws1.cell(row=i, column=1).fill = light_fill
                ws1.cell(row=i, column=2, value=val).font = body_font
                ws1.cell(row=i, column=2).fill = white_fill

            # Fixed / Floating legs side by side
            row_leg = row + len(econ) + 2
            _hdr(ws1, row_leg, 1, "FIXED LEG", fill=PatternFill("solid", start_color="1E3A5F"))
            ws1.merge_cells(f"A{row_leg}:B{row_leg}")
            _hdr(ws1, row_leg, 4, "FLOATING LEG", fill=PatternFill("solid", start_color="334155"))
            ws1.merge_cells(f"D{row_leg}:E{row_leg}")

            fixed_leg = [
                ("Payer Fixed", party_fixed_payer),
                ("Receives Floating", party_fixed_receiver),
                ("Fixed Rate", f"{zcs_fixed_rate*100:.4f}% pa"),
                ("Compounding", "Annual"),
                ("Payment", "Single payment at maturity"),
                ("Estimated Amount", f"AUD {zcs_fixed_payment:,.2f}"),
                ("Calculation", f"AUD {zcs_end_notional:,.2f}  →  AUD {zcs_start_notional:,.2f} = N_end  →  N_start"),
            ]
            float_leg = [
                ("Payer Floating", party_fixed_receiver),
                ("Receives Fixed", party_fixed_payer),
                ("Floating Index", f"{int(zcs_delta*12)}M BBSW"),
                ("Compounding", "Quarterly   —   no interim payments"),
                ("Payment", "Single payment at maturity"),
                ("Estimated Amount", f"AUD {zcs_float_payment:,.2f}"),
                ("Calculation", f"AUD {zcs_start_notional:,.2f} ≈ [├ö├¬├à(1 + BBSW├ƒ├ü├│ ≈ +├ƒ├ü├│)  →  1]"),
            ]
            for i, ((fl, fv), (ll, lv)) in enumerate(zip(fixed_leg, float_leg), start=row_leg+1):
                ws1.cell(row=i, column=1, value=fl).font = sub_font
                ws1.cell(row=i, column=1).fill = light_fill
                ws1.cell(row=i, column=2, value=fv).font = body_font
                ws1.cell(row=i, column=2).fill = white_fill
                ws1.cell(row=i, column=4, value=ll).font = sub_font
                ws1.cell(row=i, column=4).fill = mid_fill
                ws1.cell(row=i, column=5, value=lv).font = body_font
                ws1.cell(row=i, column=5).fill = white_fill

            # Final net exchange
            row_net = row_leg + len(fixed_leg) + 2
            ws1.merge_cells(f"A{row_net}:E{row_net}")
            ws1[f"A{row_net}"].value = "FINAL EXCHANGE AT MATURITY"
            ws1[f"A{row_net}"].font = hdr_font
            ws1[f"A{row_net}"].fill = red_fill
            ws1[f"A{row_net}"].alignment = centre_al

            net_rows = [
                ("Fixed Leg Amount (est.)", f"AUD {zcs_fixed_payment:,.2f}"),
                ("Floating Leg Amount (est.)", f"AUD {zcs_float_payment:,.2f}"),
                ("Net Payment (est.)", f"{'Pay ' if zcs_net_payment > 0 else 'Receive '}AUD {abs(zcs_net_payment):,.2f}"),
                ("Net Payer (est.)", party_fixed_payer if zcs_net_payment > 0 else party_fixed_receiver),
                ("Par Fixed Rate (at deal date)", f"{zcs_par_rate*100:.4f}%"),
                ("MTM (at deal date)", f"AUD {zcs_net_pv:,.2f}"),
                ("", ""),
                ("Arranging Broker", f"{broker_co}"),
                ("Broker Contact", f"{broker_name}  |  {broker_email}"),
                ("Payer Fixed LEI", f"{lei_fixed_payer}"),
                ("Payer Fixed BIC", f"{bic_fixed_payer}"),
                ("Payer Floating LEI", f"{lei_fixed_receiver}"),
                ("Payer Floating BIC", f"{bic_fixed_receiver}"),
                ("Clearing Venue", "LCH SwapClear (AUD BBSW ZCS   —   eligible up to 31Y)"),
                ("CFTC Mandate", "Mandatory cleared: 28d-30Y residual maturity"),
                ("ASIC Reporting", "Both parties to report under ASIC derivative rules"),
            ]
            for i, (lbl, val) in enumerate(net_rows, start=row_net+1):
                ws1.cell(row=i, column=1, value=lbl).font = sub_font
                ws1.cell(row=i, column=1).fill = mid_fill
                ws1.cell(row=i, column=2, value=val).font = bold_font
                ws1.cell(row=i, column=2).fill = white_fill
                ws1.merge_cells(f"B{i}:E{i}")

            # Disclaimer
            row_disc = row_net + len(net_rows) + 2
            ws1.merge_cells(f"A{row_disc}:E{row_disc+2}")
            ws1[f"A{row_disc}"].value = (
                "INDICATIVE TERMS ONLY   —   NOT A BINDING CONFIRMATION. "
                "Estimated floating amounts are based on prevailing forward BBSW rates as at trade date and will differ from actual fixings. "
                "Final amounts to be confirmed on maturity date based on published BBSW fixings. "
                "Subject to ISDA Master Agreement and Credit Support Annex. "
                "RateEdge is not an AFSL holder   —   for indicative purposes only."
            )
            ws1[f"A{row_disc}"].font = Font(name="Arial", size=8, italic=True, color="64748B")
            ws1[f"A{row_disc}"].alignment = Alignment(wrap_text=True)
            ws1.row_dimensions[row_disc].height = 45

            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            # SHEET 2   —   FLOATING LEG RESET SCHEDULE
            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            ws2 = wb.create_sheet("Floating Reset Schedule")
            col_widths = [6, 14, 14, 14, 8, 14, 22, 22, 22, 22]
            col_hdrs = ["#", "Reset Date", "Period Start", "Period End",
                        "Days", "Fwd BBSW (%)",
                        "Opening Balance (AUD)", "Period Interest (AUD)",
                        "Closing Balance (AUD)", "Accrued Interest (AUD)"]
            for ci, (w, h) in enumerate(zip(col_widths, col_hdrs), start=1):
                ws2.column_dimensions[get_column_letter(ci)].width = w
                _hdr(ws2, 1, ci, h)

            # Title
            ws2.insert_rows(1)
            ws2.merge_cells(f"A1:{get_column_letter(len(col_hdrs))}1")
            ws2["A1"].value = (
                f"FLOATING LEG   —   ESTIMATED BBSW RESET SCHEDULE  |  {client_name}  |  "
                f"Starting Notional AUD {zcs_start_notional/1e6:,.4f}mm  |  "
                f"Ending Notional AUD {zcs_end_notional/1e6:,.2f}mm  |  "
                f"No interim payments   —   all interest compounds and pays at maturity"
            )
            ws2["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            ws2["A1"].fill = navy_fill; ws2["A1"].alignment = centre_al

            _xl_prev_bal = zcs_start_notional
            for _xl_pi, p in enumerate(reset_dates):
                row_r = p["period"] + 2
                alt = light_fill if p["period"] % 2 == 0 else white_fill
                opening_b  = _xl_prev_bal
                period_i   = opening_b * p["fwd_bbsw"] * p["accrual_frac"]
                _xl_is_last = (_xl_pi == len(reset_dates) - 1)
                closing_b  = zcs_end_notional if _xl_is_last else (opening_b + period_i)
                if _xl_is_last:
                    period_i = closing_b - opening_b
                accrued_i  = closing_b - zcs_start_notional
                _cell(ws2, row_r, 1, p["period"], al=centre_al, fill=alt)
                _cell(ws2, row_r, 2, p["reset_date"].strftime("%d %b %Y"), al=centre_al, fill=alt)
                _cell(ws2, row_r, 3, p["start_date"].strftime("%d %b %Y"), al=centre_al, fill=alt)
                _cell(ws2, row_r, 4, p["end_date"].strftime("%d %b %Y"), al=centre_al, fill=alt)
                _cell(ws2, row_r, 5, p["accrual_days"], al=centre_al, fill=alt)
                _cell(ws2, row_r, 6, round(p["fwd_bbsw"]*100, 4), fmt="0.0000", fill=alt)
                _cell(ws2, row_r, 7, round(opening_b, 2), fmt='$#,##0.00', fill=alt)
                _cell(ws2, row_r, 8, round(period_i,  2), fmt='$#,##0.00', fill=alt)
                _cell(ws2, row_r, 9, round(closing_b, 2), fmt='$#,##0.00', fill=alt)
                _cell(ws2, row_r, 10, round(accrued_i, 2), fmt='$#,##0.00', fill=alt)
                _xl_prev_bal = closing_b

            # Totals row
            tot_row = len(reset_dates) + 3
            for ci in range(1, len(col_hdrs)+1):
                ws2.cell(row=tot_row, column=ci).fill = PatternFill("solid", start_color="1E3A5F")
            ws2.cell(row=tot_row, column=1, value="TOTAL").font = hdr_font
            ws2.cell(row=tot_row, column=1).alignment = centre_al
            ws2.cell(row=tot_row, column=8, value=round(zcs_float_payment, 2)).font = hdr_font
            ws2.cell(row=tot_row, column=8).number_format = '$#,##0.00'
            ws2.cell(row=tot_row, column=8).alignment = right_al
            ws2.cell(row=tot_row, column=9, value=round(_xl_prev_bal, 2)).font = hdr_font
            ws2.cell(row=tot_row, column=9).number_format = '$#,##0.00'
            ws2.cell(row=tot_row, column=9).alignment = right_al
            ws2.cell(row=tot_row, column=10, value=round(zcs_float_payment, 2)).font = hdr_font
            ws2.cell(row=tot_row, column=10).number_format = '$#,##0.00'
            ws2.cell(row=tot_row, column=10).alignment = right_al

            # Note
            note_row = tot_row + 2
            ws2.merge_cells(f"A{note_row}:{get_column_letter(len(col_hdrs))}{note_row}")
            ws2[f"A{note_row}"].value = (
                "Note: Forward BBSW rates are indicative estimates based on market curve at trade date. "
                "Actual floating payment will be determined by published 3M/6M BBSW fixings on each reset date. "
                "No interim payments are made   —   the entire compounded amount is paid as a single sum at maturity."
            )
            ws2[f"A{note_row}"].font = Font(name="Arial", size=8, italic=True, color="64748B")
            ws2[f"A{note_row}"].alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[note_row].height = 35

            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            # SHEET 3   —   FIXED LEG & FINAL EXCHANGE
            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            ws3 = wb.create_sheet("Fixed Leg & Final Exchange")
            ws3.column_dimensions["A"].width = 30
            ws3.column_dimensions["B"].width = 30
            ws3.column_dimensions["C"].width = 8
            ws3.column_dimensions["D"].width = 30
            ws3.column_dimensions["E"].width = 30

            ws3.merge_cells("A1:E1")
            ws3["A1"].value = "FIXED LEG & FINAL EXCHANGE SUMMARY"
            ws3["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
            ws3["A1"].fill = navy_fill; ws3["A1"].alignment = centre_al

            # Fixed leg calculation breakdown
            ws3.merge_cells("A3:E3")
            ws3["A3"].value = "FIXED LEG CALCULATION"
            ws3["A3"].font = hdr_font; ws3["A3"].fill = PatternFill("solid", start_color="334155")
            ws3["A3"].alignment = centre_al

            fixed_detail = [
                ("Ending Notional (Maturity Amount)", f"AUD {zcs_end_notional:,.2f}", "", ""),
                ("Starting Notional (Back-calculated)", f"AUD {zcs_start_notional:,.2f}", "", ""),
                ("Fixed Rate (pa, annual compounding)", f"{zcs_fixed_rate*100:.4f}%", "", ""),
                ("Day Count Fraction (Act/365)", f"{zcs_T:.6f} years", "", ""),
                ("Compounded Amount", f"(1 + {zcs_fixed_rate*100:.4f}%)^{zcs_T:.4f}", "=", f"{(1+zcs_fixed_rate)**zcs_T:.8f}"),
                ("Compounded Return (factor  →  1)", f"{(1+zcs_fixed_rate)**zcs_T - 1:.8f}", "", ""),
                ("Fixed Leg Payment (N_end  →  N_start)", f"AUD {zcs_fixed_payment:,.2f}", "", ""),
                ("OIS Discount Factor df(T)", f"{df_T_zcs:.8f}", "", ""),
                ("Fixed Leg PV", f"AUD {zcs_fixed_pv:,.2f}", "", ""),
            ]
            for i, (lbl, v1, sep, v2) in enumerate(fixed_detail, start=4):
                ws3.cell(row=i, column=1, value=lbl).font = sub_font
                ws3.cell(row=i, column=1).fill = light_fill
                ws3.cell(row=i, column=2, value=v1).font = body_font
                ws3.cell(row=i, column=2).fill = white_fill

            # Final exchange
            row_fx = len(fixed_detail) + 6
            ws3.merge_cells(f"A{row_fx}:E{row_fx}")
            ws3[f"A{row_fx}"].value = f"FINAL EXCHANGE   —   MATURITY DATE {zcs_mat_date.strftime('%d %B %Y')}"
            ws3[f"A{row_fx}"].font = hdr_font; ws3[f"A{row_fx}"].fill = red_fill
            ws3[f"A{row_fx}"].alignment = centre_al

            fx_rows = [
                ("Fixed Leg Payment", f"AUD {zcs_fixed_payment:,.2f}",
                 "Payer", "RateEdge / WPO" if zcs_pay_fixed else client_name),
                ("Floating Leg Payment (estimated)", f"AUD {zcs_float_payment:,.2f}",
                 "Payer", client_name if zcs_pay_fixed else dealer_name),
                ("NET PAYMENT (estimated)", f"AUD {abs(zcs_net_payment):,.2f}",
                 "Net Payer (est.)", dealer_name if (zcs_pay_fixed and zcs_net_payment > 0) else client_name),
            ]
            for i, (lbl, amt, lbl2, payer) in enumerate(fx_rows, start=row_fx+1):
                ws3.cell(row=i, column=1, value=lbl).font = bold_font if i == row_fx+3 else sub_font
                ws3.cell(row=i, column=1).fill = PatternFill("solid", start_color="DBEAFE") if i < row_fx+3 else PatternFill("solid", start_color="FEF3C7")
                ws3.cell(row=i, column=2, value=amt).font = bold_font
                ws3.cell(row=i, column=2).fill = white_fill
                ws3.cell(row=i, column=4, value=lbl2).font = sub_font
                ws3.cell(row=i, column=4).fill = light_fill
                ws3.cell(row=i, column=5, value=payer).font = body_font
                ws3.cell(row=i, column=5).fill = white_fill

            # Save to bytes
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"ZCS_{ccy}_{zcs_n:.0f}Y_{zcs_trade_date.strftime('%Y%m%d')}_{client_name.replace(' ','_')[:20]}.xlsx"

            st.download_button(
                label=f"📤 Download: {fname}",
                data=buf.getvalue(),
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="zcs_dl"
            )
            st.success(f"✅ Schedule generated   —   {len(reset_dates)} reset periods  |  "
                       f"Fixed: AUD {zcs_fixed_payment:,.0f}  |  "
                       f"Float est.: AUD {zcs_float_payment:,.0f}  |  "
                       f"Net: AUD {abs(zcs_net_payment):,.0f}")




def vol_surface_editor_tab():
    """Vol Surface Editor with mode selection - Hybrid or 3D Drag"""
    
    st.subheader(" Vol Surface Editor")
    
    # Check if new editor module is available
    if not HAS_VOL_EDITOR:
        st.warning("New vol_editor module not found. Using legacy editor.")
        _vol_surface_editor_legacy()
        return
    
    # Currency selector - use v3d_ccy from query params if coming back from Apply
    default_ccy = st.query_params.get('v3d_ccy', SUPPORTED_CURRENCIES[0])
    if default_ccy not in SUPPORTED_CURRENCIES:
        default_ccy = SUPPORTED_CURRENCIES[0]
    default_idx = SUPPORTED_CURRENCIES.index(default_ccy)
    ccy = st.selectbox("Currency", SUPPORTED_CURRENCIES, index=default_idx, key="vol_editor_ccy")
    
    # Check if we have v3d_data in query params (coming back from Apply button)
    has_v3d_data = 'v3d_data' in st.query_params and st.query_params.get('v3d_ccy') == ccy
    
    # Get current ATM surface
    atm = get_working_atm_surface(ccy)
    
    # Only block if no ATM surface AND no v3d_data to restore from
    if (atm is None or atm.empty) and not has_v3d_data:
        st.info("No ATM vol surface loaded for this currency. Go to 'Vol / SABR' tab to load data first.")
        return
    
    # If we have v3d_data but no atm, create placeholder - vol_editor will restore from query params
    if (atm is None or atm.empty) and has_v3d_data:
        atm = pd.DataFrame({"Expiry": ["1Y"], "1Y": [50.0]})
    
    # Get curve for annuity calculations
    _cc = st.session_state.get("config_curves", {}).get(ccy)
    curve = _cc if _cc is not None else get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    
    st.markdown("---")

    # Gate editor render — prevents hang on every tab visit
    if not st.session_state.get(f"_vol_editor_open_{ccy}"):
        _c1, _c2 = st.columns([2, 4])
        with _c1:
            if st.button("✏️ Open Vol Editor", key=f"open_vol_editor_{ccy}", type="primary"):
                st.session_state[f"_vol_editor_open_{ccy}"] = True
                st.rerun()
        with _c2:
            st.caption("Click to open the interactive vol surface editor.")
        return

    if st.button("✕ Close Editor", key=f"close_vol_editor_{ccy}"):
        st.session_state[f"_vol_editor_open_{ccy}"] = False
        st.rerun()

    # Render the unified editor with mode toggle (Hybrid vs 3D Drag)
    updated_surface = render_vol_surface_editor_unified(ccy, atm, curve, ois_curve)

    # Render bulk adjustment tools
    with st.expander(" Quick Adjustments (Parallel Shift, Scale, Tilt)", expanded=False):
        render_bulk_adjustment_tools(ccy)

    # ── Full SABR Recalibration (Expert Only) ────────────────────────
    st.markdown("---")
    with st.expander("🔬 Full SABR Recalibration   —   Expert Use Only", expanded=False):
        st.warning(
            "⚙️ **Expert use only.** Full recalibration fits all four SABR params (~, ~, ρ,ν, ×) "
            "simultaneously to the loaded ATM surface using a numerical optimiser. "
            "Incorrect calibration will affect all OTM swaption pricing. "
            "Only run this when the smile structure has materially changed (e.g. after a large parallel shift >50bp, "
            "curve inversion, or regime change). For daily use, run Sticky-ATM alpha recalibration in the Swaptions tab instead."
        )
        _expert_code = st.text_input("Enter expert code to unlock", type="password", key="sabr_expert_code")
        _expert_unlocked = _expert_code == "SABR2025"

        if not _expert_unlocked:
            st.caption("Contact wpo@rateedge.au for the expert recalibration code.")
        else:
            st.success("🔓 Expert mode unlocked.")
            _rc_ccy = ccy
            _rc_curve = get_ccy_curve(_rc_ccy)
            _rc_ois = get_basis_curve(_rc_ccy, "ois")
            _rc_atm = get_working_atm_surface(_rc_ccy)
            _, _rc_a, _rc_b, _rc_r, _rc_n = get_ccy_vol_data(_rc_ccy)

            if _rc_atm is None or _rc_curve is None:
                st.info("Load ATM surface and IRS curve first.")
            else:
                st.markdown("**Calibration Constraints**")
                _cc1, _cc2, _cc3, _cc4 = st.columns(4)
                with _cc1:
                    _beta_fixed = st.checkbox("Fix ~ (recommended)", value=True, key="sabr_fix_beta")
                    _beta_val = st.number_input("~ value", value=0.5, min_value=0.0, max_value=1.0, step=0.05, key="sabr_beta_fixed_val")
                with _cc2:
                    _rho_min = st.number_input("ρ,ν min", value=-0.5, min_value=-1.0, max_value=0.0, step=0.05, key="sabr_rho_min")
                    _rho_max = st.number_input("ρ,ν max", value=0.0, min_value=-1.0, max_value=1.0, step=0.05, key="sabr_rho_max")
                with _cc3:
                    _nu_min = st.number_input("× min", value=0.1, min_value=0.0, max_value=2.0, step=0.05, key="sabr_nu_min")
                    _nu_max = st.number_input("× max", value=1.0, min_value=0.0, max_value=2.0, step=0.05, key="sabr_nu_max")
                with _cc4:
                    _alpha_min = st.number_input("~ min", value=0.001, min_value=0.0, step=0.001, format="%.4f", key="sabr_alpha_min")
                    _alpha_max = st.number_input("~ max", value=0.2, min_value=0.0, step=0.01, format="%.4f", key="sabr_alpha_max")

                st.caption("Calibration fits each (expiry, tenor) cell independently using scipy.optimize.minimize with L-BFGS-B bounds.")

                if st.button("🔬 Run Full SABR Calibration", key="run_full_sabr", type="primary"):
                    from scipy.optimize import minimize
                    _expiry_col = "Expiry" if "Expiry" in _rc_atm.columns else _rc_atm.columns[0]
                    _expiries_rc = _rc_atm[_expiry_col].tolist()
                    _tenors_rc = [c for c in _rc_atm.columns if c != _expiry_col]

                    # Build new param matrices
                    _new_a = _rc_a.copy() if _rc_a is not None else _rc_atm.copy() * 0
                    _new_b = _rc_b.copy() if _rc_b is not None else _rc_atm.copy() * 0
                    _new_r = _rc_r.copy() if _rc_r is not None else _rc_atm.copy() * 0
                    _new_n = _rc_n.copy() if _rc_n is not None else _rc_atm.copy() * 0

                    _prog = st.progress(0)
                    _n_cells = len(_expiries_rc) * len(_tenors_rc)
                    _done = 0
                    _errors = 0

                    for _i, _exp_lbl in enumerate(_expiries_rc):
                        _exp_y_rc = label_to_years(str(_exp_lbl))
                        if _exp_y_rc <= 0:
                            continue
                        for _tc_lbl in _tenors_rc:
                            _ten_y_rc = label_to_years(str(_tc_lbl))
                            _atm_bp_rc = get_matrix_value(_rc_atm, str(_exp_lbl), _ten_y_rc)
                            if _atm_bp_rc is None:
                                _done += 1
                                continue
                            try:
                                _F_rc, _, _ = forward_and_annuity_from_curve(_rc_curve, _rc_ccy, _exp_y_rc, _ten_y_rc, _rc_ois)
                            except Exception:
                                _F_rc = 0.05
                            if _F_rc <= 0:
                                _F_rc = 0.05

                            _target = _atm_bp_rc / 10000.0

                            def _obj(params):
                                _a_p = params[0]
                                _b_p = _beta_val if _beta_fixed else params[1]
                                _r_p = params[1 if _beta_fixed else 2]
                                _n_p = params[2 if _beta_fixed else 3]
                                _pred = sabr_normal_atm_vol(_F_rc, _exp_y_rc, _a_p, _b_p, _r_p, _n_p)
                                return (_pred - _target) ** 2

                            _x0 = [max(_alpha_min, 0.01), _rho_min * 0.5, _nu_min + (_nu_max - _nu_min) * 0.5]
                            _bounds = [(_alpha_min, _alpha_max), (_rho_min, _rho_max), (_nu_min, _nu_max)]
                            if not _beta_fixed:
                                _x0.insert(1, 0.5)
                                _bounds.insert(1, (0.0, 1.0))

                            try:
                                _res = minimize(_obj, _x0, method="L-BFGS-B", bounds=_bounds,
                                                options={"maxiter": 200, "ftol": 1e-12})
                                _a_opt = _res.x[0]
                                _b_opt = _beta_val if _beta_fixed else _res.x[1]
                                _r_opt = _res.x[1 if _beta_fixed else 2]
                                _n_opt = _res.x[2 if _beta_fixed else 3]
                                # Write back
                                _exp_mask = _new_a[_expiry_col].astype(str).str.strip() == str(_exp_lbl).strip()
                                if _tc_lbl in _new_a.columns:
                                    _new_a.loc[_exp_mask, _tc_lbl] = _a_opt
                                if _tc_lbl in _new_b.columns:
                                    _new_b.loc[_exp_mask, _tc_lbl] = _b_opt
                                if _tc_lbl in _new_r.columns:
                                    _new_r.loc[_exp_mask, _tc_lbl] = _r_opt
                                if _tc_lbl in _new_n.columns:
                                    _new_n.loc[_exp_mask, _tc_lbl] = _n_opt
                            except Exception:
                                _errors += 1

                            _done += 1
                            _prog.progress(min(_done / max(_n_cells, 1), 1.0))

                    _old_atm_rc, _, _, _, _ = get_ccy_vol_data(_rc_ccy)
                    set_ccy_vol_data(_rc_ccy, _old_atm_rc, _new_a, _new_b, _new_r, _new_n)
                    _prog.empty()
                    # Auto-save SABR to DB
                    if HAS_POSTGRES:
                        try:
                            save_all_session_data(st.session_state.get("username", "default"))
                        except Exception:
                            pass
                    st.success(f"✅ Full SABR calibration complete. {_n_cells - _errors} cells updated, {_errors} skipped.")
                    if _errors > 0:
                        st.caption("Skipped cells had no ATM vol data or optimiser failed to converge.")

            # ── Load Calibrated ρ / ν ─────────────────────────────────
            if _expert_unlocked:
                st.markdown("---")
                st.markdown("##### 📥 Load Calibrated ρ / ν")
                _lc1, _lc2, _lc3 = st.columns([2, 2, 4])
                with _lc1:
                    if ccy in _SABR_REF and st.button("📥 Load Calibrated ρ / ν", key=f"load_sabr_cal_{ccy}", type="secondary"):
                        _n_updated = _apply_sabr_calibration(ccy)
                        if _n_updated > 0:
                            st.session_state[f"_sabr_init_{ccy}"] = True
                            st.success(f"✅ ρ / ν loaded — {_n_updated} cells. Run Recalibrate Alpha in Swaptions tab.")
                            st.rerun()
                        else:
                            st.warning("Load ATM surface first.")
                with _lc2:
                    _cal_upload = st.file_uploader("Upload calibration (.xlsx)", type=["xlsx"], key=f"sabr_cal_upload_{ccy}", label_visibility="collapsed")
                    if _cal_upload is not None:
                        try:
                            import io as _io
                            _xl = pd.read_excel(_io.BytesIO(_cal_upload.read()), sheet_name=None)
                            _rho_sheet = next((s for s in _xl if "rho" in s.lower() and ccy.lower() in s.lower()),
                                             next((s for s in _xl if "rho" in s.lower()), None))
                            _nu_sheet  = next((s for s in _xl if "nu" in s.lower() and ccy.lower() in s.lower()),
                                             next((s for s in _xl if s.lower() in ("nu","nu (ν)","nu (v)")), None))
                            if _rho_sheet and _nu_sheet:
                                _rho_df = _xl[_rho_sheet]; _nu_df = _xl[_nu_sheet]
                                _exp_col_r = _rho_df.columns[0]
                                _ten_cols_r = list(_rho_df.columns[1:])
                                _rho_arr = _rho_df[_ten_cols_r].values.astype(float)
                                _nu_arr  = _nu_df[_ten_cols_r].values.astype(float)
                                _SABR_REF[ccy] = {
                                    "expiries": [str(e) for e in _rho_df[_exp_col_r]],
                                    "tenors":   _ten_cols_r,
                                    "rho": {str(_rho_df[_exp_col_r].iloc[i]): list(_rho_arr[i]) for i in range(len(_rho_arr))},
                                    "nu":  {str(_rho_df[_exp_col_r].iloc[i]): list(_nu_arr[i])  for i in range(len(_nu_arr))},
                                }
                                st.session_state[f"_sabr_ref_updated_{ccy}"] = _cal_upload.name
                                st.success(f"✅ {_cal_upload.name} loaded — {len(_ten_cols_r)} tenors × {len(_rho_arr)} expiries. Click Load to apply.")
                            else:
                                st.error(f"Sheets not found. Available: {list(_xl.keys())}")
                        except Exception as _ue:
                            st.error(f"Upload error: {_ue}")
                with _lc3:
                    _ref_tag = st.session_state.get(f"_sabr_ref_updated_{ccy}", "31-Mar-2026 (built-in)")
                    st.caption(f"Ref: {_ref_tag} | Excel needs 'Rho' and 'Nu' sheets — expiry in col A, tenors in row 1")
    
    # Sync back to the main app's vol_data if published
    # (The vol_editor module handles publishing internally via session state)


def _vol_surface_editor_legacy():
    """Legacy vol surface editor - fallback if new module not available"""
    st.subheader("Vol surface editor (ATM  click-to-edit 3D surface)")
    if plotly_events is None:
        st.error("Install streamlit-plotly-events to use the interactive 3D editor: pip install streamlit-plotly-events")
        return

    ccy = st.selectbox("Currency", SUPPORTED_CURRENCIES, key="vol_ccy")
    atm = get_working_atm_surface(ccy)
    if atm is None or atm.empty:
        st.info("No ATM vol surface loaded for this currency yet.")
        return

    ve = st.session_state["vol_editor"]

    col_mode, col_bump = st.columns(2)
    with col_mode:
        mode = st.radio(
            "Edit mode",
            ["Local smoothing", "Global shift", "Pinned wings"],
            horizontal=False,
            key="vol_edit_mode",
        )
    with col_bump:
        bump_bp = st.slider(
            "Bump size (bp)",
            min_value=-50.0,
            max_value=50.0,
            value=5.0,
            step=1.0,
            key="vol_bump",
        )

    wdf = ve["working"].get(ccy, atm.copy())
    ve["working"][ccy] = wdf

    tenor_cols = [c for c in wdf.columns if c != "Expiry"]
    x_vals = []
    for c in tenor_cols:
        if isinstance(c, str) and c.endswith("Y") and c[:-1].isdigit():
            x_vals.append(float(c[:-1]))
        else:
            try:
                x_vals.append(float(c))
            except Exception:
                x_vals.append(0.0)

    expiry_labels = wdf["Expiry"].tolist()
    y_vals = [label_to_years(e) for e in expiry_labels]
    z_vals = wdf[tenor_cols].to_numpy().astype(float)

    fig = go.Figure()
    fig.add_surface(
        x=x_vals,
        y=y_vals,
        z=z_vals,
        colorscale="Viridis",
        opacity=0.85,
        showscale=True,
        name="surface",
    )

    xs_scatter = []
    ys_scatter = []
    zs_scatter = []
    customdata = []
    for i, exp_label in enumerate(expiry_labels):
        for j, ten_col in enumerate(tenor_cols):
            xs_scatter.append(x_vals[j])
            ys_scatter.append(y_vals[i])
            zs_scatter.append(z_vals[i, j])
            customdata.append((exp_label, ten_col))

    fig.add_trace(
        go.Scatter3d(
            x=xs_scatter,
            y=ys_scatter,
            z=zs_scatter,
            mode="markers",
            marker=dict(size=4, symbol="circle"),
            name="Nodes",
            customdata=customdata,
        )
    )

    fig.update_layout(
        title=f"{ccy} ATM vol surface (bp)  click a node to bump",
        scene=dict(
            xaxis_title="Swap Tenor (years)",
            yaxis_title="Expiry (years)",
            zaxis_title="Vol (bp)",
        ),
        height=550,
    )

    clicked_points = plotly_events(
        fig,
        select_event=True,
        override_height=550,
        override_width="100%",
        key="vol_surface_events",
    )

    if clicked_points:
        pt = clicked_points[0]
        exp_label = None
        ten_col = None
        if "customdata" in pt and pt["customdata"]:
            exp_label, ten_col = pt["customdata"]
        else:
            x_click = pt.get("x", None)
            y_click = pt.get("y", None)
            if x_click is not None and y_click is not None:
                y_arr = np.array(y_vals)
                x_arr = np.array(x_vals)
                exp_idx = int(np.argmin(np.abs(y_arr - y_click)))
                ten_idx = int(np.argmin(np.abs(x_arr - x_click)))
                exp_label = expiry_labels[exp_idx]
                ten_col = tenor_cols[ten_idx]

        if exp_label is not None and ten_col is not None:
            push_vol_history(ccy)
            w = ve["working"].get(ccy, atm.copy())
            mask = w["Expiry"] == exp_label
            if not mask.any():
                st.error("Clicked expiry not found in surface.")
            elif ten_col not in w.columns:
                st.error("Clicked tenor column not found in surface.")
            else:
                idx = w[mask].index[0]
                delta = bump_bp

                if mode == "Local smoothing":
                    w.at[idx, ten_col] = w.at[idx, ten_col] + delta
                    if idx - 1 in w.index:
                        w.at[idx - 1, ten_col] = w.at[idx - 1, ten_col] + 0.5 * delta
                    if idx + 1 in w.index:
                        w.at[idx + 1, ten_col] = w.at[idx + 1, ten_col] + 0.5 * delta
                elif mode == "Global shift":
                    w[ten_col] = w[ten_col] + delta
                elif mode == "Pinned wings":
                    if len(w) > 2:
                        inner_idx = w.index[1:-1]
                        w.loc[inner_idx, ten_col] = w.loc[inner_idx, ten_col] + delta

                ve["working"][ccy] = w
                st.success(f"Bumped node ({exp_label}, {ten_col}) by {delta:+.0f} bp.")

    col_u, col_r, col_p = st.columns(3)
    with col_u:
        if st.button("Undo last", key="vol_undo"):
            undo_vol(ccy)
    with col_r:
        if st.button("Redo", key="vol_redo"):
            redo_vol(ccy)
    with col_p:
        if st.button("Publish to pricing", key="vol_publish"):
            publish_vol(ccy)
            st.success("Published  swaptions & caps/floors now use edited surface.")

    st.markdown("#### Current working ATM surface (bp)")
    st.dataframe(ve["working"][ccy], use_container_width=True)


def multi_ccy_tab(vol_mode: str):
    st.subheader("Multi-currency comparison")

    col1, col2 = st.columns(2)
    with col1:
        ccy1 = st.selectbox("Currency 1", SUPPORTED_CURRENCIES, index=0, key="mc_ccy1")
    with col2:
        ccy2 = st.selectbox("Currency 2", SUPPORTED_CURRENCIES, index=1, key="mc_ccy2")

    expiry = st.selectbox(
        "Expiry",
        ["1w","1m","2m","3m","6m","9m","1y","18m","2y","3y","4y","5y","6y","7y","8y","9y","10y","12y","15y","20y","25y","30y"],
        index=5,
        key="mc_exp",
    )
    tenor_str = st.selectbox(
        "Swap tenor",
        ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","15Y","20Y"],
        index=2,
        key="mc_tenor",
    )
    tenor_y = float(tenor_str[:-1])
    expiry_y = label_to_years(expiry)

    side = st.radio("Side", ["Payer", "Receiver"], horizontal=True, key="mc_side")
    notional_mm = st.number_input("Notional (mm, per ccy)", 0.0, 1e6, 100.0, step=10.0, key="mc_not")
    disc = st.number_input("Flat discount rate (%)", 0.0, 20.0, 4.0, step=0.1, key="mc_disc") / 100.0

    def price_leg(ccy: str) -> Optional[dict]:
        _cc = st.session_state.get("config_curves", {}).get(ccy)
        curve = _cc if _cc is not None else get_ccy_curve(ccy)
        _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
        ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
        atm = get_working_atm_surface(ccy)
        _, a, b, r, n = get_ccy_vol_data(ccy)
        if curve is None or atm is None:
            return None
        fwd, ann, _ = forward_and_annuity_from_curve(curve, ccy, expiry_y, tenor_y, ois_curve)
        expiry_label = expiry
        atm_val = get_matrix_value(atm, expiry_label, tenor_y)
        if atm_val is None:
            atm_val = 35.0
        if vol_mode.startswith("Normal"):
            vol = atm_val / 10000.0
            model = "Normal"
        else:
            sabr = get_sabr_params_from_matrices(a, b, r, n, expiry_label, tenor_y)
            if sabr:
                vol = sabr_implied_vol_black(
                    fwd,
                    fwd,
                    expiry_y,
                    sabr["alpha"],
                    sabr["beta"],
                    sabr["rho"],
                    sabr["nu"],
                )
            else:
                vol = atm_val / 100.0
            model = "Black"
        ticket = SwaptionTicket(
            side=side,
            payoff_type="vanilla",
            notional=notional_mm * 1e6,
            currency=ccy,
            expiry_years=expiry_y,
            swap_tenor_years=tenor_y,
            forward=fwd,
            strike=fwd,
            vol=vol,
            discount_rate=disc,
            annuity=ann,
            model=model,
            label=f"{side} {expiry}x{tenor_str}",
            use_curve=True,
        )
        return price_swaption(ticket)

    if st.button("Compare", key="mc_compare"):
        res1 = price_leg(ccy1)
        res2 = price_leg(ccy2)
        if res1 is None or res2 is None:
            st.warning("Need curves and ATM vols loaded for both currencies.")
        else:
            st.markdown(
                f"**{ccy1} PV:** {res1['pv']:,.0f} ({res1['pv_bp']:,.1f} bp)  \n"
                f"**{ccy2} PV:** {res2['pv']:,.0f} ({res2['pv_bp']:,.1f} bp)  \n"
                f"**Spread ({ccy1}-{ccy2}) PV (bp):** {res1['pv_bp'] - res2['pv_bp']:.1f}"
            )


def credit_xva_tab():
    st.subheader("Credit / CVA overlay")

    if not st.session_state.get("portfolio"):
        st.info("Price some swaptions or caps/floors first  portfolio is empty.")
        return

    dfp = pd.DataFrame(st.session_state["portfolio"])
    # Use columns that actually exist in the portfolio
    display_cols = [c for c in ["instrument_type", "currency", "structure", "expiry", "tenor", "pv", "pv_bp", "vega"] if c in dfp.columns]
    st.dataframe(dfp[display_cols] if display_cols else dfp)

    disc = st.number_input("Flat discount rate for CVA (%)", 0.0, 20.0, 4.0, step=0.25, key="cva_disc") / 100.0
    hazard = st.number_input("Flat hazard rate (% per year)", 0.0, 20.0, 1.0, step=0.1, key="cva_hz") / 100.0
    lgd = st.number_input("LGD (%)", 0.0, 100.0, 60.0, step=5.0, key="cva_lgd") / 100.0
    horizon = st.number_input("CVA horizon (years)", 0.5, 30.0, 10.0, step=0.5, key="cva_hor")

    if st.button("Compute crude CVA on portfolio", key="cva_btn"):
        pv0 = dfp["pv"].sum()
        vega_total = dfp["vega"].sum()
        vol_ref = 0.01
        ts, ee = simple_exposure_profile(pv0, vega_total, vol_ref, horizon_years=horizon)
        cva = cva_from_hazard(ts, ee, hazard_rate=hazard, lgd=lgd, discount_rate=disc)
        st.markdown(
            f"**Clean PV:** {pv0:,.0f}  \n"
            f"**CVA:** {cva:,.0f}  \n"
            f"**CVA-adjusted PV:** {pv0 + cva:,.0f}"
        )
        chart_df = pd.DataFrame({"t": ts, "EE": ee})
        st.line_chart(chart_df.set_index("t"))



# ── Historical Vol Analysis helpers ──────────────────────────────────────────

_EXPIRY_YEARS_MAP = {
    "1w": 1/52, "1m": 1/12, "2m": 2/12, "3m": 3/12, "6m": 0.5, "9m": 0.75,
    "1y": 1.0, "18m": 1.5, "2y": 2.0, "3y": 3.0, "4y": 4.0, "5y": 5.0,
    "6y": 6.0, "7y": 7.0, "8y": 8.0, "9y": 9.0, "10y": 10.0, "12y": 12.0,
    "15y": 15.0, "20y": 20.0, "25y": 25.0, "30y": 30.0,
}

@st.cache_data(ttl=120, show_spinner=False)
def _load_vol_snapshots_for_viz(ccy: str, start_date: str, end_date: str) -> list:
    """Load vol snapshots from vol_history within date range. Returns list of dicts."""
    if not HAS_POSTGRES:
        return []
    try:
        conn = get_db_connection()
        if conn is None:
            return []
        cur = conn.cursor()
        cur.execute(
            """SELECT id, snapshot_date, label, atm_vols
               FROM vol_history
               WHERE currency = %s AND atm_vols IS NOT NULL
                 AND snapshot_date::date BETWEEN %s AND %s
               ORDER BY snapshot_date ASC
               LIMIT 60""",
            (ccy, start_date, end_date)
        )
        rows = cur.fetchall()
        conn.close()
        results = []
        for row_id, snap_date, label, atm_vols in rows:
            if not atm_vols or "values" not in atm_vols:
                continue
            try:
                df = pd.DataFrame(atm_vols["values"])
                if "Expiry" in df.columns:
                    df = df.set_index("Expiry")
                # Normalise index to lowercase
                df.index = df.index.str.lower().str.strip()
                df = df.apply(pd.to_numeric, errors="coerce")
                results.append({
                    "id": row_id,
                    "date": pd.to_datetime(snap_date),
                    "label": label or str(snap_date)[:10],
                    "df": df,
                })
            except Exception:
                pass
        return results
    except Exception:
        return []

@st.cache_data(ttl=120, show_spinner=False)
def _load_fwd_rates_for_viz(ccy: str, start_date: str, end_date: str,
                             floating_rate: str = "6M BBSW") -> pd.DataFrame:
    """Load par swap rates from swap_rates table, return wide DataFrame: date×tenor."""
    if not HAS_POSTGRES:
        return pd.DataFrame()
    try:
        conn = get_db_connection()
        if conn is None:
            return pd.DataFrame()
        cur = conn.cursor()
        cur.execute(
            """SELECT date, tenor, rate FROM swap_rates
               WHERE currency = %s AND floating_rate = %s
                 AND date BETWEEN %s AND %s
               ORDER BY date ASC, tenor ASC""",
            (ccy, floating_rate, start_date, end_date)
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=["date", "tenor", "rate"])
        df["date"] = pd.to_datetime(df["date"])
        pivot = df.pivot_table(index="date", columns="tenor", values="rate", aggfunc="mean")
        # Sort columns by tenor years
        def _tenor_years(t):
            try:
                return float(t.replace("Y","").replace("y",""))
            except Exception:
                return 999
        pivot = pivot[sorted(pivot.columns, key=_tenor_years)]
        return pivot
    except Exception:
        return pd.DataFrame()

def _build_vol_surface_arrays(snap: dict):
    """Convert snapshot dict to (X tenors, Y expiry_years, Z vol matrix) arrays."""
    df = snap["df"]
    tenor_cols = [c for c in df.columns]
    tenor_x = []
    for t in tenor_cols:
        try:
            tenor_x.append(float(str(t).replace("Y","").replace("y","")))
        except Exception:
            tenor_x.append(0)
    expiry_y = [_EXPIRY_YEARS_MAP.get(str(e).lower().strip(), 0) for e in df.index]
    # Filter rows with valid expiry mapping
    valid = [i for i, v in enumerate(expiry_y) if v > 0]
    expiry_y = [expiry_y[i] for i in valid]
    expiry_labels = [list(df.index)[i] for i in valid]
    z = df.values[valid, :]  # shape: (n_expiry, n_tenor)
    return tenor_x, expiry_y, expiry_labels, tenor_cols, z

def _make_vol_surface_fig(snapshots: list, title: str = "ATM Vol Surface (bp)",
                          track_exp: str = None, track_ten: str = None):
    """Build animated Plotly 3D surface figure from list of snapshots."""
    import plotly.graph_objects as go

    if not snapshots:
        return None

    # Build frames
    frames = []
    dates = []
    for snap in snapshots:
        tenor_x, expiry_y, exp_labels, tenor_labels, z = _build_vol_surface_arrays(snap)
        if z.size == 0:
            continue
        lbl = snap["label"] if snap["label"] else snap["date"].strftime("%Y-%m-%d")
        dates.append(lbl)
        # Build hover text: "3m5y = 92.3bp" (tenor=X, expiry=Y, z shape=n_expiry×n_tenor)
        _hover = []
        for ei, el in enumerate(exp_labels):
            _row = []
            for ti, tl in enumerate(tenor_labels):
                _row.append(f"{el}{tl.lower()}<br>{z[ei,ti]:.1f}bp")
            _hover.append(_row)
        # Build tracking point if requested
        _track_traces = []
        if track_exp and track_ten:
            try:
                from scipy.interpolate import interp1d as _i1d
                _te_y = label_to_years(track_exp)
                _tt_y = label_to_years(track_ten)
                # Find vol at tracking point via interpolation
                _exp_arr = expiry_y  # already in years
                _ten_arr = tenor_x
                _te_idx = np.argmin(np.abs(np.array(_exp_arr) - _te_y))
                _tt_idx = np.argmin(np.abs(np.array(_ten_arr) - _tt_y))
                _track_z = float(z[_te_idx, _tt_idx])
                _track_traces = [go.Scatter3d(
                    x=[_tt_y], y=[_te_y], z=[_track_z + 2],
                    mode="markers+text",
                    marker=dict(size=8, color="#FFD700", symbol="diamond",
                                line=dict(color="#ffffff", width=1)),
                    text=[f"{track_exp}×{track_ten}<br>{_track_z:.1f}bp"],
                    textposition="top center",
                    textfont=dict(color="#FFD700", size=11, family="Arial Bold"),
                    showlegend=False,
                    hovertemplate=f"<b>{track_exp}×{track_ten}</b><br>{_track_z:.1f} bp<extra></extra>",
                )]
            except Exception:
                _track_traces = []

        frames.append(go.Frame(
            data=[go.Surface(
                x=tenor_x, y=expiry_y, z=z.tolist(),
                colorscale="RdYlGn_r",
                cmin=50, cmax=130,
                showscale=True,
                colorbar=dict(title="bp", thickness=12, len=0.6),
                hovertext=_hover,
                hovertemplate="<b>%{hovertext}</b><extra></extra>",
                hoverlabel=dict(bgcolor="#1e3a5f", font=dict(color="#f1f5f9", size=13, family="Arial")),
            )] + _track_traces,
            name=lbl,
        ))

    if not frames:
        return None

    # Format date label e.g. "2026-02-03" → "03-Feb-26"
    def _fmt_snap_date(s):
        try:
            from datetime import datetime as _dt2
            return _dt2.strptime(s, "%Y-%m-%d").strftime("%d-%b-%y")
        except Exception:
            return s

    # Initial surface = first frame
    first = frames[0].data[0]
    first_extras = list(frames[0].data[1:]) if len(frames[0].data) > 1 else []
    fig = go.Figure(
        data=[first] + first_extras,
        frames=frames,
        layout=go.Layout(
            title=dict(text=title, font=dict(color="#f1f5f9", size=14)),
            height=720,
            paper_bgcolor="rgba(15,23,42,0.95)",
            plot_bgcolor="rgba(15,23,42,0.95)",
            font=dict(color="#94a3b8", family="Arial", size=11),
            scene=dict(
                bgcolor="rgba(15,23,42,0.95)",
                xaxis=dict(
                    title=dict(text="Swap Tenor (Y)", font=dict(color="#c8d8e8", size=12)),
                    tickmode="array", tickvals=[1,2,3,5,7,10,15,20,30],
                    ticktext=["1","2","3","5","7","10","15","20","30"],
                    tickfont=dict(color="#c8d8e8", size=10),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                    autorange="reversed",
                ),
                yaxis=dict(
                    title=dict(text="Option Expiry", font=dict(color="#c8d8e8", size=12)),
                    tickmode="array",
                    tickvals=[0.08,0.25,0.5,1,2,3,5,7,10,15,20,25,30],
                    ticktext=["1m","3m","6m","1y","2y","3y","5y","7y","10y","15y","20y","25y","30y"],
                    tickfont=dict(color="#c8d8e8", size=10),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                    autorange="reversed",
                ),
                zaxis=dict(
                    title=dict(text="Vol (bp)", font=dict(color="#c8d8e8", size=12)),
                    tickfont=dict(color="#c8d8e8", size=10),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                ),
                camera=dict(eye=dict(x=1.8, y=1.6, z=0.8),
                            up=dict(x=0, y=0, z=1)),
                aspectmode="manual",
                aspectratio=dict(x=1.4, y=1.0, z=0.7),
            ),
            margin=dict(l=0, r=0, b=80, t=50),
            updatemenus=[dict(
                type="buttons", showactive=False, y=0.02, x=0.12,
                xanchor="right", yanchor="bottom",
                buttons=[
                    dict(label="▶ Play", method="animate",
                         args=[None, dict(frame=dict(duration=600, redraw=True),
                                          fromcurrent=True, mode="immediate")]),
                    dict(label="⏸ Pause", method="animate",
                         args=[[None], dict(frame=dict(duration=0, redraw=False),
                                             mode="immediate")]),
                ]
            )],
            sliders=[dict(
                currentvalue=dict(
                    prefix="",
                    visible=True,
                    xanchor="center",
                    font=dict(color="#f1f5f9", size=14, family="Arial"),
                ),
                pad=dict(t=20, b=10),
                steps=[dict(
                    method="animate",
                    args=[[d], dict(mode="immediate", frame=dict(duration=0, redraw=True))],
                    # Show every 5th label to avoid crowding; blank the rest
                    label=(_fmt_snap_date(d.split()[-1] if " " in d else d)
                           if _i % 3 == 0 else ""),
                ) for _i, d in enumerate(dates)],
                len=0.90, x=0.05, y=0.0,
                font=dict(color="#94a3b8", size=10, family="Arial"),
                tickwidth=2,
                tickcolor="#334155",
                bgcolor="#0f172a",
                bordercolor="#334155",
                borderwidth=1,
            )],
        )
    )
    return fig

def _make_fwd_matrix_surface_fig(pivot: pd.DataFrame, date_range_dates: list,
                                  title: str = "AUD Par Swap Rates (%)"):
    """Build animated Plotly 3D surface from par rate pivot (date × tenor)."""
    import plotly.graph_objects as go

    if pivot.empty:
        return None

    def _tenor_years(t):
        try:
            return float(str(t).replace("Y","").replace("y",""))
        except Exception:
            return 0

    tenor_x = [_tenor_years(c) for c in pivot.columns]
    dates = [d.strftime("%Y-%m-%d") for d in pivot.index]

    frames = []
    for i, (dt, row) in enumerate(pivot.iterrows()):
        z_row = row.values.tolist()
        frames.append(go.Frame(
            data=[go.Surface(
                x=tenor_x,
                y=[0],  # single date slice = flat surface showing curve shape
                z=[z_row],
                colorscale="Blues",
                cmin=float(pivot.min().min()) * 0.98,
                cmax=float(pivot.max().max()) * 1.02,
                showscale=True,
                colorbar=dict(title="%", thickness=12, len=0.6),
            )],
            name=dates[i],
        ))

    # For a proper 2D surface over time, use date index as Y axis
    date_nums = list(range(len(pivot)))
    z_full = pivot.values.tolist()

    fig = go.Figure(
        data=[go.Surface(
            x=tenor_x,
            y=date_nums,
            z=z_full,
            colorscale="Blues",
            showscale=True,
            colorbar=dict(title="%", thickness=12, len=0.6),
        )],
        layout=go.Layout(
            title=dict(text=title, font=dict(color="#f1f5f9", size=14)),
            height=720,
            paper_bgcolor="rgba(15,23,42,0.95)",
            plot_bgcolor="rgba(15,23,42,0.95)",
            font=dict(color="#94a3b8", family="Arial"),
            scene=dict(
                bgcolor="rgba(15,23,42,0.95)",
                xaxis=dict(
                    title=dict(text="Tenor (Y)", font=dict(color="#c8d8e8", size=12)),
                    tickfont=dict(color="#c8d8e8", size=10),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                ),
                yaxis=dict(
                    title=dict(text="Date", font=dict(color="#94a3b8", size=11)),
                    tickmode="array",
                    tickvals=list(range(0, len(dates), max(1, len(dates)//8))),
                    ticktext=[dates[i] for i in range(0, len(dates), max(1, len(dates)//8))],
                    tickfont=dict(color="#94a3b8", size=9),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                ),
                zaxis=dict(
                    title=dict(text="Rate (%)", font=dict(color="#94a3b8", size=11)),
                    tickfont=dict(color="#94a3b8", size=9),
                    gridcolor="#334155", showbackground=True,
                    backgroundcolor="rgba(15,23,42,0.5)",
                ),
                camera=dict(eye=dict(x=-1.6, y=-1.8, z=0.8),
                            up=dict(x=0, y=0, z=1)),
                aspectmode="manual",
                aspectratio=dict(x=1.4, y=1.0, z=0.7),
            ),
            margin=dict(l=0, r=0, b=40, t=50),
        )
    )
    return fig

def _make_overlay_fig(snap_a: dict, snap_b: dict):
    """Three-panel: surface A, surface B, delta B-A."""
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    def _to_arrays(snap):
        tx, ey, _, _, z = _build_vol_surface_arrays(snap)
        return tx, ey, z

    tx_a, ey_a, z_a = _to_arrays(snap_a)
    tx_b, ey_b, z_b = _to_arrays(snap_b)

    # Align shapes
    if z_a.shape != z_b.shape:
        return None

    import numpy as np
    z_delta = np.array(z_b) - np.array(z_a)

    lbl_a = snap_a["date"].strftime("%Y-%m-%d")
    lbl_b = snap_b["date"].strftime("%Y-%m-%d")

    fig = make_subplots(
        rows=1, cols=3,
        specs=[[{"type": "surface"}, {"type": "surface"}, {"type": "surface"}]],
        subplot_titles=[f"Date A: {lbl_a}", f"Date B: {lbl_b}", f"Δ  B − A"],
        horizontal_spacing=0.02,
    )

    cam = dict(eye=dict(x=1.4, y=-1.6, z=0.85))
    shared_scene = dict(
        xaxis=dict(title="Tenor"),
        yaxis=dict(title="Expiry"),
        zaxis=dict(title="Vol (bp)"),
        camera=cam,
    )

    vmin = min(z_a.min(), z_b.min())
    vmax = max(z_a.max(), z_b.max())

    fig.add_trace(go.Surface(x=tx_a, y=ey_a, z=z_a.tolist(),
                             colorscale="RdYlGn_r", cmin=vmin, cmax=vmax,
                             showscale=False), row=1, col=1)
    fig.add_trace(go.Surface(x=tx_b, y=ey_b, z=z_b.tolist(),
                             colorscale="RdYlGn_r", cmin=vmin, cmax=vmax,
                             showscale=False), row=1, col=2)
    fig.add_trace(go.Surface(x=tx_a, y=ey_a, z=z_delta.tolist(),
                             colorscale="RdBu_r", cmid=0,
                             showscale=True,
                             colorbar=dict(title="Δbp", thickness=10, len=0.6)), row=1, col=3)

    fig.update_layout(
        height=480,
        margin=dict(l=0, r=0, b=20, t=50),
        scene=shared_scene,
        scene2=shared_scene,
        scene3=dict(
            xaxis=dict(title="Tenor"),
            yaxis=dict(title="Expiry"),
            zaxis=dict(title="Δ bp"),
            camera=cam,
        ),
    )
    return fig


def backtesting_tab():
    st.subheader("📊 Historical VOL Analysis")

    ccy = st.session_state.get("selected_ccy", "AUD")

    # ── Section 1: Vol Surface History ───────────────────────────────────────
    st.markdown("### 🌊 Vol Surface History")

    if not HAS_POSTGRES:
        st.warning("Database not connected — vol history unavailable.")
    else:
        c1, c2, c3 = st.columns([2, 2, 2])
        with c1:
            _vs_start = st.date_input("From", value=pd.Timestamp.now() - pd.Timedelta(days=90),
                                       key="hviz_vol_start")
        with c2:
            _vs_end = st.date_input("To", value=pd.Timestamp.now(), key="hviz_vol_end")
        with c3:
            _vs_mode = st.selectbox("View", ["Animated Timeline", "Single Date", "Overlay A vs B"],
                                     key="hviz_vol_mode")

        if st.button("🔄 Load Vol Snapshots", key="hviz_load_vol"):
            _load_vol_snapshots_for_viz.clear()
            st.session_state["hviz_snaps_loaded"] = True

        snaps = _load_vol_snapshots_for_viz(ccy, str(_vs_start), str(_vs_end))

        if not snaps:
            st.info("No vol snapshots in this date range. Save EOD snapshots from the Vol Export tab.")

            # ── 5-day AUD seed (mirrors USD SOD seed) ────────────────────────
            with st.expander("🌱 Seed AUD Vol History (backfill from current surface)", expanded=True):
                st.caption("Creates snapshots for the last N business days using the current "
                           "AUD vol surface with small random daily moves. Useful for seeding "
                           "the DB before live EOD saving is established.")
                _sc1, _sc2 = st.columns(2)
                with _sc1:
                    _seed_n = st.number_input("Days to backfill", 2, 30, 5, key="hviz_seed_days")
                with _sc2:
                    _seed_sigma = st.number_input("Daily move σ (bp)", 0.5, 10.0, 1.5,
                                                   step=0.5, key="hviz_seed_sigma")
                if st.button("🌱 Seed AUD Snapshots", key="hviz_seed_btn", type="primary"):
                    _aud_atm = st.session_state.get("vol_data", {}).get("AUD", {}).get("atm")
                    if _aud_atm is None:
                        st.error("No AUD ATM vol surface loaded. Upload config in Vol/SABR tab first.")
                    else:
                        import random as _rnd
                        from datetime import datetime as _dt2, timedelta as _td2
                        _atm_base = _aud_atm.copy()
                        if "Expiry" in _atm_base.columns:
                            _atm_base = _atm_base.set_index("Expiry")
                        _atm_base = _atm_base.apply(pd.to_numeric, errors="coerce")
                        _exp_rows = list(_atm_base.index)
                        _t_cols = list(_atm_base.columns)
                        _conn = get_db_connection()
                        if not _conn:
                            st.error("Cannot connect to database.")
                        else:
                            _cur = _conn.cursor()
                            _rnd.seed(None)
                            _running = _atm_base.copy().astype(float)
                            _day = _dt2.now().replace(hour=17, minute=0, second=0, microsecond=0)
                            _seeded = 0
                            for _d in range(int(_seed_n)):
                                _day -= _td2(days=1)
                                while _day.weekday() >= 5:
                                    _day -= _td2(days=1)
                                for _e in _exp_rows:
                                    for _t in _t_cols:
                                        try:
                                            _mv = _rnd.gauss(0, _seed_sigma)
                                            _running.loc[_e, _t] = max(
                                                float(_running.loc[_e, _t]) + _mv, 1.0)
                                        except Exception:
                                            pass
                                _recs = (_running.reset_index()
                                         .rename(columns={"index": "Expiry"})
                                         .to_dict(orient="records"))
                                _lbl = f"AUD EOD {_day.strftime('%Y-%m-%d')} [SEEDED]"
                                _uid = st.session_state.get("username", "default")
                                try:
                                    _cur.execute(
                                        """INSERT INTO vol_history
                                           (user_id, currency, snapshot_date, label, atm_vols, notes)
                                           VALUES (%s,%s,%s,%s,%s,%s)
                                           ON CONFLICT DO NOTHING""",
                                        (_uid, "AUD", _day, _lbl,
                                         Json({"values": _recs}),
                                         "Seeded backfill for historical viz"))
                                    _seeded += 1
                                except Exception:
                                    pass
                            _conn.commit()
                            _cur.close()
                            _conn.close()
                            st.success(f"✅ Seeded {_seeded} AUD snapshots. Reload to view.")
                            _load_vol_snapshots_for_viz.clear()
                            st.rerun()
        else:
            st.caption(f"Found **{len(snaps)}** snapshots  ·  "
                       f"{snaps[0]['date'].strftime('%Y-%m-%d')} → {snaps[-1]['date'].strftime('%Y-%m-%d')}")

            if _vs_mode == "Animated Timeline":
                # Click-to-track: user clicks on surface, marker jumps to that point
                _track_exp = st.session_state.get("hviz_track_exp", "2y")
                _track_ten = st.session_state.get("hviz_track_ten", "5Y")
                st.caption(f"📍 Tracking: **{_track_exp}×{_track_ten}** — click any point on the surface to move the tracker")
                _fig = _make_vol_surface_fig(snaps, f"{ccy} ATM Vol Surface — bp (animated)",
                                              track_exp=_track_exp, track_ten=_track_ten)
                if _fig:
                    _sel = st.plotly_chart(_fig, use_container_width=True,
                                           on_select="rerun", key="hviz_vol_chart")
                    # Read click data and update tracker position
                    if _sel and hasattr(_sel, "selection") and _sel.selection:
                        _pts = _sel.selection.get("points", [])
                        if _pts:
                            _pt = _pts[0]
                            _clicked_x = _pt.get("x")  # tenor years
                            _clicked_y = _pt.get("y")  # expiry years
                            if _clicked_x is not None and _clicked_y is not None:
                                # Snap to nearest grid labels
                                _exp_opts = ["1w","2w","1m","3m","6m","9m","1y","18m","2y","3y","4y","5y","6y","7y","8y","9y","10y","12y","15y","20y","25y","30y"]
                                _ten_opts = ["1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"]
                                import numpy as _np2
                                _exp_yrs = _np2.array([label_to_years(e) for e in _exp_opts])
                                _ten_yrs = _np2.array([label_to_years(t) for t in _ten_opts])
                                _best_exp = _exp_opts[int(_np2.argmin(_np2.abs(_exp_yrs - float(_clicked_y))))]
                                _best_ten = _ten_opts[int(_np2.argmin(_np2.abs(_ten_yrs - float(_clicked_x))))]
                                st.session_state["hviz_track_exp"] = _best_exp
                                st.session_state["hviz_track_ten"] = _best_ten
                                st.rerun()
                else:
                    st.warning("Could not build surface — check snapshot data format.")

            elif _vs_mode == "Single Date":
                _snap_labels = [f"{s['date'].strftime('%Y-%m-%d')}  {s['label']}" for s in snaps]
                _sel_idx = st.selectbox("Select snapshot", range(len(_snap_labels)),
                                         format_func=lambda i: _snap_labels[i],
                                         key="hviz_single_sel")
                _fig = _make_vol_surface_fig([snaps[_sel_idx]],
                                              f"{ccy} ATM Vol — {snaps[_sel_idx]['date'].strftime('%Y-%m-%d')}")
                if _fig:
                    st.plotly_chart(_fig, use_container_width=True)

            elif _vs_mode == "Overlay A vs B":
                _snap_labels = [f"{s['date'].strftime('%Y-%m-%d')}  {s['label']}" for s in snaps]
                _oa, _ob = st.columns(2)
                with _oa:
                    _idx_a = st.selectbox("Date A", range(len(_snap_labels)),
                                           format_func=lambda i: _snap_labels[i],
                                           index=0, key="hviz_ov_a")
                with _ob:
                    _idx_b = st.selectbox("Date B", range(len(_snap_labels)),
                                           format_func=lambda i: _snap_labels[i],
                                           index=min(len(snaps)-1, len(snaps)-1),
                                           key="hviz_ov_b")
                if _idx_a != _idx_b:
                    _fig = _make_overlay_fig(snaps[_idx_a], snaps[_idx_b])
                    if _fig:
                        st.plotly_chart(_fig, use_container_width=True)
                    else:
                        st.warning("Snapshot grids don't match — select two snapshots with identical expiry/tenor structure.")
                else:
                    st.info("Select two different dates to compare.")

    # ── Section 2: Forward Swap Rate Matrix History ───────────────────────────
    st.markdown("---")
    st.markdown("### 📈 Par Swap Rate History")

    if not HAS_POSTGRES:
        st.warning("Database not connected.")
    else:
        _fr1, _fr2, _fr3 = st.columns([2, 2, 2])
        with _fr1:
            _fr_start = st.date_input("From", value=pd.Timestamp.now() - pd.Timedelta(days=90),
                                       key="hviz_fwd_start")
        with _fr2:
            _fr_end = st.date_input("To", value=pd.Timestamp.now(), key="hviz_fwd_end")
        with _fr3:
            _fr_type = st.selectbox("Floating Rate", ["6M BBSW", "3M BBSW", "OIS"],
                                     key="hviz_fwd_type")

        if st.button("🔄 Load Rate History", key="hviz_load_fwd"):
            _load_fwd_rates_for_viz.clear()

        _pivot = _load_fwd_rates_for_viz(ccy, str(_fr_start), str(_fr_end), _fr_type)

        if _pivot.empty:
            st.info(f"No {_fr_type} data in this date range. Load data into swap_rates table first.")
        else:
            st.caption(f"Loaded **{len(_pivot)}** daily curves  ·  "
                       f"Tenors: {', '.join(list(_pivot.columns)[:6])}{'...' if len(_pivot.columns) > 6 else ''}")

            _fwd_mode = st.radio("View", ["3D Surface (time × tenor)", "Single Date Curve",
                                           "Overlay A vs B"],
                                  horizontal=True, key="hviz_fwd_mode")

            if _fwd_mode == "3D Surface (time × tenor)":
                _fig2 = _make_fwd_matrix_surface_fig(
                    _pivot, list(_pivot.index),
                    f"{ccy} {_fr_type} Par Rates — {str(_fr_start)} to {str(_fr_end)}")
                if _fig2:
                    st.plotly_chart(_fig2, use_container_width=True)

            elif _fwd_mode == "Single Date Curve":
                import plotly.graph_objects as go
                _avail_dates = [d.strftime("%Y-%m-%d") for d in _pivot.index]
                _sel_date = st.selectbox("Date", _avail_dates,
                                          index=len(_avail_dates)-1, key="hviz_fwd_single")
                _row = _pivot.loc[_pivot.index.strftime("%Y-%m-%d") == _sel_date]
                if not _row.empty:
                    _tx = [float(str(c).replace("Y","")) for c in _row.columns]
                    _ry = _row.values[0].tolist()
                    _fig3 = go.Figure(go.Scatter(x=_tx, y=_ry, mode="lines+markers",
                                                  line=dict(color="#00B4C8", width=2),
                                                  marker=dict(size=6)))
                    _fig3.update_layout(
                        title=dict(text=f"{ccy} {_fr_type}  {_sel_date}", font=dict(color="#f1f5f9")),
                        xaxis_title="Tenor (Y)", yaxis_title="Rate (%)",
                        height=350, template="plotly_dark",
                        paper_bgcolor="rgba(15,23,42,0.95)",
                        plot_bgcolor="rgba(15,23,42,0.8)",
                        font=dict(color="#94a3b8"),
                        xaxis=dict(gridcolor="#334155"),
                        yaxis=dict(gridcolor="#334155"),
                        margin=dict(l=40, r=20, t=50, b=40))
                    st.plotly_chart(_fig3, use_container_width=True)

            elif _fwd_mode == "Overlay A vs B":
                import plotly.graph_objects as go
                _avail_dates = [d.strftime("%Y-%m-%d") for d in _pivot.index]
                _fo1, _fo2 = st.columns(2)
                with _fo1:
                    _fd_a = st.selectbox("Date A", _avail_dates, index=0, key="hviz_fo_a")
                with _fo2:
                    _fd_b = st.selectbox("Date B", _avail_dates,
                                          index=len(_avail_dates)-1, key="hviz_fo_b")
                _row_a = _pivot.loc[_pivot.index.strftime("%Y-%m-%d") == _fd_a]
                _row_b = _pivot.loc[_pivot.index.strftime("%Y-%m-%d") == _fd_b]
                if not _row_a.empty and not _row_b.empty:
                    import numpy as np
                    _tx = [float(str(c).replace("Y","")) for c in _pivot.columns]
                    _ra = _row_a.values[0]
                    _rb = _row_b.values[0]
                    _delta = _rb - _ra
                    _fig4 = go.Figure()
                    _fig4.add_trace(go.Scatter(x=_tx, y=_ra.tolist(), name=f"A: {_fd_a}",
                                               mode="lines+markers", line=dict(color="#00B4C8", width=2)))
                    _fig4.add_trace(go.Scatter(x=_tx, y=_rb.tolist(), name=f"B: {_fd_b}",
                                               mode="lines+markers", line=dict(color="#F0A500", width=2)))
                    _fig4.add_trace(go.Bar(x=_tx, y=(_delta * 100).tolist(),
                                           name="Δ B−A (bp)", yaxis="y2",
                                           marker_color=["#18A96A" if v >= 0 else "#DC3545"
                                                         for v in _delta],
                                           opacity=0.5))
                    _fig4.update_layout(
                        title=dict(text=f"{ccy} {_fr_type}  Overlay: {_fd_a} vs {_fd_b}", font=dict(color="#f1f5f9")),
                        xaxis_title="Tenor (Y)", yaxis_title="Rate (%)",
                        template="plotly_dark",
                        paper_bgcolor="rgba(15,23,42,0.95)",
                        plot_bgcolor="rgba(15,23,42,0.8)",
                        font=dict(color="#94a3b8"),
                        xaxis=dict(gridcolor="#334155"),
                        yaxis=dict(gridcolor="#334155"),
                        yaxis2=dict(title="Δ bp", overlaying="y", side="right", gridcolor="#334155"),
                        legend=dict(orientation="h", y=-0.2, font=dict(color="#e2e8f0")),
                        height=380, margin=dict(l=40, r=60, t=50, b=60))
                    st.plotly_chart(_fig4, use_container_width=True)

    # ── Section 3: What-if Scenarios (retained) ───────────────────────────────
    st.markdown("---")
    st.markdown("### 🎛️ What-if Scenarios")
    st.caption("Placeholder — will clone surfaces, apply shocks, and reprice portfolio with RV breakdowns.")
    col1, col2 = st.columns(2)
    with col1:
        st.slider("Parallel curve shift (bp)", -200, 200, 0, key="bt_curve_shift")
        st.slider("ATM vol shift (bp)", -50, 50, 0, key="bt_vol_shift")
    with col2:
        st.slider("Wing vol shift (bp)", -50, 50, 0, key="bt_wing_shift")
        st.slider("Vega flatten (%)", -20, 20, 0, key="bt_vega_flat")


# ─── RV Historical Data ──────────────────────────────────────────────────────
_RV_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rv_historical_data.json")

@st.cache_data(ttl=3600, show_spinner=False)
def _load_rv_data():
    """Load local dummy historical data. Swap for DB query when licensed."""
    try:
        with open(_RV_DATA_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return {"swap_rates": [], "vol_surface": []}

@st.cache_data(ttl=300, show_spinner=False)
def _load_rv_vols_from_db(ccy: str = "AUD", limit: int = 60) -> pd.DataFrame:
    """Load vol history from Supabase vol_history table. Returns long-format DataFrame."""
    if not HAS_POSTGRES:
        return pd.DataFrame()
    try:
        conn = get_db_connection()
        if conn is None:
            return pd.DataFrame()
        cur = conn.cursor()
        cur.execute(
            """SELECT snapshot_date, label, atm_vols FROM vol_history
               WHERE user_id = %s AND currency = %s AND atm_vols IS NOT NULL
               ORDER BY snapshot_date DESC LIMIT %s""",
            (get_db_url() and "wpo70@icloud.com", ccy, limit)
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return pd.DataFrame()
        records = []
        for snap_date, label, atm_vols in rows:
            if not atm_vols or "values" not in atm_vols:
                continue
            for row in atm_vols["values"]:
                exp = row.get("Expiry", "")
                for tenor, val in row.items():
                    if tenor == "Expiry":
                        continue
                    try:
                        records.append({
                            "date": pd.to_datetime(snap_date),
                            "label": label,
                            "expiry": exp,
                            "tenor": tenor,
                            "vol_bp": float(val)
                        })
                    except Exception:
                        pass
        if not records:
            return pd.DataFrame()
        return pd.DataFrame(records).sort_values("date").reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300, show_spinner=False)
def _load_rv_vols_snapshots_list(ccy: str = "AUD") -> list:
    """List available vol snapshots for RV dropdown."""
    if not HAS_POSTGRES:
        return []
    try:
        conn = get_db_connection()
        if conn is None:
            return []
        cur = conn.cursor()
        cur.execute(
            """SELECT id, snapshot_date, label FROM vol_history
               WHERE currency = %s AND atm_vols IS NOT NULL
               ORDER BY snapshot_date DESC LIMIT 90""",
            (ccy,)
        )
        rows = cur.fetchall()
        conn.close()
        return [{"id": r[0], "snapshot_date": str(r[1]), "label": r[2]} for r in rows]
    except Exception:
        return []

def _rv_get_rates(data: dict, ccy: str = "AUD") -> pd.DataFrame:
    """Return DataFrame: date, tenor_cols in % pa."""
    rows = [r for r in data["swap_rates"] if r["ccy"] == ccy]
    if not rows:
        return pd.DataFrame()
    records = [{"date": r["date"], **{f"r{k}": float(v) for k, v in r["rates"].items()}} for r in rows]
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values("date").reset_index(drop=True)

def _rv_get_vols(data: dict, ccy: str = "AUD") -> pd.DataFrame:
    """Return long-format vol DataFrame: date, expiry, tenor, vol_bp."""
    rows = [r for r in data["vol_surface"] if r["ccy"] == ccy]
    if not rows:
        return pd.DataFrame()
    records = []
    for r in rows:
        dt = r["date"]
        for exp_k, tenors in r["vols"].items():
            for tn, v in tenors.items():
                records.append({"date": dt, "expiry": exp_k, "tenor": int(tn), "vol_bp": float(v)})
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values("date").reset_index(drop=True)

def rv_tab():
    st.subheader("📊 Relative Value   —   Swaption & Cap/Floor Trade Ideas")
    st.caption("Live vol surface + IRS curve for richness/cheapness signals.")

    ccy = "AUD"
    curve     = get_ccy_curve(ccy)
    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")
    atm       = get_working_atm_surface(ccy)
    _, a_m, b_m, r_m, n_m = get_ccy_vol_data(ccy)

    # ── Historical data source selector ──────────────────────────────────────
    st.markdown("---")
    _rv_src_col, _rv_snap_col = st.columns([2, 4])
    with _rv_src_col:
        _rv_src = st.radio("Historical Data", ["Saved Database", "LIVE (Realtime Feed)", "Local (Demo)"],
                           horizontal=False, key="rv_data_src")
    
    df_vols_hist = pd.DataFrame()
    _snap_label_used = None

    if _rv_src == "Saved Database" and HAS_POSTGRES:
        _rv_snaps = _load_rv_vols_snapshots_list(ccy)
        if _rv_snaps:
            with _rv_snap_col:
                _snap_opts = ["All snapshots (last 60)"] + [f"{s['snapshot_date']} — {s['label']}" for s in _rv_snaps]
                _rv_snap_sel = st.selectbox("Snapshot", _snap_opts, key="rv_snap_sel")
            df_vols_hist = _load_rv_vols_from_db(ccy, limit=60)
            if _rv_snap_sel != "All snapshots (last 60)":
                _snap_dt = _rv_snap_sel.split(" — ")[0]
                df_vols_hist = df_vols_hist[df_vols_hist["date"].dt.strftime("%Y-%m-%d") == _snap_dt]
                _snap_label_used = _rv_snap_sel
        else:
            with _rv_snap_col:
                st.info("No snapshots found. Save EOD snapshots from the Vol Export tab.")
    elif _rv_src == "LIVE (Realtime Feed)":
        with _rv_snap_col:
            st.info("📡 Realtime feed coming soon — will pull intraday vol marks.")
    else:
        _rv_data_local = _load_rv_data()
        df_vols_hist = _rv_get_vols(_rv_data_local, ccy)

    has_hist_vols = len(df_vols_hist) > 0

    rv_data   = _load_rv_data()
    df_rates  = _rv_get_rates(rv_data)
    df_vols   = df_vols_hist if has_hist_vols else _rv_get_vols(rv_data)
    has_hist  = len(df_rates) > 10 or has_hist_vols
    st.markdown("---")

    rv_tabs = st.tabs([
        "📊 Vol Surface RV",
        "📈 Curve RV & Spread Analysis",
        "💡 Swaption Trade Ideas",
        "💡 Cap/Floor Trade Ideas",
        "🔮 What-If Scenarios",
    ])

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 1   —   VOL SURFACE RV
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with rv_tabs[0]:
        st.markdown("### Vol Surface Richness / Cheapness")

        if atm is None:
            st.warning("Load your AUD ATM vol surface first (Vol/SABR tab).")
        else:
            # ── Butterfly richness ────────────────────────────────────
            st.markdown("#### Volatility Butterfly (Smile Curvature)")
            st.caption("Fly = ATM vol  →  0.5≈(short-expiry vol + long-expiry vol). "
                       "Positive = ATM rich vs wings. Negative = wings rich vs ATM.")

            expiry_list = ["1m","2m","3m","6m","9m","1y","18m","2y","3y","5y","7y","10y"]
            tenor_list  = [1,2,3,5,7,10,15,20]

            fly_rows = []
            for tn in tenor_list:
                row = {"Tenor": f"{tn}Y"}
                for i in range(1, len(expiry_list)-1):
                    mid_lbl = expiry_list[i]
                    lo_lbl  = expiry_list[i-1]
                    hi_lbl  = expiry_list[i+1]
                    v_mid = get_matrix_value(atm, mid_lbl, float(tn))
                    v_lo  = get_matrix_value(atm, lo_lbl,  float(tn))
                    v_hi  = get_matrix_value(atm, hi_lbl,  float(tn))
                    if all(x is not None for x in (v_mid, v_lo, v_hi)):
                        fly = v_mid - 0.5*(v_lo + v_hi)
                        row[mid_lbl] = round(fly, 2)
                fly_rows.append(row)

            df_fly = pd.DataFrame(fly_rows).set_index("Tenor")
            valid_cols = [c for c in df_fly.columns if df_fly[c].notna().any()]
            if valid_cols:
                df_fly_disp = df_fly[valid_cols]
                st.dataframe(
                    df_fly_disp.style.background_gradient(cmap="RdYlGn_r", axis=None)
                                     .format("{:+.2f}"),
                    use_container_width=True)
                st.caption("Units: bp. Red = ATM vol expensive vs calendar wings. Green = ATM cheap.")

            # ── Calendar spread ───────────────────────────────────────
            st.markdown("#### Calendar Vol Spread   —   Term Structure Shape")
            st.caption("Short/Long ratio. >1 = inverted (short-dated expensive). <1 = normal (long-dated expensive).")

            cal_pairs = [("1m","3m"), ("3m","6m"), ("6m","1y"), ("1y","2y"),
                         ("3m","1y"), ("3m","2y"), ("6m","2y"), ("1y","3y"), ("1y","5y"), ("2y","5y")]
            cal_rows  = []
            for tn in tenor_list:
                row = {"Tenor": f"{tn}Y"}
                for (short_e, long_e) in cal_pairs:
                    vs = get_matrix_value(atm, short_e, float(tn))
                    vl = get_matrix_value(atm, long_e,  float(tn))
                    if vs and vl and vl > 0:
                        row[f"{short_e}/{long_e}"] = round(vs/vl, 3)
                cal_rows.append(row)

            df_cal = pd.DataFrame(cal_rows).set_index("Tenor")
            valid_cal = [c for c in df_cal.columns if df_cal[c].notna().any()]
            if valid_cal:
                st.dataframe(
                    df_cal[valid_cal].style.background_gradient(cmap="RdYlGn_r", axis=None)
                                          .format("{:.3f}"),
                    use_container_width=True)

            # ── Vol cone   —   current vs historical percentile ───────────
            if has_hist and not df_vols.empty:
                st.markdown("#### Vol Cone   —   Current vs Historical Percentile")
                st.caption("Compares current vol surface to 2022-2025 distribution. "
                           "Percentile = % of history where vol was BELOW current.")

                cone_rows = []
                focus_tenors = [2, 5, 10, 20]
                focus_expiries = ["3m","6m","1y","2y","3y","5y"]
                for tn in focus_tenors:
                    row = {"Tenor": f"{tn}Y"}
                    for exp in focus_expiries:
                        curr_v = get_matrix_value(atm, exp, float(tn))
                        if curr_v is None:
                            continue
                        hist = df_vols[(df_vols["tenor"] == tn) & (df_vols["expiry"] == exp)]["vol_bp"].dropna()
                        if len(hist) < 5:
                            continue
                        pct = float((hist < curr_v).mean() * 100)
                        row[exp] = round(pct, 0)
                    cone_rows.append(row)

                df_cone = pd.DataFrame(cone_rows).set_index("Tenor")
                valid_cone = [c for c in df_cone.columns if df_cone[c].notna().any()]
                if valid_cone:
                    st.dataframe(
                        df_cone[valid_cone].style.background_gradient(cmap="RdYlGn_r", axis=None)
                                                  .format("{:.0f}%"),
                        use_container_width=True)
                    st.caption("Red = high percentile (vol expensive vs history). Green = cheap.")

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 2   —   CURVE RV & SPREAD ANALYSIS
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with rv_tabs[1]:
        st.markdown("### Curve Shape & Forward Spread RV")

        if curve is None:
            st.warning("Load AUD IRS curve first (Curves tab).")
        else:
            # Current curve shape
            xs_c = curve["MaturityY"].to_numpy().astype(float)
            ys_c = curve["ZeroRatePct"].to_numpy().astype(float)

            # Key spreads live
            def _par_rate(t):
                return float(np.interp(t, xs_c, ys_c))

            # Basis adjustment: AUD curve ≤3Y is Q/Q, ≥4Y is S/S
            # For spreads crossing the Q/Q-S/S boundary, convert Q/Q rates to S/S
            # by ADDING the 6v3 basis (Q/Q payer pays more = S/S equivalent is higher)
            _basis_6v3 = get_basis_curve(ccy, "6v3")
            def _basis_at(t):
                """6v3 basis in % at maturity t (only applies to Q/Q tenors <=3Y)"""
                if _basis_6v3 is None or t > 3.0:
                    return 0.0
                bx = _basis_6v3["MaturityY"].to_numpy().astype(float)
                by = _basis_6v3["BasisBp"].to_numpy().astype(float) / 100.0
                return float(np.interp(t, bx, by))

            def _par_rate_ss(t):
                """Par rate converted to S/S equivalent for cross-boundary spreads"""
                r = _par_rate(t)
                if t <= 3.0:
                    r += _basis_at(t)  # add basis to convert Q/Q -> S/S equivalent
                return r

            spreads_live = {
                "2s10s":  (_par_rate_ss(10) - _par_rate_ss(2)),
                "3s10s":  (_par_rate_ss(10) - _par_rate_ss(3)),
                "5s10s":  (_par_rate(10) - _par_rate(5)),  # both S/S, no adjustment
                "10s30s": (_par_rate(30) - _par_rate(10)) if max(xs_c) >= 25 else None,
                "3s30s":  (_par_rate(30) - _par_rate_ss(3)) if max(xs_c) >= 25 else None,
                "2s5s":   (_par_rate(5)  - _par_rate_ss(2)),  # 2Y Q/Q vs 5Y S/S
                "2s30s":  (_par_rate(30) - _par_rate_ss(2))  if max(xs_c) >= 25 else None,
            }

            # ── Historical context from BlueGamma swap data ──────────
            _hist_ctx = {}
            try:
                import io as _io
                _bg_data = {
                    "2s10s": {"mean": 0.64, "p10": 0.47, "p90": 0.79},
                    "2s5s":  {"mean": 0.21, "p10": 0.11, "p90": 0.27},
                    "5s10s": {"mean": 0.43, "p10": 0.35, "p90": 0.52},
                }
                for _k, _h in _bg_data.items():
                    _cur = spreads_live.get(_k)
                    if _cur is not None:
                        _pct = (_cur - _h["p10"]) / (_h["p90"] - _h["p10"])
                        _pct = max(0.0, min(1.0, _pct))
                        _hist_ctx[_k] = {
                            "mean": _h["mean"],
                            "p10":  _h["p10"],
                            "p90":  _h["p90"],
                            "current": _cur,
                            "pct": _pct * 100,  # 0-100
                        }
            except Exception:
                pass

            def _hist_note(spread_key):
                h = _hist_ctx.get(spread_key)
                if not h:
                    return ""
                return (f" Historically {h['pct']:.0f}th percentile "
                        f"(range {h['p10']*100:.0f}-{h['p90']*100:.0f}bp, mean {h['mean']*100:.0f}bp, 2018-2025).")

            # Fwd rates — use same pure QQ/SS zero curves as the forward matrix
            # This ensures RV signals use identical rates to what's quoted in the matrix
            _rv_zc_qq = st.session_state.get("_aud_zc_qq")
            _rv_zc_ss = st.session_state.get("_aud_zc_ss")

            def _fwd_rate(t1, t2):
                """Forward swap rate using pure QQ/SS zero curves (matches forward matrix)."""
                tenor = t2 - t1
                if tenor <= 0:
                    return None
                try:
                    if tenor <= 3.0 and _rv_zc_qq:
                        return _fwd_from_zc(_rv_zc_qq, float(t1), float(tenor), 0.25)
                    if _rv_zc_ss:
                        return _fwd_from_zc(_rv_zc_ss, float(t1), float(tenor), 0.50)
                except Exception:
                    pass
                return None

            fwds_live = {
                "1y1y":  _fwd_rate(1, 2),
                "2y1y":  _fwd_rate(2, 3),
                "2y3y":  _fwd_rate(2, 5),
                "2y5y":  _fwd_rate(2, 7),
                "2y8y":  _fwd_rate(2, 10),
                "5y5y":  _fwd_rate(5, 10),
                "5y10y": _fwd_rate(5, 15),
                "10y10y":_fwd_rate(10,20),
            }

            # Current curve metrics
            cv1, cv2, cv3, cv4 = st.columns(4)
            cv1.metric("2s10s", f"{spreads_live['2s10s']:.2f}bp")
            cv2.metric("3s10s", f"{spreads_live['3s10s']:.2f}bp")
            cv3.metric("5y5y Fwd", f"{fwds_live['5y5y']:.4f}%" if fwds_live['5y5y'] else "N/A")
            cv4.metric("2y5y Fwd", f"{fwds_live['2y5y']:.4f}%" if fwds_live['2y5y'] else "N/A")

            # Current curve plot
            fig_curve = go.Figure()
            fig_curve.add_trace(go.Scatter(
                x=list(xs_c), y=list(ys_c),
                name="Current AUD Curve", line=dict(color="#3b82f6", width=2.5)))
            # Add key fwd rates as markers
            fwd_pts = [(2, fwds_live["2y1y"], "2y1y"), (2, fwds_live["2y3y"], "2y3y"),
                       (5, fwds_live["5y5y"], "5y5y")]
            for t, v, lbl in fwd_pts:
                if v:
                    fig_curve.add_annotation(x=t, y=v,
                        text=f"{lbl}: {v:.2f}%",
                        showarrow=True, arrowhead=2, arrowcolor="#f59e0b",
                        font=dict(color="#f59e0b", size=10))
            fig_curve.update_layout(title="AUD IRS Curve   —   Current", xaxis_title="Tenor (y)",
                yaxis_title="Rate (%)", template="plotly_dark", height=300)
            st.plotly_chart(fig_curve, use_container_width=True)

            # Historical spread analysis
            if has_hist:
                st.markdown("#### Spread History & Percentile Positioning")

                sa1, sa2 = st.columns(2)
                with sa1:
                    spread_sel = st.selectbox("Spread",
                        ["2s10s","3s10s","5s10s","2s5s","3s30s","10s30s"],
                        key="rv_spread_sel")
                with sa2:
                    lookback = st.selectbox("Lookback", ["6M","1Y","2Y","All"], index=2, key="rv_lb")

                t1_map = {"2s10s":(2,10),"3s10s":(3,10),"5s10s":(5,10),
                          "2s5s":(2,5),"3s30s":(3,30),"10s30s":(10,30)}
                t1, t2 = t1_map[spread_sel]

                col1_key = f"r{t1}"
                col2_key = f"r{t2}"
                if col1_key in df_rates.columns and col2_key in df_rates.columns:
                    df_sp = df_rates[["date", col1_key, col2_key]].dropna()
                    df_sp["spread"] = df_sp[col2_key] - df_sp[col1_key]

                    lb_days = {"6M":126,"1Y":252,"2Y":504,"All":9999}[lookback]
                    df_sp_lb = df_sp.tail(lb_days)

                    curr_spread = spreads_live.get(spread_sel)
                    if curr_spread is not None:
                        pct = float((df_sp_lb["spread"] < curr_spread).mean() * 100)
                        mn, mx = df_sp_lb["spread"].min(), df_sp_lb["spread"].max()
                        med = df_sp_lb["spread"].median()

                        p1,p2,p3,p4 = st.columns(4)
                        p1.metric("Current", f"{curr_spread:.2f}bp")
                        p2.metric("Percentile", f"{pct:.0f}th",
                                  delta="Rich" if pct > 70 else ("Cheap" if pct < 30 else "Fair"))
                        p3.metric(f"Median ({lookback})", f"{med:.2f}bp")
                        p4.metric("Range", f"{mn:.1f} / {mx:.1f}bp")

                        fig_hist = go.Figure()
                        fig_hist.add_trace(go.Scatter(
                            x=df_sp_lb["date"], y=df_sp_lb["spread"],
                            name=spread_sel, line=dict(color="#3b82f6", width=1.5)))
                        fig_hist.add_hline(y=curr_spread, line_dash="dash",
                                           line_color="#f59e0b",
                                           annotation_text=f"Current {curr_spread:.1f}bp")
                        fig_hist.add_hline(y=med, line_dash="dot",
                                           line_color="#94a3b8",
                                           annotation_text=f"Median {med:.1f}bp")
                        fig_hist.update_layout(
                            title=f"AUD {spread_sel} Spread History ({lookback})",
                            xaxis_title="Date", yaxis_title="bp",
                            template="plotly_dark", height=300)
                        st.plotly_chart(fig_hist, use_container_width=True)

            # Fwd spread scatter (replicated from swap_relative_value.py)
            if has_hist:
                st.markdown("#### Forward Spread vs Curve Spread   —   Scatter RV")
                sc1, sc2, sc3, sc4 = st.columns(4)
                with sc1:
                    fwd_sel_start = st.selectbox("Fwd Start", ["1y","2y","3y","5y"], index=1, key="rv_fwd_s")
                    fwd_sel_tenor = st.selectbox("Fwd Tenor", ["1Y","2Y","3Y","5Y"], index=1, key="rv_fwd_t")
                with sc2:
                    fwd2_sel_start = st.selectbox("Fwd2 Start", ["2y","3y","4y","5y","7y"], index=2, key="rv_fwd2_s")
                    fwd2_sel_tenor = st.selectbox("Fwd2 Tenor", ["1Y","2Y","3Y","5Y"], index=1, key="rv_fwd2_t")
                with sc3:
                    curve_t1 = st.selectbox("Curve T1", ["2Y","3Y","5Y"], index=0, key="rv_ct1")
                    curve_t2 = st.selectbox("Curve T2", ["10Y","15Y","20Y","30Y"], index=0, key="rv_ct2")
                with sc4:
                    scatter_lb = st.selectbox("Lookback", ["6M","1Y","2Y","All"], index=2, key="rv_sc_lb")

                f1_s = label_to_years(fwd_sel_start); f1_t = float(fwd_sel_tenor[:-1])
                f2_s = label_to_years(fwd2_sel_start); f2_t = float(fwd2_sel_tenor[:-1])
                c_t1 = float(curve_t1[:-1]); c_t2 = float(curve_t2[:-1])

                c1k = f"r{int(c_t1)}"; c2k = f"r{int(c_t2)}"
                f1k = f"r{int(f1_s+f1_t)}"; f1sk = f"r{int(f1_s)}"
                f2k = f"r{int(f2_s+f2_t)}"; f2sk = f"r{int(f2_s)}"

                needed = [c1k, c2k, f1k, f1sk, f2k, f2sk]
                if all(k in df_rates.columns for k in needed):
                    lb_d = {"6M":126,"1Y":252,"2Y":504,"All":9999}[scatter_lb]
                    df_sc = df_rates[["date"] + needed].dropna().tail(lb_d).copy()
                    # Forward rate approximation: simple interpolation
                    def _fwd_simple(r_start, r_end, t1_, t2_):
                        try:
                            if t2_ <= t1_: return float('nan')
                            return ((1 + r_end/100)**t2_ / (1 + r_start/100)**t1_)**(1/(t2_-t1_)) - 1
                        except Exception:
                            return float('nan')
                    try:
                        df_sc["fwd1"] = df_sc.apply(lambda r: float(_fwd_simple(float(r[f1sk]), float(r[f1k]), f1_s, f1_s+f1_t))*100, axis=1)
                        df_sc["fwd2"] = df_sc.apply(lambda r: float(_fwd_simple(float(r[f2sk]), float(r[f2k]), f2_s, f2_s+f2_t))*100, axis=1)
                        df_sc = df_sc.dropna(subset=["fwd1","fwd2"])
                    except Exception as _e:
                        st.warning(f"Could not compute forward rates: {_e}")
                        df_sc = pd.DataFrame()
                    if df_sc.empty:
                        st.info("No data for selected forward start/tenor combination.")
                    else:
                        df_sc["fwd_spread"] = (df_sc["fwd1"] - df_sc["fwd2"]) * 100  # bp
                        df_sc["curve_spread"] = df_sc[c2k] - df_sc[c1k]

                        # Regression
                        from scipy import stats as _stats
                        slope, intercept, r_val, _, _ = _stats.linregress(
                            df_sc["curve_spread"], df_sc["fwd_spread"])
                        df_sc["predicted_fwd"] = slope * df_sc["curve_spread"] + intercept
                        df_sc["residual"] = df_sc["fwd_spread"] - df_sc["predicted_fwd"]
                        curr_curve_sp = _par_rate(c_t2) - _par_rate(c_t1)
                        curr_fwd_sp = _fwd_rate(f1_s, f1_s+f1_t) and _fwd_rate(f2_s, f2_s+f2_t)
                        if curr_fwd_sp is not None and _fwd_rate(f1_s, f1_s+f1_t) is not None:
                            curr_fwd_spread = (_fwd_rate(f1_s, f1_s+f1_t) - _fwd_rate(f2_s, f2_s+f2_t)) * 100 if _fwd_rate(f2_s, f2_s+f2_t) else 0
                        else:
                            curr_fwd_spread = None

                        fig_sc = go.Figure()
                        # Colour points by recency
                        n_pts = len(df_sc)
                        colours = [f"rgba(59,130,246,{0.3 + 0.7*i/n_pts})" for i in range(n_pts)]
                        fig_sc.add_trace(go.Scatter(
                            x=df_sc["curve_spread"], y=df_sc["fwd_spread"],
                            mode="markers",
                            marker=dict(color=colours, size=5),
                            name="Historical",
                            customdata=df_sc["date"].dt.strftime("%d %b %Y"),
                            hovertemplate="%{customdata}<br>Curve: %{x:.1f}bp<br>Fwd: %{y:.1f}bp<extra></extra>"))
                        # Regression line
                        x_line = [df_sc["curve_spread"].min(), df_sc["curve_spread"].max()]
                        y_line = [slope*x + intercept for x in x_line]
                        fig_sc.add_trace(go.Scatter(x=x_line, y=y_line, name=f"Regression (R~={r_val**2:.2f})",
                                                    line=dict(color="#f59e0b", dash="dash")))
                        # Current
                        if curr_fwd_spread is not None:
                            pred = slope * curr_curve_sp + intercept
                            resid = curr_fwd_spread - pred
                            fig_sc.add_trace(go.Scatter(
                                x=[curr_curve_sp], y=[curr_fwd_spread],
                                mode="markers", marker=dict(color="#ef4444", size=14, symbol="star"),
                                name=f"Now (resid {resid:+.1f}bp)"))
                            fig_sc.add_annotation(x=curr_curve_sp, y=curr_fwd_spread,
                                text=f" NOW: {resid:+.1f}bp vs model",
                                font=dict(color="#ef4444", size=11), showarrow=False, xshift=60)
                        fig_sc.update_layout(
                            title=f"Fwd Spread ({fwd_sel_start}≈{fwd_sel_tenor}  →  {fwd2_sel_start}≈{fwd2_sel_tenor}) vs {curve_t1}/{curve_t2} Curve",
                            xaxis_title=f"{curve_t1}/{curve_t2} Curve Spread (bp)",
                            yaxis_title="Forward Spread (bp)",
                            template="plotly_dark", height=400,
                            legend=dict(orientation="h", y=1.05))
                        st.plotly_chart(fig_sc, use_container_width=True)

                        # Residual time series
                        fig_resid = go.Figure()
                        fig_resid.add_trace(go.Scatter(x=df_sc["date"], y=df_sc["residual"],
                                                        line=dict(color="#a78bfa", width=1.5), name="Residual"))
                        fig_resid.add_hline(y=0, line_dash="dash", line_color="#94a3b8")
                        fig_resid.add_hline(y=df_sc["residual"].std(), line_dash="dot",
                                            line_color="#22c55e", annotation_text="+1┬ñ├ó")
                        fig_resid.add_hline(y=-df_sc["residual"].std(), line_dash="dot",
                                            line_color="#ef4444", annotation_text=" → 1┬ñ├ó")
                        fig_resid.update_layout(title="Residual vs Regression (Rich/Cheap Signal)",
                                                template="plotly_dark", height=220)
                        st.plotly_chart(fig_resid, use_container_width=True)
                        st.caption("Residual > +1┬ñ├ó = forward spread RICH vs curve → fade. Residual <  → 1┬ñ├ó = CHEAP → buy.")

    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    # TAB 3   —   SWAPTION TRADE IDEAS
    # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
    with rv_tabs[2]:
        st.markdown("### Swaption RV Trade Recommendations")
        st.caption("Gamma/vega-optimised ideas from current vol surface + curve.")

        if atm is None or curve is None:
            st.warning("Load both AUD ATM vol surface and IRS curve to generate ideas.")
        else:
            ideas = []

            # ── Idea engine ──────────────────────────────────────────
            # 1. Vol butterfly   —   ATM vs wings in expiry dim
            for tn in [2, 5, 10]:
                for mid_e, lo_e, hi_e in [("3m","1m","6m"),("6m","3m","1y"),
                                           ("1y","6m","2y"),("2y","1y","3y")]:
                    v_mid = get_matrix_value(atm, mid_e, float(tn))
                    v_lo  = get_matrix_value(atm, lo_e,  float(tn))
                    v_hi  = get_matrix_value(atm, hi_e,  float(tn))
                    if not all(x for x in (v_mid, v_lo, v_hi)):
                        continue
                    fly = v_mid - 0.5*(v_lo + v_hi)
                    if abs(fly) > 2.5:
                        direction = "Sell" if fly > 0 else "Buy"
                        ideas.append({
                            "Type": "Vol Butterfly",
                            "Structure": f"{mid_e}≈{tn}Y fly",
                            "Signal": f"Fly = {fly:+.2f}bp",
                            "Trade": f"{direction} {mid_e}≈{tn}Y, buy wings ({lo_e} + {hi_e})" if direction == "Sell"
                                     else f"{direction} {mid_e}≈{tn}Y, sell wings ({lo_e} + {hi_e})",
                            "Rationale": f"ATM {'rich' if fly>0 else 'cheap'} vs calendar wings by {abs(fly):.1f}bp",
                            "Risk": "Vol mean reversion timing",
                            "Score": abs(fly),
                        })

            # 2. Calendar spread   —   term structure steepness
            for tn in [2, 5, 10]:
                v_short = get_matrix_value(atm, "3m", float(tn))
                v_long  = get_matrix_value(atm, "2y", float(tn))
                if v_short and v_long and v_long > 0:
                    ratio = v_short / v_long
                    if ratio > 1.25:
                        ideas.append({
                            "Type": "Calendar Spread",
                            "Structure": f"3m/2y≈{tn}Y calendar",
                            "Signal": f"Ratio = {ratio:.3f}",
                            "Trade": f"Sell 3m≈{tn}Y / Buy 2y≈{tn}Y (sell calendar)",
                            "Rationale": f"3m vol {(ratio-1)*100:.0f}% above 2y vol   —   inverted ts, mean-reverts",
                            "Risk": "Short gamma if rates move sharply near-term",
                            "Score": (ratio - 1.0) * 100,
                        })
                    elif ratio < 0.85:
                        ideas.append({
                            "Type": "Calendar Spread",
                            "Structure": f"3m/2y≈{tn}Y calendar",
                            "Signal": f"Ratio = {ratio:.3f}",
                            "Trade": f"Buy 3m≈{tn}Y / Sell 2y≈{tn}Y (buy calendar)",
                            "Rationale": f"Short-dated vol cheap vs 2Y   —   normal structure unusually steep",
                            "Risk": "Carry negative; vol may stay low near-term",
                            "Score": (1.0 - ratio) * 100,
                        })

            # 3. Curve-driven swaption direction
            if curve is not None:
                r2  = _par_rate(2)
                r5  = _par_rate(5)
                r10 = _par_rate(10)
                fwd_2y5y = _fwd_rate(2, 7)
                fwd_5y5y = _fwd_rate(5, 10)
                inv_2s10s = r2 > r10

                if fwd_5y5y and fwd_5y5y > r5 + 0.20:
                    v_5y5y = get_matrix_value(atm, "5y", 5.0)
                    ideas.append({
                        "Type": "Curve / Directional",
                        "Structure": "5y≈5Y Payer Swaption",
                        "Signal": f"5y5y fwd {fwd_5y5y:.3f}% vs spot 5Y {r5:.3f}%",
                        "Trade": "Buy 5y≈5Y ATM Payer",
                        "Rationale": f"Curve pricing {(fwd_5y5y-r5)*100:.0f}bp of steepening by 5Y point   —   asymmetric "
                                     f"risk if RBA easier than fwd. Vol at {v_5y5y:.1f}bp.",
                        "Risk": "Pays premium; loses if rates fall or stay flat",
                        "Score": (fwd_5y5y - r5) * 80,
                    })
                elif fwd_5y5y and fwd_5y5y < r5 - 0.10:
                    v_5y5y = get_matrix_value(atm, "5y", 5.0)
                    ideas.append({
                        "Type": "Curve / Directional",
                        "Structure": "5y≈5Y Receiver Swaption",
                        "Signal": f"5y5y fwd {fwd_5y5y:.3f}% vs spot 5Y {r5:.3f}%",
                        "Trade": "Buy 5y≈5Y ATM Receiver",
                        "Rationale": f"Inverted fwd curve pricing easing   —   if RBA cuts more aggressively, receiver pays well.",
                        "Risk": "Pays premium; loses if easing is less than priced",
                        "Score": abs(fwd_5y5y - r5) * 80,
                    })

            # 4. Gamma vs theta   —   short-dated high gamma
            for tn in [2, 5, 10]:
                v1m = get_matrix_value(atm, "1m", float(tn))
                v3m = get_matrix_value(atm, "3m", float(tn))
                v1y = get_matrix_value(atm, "1y", float(tn))
                if v1m and v3m and v1y:
                    # Normalised gamma proxy: vol ≈ sqrt(T) should scale with sqrt(T)
                    # If 1m vol >> 1y vol / sqrt(12), 1m gamma is expensive
                    gamma_fair = v1y / math.sqrt(12)
                    gamma_actual = v1m
                    gamma_ratio = gamma_actual / gamma_fair if gamma_fair > 0 else 1.0
                    if gamma_ratio > 1.30:
                        ideas.append({
                            "Type": "Gamma/Theta",
                            "Structure": f"1m≈{tn}Y short gamma",
                            "Signal": f"> ratio = {gamma_ratio:.2f}x fair",
                            "Trade": f"Sell 1m≈{tn}Y straddle (short gamma)",
                            "Rationale": f"1m vol {gamma_actual:.0f}bp vs fair {gamma_fair:.0f}bp   —   "
                                         f"gamma {(gamma_ratio-1)*100:.0f}% expensive on ±-adj basis",
                            "Risk": "Large near-term rate move would hurt",
                            "Score": (gamma_ratio - 1) * 60,
                        })

            # ── Curve Steepener / Flattener   —   IRS ───────────────────
            if curve is not None:
                r2  = _par_rate(2)
                r5  = _par_rate(5)
                r10 = _par_rate(10)
                slope_2s10s = (r10 - r2) if r2 and r10 else None
                slope_2s5s  = (r5  - r2) if r2 and r5  else None
                slope_5s10s = (r10 - r5) if r5 and r10 else None

                if slope_2s10s is not None:
                    if slope_2s10s > 0.50:
                        ideas.append({
                            "Type": "IRS Curve",
                            "Structure": "2s10s Flattener",
                            "Signal": f"2s10s = {slope_2s10s*100:.0f}bp steep",
                            "Trade": "Pay 2Y / Receive 10Y (IRS flattener)",
                            "Rationale": f"2s10s at {slope_2s10s*100:.0f}bp   —   historically elevated. "
                                         f"Flattener profits if curve reverts toward fair." + _hist_note("2s10s"),
                            "Risk": "Carry negative in steep curve; stop if curve steepens further",
                            "Score": slope_2s10s * 120,
                        })
                    elif slope_2s10s < -0.20:
                        ideas.append({
                            "Type": "IRS Curve",
                            "Structure": "2s10s Steepener",
                            "Signal": f"2s10s = {slope_2s10s*100:.0f}bp inverted",
                            "Trade": "Receive 2Y / Pay 10Y (IRS steepener)",
                            "Rationale": f"2s10s inverted {abs(slope_2s10s)*100:.0f}bp   —   "
                                         f"steepener profits on RBA pivot / normalisation." + _hist_note("2s10s"),
                            "Risk": "Carry positive but inversion can persist",
                            "Score": abs(slope_2s10s) * 100,
                        })

                if slope_2s5s is not None and abs(slope_2s5s) > 0.30:
                    direction = "Flattener" if slope_2s5s > 0 else "Steepener"
                    trade = "Pay 2Y / Receive 5Y" if slope_2s5s > 0 else "Receive 2Y / Pay 5Y"
                    ideas.append({
                        "Type": "IRS Curve",
                        "Structure": f"2s5s {direction}",
                        "Signal": f"2s5s = {slope_2s5s*100:.0f}bp",
                        "Trade": f"{trade} (IRS {direction.lower()})",
                        "Rationale": f"2s5s at {slope_2s5s*100:.0f}bp   —   "
                                     f"{'steep relative to history' if slope_2s5s>0 else 'inverted   —   cuts priced'}." + _hist_note("2s5s"),
                        "Risk": "Mark-to-market vol on DV01 mismatch",
                        "Score": abs(slope_2s5s) * 90,
                    })

                if slope_5s10s is not None and abs(slope_5s10s) > 0.25:
                    direction = "Flattener" if slope_5s10s > 0 else "Steepener"
                    trade = "Pay 5Y / Receive 10Y" if slope_5s10s > 0 else "Receive 5Y / Pay 10Y"
                    ideas.append({
                        "Type": "IRS Curve",
                        "Structure": f"5s10s {direction}",
                        "Signal": f"5s10s = {slope_5s10s*100:.0f}bp",
                        "Trade": f"{trade} (IRS {direction.lower()})",
                        "Rationale": f"5s10s at {slope_5s10s*100:.0f}bp." + _hist_note("5s10s"),
                        "Risk": "Basis and carry risk",
                        "Score": abs(slope_5s10s) * 80,
                    })

            # ── Curve Steepener / Flattener   —   Vol Expression ─────────
            if curve is not None and atm is not None:
                r2  = _par_rate(2)
                r5  = _par_rate(5)
                r10 = _par_rate(10)
                slope_2s10s = (r10 - r2) if r2 and r10 else None

                if slope_2s10s is not None and slope_2s10s > 0.50:
                    v_pay_2y10 = get_matrix_value(atm, "1y", 10.0)
                    v_rec_2y2  = get_matrix_value(atm, "1y", 2.0)
                    ideas.append({
                        "Type": "Vol Curve",
                        "Structure": "2s10s Flattener via Swaptions",
                        "Signal": f"2s10s = {slope_2s10s*100:.0f}bp",
                        "Trade": "Buy 1y≈2Y Receiver / Buy 1y≈10Y Payer (vol flattener)",
                        "Rationale": f"Express curve flattener via swaptions   —   receiver on front "
                                     f"(short rates fall), payer on back (long rates rise or stay). "
                                     f"Vols: 2Y={v_rec_2y2:.0f}bp, 10Y={v_pay_2y10:.0f}bp." if v_rec_2y2 and v_pay_2y10 else
                                     f"Express curve flattener via swaptions.",
                        "Risk": "Pays two premiums; needs curve to move",
                        "Score": slope_2s10s * 80,
                    })
                elif slope_2s10s is not None and slope_2s10s < -0.20:
                    v_pay_2y2  = get_matrix_value(atm, "1y", 2.0)
                    v_rec_2y10 = get_matrix_value(atm, "1y", 10.0)
                    ideas.append({
                        "Type": "Vol Curve",
                        "Structure": "2s10s Steepener via Swaptions",
                        "Signal": f"2s10s = {slope_2s10s*100:.0f}bp inverted",
                        "Trade": "Buy 1y≈2Y Payer / Buy 1y≈10Y Receiver (vol steepener)",
                        "Rationale": f"Express curve steepener   —   payer on 2Y (front rates up on inflation), "
                                     f"receiver on 10Y (longs rally on flight to quality). "
                                     f"Vols: 2Y={v_pay_2y2:.0f}bp, 10Y={v_rec_2y10:.0f}bp." if v_pay_2y2 and v_rec_2y10 else
                                     f"Vol expression of curve steepening.",
                        "Risk": "Pays two premiums; needs asymmetric rate moves",
                        "Score": abs(slope_2s10s) * 70,
                    })

            # ── Calendar Vol Spreads ──────────────────────────────────
            if atm is not None:
                for tn in [2, 5, 10]:
                    for short_e, long_e in [("1m","3m"),("3m","6m"),("6m","1y"),("1y","2y"),
                                         ("3m","6m"),("6m","1y"),("1y","2y")]:
                        v_short = get_matrix_value(atm, short_e, float(tn))
                        v_long  = get_matrix_value(atm, long_e,  float(tn))
                        if v_short and v_long and v_long > 0:
                            ratio = v_short / v_long
                            # Normalise by sqrt(T)   —   fair ratio should be ~sqrt(T_short/T_long)
                            import re as _re
                            def _e2y(e):
                                m = _re.match(r"(\d+)(m|y)", e)
                                if m:
                                    return float(m.group(1))/12 if m.group(2)=="m" else float(m.group(1))
                                return 1.0
                            fair_ratio = math.sqrt(_e2y(short_e) / _e2y(long_e))
                            rich_cheap = ratio / fair_ratio
                            if rich_cheap > 1.30:
                                ideas.append({
                                    "Type": "Calendar Vol Spread",
                                    "Structure": f"Sell {short_e} / Buy {long_e} ≈{tn}Y",
                                    "Signal": f"Ratio {ratio:.2f}x vs fair {fair_ratio:.2f}x",
                                    "Trade": f"Sell {short_e}≈{tn}Y straddle, Buy {long_e}≈{tn}Y straddle (vega-neutral)",
                                    "Rationale": f"{short_e} vol {(rich_cheap-1)*100:.0f}% rich vs {long_e} on sqrt(T) basis. "
                                                 f"Sell expensive short-dated gamma, buy cheap long-dated vega.",
                                    "Risk": "Short near-term gamma; large move hurts",
                                    "Score": (rich_cheap - 1) * 80,
                                })
                            elif rich_cheap < 0.80:
                                ideas.append({
                                    "Type": "Calendar Vol Spread",
                                    "Structure": f"Buy {short_e} / Sell {long_e} ≈{tn}Y",
                                    "Signal": f"Ratio {ratio:.2f}x vs fair {fair_ratio:.2f}x",
                                    "Trade": f"Buy {short_e}≈{tn}Y straddle, Sell {long_e}≈{tn}Y straddle",
                                    "Rationale": f"{short_e} vol {(1-rich_cheap)*100:.0f}% cheap vs {long_e}. "
                                                 f"Buy cheap near-dated gamma vs expensive long-dated vol.",
                                    "Risk": "Negative carry on long-dated short",
                                    "Score": (1 - rich_cheap) * 60,
                                })

            # Sort by score
            ideas.sort(key=lambda x: x["Score"], reverse=True)

            if not ideas:
                st.info("No strong signals at current vol levels. Surface appears fair.")
            else:
                st.markdown(f"**{len(ideas)} trade ideas generated**")

                # ── Select all / copy all ──────────────────────────────
                if "rv_selected" not in st.session_state:
                    st.session_state["rv_selected"] = set()

                col_sel1, col_sel2, col_sel3, col_sel4 = st.columns([1, 1, 1, 3])
                with col_sel1:
                    if st.button("☑️ Select All", key="rv_sel_all"):
                        st.session_state["rv_selected"] = set(range(len(ideas[:12])))
                with col_sel2:
                    if st.button("✖ Clear", key="rv_sel_none"):
                        st.session_state["rv_selected"] = set()

                selected = st.session_state["rv_selected"]

                # Build copy text (all ideas if none selected, else selected)
                _copy_pool = sorted(selected) if selected else range(len(ideas[:12]))
                _copy_lines = []
                for idx in _copy_pool:
                    if idx < len(ideas):
                        idea = ideas[idx]
                        _copy_lines.append(
                            f"[{idea['Type']}] {idea['Structure']}\n"
                            f"Trade: {idea['Trade']}\n"
                            f"Rationale: {idea['Rationale']}\n"
                            f"Risk: {idea['Risk']}\n"
                            f"Score: {idea['Score']:.1f}\n"
                        )
                _copy_text = "\n".join(_copy_lines)
                with col_sel3:
                    # Use text_area + clipboard JS instead of download_button to avoid rerun freezing P&L
                    if st.button("📅 Copy Ideas", key="rv_copy_btn"):
                        st.session_state["rv_copy_text"] = _copy_text
                        st.session_state["rv_copy_show"] = True

                if st.session_state.get("rv_copy_show"):
                    st.text_area("📅 Select all and copy (Ctrl+A, Ctrl+C)",
                                 st.session_state.get("rv_copy_text", ""),
                                 height=120, key="rv_copy_area")
                    if st.button("➖ Close", key="rv_copy_close"):
                        st.session_state["rv_copy_show"] = False
                        st.rerun()

                for i, idea in enumerate(ideas[:12]):
                    score_color = "#22c55e" if idea["Score"] > 5 else "#f59e0b"
                    is_selected = i in selected
                    with st.expander(
                        f"{'✅' if is_selected else '  '} **{idea['Type']}**   —   {idea['Structure']}  |  {idea['Signal']}  "
                        f"| Score: {idea['Score']:.1f}", expanded=i < 2):
                        c1, c2 = st.columns([3, 1])
                        with c1:
                            st.markdown(f"**Trade:** {idea['Trade']}")
                            st.markdown(f"**Rationale:** {idea['Rationale']}")
                            st.markdown(f"**Risk:** {idea['Risk']}")
                        with c2:
                            st.metric("Signal Score", f"{idea['Score']:.1f}")
                            cb_key = f"rv_chk_{i}"
                            checked = st.checkbox("Select", value=is_selected, key=cb_key)
                            if checked and i not in selected:
                                st.session_state["rv_selected"].add(i)
                                st.rerun()
                            elif not checked and i in selected:
                                st.session_state["rv_selected"].discard(i)
                                st.rerun()
                            idea_text = (
                                f"[{idea['Type']}] {idea['Structure']}\n"
                                f"Trade: {idea['Trade']}\n"
                                f"Rationale: {idea['Rationale']}\n"
                                f"Risk: {idea['Risk']}\n"
                                f"Score: {idea['Score']:.1f}"
                            )
                            st.code(idea_text, language=None)

            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            # ── Theoretical P&L at a Given Forward Rate ───────────────
            # ├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë├ö├▓├ë
            st.markdown("---")
            st.markdown("#### 📈 Theoretical P&L at a Given Forward Rate")
            st.caption("Each trade uses its own specific rate exposure   —   not a single 5Y proxy.")

            # ── Helper: parse expiry + tenor from idea Structure string ─
            import re as _re_pnl
            def _parse_structure(struct: str):
                """Extract (expiry_str, tenor_y) from strings like '1m≈10Y fly', '5y≈5Y', '2s10s Flattener'."""
                # Pattern: XeYY e.g. 1m≈10Y, 3m≈5Y
                m = _re_pnl.search(r"(\d+(?:\.\d+)?)(m|y|w)\s*[≈x*/]\s*(\d+(?:\.\d+)?)Y", struct, _re_pnl.IGNORECASE)
                if m:
                    qty, unit, tenor = m.group(1), m.group(2).lower(), float(m.group(3))
                    exp_y = float(qty)/12 if unit == "m" else float(qty)/52 if unit == "w" else float(qty)
                    return exp_y, tenor
                # Pattern: NsMs e.g. 2s10s → expiry=1y, tenor=10Y (curve spread)
                m2 = _re_pnl.search(r"(\d+)s(\d+)s", struct)
                if m2:
                    short_t = float(m2.group(1))
                    long_t  = float(m2.group(2))
                    return 1.0, long_t  # use 1y expiry, long-end tenor for DV01
                # Pattern: Ny≈NY e.g. 5y≈5Y, 2y≈10Y
                m3 = _re_pnl.search(r"(\d+)y\s*[≈x*/]\s*(\d+)Y", struct, _re_pnl.IGNORECASE)
                if m3:
                    return float(m3.group(1)), float(m3.group(2))
                return 1.0, 5.0  # fallback

            # ── Helper: AFMA modified-following year-fraction ───────────
            def _to_yearfrac(val_date, exp_y: float) -> float:
                """Convert expiry year fraction from val_date, approx AFMA modified following."""
                from datetime import timedelta as _td
                import calendar as _cal
                d = val_date + _td(days=round(exp_y * 365))
                # Roll to Monday if weekend
                if d.weekday() == 5: d += _td(days=2)
                if d.weekday() == 6: d += _td(days=1)
                # Roll back if month-end crossed   —   simplified
                from datetime import date as _dt2
                _, last = _cal.monthrange(d.year, d.month)
                if d.day > last:
                    d = d.replace(day=last)
                    while d.weekday() >= 5:
                        d -= _td(days=1)
                return (d - val_date).days / 365.0

            # ── Build forward rate lookup from matrix ──────────────────
            _fwd_matrix_ss = st.session_state.get("fwd_matrix", {}).get("AUD")
            _has_matrix = _fwd_matrix_ss is not None and not _fwd_matrix_ss.empty

            # Matrix expiry labels and their year fractions
            _matrix_exp_labels = list(_fwd_matrix_ss.index) if _has_matrix else []
            _matrix_exp_yf     = [label_to_years(e) for e in _matrix_exp_labels] if _has_matrix else []
            _matrix_tenors     = [float(c[:-1]) for c in _fwd_matrix_ss.columns] if _has_matrix else []

            def _matrix_rate_at(exp_y: float, tenor_y: float) -> float | None:
                """Interpolate fwd matrix at (expiry_yf_from_today, tenor_y).
                exp_y is already the year fraction from today   —   no date conversion needed here."""
                if not _has_matrix:
                    return None
                if len(_matrix_exp_yf) < 2 or len(_matrix_tenors) < 2:
                    return None
                _ey = max(exp_y, _matrix_exp_yf[0])
                try:
                    from scipy.interpolate import CubicSpline as _CS
                    # Interpolate across tenor columns first, then spline across expiry axis
                    if tenor_y <= _matrix_tenors[0]:
                        rates = _fwd_matrix_ss[_fwd_matrix_ss.columns[0]].values.astype(float)
                    elif tenor_y >= _matrix_tenors[-1]:
                        rates = _fwd_matrix_ss[_fwd_matrix_ss.columns[-1]].values.astype(float)
                    else:
                        col_idx = int(np.searchsorted(_matrix_tenors, tenor_y))
                        col_lo  = _fwd_matrix_ss.columns[col_idx - 1]
                        col_hi  = _fwd_matrix_ss.columns[col_idx]
                        w = (tenor_y - _matrix_tenors[col_idx-1]) / (_matrix_tenors[col_idx] - _matrix_tenors[col_idx-1])
                        rates = ((1-w) * _fwd_matrix_ss[col_lo].values +
                                    w  * _fwd_matrix_ss[col_hi].values).astype(float)
                    cs = _CS(_matrix_exp_yf, rates, extrapolate=True)
                    return float(cs(_ey))
                except Exception:
                    return float(np.interp(_ey, _matrix_exp_yf,
                               [float(np.interp(tenor_y, _matrix_tenors,
                                [float(_fwd_matrix_ss[c].iloc[i]) for c in _fwd_matrix_ss.columns]))
                                for i in range(len(_matrix_exp_yf))]))

            # ── UI ─────────────────────────────────────────────────────
            _src_col, _notional_col = st.columns([5, 2])
            with _src_col:
                _pnl_src = st.radio("Rate source", ["Manual shift (bp)", "Fwd Curve from Matrix"],
                                    horizontal=True, key="rv_pnl_src")
            with _notional_col:
                _notional_mm = st.number_input("Notional (AUD mm)", min_value=1.0, max_value=5000.0,
                                               value=100.0, step=25.0, key="rv_pnl_notional")

            if _pnl_src == "Manual shift (bp)":
                _shift_col, _info_col = st.columns([3, 3])
                with _shift_col:
                    _manual_shift_bp = st.number_input("Parallel shift (bp)", min_value=-300.0, max_value=300.0,
                                                       value=0.0, step=5.0, key="rv_pnl_shift",
                                                       help="Applied uniformly to every trade's specific rate")
                with _info_col:
                    st.info(f"Each trade's P&L uses its own expiry/tenor rate from the current curve, "
                            f"then applies this {_manual_shift_bp:+.0f}bp shift.")

                def _get_move_for_idea(exp_y, tenor_y, val_date=None):
                    """Manual: shift is same for all."""
                    spot = (_par_rate(tenor_y) or 4.5) if curve is not None else 4.5
                    return _manual_shift_bp, spot, spot + _manual_shift_bp / 100

                _val_date_used = None
                _show_fwd_curve = False

            else:  # Fwd Curve from Matrix
                if not _has_matrix:
                    st.warning("No fwd matrix loaded   —   go to Rate/Vol Matrix tab and click 'Generate All Matrices' first.")
                    _get_move_for_idea = lambda e, t, d=None: (0.0, 4.5, 4.5)
                    _val_date_used = None
                    _show_fwd_curve = False
                else:
                    _dc1, _dc2, _dc3 = st.columns([2, 2, 2])
                    with _dc1:
                        from datetime import date as _today_cls
                        _val_date = st.date_input("Valuation date", value=_today_cls.today(),
                                                  key="rv_pnl_valdate", format="DD/MM/YYYY",
                                                  help="Forward rates interpolated to this date using Modified Following")
                    with _dc2:
                        _show_fwd_curve = st.checkbox("Show interpolated curve for this date", value=True,
                                                      key="rv_pnl_show_curve")
                    with _dc3:
                        st.caption(f"Matrix has {len(_matrix_exp_labels)} expiry rows ≈ {len(_matrix_tenors)} tenor cols. "
                                   f"Cubic spline across expiry axis, linear across tenors.")

                    _val_date_used = _val_date

                    # Compute year fraction from today to val_date (the base offset)
                    from datetime import date as _today_ref
                    _val_exp_y = max((_val_date - _today_ref.today()).days / 365.0, 0.0)

                    # Show interpolated curve for selected date
                    if _show_fwd_curve:
                        _curve_rows = []
                        for _tn_y in _matrix_tenors:
                            # Curve AT val_date = matrix lookup at expiry = val_exp_y
                            _r = _matrix_rate_at(_val_exp_y, _tn_y)
                            if _r is not None:
                                _curve_rows.append({"Tenor (Y)": _tn_y, "Fwd Rate (%)": round(_r, 4)})
                        if _curve_rows:
                            _fc1, _fc2 = st.columns([2, 3])
                            with _fc1:
                                st.dataframe(pd.DataFrame(_curve_rows), use_container_width=True,
                                             hide_index=True, height=220)
                            with _fc2:
                                _fig_fc = go.Figure()
                                _fig_fc.add_trace(go.Scatter(
                                    x=[r["Tenor (Y)"] for r in _curve_rows],
                                    y=[r["Fwd Rate (%)"] for r in _curve_rows],
                                    mode="lines+markers", line=dict(color="#3b82f6", width=2),
                                    marker=dict(size=7), name=f"Fwd curve {_val_date}"))
                                # Also show today's curve for comparison
                                _today_rows = [_matrix_rate_at(0.0, t) for t in _matrix_tenors]
                                if any(r is not None for r in _today_rows):
                                    _fig_fc.add_trace(go.Scatter(
                                        x=_matrix_tenors,
                                        y=[r for r in _today_rows if r is not None],
                                        mode="lines", line=dict(color="#94a3b8", width=1.5, dash="dot"),
                                        name="Today"))
                                _fig_fc.update_layout(
                                    title=f"Forward Curve at {_val_date} vs Today (Modified Following approx.)",
                                    xaxis_title="Tenor (Y)", yaxis_title="Rate (%)",
                                    template="plotly_dark", height=240,
                                    margin=dict(t=40, b=35, l=50, r=20),
                                    legend=dict(orientation="h", y=1.05))
                                st.plotly_chart(_fig_fc, use_container_width=True)

                    def _get_move_for_idea(exp_y, tenor_y, val_date=None):
                        """Fwd rate at (val_date offset + trade expiry), vs today's spot."""
                        spot = (_par_rate(tenor_y) or 4.5) if curve is not None else 4.5
                        # Total expiry from today = time to val_date + trade's own expiry
                        total_exp_y = _val_exp_y + exp_y
                        fwd = _matrix_rate_at(total_exp_y, tenor_y)
                        if fwd is None:
                            fwd = spot
                        move_bp = round((fwd - spot) * 100, 2)
                        return move_bp, spot, fwd

            # ── Per-trade P&L calculation ──────────────────────────────
            if ideas:
                _notional = _notional_mm * 1e6
                _pnl_rows = []

                for idx, idea in enumerate(ideas[:12]):
                    is_sel = idx in selected if selected else True
                    if not is_sel:
                        continue

                    exp_y, tenor_y = _parse_structure(idea["Structure"])
                    _move_bp, _spot_rate, _fwd_rate_val = _get_move_for_idea(exp_y, tenor_y, _val_date_used)

                    # DV01 scaled to actual tenor of the underlying swap
                    _dv01 = _notional * 0.0001 * tenor_y  # $1bp per unit notional ≈ tenor years
                    _est_pnl = None
                    _basis = f"exp {exp_y:.2f}y ≈ {tenor_y:.0f}Y swap"

                    _trade = idea["Trade"]
                    _type  = idea["Type"]

                    if "Calendar" in _type or "Spread" in _type or "straddle" in _trade.lower():
                        # Long straddle / calendar spread   —   profits from vol change + abs move
                        _est_pnl = abs(_move_bp) * _dv01 / 100 * 0.5
                        _basis += f" | |{_move_bp:.1f}bp| >├ö├½├¬50%"
                    elif "Payer" in _trade and "Receiver" not in _trade:
                        _est_pnl = _move_bp * _dv01 / 100
                        _basis += f" | +{_move_bp:+.1f}bp"
                    elif "Receiver" in _trade and "Payer" not in _trade:
                        _est_pnl = -_move_bp * _dv01 / 100
                        _basis += f" | -{_move_bp:+.1f}bp"
                    elif "Flattener" in _trade:
                        # Flattener: gains if short end rises vs long end   —   use slope move
                        _short_y = max(tenor_y - 8, 2.0)
                        _, _, _fwd_short = _get_move_for_idea(exp_y, _short_y, _val_date_used)
                        _, _, _fwd_long  = _get_move_for_idea(exp_y, tenor_y, _val_date_used)
                        _slope_move = ((_fwd_long - (_par_rate(tenor_y) or 4.5)) -
                                       (_fwd_short - (_par_rate(_short_y) or 4.0))) * 100
                        _est_pnl = -_slope_move * _notional * 0.0001 * tenor_y / 100
                        _basis += f" | slope {_slope_move:+.1f}bp"
                    elif "Steepener" in _trade:
                        _short_y = max(tenor_y - 8, 2.0)
                        _, _, _fwd_short = _get_move_for_idea(exp_y, _short_y, _val_date_used)
                        _, _, _fwd_long  = _get_move_for_idea(exp_y, tenor_y, _val_date_used)
                        _slope_move = ((_fwd_long - (_par_rate(tenor_y) or 4.5)) -
                                       (_fwd_short - (_par_rate(_short_y) or 4.0))) * 100
                        _est_pnl = _slope_move * _notional * 0.0001 * tenor_y / 100
                        _basis += f" | slope {_slope_move:+.1f}bp"
                    else:
                        # Catch-all: treat as delta-one to the move
                        _est_pnl = _move_bp * _dv01 / 100 * 0.3
                        _basis += f" | {_move_bp:+.1f}bp ≈30%"

                    if _est_pnl is not None:
                        _pnl_rows.append({
                            "Structure": idea["Structure"],
                            "Exp": f"{exp_y:.2f}y",
                            "Tenor": f"{tenor_y:.0f}Y",
                            "Spot (%)": f"{_spot_rate:.3f}",
                            "Fwd (%)": f"{_fwd_rate_val:.3f}",
                            "Move": f"{_move_bp:+.1f}bp",
                            "DV01 (k)": f"{_dv01/1e3:.1f}",
                            "Est. P&L": f"{'+'if _est_pnl>=0 else ''}{_est_pnl/1e3:.1f}k",
                        })

                if _pnl_rows:
                    st.dataframe(pd.DataFrame(_pnl_rows), use_container_width=True, hide_index=True)
                    st.caption("⚙️ Estimates only   —   DV01 ≈ rate move proxy. "
                               "Use Swaptions/Caps tabs for full vol-adjusted pricing.")

                    _vals = [float(r["Est. P&L"].replace("k","").replace("+","")) for r in _pnl_rows]
                    _lbls = [r["Structure"] for r in _pnl_rows]
                    _moves = [r["Move"] for r in _pnl_rows]
                    _colors = ["#22c55e" if v >= 0 else "#ef4444" for v in _vals]
                    _fig_pnl = go.Figure(go.Bar(
                        x=_lbls, y=_vals, marker_color=_colors,
                        text=[f"{v:+.1f}k\n{m}" for v, m in zip(_vals, _moves)],
                        textposition="outside"))
                    _title_date = str(_val_date_used) if _val_date_used else "manual shift"
                    _fig_pnl.update_layout(
                        title=f"Est. P&L   —   {_notional_mm:.0f}mm notional | {_title_date}",
                        yaxis_title="Est. P&L (AUD '000)",
                        xaxis_tickangle=-35,
                        template="plotly_dark", height=360,
                        margin=dict(t=50, b=100))
                    st.plotly_chart(_fig_pnl, use_container_width=True)
                else:
                    st.info("Generate ideas above to see P&L estimates.")
    with rv_tabs[3]:
        st.markdown("### Cap/Floor RV Trade Recommendations")
        st.caption("Forward BBSW path vs caplet vol   —   find richness/cheapness by strike and maturity.")

        if curve is None:
            st.warning("Load AUD IRS curve to generate cap/floor ideas.")
        else:
            # ── Forward BBSW path ────────────────────────────────────
            st.markdown("#### Implied Forward BBSW Path vs Current Level")

            fwd_bbsw_pts = []
            for t_start in [0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 7.0]:
                r = _fwd_rate(t_start, t_start + 0.25)
                if r:
                    fwd_bbsw_pts.append({"Start (y)": t_start, "Fwd 3m BBSW (%)": round(r, 4)})

            if fwd_bbsw_pts:
                fig_fwd = go.Figure()
                t_pts = [p["Start (y)"] for p in fwd_bbsw_pts]
                r_pts = [p["Fwd 3m BBSW (%)"] for p in fwd_bbsw_pts]
                fig_fwd.add_trace(go.Scatter(x=t_pts, y=r_pts,
                    mode="lines+markers",
                    line=dict(color="#3b82f6", width=2),
                    marker=dict(size=8), name="Fwd 3m BBSW"))
                # Add current 3m rate
                r_spot = _par_rate(0.25)
                fig_fwd.add_hline(y=r_spot, line_dash="dot", line_color="#94a3b8",
                                  annotation_text=f"Spot 3m: {r_spot:.3f}%")
                fig_fwd.update_layout(
                    title="Implied Forward 3m BBSW Path",
                    xaxis_title="Forward Start (y)", yaxis_title="Rate (%)",
                    template="plotly_dark", height=280)
                st.plotly_chart(fig_fwd, use_container_width=True)

            # ── Cap/floor ideas ──────────────────────────────────────
            cf_ideas = []
            spot_3m = _par_rate(0.25)
            peak_fwd = max((p["Fwd 3m BBSW (%)"] for p in fwd_bbsw_pts), default=spot_3m)
            trough_fwd = min((p["Fwd 3m BBSW (%)"] for p in fwd_bbsw_pts), default=spot_3m)
            t_pts = [p["Start (y)"] for p in fwd_bbsw_pts] if fwd_bbsw_pts else []
            r_pts = [p["Fwd 3m BBSW (%)"] for p in fwd_bbsw_pts] if fwd_bbsw_pts else []

            # Idea 1: curve shape → cap vs floor preference
            curve_slope_2s5s = _par_rate(5) - _par_rate_ss(2)  # 5Y S/S vs 2Y Q/Q adjusted to S/S equivalent
            fwd_peak_t = t_pts[r_pts.index(max(r_pts))] if r_pts else 0

            if curve_slope_2s5s > 0.20:  # steep → rates going up
                v_2y_cap = get_matrix_value(atm, "2y", 2.0)
                cf_ideas.append({
                    "Type": "Cap",
                    "Structure": "2Y ATM Cap",
                    "Signal": f"2s5s = {curve_slope_2s5s*100:.0f}bp steep",
                    "Trade": "Buy 2Y ATM Cap (receive if BBSW > strike at each reset)",
                    "Rationale": f"Steep 2s5s ({curve_slope_2s5s*100:.0f}bp) implies market pricing rate rises. "
                                 f"Cap protects / benefits from realised hikes. "
                                 f"{'Vol at ' + str(round(v_2y_cap,1)) + 'bp.' if v_2y_cap else ''}",
                    "Risk": "Pays premium; loses if rates stay flat or fall",
                    "Score": curve_slope_2s5s * 300,
                })
            elif curve_slope_2s5s < -0.10:  # inverted → cuts priced
                v_2y_floor = get_matrix_value(atm, "2y", 2.0)
                cf_ideas.append({
                    "Type": "Floor",
                    "Structure": "2Y ATM Floor",
                    "Signal": f"2s5s = {curve_slope_2s5s*100:.0f}bp inverted",
                    "Trade": "Buy 2Y ATM Floor (receive if BBSW < strike)",
                    "Rationale": f"Inverted 2s5s   —   market pricing cuts. Floor benefits from deeper/faster cuts.",
                    "Risk": "Pays premium; loses if cuts slower than priced",
                    "Score": abs(curve_slope_2s5s) * 250,
                })

            # Idea 2: Forward peak timing
            if fwd_peak_t > 0:
                peak_rate = peak_fwd
                v_exp = get_matrix_value(atm, "1y", 2.0)
                cf_ideas.append({
                    "Type": "Cap",
                    "Structure": f"2Y Cap struck at fwd peak ({peak_rate:.3f}%)",
                    "Signal": f"Fwd peak {peak_rate:.3f}% at {fwd_peak_t:.1f}y",
                    "Trade": f"Buy 2Y Cap struck at {peak_rate:.3f}%   —   OTM cap on peak BBSW",
                    "Rationale": f"Curve implies BBSW peaks at {peak_rate:.3f}% around {fwd_peak_t:.1f}y. "
                                 f"OTM cap cheap if realised path overshoots. "
                                 f"{'Cap vol ~' + str(round(v_exp,1)) + 'bp.' if v_exp else ''}",
                    "Risk": "OTM   —   needs BBSW to exceed forward peak",
                    "Score": 15,
                })

            # Idea 3: Vol cone percentile for caps
            if has_hist and not df_vols.empty:
                for tn in [1, 2, 3]:
                    curr_v = get_matrix_value(atm, "1y", float(tn))
                    if not curr_v:
                        continue
                    hist_v = df_vols[(df_vols["tenor"] == tn) & (df_vols["expiry"] == "1y")]["vol_bp"]
                    if len(hist_v) < 10:
                        continue
                    pct = float((hist_v < curr_v).mean() * 100)
                    if pct < 25:
                        cf_ideas.append({
                            "Type": "Cap/Floor",
                            "Structure": f"1y≈{tn}Y vol cheap",
                            "Signal": f"Vol at {pct:.0f}th percentile",
                            "Trade": f"Buy 1y≈{tn}Y straddle (long gamma + vol)",
                            "Rationale": f"1y caplet vol at {curr_v:.0f}bp = {pct:.0f}th percentile historically. "
                                         f"Cheap entry for long vol / long gamma position.",
                            "Risk": "Theta drag if rates stay range-bound",
                            "Score": (25 - pct) * 2,
                        })
                    elif pct > 75:
                        cf_ideas.append({
                            "Type": "Cap/Floor",
                            "Structure": f"1y≈{tn}Y vol rich",
                            "Signal": f"Vol at {pct:.0f}th percentile",
                            "Trade": f"Sell 1y≈{tn}Y cap or floor vs delta hedge",
                            "Rationale": f"1y caplet vol at {curr_v:.0f}bp = {pct:.0f}th percentile. Rich historically.",
                            "Risk": "Short gamma; realised vol could exceed implied",
                            "Score": (pct - 75) * 1.5,
                        })

            cf_ideas.sort(key=lambda x: x["Score"], reverse=True)

            if not cf_ideas:
                st.info("No strong cap/floor signals at current levels.")
            else:
                st.markdown(f"**{len(cf_ideas)} cap/floor ideas generated**")
                for i, idea in enumerate(cf_ideas[:6]):
                    with st.expander(
                        f"**{idea['Type']}**   —   {idea['Structure']}  |  {idea['Signal']}",
                        expanded=i < 2):
                        c1, c2 = st.columns([3, 1])
                        with c1:
                            st.markdown(f"**Trade:** {idea['Trade']}")
                            st.markdown(f"**Rationale:** {idea['Rationale']}")
                            st.markdown(f"**Risk:** {idea['Risk']}")
                        with c2:
                            st.metric("Score", f"{idea['Score']:.1f}")





    with rv_tabs[4]:
        st.markdown("### 🔮 What-If Scenarios")
        st.caption("Stress-test and scenario analysis using 2+ months of AUD vol history.")

        _wi_atm = get_working_atm_surface("AUD")
        if _wi_atm is None:
            st.warning("Load AUD ATM vol surface first.")
        else:
            _wi_snaps = _load_vol_snapshots_for_viz("AUD", "2026-02-01", "2026-12-31")
            _wi_n = len(_wi_snaps)
            st.caption(f"Using {_wi_n} historical snapshots in database.")

            wi_tab1, wi_tab2, wi_tab3 = st.tabs([
                "📈 Vol Shock / Parallel Shift",
                "📅 Historical Scenario Replay",
                "🚨 Vol Percentile Alerts",
            ])

            # ── TAB 1: Vol Shock ─────────────────────────────────────────────
            with wi_tab1:
                st.markdown("#### Apply Vol Shock to Current Surface")
                st.caption("Shift the current ATM vol surface and see the impact on key swaption prices.")

                _sh_c1, _sh_c2, _sh_c3 = st.columns(3)
                with _sh_c1:
                    _shock_parallel = st.number_input("Parallel Shift (bp)", min_value=-50, max_value=50,
                                                       value=0, step=1, key="wi_shock_parallel")
                with _sh_c2:
                    _shock_exp_tilt = st.number_input("Expiry Tilt (bp/yr, + = longer higher)",
                                                       min_value=-5.0, max_value=5.0,
                                                       value=0.0, step=0.5, key="wi_shock_exp_tilt")
                with _sh_c3:
                    _shock_ten_tilt = st.number_input("Tenor Tilt (bp/yr, + = longer higher)",
                                                       min_value=-5.0, max_value=5.0,
                                                       value=0.0, step=0.5, key="wi_shock_ten_tilt")

                if st.button("▶ Apply Shock & Show Δ", key="wi_apply_shock", type="primary"):
                    import plotly.graph_objects as _go_wi
                    _av = _wi_atm.copy()
                    if "Expiry" in _av.columns: _av = _av.set_index("Expiry")
                    _tenor_cols = [c for c in _av.columns]

                    _shocked = _av.copy()
                    _delta   = pd.DataFrame(index=_av.index, columns=_av.columns, dtype=float)

                    for _exp_lbl in _av.index:
                        _exp_y = label_to_years(str(_exp_lbl))
                        for _ten_col in _tenor_cols:
                            try:
                                _ten_y = float(str(_ten_col).replace("Y","").replace("y",""))
                                _base = float(_av.loc[_exp_lbl, _ten_col])
                                _adj = _shock_parallel + _shock_exp_tilt * _exp_y + _shock_ten_tilt * _ten_y
                                _shocked.loc[_exp_lbl, _ten_col] = round(_base + _adj, 2)
                                _delta.loc[_exp_lbl, _ten_col] = round(_adj, 2)
                            except Exception:
                                pass

                    # Show delta heatmap
                    _nc = [c for c in _delta.columns]
                    _fmt = {c: "{:+.1f}" for c in _nc}
                    st.markdown("**Vol Change (bp):**")
                    st.dataframe(_delta.style.format(_fmt).background_gradient(
                        "RdYlGn", axis=None, subset=_nc,
                        vmin=-max(abs(_shock_parallel)+5, 1), vmax=max(abs(_shock_parallel)+5, 1)),
                        use_container_width=True)

                    # Key swaption pricer cells
                    st.markdown("**Shocked Surface ATM Vols (bp):**")
                    _nc2 = [c for c in _shocked.columns]
                    st.dataframe(_shocked.style.format({c: "{:.2f}" for c in _nc2}).background_gradient(
                        "RdYlGn_r", axis=None, subset=_nc2),
                        use_container_width=True, height=350)

                    # Store shocked surface in session for further use
                    st.session_state["_wi_shocked_surface"] = _shocked
                    st.success(f"Shocked surface stored. Parallel: {_shock_parallel:+.0f}bp | "
                               f"Expiry tilt: {_shock_exp_tilt:+.1f}bp/yr | Tenor tilt: {_shock_ten_tilt:+.1f}bp/yr")

            # ── TAB 2: Historical Scenario Replay ────────────────────────────
            with wi_tab2:
                st.markdown("#### Historical Scenario Replay")
                st.caption("Load any historical vol snapshot as the working scenario surface. "
                           "Compares to today and shows the vol change.")

                if _wi_n == 0:
                    st.info("No historical snapshots found — run Load Vol Snapshots in Historical VOL Analysis tab first.")
                else:
                    _snap_opts = {f"{s['label']} ({s['date'].strftime('%d-%b-%Y')})": s
                                  for s in _wi_snaps}
                    _sel_snap = st.selectbox("Select Historical Date", list(_snap_opts.keys()),
                                              key="wi_hist_snap")

                    if st.button("▶ Load Scenario", key="wi_load_scenario", type="primary"):
                        _hs = _snap_opts[_sel_snap]
                        # snap dict has keys: id, date, label, df (DataFrame already indexed)
                        _hist_df = _hs["df"].copy()

                        _curr_df = _wi_atm.copy()
                        if "Expiry" in _curr_df.columns: _curr_df = _curr_df.set_index("Expiry")

                        _common_exp = [e for e in _hist_df.index if e in _curr_df.index]
                        _common_ten = [c for c in _hist_df.columns if c in _curr_df.columns]

                        if _common_exp and _common_ten:
                            _hist_sub = _hist_df.loc[_common_exp, _common_ten].apply(pd.to_numeric, errors="coerce")
                            _curr_sub = _curr_df.loc[_common_exp, _common_ten].apply(pd.to_numeric, errors="coerce")
                            _chg = _curr_sub - _hist_sub

                            col_a, col_b = st.columns(2)
                            with col_a:
                                st.markdown(f"**Historical: {_sel_snap}**")
                                st.dataframe(_hist_sub.style.format("{:.2f}").background_gradient(
                                    "RdYlGn_r", axis=None), use_container_width=True, height=320)
                            with col_b:
                                st.markdown("**Change vs Today (bp):**")
                                _fmt_chg = {c: "{:+.2f}" for c in _common_ten}
                                st.dataframe(_chg.style.format(_fmt_chg).background_gradient(
                                    "RdYlGn", axis=None,
                                    vmin=-10, vmax=10), use_container_width=True, height=320)

                            # Summary stats
                            _avg_chg = float(_chg.mean().mean())
                            _max_chg = float(_chg.max().max())
                            _min_chg = float(_chg.min().min())
                            m1, m2, m3 = st.columns(3)
                            m1.metric("Avg Change", f"{_avg_chg:+.1f}bp")
                            m2.metric("Max Increase", f"{_max_chg:+.1f}bp")
                            m3.metric("Max Decrease", f"{_min_chg:+.1f}bp")

                            st.session_state["_wi_scenario_surface"] = _hist_df
                            st.success(f"Historical scenario loaded: {_sel_snap}")

            # ── TAB 3: Vol Percentile Alerts ─────────────────────────────────
            with wi_tab3:
                st.markdown("#### Vol Percentile Alerts")
                st.caption("Shows where current vols sit in the historical distribution. "
                           "Red = top decile (expensive), Green = bottom decile (cheap).")

                if _wi_n < 5:
                    st.info(f"Need at least 5 snapshots for percentile analysis (have {_wi_n}). "
                            "Load more data in Historical VOL Analysis.")
                else:
                    _av2 = _wi_atm.copy()
                    if "Expiry" in _av2.columns: _av2 = _av2.set_index("Expiry")
                    _tenor_cols2 = [c for c in _av2.columns]

                    # Build historical distribution per cell
                    _hist_data = {}  # (exp, ten) -> list of historical vols
                    for _hs2 in _wi_snaps:
                        _hdf2 = _hs2.get("df")
                        if _hdf2 is None:
                            continue
                        for _exp2 in _hdf2.index:
                            for _ten2 in _hdf2.columns:
                                try:
                                    _v2 = float(_hdf2.loc[_exp2, _ten2])
                                    if not math.isnan(_v2):
                                        _hist_data.setdefault((_exp2, _ten2), []).append(_v2)
                                except Exception:
                                    pass

                    # Build percentile table
                    _pct_rows = []
                    for _exp2 in _av2.index:
                        _row2 = {"Expiry": str(_exp2)}
                        for _ten2 in _tenor_cols2:
                            try:
                                _curr_v = float(_av2.loc[_exp2, _ten2])
                                _hist_v = _hist_data.get((_exp2, _ten2), [])
                                if len(_hist_v) >= 3:
                                    _pct = float(np.mean(np.array(_hist_v) <= _curr_v) * 100)
                                    _row2[_ten2] = round(_pct, 0)
                                else:
                                    _row2[_ten2] = None
                            except Exception:
                                _row2[_ten2] = None
                        _pct_rows.append(_row2)

                    _pct_df = pd.DataFrame(_pct_rows).set_index("Expiry")
                    _nc3 = [c for c in _pct_df.columns]

                    st.markdown(f"**Percentile of current vol vs {_wi_n}-day history "
                                f"(100=all-time high, 0=all-time low):**")
                    st.dataframe(
                        _pct_df.style.format("{:.0f}%", na_rep="—").background_gradient(
                            "RdYlGn_r", axis=None, subset=_nc3, vmin=0, vmax=100),
                        use_container_width=True, height=400)

                    # Alerts for extreme cells
                    _alerts = []
                    for _exp2 in _pct_df.index:
                        for _ten2 in _pct_df.columns:
                            _p = _pct_df.loc[_exp2, _ten2]
                            if _p is not None and not math.isnan(float(_p if _p is not None else float('nan'))):
                                if float(_p) >= 90:
                                    _alerts.append(f"🔴 **{_exp2}{_ten2.lower()}** vol at {float(_p):.0f}th pct — EXPENSIVE")
                                elif float(_p) <= 10:
                                    _alerts.append(f"🟢 **{_exp2}{_ten2.lower()}** vol at {float(_p):.0f}th pct — CHEAP")
                    if _alerts:
                        st.markdown("**Extreme Vol Alerts:**")
                        for _a in _alerts[:20]:
                            st.markdown(_a)
                    else:
                        st.success("No extreme percentile readings (all cells 10th–90th percentile).")


# ─────────────────────────────────────────────────────────────────────────────
# AGB BOND OPTION PRICER   —   OTC Physical Bond Options
# ─────────────────────────────────────────────────────────────────────────────

# Current AGB lines on issue (coupon %, maturity date, label)
AGB_LINES = [
    # Active bonds as of Mar-2026   —   remove matured: Nov-24, Apr-25, Oct-25
    (4.50, "2026-04-21", "4.50% Apr-26"),
    (1.75, "2026-11-21", "1.75% Nov-26"),
    (4.75, "2027-04-21", "4.75% Apr-27"),
    (2.75, "2027-11-21", "2.75% Nov-27"),
    (2.25, "2028-05-21", "2.25% May-28"),
    (3.25, "2029-04-21", "3.25% Apr-29"),
    (2.75, "2029-11-21", "2.75% Nov-29"),
    (1.00, "2030-11-21", "1.00% Nov-30"),
    (2.25, "2031-05-21", "2.25% May-31"),
    (1.75, "2032-11-21", "1.75% Nov-32"),
    (3.75, "2034-04-21", "3.75% Apr-34"),
    (2.75, "2035-11-21", "2.75% Nov-35"),
    (4.25, "2036-04-21", "4.25% Apr-36"),
    (3.00, "2047-03-21", "3.00% Mar-47"),
]


def _coupon_dates(maturity: date, settlement: date) -> list:
    """
    Generate all semi-annual coupon dates for an AGB from settlement forward.
    AGB pays on the 21st of the coupon month (6 months apart).
    """
    mat_month = maturity.month
    mat_day = maturity.day
    # coupon months are 6 apart: mat_month and mat_month+6 (mod 12)
    months = [mat_month, (mat_month + 6 - 1) % 12 + 1]
    cpn_dates = []
    year = settlement.year - 1
    for _ in range(80):  # scan 40 years forward
        for m in sorted(months):
            try:
                d = date(year, m, mat_day)
            except ValueError:
                d = date(year, m, 28)
            if d > settlement:
                cpn_dates.append(d)
        year += 1
        if cpn_dates and cpn_dates[-1] >= maturity:
            break
    return sorted([d for d in cpn_dates if d <= maturity])


def aofm_dirty_price(coupon_pct: float, maturity: date, yield_pct: float, settlement: date) -> float:
    """
    AOFM/RBA formula for AGB dirty price per $100 face value.
    Handles standard case (Formula 1). Returns price per $100.
    """
    from datetime import date as _date
    y = yield_pct / 100.0
    i = y / 2.0  # semi-annual yield
    g = coupon_pct / 2.0  # half-yearly coupon per $100

    # Find next coupon date after settlement
    all_cpdates = _coupon_dates(maturity, settlement)
    if not all_cpdates:
        # Bond matured or near maturity   —   simple formula
        if i == 0:
            return 100.0 + g
        f_days = (maturity - settlement).days
        return (100.0 + g) / (1.0 + (f_days / 365.0) * y)

    next_cpn = all_cpdates[0]

    # f = days from settlement to next coupon
    f = (next_cpn - settlement).days
    # d = days in half-year ending on next coupon (prev coupon to next coupon)
    # prev coupon is 6 months before next_cpn
    if next_cpn.month <= 6:
        prev_cpn = date(next_cpn.year - 1, next_cpn.month + 6, next_cpn.day)
    else:
        prev_cpn = date(next_cpn.year, next_cpn.month - 6, next_cpn.day)
    d = (next_cpn - prev_cpn).days

    # n = number of full half-years from next coupon to maturity
    remaining = [dt for dt in all_cpdates if dt >= next_cpn]
    n = len(remaining)

    if i == 0:
        ann = n
    else:
        v = 1.0 / (1.0 + i)
        vn = round(v ** n, 8)
        ann = round((1.0 - vn) / i, 8)
        # Final price
        P = (v ** (f / d)) * (g * ann + 100.0 * vn)
        return P

    v = 1.0 / (1.0 + i)
    vn = v ** n
    ann = n  # i==0 case already returned above
    P = (v ** (f / d)) * (g * ann + 100.0 * vn)
    return P


def aofm_yield_from_price(price: float, coupon_pct: float, maturity: date, settlement: date) -> float:
    """Newton-Raphson inversion of AOFM dirty price formula."""
    import scipy.optimize as _opt
    def objective(y):
        return aofm_dirty_price(coupon_pct, maturity, y, settlement) - price
    try:
        return _opt.brentq(objective, 0.0001, 30.0, xtol=1e-8)
    except Exception:
        return float('nan')


def agb_dv01(coupon_pct: float, maturity: date, yield_pct: float, settlement: date, face: float = 1e6) -> float:
    """DV01 per face_value notional (bump 1bp)."""
    p_up = aofm_dirty_price(coupon_pct, maturity, yield_pct - 0.01, settlement)
    p_dn = aofm_dirty_price(coupon_pct, maturity, yield_pct + 0.01, settlement)
    dirty = aofm_dirty_price(coupon_pct, maturity, yield_pct, settlement)
    dv01_per100 = (p_up - p_dn) / 2.0
    return dv01_per100 * face / 100.0


def agb_modified_duration(coupon_pct: float, maturity: date, yield_pct: float, settlement: date) -> float:
    """Modified duration (years)."""
    p = aofm_dirty_price(coupon_pct, maturity, yield_pct, settlement)
    p_up = aofm_dirty_price(coupon_pct, maturity, yield_pct - 0.01, settlement)
    p_dn = aofm_dirty_price(coupon_pct, maturity, yield_pct + 0.01, settlement)
    dp_dy = (p_up - p_dn) / (2.0 * 0.0001)  # d/dy in decimal units: yield moved 0.01% = 0.0001
    return -dp_dy / p


def agb_forward_price(
    coupon_pct: float,
    maturity: date,
    yield_pct: float,
    settle: date,
    expiry: date,
    repo_rate_pct: float,
    ois_curve: Optional[pd.DataFrame],
) -> Tuple[float, float, list]:
    """
    Forward dirty price of AGB at option expiry date.
    Returns (forward_dirty_price, forward_yield, coupons_stripped).
    """
    dirty_today = aofm_dirty_price(coupon_pct, maturity, yield_pct, settle)
    T_days = (expiry - settle).days
    T = T_days / 365.0

    # Coupons paid between settle and expiry   —   strip and PV at OIS rate
    all_cpdates = _coupon_dates(maturity, settle)
    g = coupon_pct / 2.0
    coupons_stripped = []
    pv_coupons = 0.0
    for cd in all_cpdates:
        if settle < cd <= expiry:
            t_cpn = (cd - settle).days / 365.0
            if ois_curve is not None:
                xs = ois_curve["MaturityY"].to_numpy().astype(float)
                ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                r = float(np.interp(t_cpn, xs, ys))
            else:
                r = repo_rate_pct / 100.0  # fallback
            df = math.exp(-r * t_cpn)
            pv_coupons += g * df
            coupons_stripped.append((cd, g, df))

    # Carry forward at repo rate
    repo = repo_rate_pct / 100.0
    fwd_dirty = (dirty_today - pv_coupons) * math.exp(repo * T)

    # Derive forward yield
    fwd_yield = aofm_yield_from_price(fwd_dirty, coupon_pct, maturity, expiry)
    return fwd_dirty, fwd_yield, coupons_stripped


def black76_bond_option(
    F: float, K: float, sigma: float, T: float, df_T: float, is_call: bool
) -> dict:
    """
    Black-76 for European bond option.
    F, K = forward/strike dirty price per $100.
    sigma = lognormal price vol (decimal, e.g. 0.06 for 6%).
    Returns premium per $100 face, plus Greeks.
    """
    from statistics import NormalDist as _N
    _nd = _N()

    if T <= 0 or sigma <= 0 or F <= 0 or K <= 0:
        return {"prem_per100": 0.0, "delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0}

    sqrt_T = math.sqrt(T)
    d1 = (math.log(F / K) + 0.5 * sigma ** 2 * T) / (sigma * sqrt_T)
    d2 = d1 - sigma * sqrt_T

    Nd1 = _nd.cdf(d1)
    Nd2 = _nd.cdf(d2)
    phi_d1 = _nd.pdf(d1)

    if is_call:
        prem = df_T * (F * Nd1 - K * Nd2)
        delta = df_T * Nd1
    else:
        prem = df_T * (K * _nd.cdf(-d2) - F * _nd.cdf(-d1))
        delta = -df_T * _nd.cdf(-d1)

    gamma = df_T * phi_d1 / (F * sigma * sqrt_T)
    vega = df_T * F * phi_d1 * sqrt_T  # per unit vol
    theta = -(df_T * F * phi_d1 * sigma / (2 * sqrt_T)) / 365.0  # per day

    return {
        "prem_per100": prem,
        "delta_price": delta,       # ├ö├¬├®V/├ö├¬├®F (price)
        "gamma": gamma,
        "vega_per1pct": vega / 100.0,  # per 1% vol move
        "theta_per_day": theta,
    }


def bachelier_bond_option_yield(
    F_yield: float, K_yield: float, sigma_bp: float, T: float,
    df_T: float, mod_dur: float, F_price: float, is_payer: bool, face: float
) -> dict:
    """
    Normal/Bachelier option in yield space.
    is_payer=True: receiver of bond (payer of yield = profits when yield rises).
    Returns dollar premium and Greeks.
    """
    from statistics import NormalDist as _N
    _nd = _N()
    sigma = sigma_bp / 10000.0  # convert bp to decimal
    if T <= 0 or sigma <= 0:
        return {"prem_dollar": 0.0, "pv_bp": 0.0, "delta_yield_bp": 0.0, "vega": 0.0}

    sqrt_T = math.sqrt(T)
    z = (F_yield - K_yield) / (sigma * sqrt_T)

    phi_z = _nd.pdf(z)
    Nz = _nd.cdf(z)

    # Bachelier formula: prem in yield units
    if is_payer:  # benefits from yield rise
        prem_yield = sigma * sqrt_T * (-z * _nd.cdf(-z) + phi_z)
    else:  # benefits from yield fall
        prem_yield = sigma * sqrt_T * (z * Nz + phi_z)

    # Convert yield premium to price premium via duration
    # prem_price_per100 ├ö├½├¬ prem_yield * mod_dur * F_price / yield_decimal_unit
    prem_per100 = prem_yield * 100.0 * mod_dur * F_price / 100.0
    prem_dollar = prem_per100 * face / 100.0
    pv_bp = prem_per100 * 100.0  # in bp of face value

    # Delta in bp (yield sensitivity)
    delta_yield = _nd.cdf(-z) if is_payer else Nz
    delta_yield_bp = delta_yield * mod_dur * face / 1e4  # $ per bp

    # Vega per bp vol
    vega = df_T * sqrt_T * phi_z * mod_dur * F_price / 100.0 * face / 10000.0

    return {
        "prem_dollar": prem_dollar * df_T,
        "pv_bp": pv_bp,
        "delta_yield_bp": delta_yield_bp,
        "vega_per_bp": vega,
    }


def bond_option_tab():
    """OTC AGB Bond Option Pricer."""
    st.subheader("📜 Bond Options   —   OTC Physical AGB")

    # ── Licence gate ────────────────────────────────────────────────────
    _BOND_PW = "REBO2024"   # change this to your chosen licence key
    if not st.session_state.get("bond_options_unlocked"):
        st.info("⚙️ Bond Options is a licensed add-on. Enter your access code to continue.")
        _pw_col, _btn_col = st.columns([3, 1])
        with _pw_col:
            _pw = st.text_input("Access Code", type="password", key="bond_pw_input",
                                placeholder="Enter licence key")
        with _btn_col:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("Unlock", key="bond_pw_btn", type="primary"):
                if _pw == _BOND_PW:
                    st.session_state["bond_options_unlocked"] = True
                    st.rerun()
                else:
                    st.error("Incorrect access code. Contact wpo@rateedge.au to licence Bond Options.")
        return

    from datetime import date as _date, timedelta

    ois_curve = get_basis_curve("AUD", "ois")

    # ── Bond selection ───────────────────────────────────────────────
    col_bond, col_face, col_model = st.columns([3, 2, 2])
    with col_bond:
        bond_labels = [b[2] for b in AGB_LINES]
        bond_sel = st.selectbox("Bond Line", bond_labels, index=min(8, len(bond_labels)-1), key="bo_bond")
        bond_idx = bond_labels.index(bond_sel)
        coupon_pct, mat_str, _ = AGB_LINES[bond_idx]
        maturity = _date.fromisoformat(mat_str)
    with col_face:
        face_mm = st.number_input("Face Value (AUD mm)", min_value=1.0, max_value=5000.0,
                                   value=50.0, step=10.0, key="bo_face")
        face = face_mm * 1e6
    with col_model:
        model = st.radio("Vol Model", ["Price Log-Normal (Black-76)", "Yield Normal (Bachelier)"],
                         key="bo_model")

    # ── Option structure ─────────────────────────────────────────────
    col_type, col_exp, col_strike, col_vol = st.columns([2, 2, 2, 2])
    with col_type:
        opt_type = st.selectbox("Option Type", ["Receiver (Buy Bond)", "Payer (Sell Bond)", "Straddle"],
                                key="bo_type")
    with col_exp:
        today = _date.today()
        expiry_presets = {"1m": 30, "2m": 60, "3m": 91, "6m": 182, "9m": 273, "1y": 365,
                          "2y": 730, "3y": 1095, "5y": 1825}
        exp_sel = st.selectbox("Expiry", list(expiry_presets.keys()) + ["ƒôÇ Custom..."],
                               index=3, key="bo_expiry")
        if exp_sel == "ƒôÇ Custom...":
            exp_date = st.date_input("Expiry Date", value=today + timedelta(days=182),
                                     format="DD/MM/YYYY", key="bo_exp_custom")
        else:
            exp_date = today + timedelta(days=expiry_presets[exp_sel])
        T = (exp_date - today).days / 365.0
    with col_strike:
        strike_type = st.radio("Strike", ["ATM (Fwd)", "Manual"], horizontal=True, key="bo_stk_type")
    with col_vol:
        if "Price" in model:
            vol_input = st.number_input("Price Vol (%)", min_value=0.1, max_value=30.0,
                                         value=6.0, step=0.1, key="bo_vol_price")
        else:
            vol_input = st.number_input("Yield Vol (bp pa)", min_value=1.0, max_value=2000.0,
                                         value=80.0, step=1.0, key="bo_vol_yield")

    # ── Yield / Repo inputs ──────────────────────────────────────────
    col_yield, col_repo, col_spread = st.columns([2, 2, 2])
    with col_yield:
        spot_yield = st.number_input("Spot Yield (% pa)", min_value=0.01, max_value=20.0,
                                      value=4.50, step=0.01, format="%.4f", key="bo_yield")
    with col_repo:
        repo_rate = st.number_input("Repo Rate (% pa)", min_value=0.01, max_value=20.0,
                                     value=4.30, step=0.01, format="%.3f", key="bo_repo")
    with col_spread:
        st.caption("GC AGB repo ├ö├½├¬ AONIA + 5-15bp")
        aonia_proxy = repo_rate  # user sets directly

    # ─── Pricing engine ──────────────────────────────────────────────
    settle = add_au_bds(today, 2)  # T+2 AUD BD (AFMA calendar)

    # Spot dirty price
    try:
        dirty_spot = aofm_dirty_price(coupon_pct, maturity, spot_yield, settle)
        accrued = coupon_pct / 2.0 * (1.0 - ((exp_date if False else settle) - settle).days / 182.5)
        # proper accrued
        all_cpdates = _coupon_dates(maturity, settle)
        g = coupon_pct / 2.0
        if all_cpdates:
            next_cpn = all_cpdates[0]
            if next_cpn.month <= 6:
                prev_cpn = _date(next_cpn.year - 1, next_cpn.month + 6, next_cpn.day)
            else:
                prev_cpn = _date(next_cpn.year, next_cpn.month - 6, next_cpn.day)
            d_period = (next_cpn - prev_cpn).days
            f_days = (next_cpn - settle).days
            accrued = g * (d_period - f_days) / d_period
        else:
            accrued = 0.0
        clean_spot = dirty_spot - accrued
    except Exception as e:
        st.error(f"Spot price error: {e}")
        return

    # Forward dirty price
    try:
        fwd_dirty, fwd_yield, coupons_stripped = agb_forward_price(
            coupon_pct, maturity, spot_yield, settle, exp_date, repo_rate, ois_curve
        )
    except Exception as e:
        st.error(f"Forward price error: {e}")
        return

    # Strike
    if strike_type == "ATM (Fwd)":
        K_yield = fwd_yield
        K_price = fwd_dirty
    else:
        K_yield = st.number_input("Strike Yield (% pa)", min_value=0.01, max_value=20.0,
                                   value=round(fwd_yield, 3), step=0.01, format="%.4f",
                                   key="bo_strike_yield")
        K_price = aofm_dirty_price(coupon_pct, maturity, K_yield, exp_date)

    # OIS discount factor to expiry
    if ois_curve is not None:
        xs = ois_curve["MaturityY"].to_numpy().astype(float)
        ys = ois_curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
        r_T = float(np.interp(T, xs, ys))
        df_T = math.exp(-r_T * T)
    else:
        df_T = math.exp(-repo_rate / 100.0 * T)

    # Duration for yield-space conversion
    mod_dur = agb_modified_duration(coupon_pct, maturity, fwd_yield, exp_date)
    dv01 = agb_dv01(coupon_pct, maturity, fwd_yield, exp_date, face)

    # ─── Metrics bar ─────────────────────────────────────────────────
    st.markdown("---")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Spot Clean", f"{clean_spot:.4f}")
    m2.metric("Spot Dirty", f"{dirty_spot:.4f}")
    m3.metric("Fwd Dirty", f"{fwd_dirty:.4f}")
    m4.metric("Fwd Yield", f"{fwd_yield:.4f}%")
    m5.metric("Mod Duration", f"{mod_dur:.3f}y")
    m6.metric("DV01 ($)", f"${dv01:,.0f}")
    if coupons_stripped:
        cpn_str = "  |  ".join([f"{cd.strftime('%b-%y')} ${pv*face/100:.0f}" for cd, g_, pv in coupons_stripped])
        st.caption(f"Coupons stripped: {cpn_str}")

    # ─── Option pricing ──────────────────────────────────────────────
    st.markdown("---")

    def price_one(is_call_price: bool, is_payer_yield: bool):
        if "Price" in model:
            sigma = vol_input / 100.0
            return black76_bond_option(fwd_dirty, K_price, sigma, T, df_T, is_call=is_call_price)
        else:
            return bachelier_bond_option_yield(
                fwd_yield / 100.0, K_yield / 100.0,
                vol_input, T, df_T, mod_dur, fwd_dirty, is_payer_yield, face
            )

    if opt_type == "Receiver (Buy Bond)":
        res = price_one(is_call_price=True, is_payer_yield=False)
        results = {"Receiver": res}
    elif opt_type == "Payer (Sell Bond)":
        res = price_one(is_call_price=False, is_payer_yield=True)
        results = {"Payer": res}
    else:  # Straddle
        r_recv = price_one(is_call_price=True, is_payer_yield=False)
        r_payr = price_one(is_call_price=False, is_payer_yield=True)
        results = {"Receiver": r_recv, "Payer": r_payr}

    # ─── Results display ─────────────────────────────────────────────
    for name, res in results.items():
        st.markdown(f"#### {name}")
        if "Price" in model:
            prem_per100 = res["prem_per100"]
            prem_dollar = prem_per100 * face / 100.0
            prem_bp = prem_per100 * 100.0  # bp of face

            r1, r2, r3, r4, r5 = st.columns(5)
            r1.metric("Premium (per $100)", f"{prem_per100:.4f}")
            r2.metric("Premium (AUD)", f"${prem_dollar:,.0f}")
            r3.metric("Premium (bp face)", f"{prem_bp:.2f}bp")
            r4.metric("Delta (per $100 fwd)", f"{res['delta_price']:.4f}")
            r5.metric("Vega (per 1% vol)", f"${res['vega_per1pct'] * face / 100:,.0f}")

            st.caption(f"Gamma: {res['gamma']:.6f}  |  "
                       f"Theta: ${res['theta_per_day'] * face / 100:,.0f}/day  |  "
                       f"Strike: {K_price:.4f} (price)  |  df(T): {df_T:.6f}")
        else:
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Premium (AUD)", f"${res['prem_dollar']:,.0f}")
            r2.metric("PV (bp face)", f"{res['pv_bp']:.2f}bp")
            r3.metric("Delta ($/bp yield)", f"${res['delta_yield_bp']:,.0f}")
            r4.metric("Vega ($/bp vol)", f"${res['vega_per_bp']:,.0f}")
            st.caption(f"Strike: {K_yield:.4f}% (yield)  |  Fwd: {fwd_yield:.4f}%  |  "
                       f"┬ñ├ó: {vol_input:.1f}bp  |  T: {T:.4f}y  |  df(T): {df_T:.6f}")

    # ─── Vol smile / payoff chart ────────────────────────────────────
    st.markdown("---")
    chart_tab1, chart_tab2 = st.tabs(["📈 Premium vs Strike", "📊 Vol Sensitivity"])

    with chart_tab1:
        strikes_y = np.linspace(max(fwd_yield - 1.5, 0.1), fwd_yield + 1.5, 60)
        prems_recv, prems_payr = [], []
        for ky in strikes_y:
            kp = aofm_dirty_price(coupon_pct, maturity, ky, exp_date)
            if "Price" in model:
                s = vol_input / 100.0
                r_ = black76_bond_option(fwd_dirty, kp, s, T, df_T, is_call=True)
                p_ = black76_bond_option(fwd_dirty, kp, s, T, df_T, is_call=False)
                prems_recv.append(r_["prem_per100"] * face / 100.0)
                prems_payr.append(p_["prem_per100"] * face / 100.0)
            else:
                r_ = bachelier_bond_option_yield(fwd_yield/100, ky/100, vol_input, T, df_T, mod_dur, fwd_dirty, False, face)
                p_ = bachelier_bond_option_yield(fwd_yield/100, ky/100, vol_input, T, df_T, mod_dur, fwd_dirty, True, face)
                prems_recv.append(r_["prem_dollar"])
                prems_payr.append(p_["prem_dollar"])

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=list(strikes_y), y=prems_recv, name="Receiver",
                                  line=dict(color="#22c55e", width=2)))
        fig.add_trace(go.Scatter(x=list(strikes_y), y=prems_payr, name="Payer",
                                  line=dict(color="#ef4444", width=2)))
        fig.add_vline(x=fwd_yield, line_dash="dash", line_color="#94a3b8",
                      annotation_text=f"Fwd {fwd_yield:.3f}%")
        if strike_type == "Manual":
            fig.add_vline(x=K_yield, line_dash="dot", line_color="#f59e0b",
                          annotation_text=f"Strike {K_yield:.3f}%")
        fig.update_layout(title="Premium vs Strike Yield", xaxis_title="Strike Yield (%)",
                          yaxis_title="Premium (AUD)", template="plotly_dark", height=380,
                          legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig, use_container_width=True)

    with chart_tab2:
        # Vega profile vs vol
        vol_range = np.linspace(max(vol_input * 0.3, 0.5), vol_input * 2.5, 50)
        prems_v = []
        for v_ in vol_range:
            if "Price" in model:
                r_ = black76_bond_option(fwd_dirty, K_price, v_/100, T, df_T, is_call=True)
                prems_v.append(r_["prem_per100"] * face / 100.0)
            else:
                r_ = bachelier_bond_option_yield(fwd_yield/100, K_yield/100, v_, T, df_T, mod_dur, fwd_dirty, False, face)
                prems_v.append(r_["prem_dollar"])

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=list(vol_range), y=prems_v, name="Receiver Premium",
                                   line=dict(color="#3b82f6", width=2)))
        fig2.add_vline(x=vol_input, line_dash="dash", line_color="#94a3b8",
                       annotation_text="Current Vol")
        x_label = "Price Vol (%)" if "Price" in model else "Yield Vol (bp)"
        fig2.update_layout(title="Premium vs Volatility", xaxis_title=x_label,
                           yaxis_title="Premium (AUD)", template="plotly_dark", height=380)
        st.plotly_chart(fig2, use_container_width=True)

    # ─── Bond schedule ───────────────────────────────────────────────
    with st.expander("📅 Bond & Option Details"):
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**Bond Cashflows to Expiry (stripped)**")
            if coupons_stripped:
                df_cpn = pd.DataFrame([
                    {"Date": cd.strftime("%d %b %Y"), "Coupon (per $100)": f"{g_:.4f}",
                     "OIS DF": f"{pv_:.6f}", "PV (per $100)": f"{g_*pv_:.4f}"}
                    for cd, g_, pv_ in coupons_stripped
                ])
                st.dataframe(df_cpn, use_container_width=True, hide_index=True)
            else:
                st.info("No coupons between today and expiry.")
        with col_b:
            st.markdown("**Pricing Summary**")
            summary = {
                "Bond": f"{coupon_pct}% {maturity.strftime('%d %b %Y')}",
                "Face Value": f"AUD {face_mm:.0f}mm",
                "Settlement": settle.strftime("%d %b %Y"),
                "Option Expiry": exp_date.strftime("%d %b %Y"),
                "Time to Expiry": f"{T:.4f}y ({(exp_date-today).days}d)",
                "Spot Yield": f"{spot_yield:.4f}%",
                "Spot Dirty": f"{dirty_spot:.4f}",
                "Spot Clean": f"{clean_spot:.4f}",
                "Repo Rate": f"{repo_rate:.3f}%",
                "Fwd Dirty Price": f"{fwd_dirty:.4f}",
                "Fwd Yield": f"{fwd_yield:.4f}%",
                "Strike Yield": f"{K_yield:.4f}%",
                "Strike Price": f"{K_price:.4f}",
                "Mod Duration": f"{mod_dur:.3f}y",
                "DV01 (AUD)": f"${dv01:,.0f}",
                "OIS df(T)": f"{df_T:.6f}",
            }
            for k, v in summary.items():
                st.markdown(f"**{k}:** {v}")



def portfolio_tab():
    st.subheader("Trade Blotter  —  Swaptions + Caps & Floors")

    portfolio = st.session_state.get("portfolio", [])
    if not portfolio:
        st.info("Trade Blotter is empty. Price swaptions or caps/floors to add trades.")
        return

    _sw_port  = [t for t in portfolio if t.get("instrument_type","Swaption") == "Swaption"]
    _cf_port  = [t for t in portfolio if t.get("instrument_type") == "Cap/Floor"]

    _ptf_view = st.radio("View", ["📊 Swaptions", "🔔 Caps/Floors", "📋 Composite"],
                         horizontal=True, key="ptf_view_sel")

    if _ptf_view == "📊 Swaptions":
        if not _sw_port:
            st.info("No swaptions in portfolio.")
        else:
            df_sw = pd.DataFrame(_sw_port)
            df_sw["_expiry_sort"] = df_sw["expiry"].apply(lambda e: label_to_years(str(e)))
            df_sw = df_sw.sort_values("_expiry_sort").reset_index(drop=True)

            st.markdown("---")

            # Blotter — fast dataframe view
            _sw_disp_rows = []
            for idx, row in df_sw.iterrows():
                _sw_disp_rows.append({
                    "#": idx+1,
                    "Structure": row.get("structure",""),
                    "Expiry": row.get("expiry",""),
                    "Tenor": row.get("tenor",""),
                    "Notl (mm)": float(row.get("notional_mm",100)),
                    "Strike (%)": f"{float(row.get('strike',0)):.4f}",
                    "Fwd (%)": f"{float(row.get('forward',0)):.4f}",
                    "PV (bp)": round(float(row.get("pv_bp",0)),2),
                    "PV ($k)": round(float(row.get("pv",0))/1000,1),
                    "Δ ($k)": round(float(row.get("delta",0))/1000,1),
                    "Vega ($k)": round(float(row.get("vega",0))/1000,1),
                    "CCY": row.get("currency","AUD"),
                })
            _sw_disp_df = pd.DataFrame(_sw_disp_rows)
            _th = min(max(150, len(_sw_disp_df)*35+38), 500)
            st.dataframe(_sw_disp_df, use_container_width=True, height=_th, hide_index=True)

            # Single row detail + Quick Tix + delete
            _sw_sel_opts = [f"{i+1}: {t.get('label', t.get('expiry','?')+'x'+str(t.get('tenor','?')))}" for i,t in enumerate(_sw_port)]
            _sw_sel = st.selectbox("Select row for detail / Quick Tix / delete", ["  —  "] + _sw_sel_opts, key="ptf_sw_detail_sel")
            if _sw_sel != "  —  ":
                _si = int(_sw_sel.split(":")[0]) - 1
                _srow = _sw_port[_si]
                _struct = _srow.get("structure","")
                _legs   = _srow.get("legs",[]) if isinstance(_srow.get("legs",[]),list) else []
                _label  = _srow.get("label","")
                _pv_bp  = float(_srow.get("pv_bp",0))
                _pv     = float(_srow.get("pv",0))
                _delta  = float(_srow.get("delta",0))
                _ccy    = _srow.get("currency","AUD")
                _notl   = float(_srow.get("notional_mm",100))
                _strike = float(_srow.get("strike",0))
                _fwd    = float(_srow.get("forward",0))
                _expiry = _srow.get("expiry","")
                _tenor  = _srow.get("tenor","")

                _dc1, _dc2 = st.columns([4,1])
                with _dc2:
                    if st.button("🗑️ Delete", key="ptf_sw_del_btn"):
                        st.session_state["portfolio"] = [p for p in portfolio if p.get("label")!=_label or p.get("instrument_type")!="Swaption"]
                        st.session_state["swaption_portfolio"] = [p for p in st.session_state["swaption_portfolio"] if p.get("label")!=_label]
                        _save_portfolio(); st.rerun()

                if _struct in ["Risk Reversal","Payer Ladder","Receiver Ladder"] and _legs:
                    _ldf = pd.DataFrame([{
                        "Leg": l.get("name",""), "B/S": "Buy" if int(l.get("qty",1))>0 else "Sell",
                        "Strike (%)": f"{float(l.get('strike',0)):.4f}",
                        "PV ($k)": f"{float(l.get('pv',0))/1000:,.1f}",
                        "Delta ($k)": f"{float(l.get('delta',0))/1000:,.1f}",
                    } for l in _legs])
                    st.dataframe(_ldf, use_container_width=True, hide_index=True)

                if can_quick_tix():
                    with st.expander("📋 Quick Tix", expanded=False):
                        from datetime import date as _qdate
                        from dateutil.relativedelta import relativedelta as _qrd
                        try:
                            _qtoday = _qdate.today()
                            _exp_y  = label_to_years(str(_expiry))
                            _ten_y  = float(str(_tenor).replace("Y","").replace("y",""))
                            _exp_dt   = _qtoday + _qrd(days=int(_exp_y*365.25))
                            _start_dt = _exp_dt + _qrd(days=2)
                            _end_dt   = _start_dt + _qrd(months=int(_ten_y*12))
                            _rolls = [(_start_dt+_qrd(months=3*(i+1))).strftime('%d-%b-%Y') for i in range(min(4,int(_ten_y*4)))]
                            _delta_dir = "Pay Fixed" if _delta>0 else "Rec Fixed"
                            _tix = f"=== {_ccy} {_struct.upper()} ===\n"
                            _tix += f"Expiry:      {_exp_dt.strftime('%d-%b-%Y')} ({_expiry})\n"
                            _tix += f"Swap Start:  {_start_dt.strftime('%d-%b-%Y')}\n"
                            _tix += f"Swap End:    {_end_dt.strftime('%d-%b-%Y')}\n"
                            _tix += f"Rolls:       {', '.join(_rolls)}\n"
                            _tix += f"Fwd Rate:    {_fwd:.4f}%\n"
                            if _legs and _struct in ["Risk Reversal","Payer Ladder","Receiver Ladder"]:
                                for _lg in _legs:
                                    _s = "BUY " if int(_lg.get("qty",1))>0 else "SELL"
                                    _tix += f"{_s} {_lg.get('name',''):20s} K={float(_lg.get('strike',0)):.4f}%\n"
                            else:
                                _tix += f"Strike:      {_strike:.4f}%\n"
                            _tix += f"Premium:     {_pv_bp:.2f} bp  ({_pv/1000:+,.1f}k)\n"
                            _tix += f"Net Delta:   {_delta/1000:,.1f}k  →  {_delta_dir}\n"
                            _tix += f"Notional:    {_notl:.0f}mm {_ccy}"
                            st.code(_tix, language=None)
                        except Exception as _qe:
                            st.caption(f"Tix error: {_qe}")

            # Reload into pricer
            st.markdown("##### Reload into Pricer")
            _sw_opts = [f"{i}: {t.get('label','?')}" for i,t in enumerate(_sw_port)]
            _rsel = st.selectbox("Ticket", ["  —  "] + _sw_opts, key="ptf_sw_reload")
            if st.button("🔄 Reload", key="ptf_sw_reload_btn") and _rsel != "  —  ":
                _ri = int(_rsel.split(":")[0])
                _rrow = _sw_port[_ri]
                _EP = ["1w","2w","1m","2m","3m","6m","9m","1y","18m","2y","3y","5y","7y","10y","12y","15y","20y","📅 Custom Date..."]
                _TO = ["1Y","2Y","3Y","4Y","5Y","6Y","7Y","8Y","9Y","10Y","12Y","15Y","20Y","25Y","30Y"]
                _ev = str(_rrow.get("expiry","5y")); _tv = str(_rrow.get("tenor","5Y"))
                st.session_state["sw_pending_reload"] = {
                    "expiry": _ev if _ev in _EP else "5y",
                    "tenor": _tv if _tv in _TO else "5Y",
                    "structure": str(_rrow.get("structure","ATM Straddle")),
                    "notional_mm": float(_rrow.get("notional_mm",100)),
                    "strike": float(_rrow["strike"]) if "strike" in _rrow else None,
                }
                st.rerun()

    elif _ptf_view == "🔔 Caps/Floors":
        if not _cf_port:
            st.info("No caps/floors in portfolio.")
        else:
            df_cf = pd.DataFrame(_cf_port)
            st.markdown("---")
            _cf_cols = ["instrument_type","currency","structure","first_fixing","tenor","notional_mm",
                        "strike","pv_bp","pv","delta","vega"]
            _cf_disp = df_cf[[c for c in _cf_cols if c in df_cf.columns]].copy()
            if "pv" in _cf_disp: _cf_disp["pv"] = (_cf_disp["pv"]/1000).round(1)
            if "delta" in _cf_disp: _cf_disp["delta"] = (_cf_disp["delta"]/1000).round(1)
            if "vega" in _cf_disp: _cf_disp["vega"] = (_cf_disp["vega"]/1000).round(1)
            if "pv_bp" in _cf_disp: _cf_disp["pv_bp"] = _cf_disp["pv_bp"].round(2)
            st.dataframe(_cf_disp, use_container_width=True)
            if st.button("🗑️ Clear Cap/Floor", key="ptf_cf_clear"):
                st.session_state["portfolio"] = [p for p in portfolio if p.get("instrument_type")!="Cap/Floor"]
                _save_portfolio(); st.rerun()

    elif _ptf_view == "📋 Composite":
        st.markdown("### Combined Portfolio — Swaptions + Caps/Floors")
        st.markdown("---")

        # Full blotter
        _all_rows = []
        for t in portfolio:
            _all_rows.append({
                "Type": t.get("instrument_type","Swaption"),
                "CCY": t.get("currency","AUD"),
                "Structure": t.get("structure",""),
                "Expiry": t.get("expiry","") or t.get("first_fixing",""),
                "Tenor": t.get("tenor",""),
                "Notl (mm)": float(t.get("notional_mm",100)),
                "Strike (%)": f"{float(t.get('strike',0)):.4f}",
                "Fwd (%)": f"{float(t.get('forward',0)):.4f}",
                "PV (bp)": f"{float(t.get('pv_bp',0)):.2f}",
                "PV ($k)": f"{float(t.get('pv',0))/1000:,.1f}",
                "Δ ($k)": f"{float(t.get('delta',0))/1000:,.1f}",
                "Vega ($k)": f"{float(t.get('vega',0))/1000:,.1f}",
            })
        _comp_df = pd.DataFrame(_all_rows)
        _th = min(max(200, len(_comp_df)*35+38), 600)
        st.dataframe(_comp_df, use_container_width=True, height=_th)

        # Download
        _csv = _comp_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download CSV", data=_csv,
                           file_name="RateEdge_portfolio.csv", mime="text/csv", key="ptf_comp_dl")

        # Clear all
        if st.button("🗑️ Clear All", key="ptf_comp_clear"):
            st.session_state["portfolio"] = []
            st.session_state["swaption_portfolio"] = []
            _save_portfolio(); st.rerun()


def home_tab():
    # Get theme colors for inline styling
    is_dark = st.session_state.get("theme_name", "Dealer Dark") == "Dealer Dark"
    text_color = "#f1f5f9" if is_dark else "#1e3a5f"
    muted_color = "#94a3b8" if is_dark else "#64748b"
    accent_color = "#ef4444" if is_dark else "#dc2626"
    card_bg = "#1e293b" if is_dark else "#ffffff"
    border_color = "#334155" if is_dark else "#e2e8f0"
    
    st.markdown(
        f"""
        <div style="background:{card_bg};border:1px solid {border_color};border-radius:16px;padding:2rem;text-align:center;margin-bottom:2rem;">
            <div style="font-size:2.5rem;font-weight:700;color:{text_color};margin-bottom:0.5rem;">
                <span style="color:#1e3a5f;">Rate</span><span style="color:{accent_color};">Edge</span> Options
            </div>
            <div style="font-size:1.1rem;color:{muted_color};margin-bottom:1rem;">
                Professional Interest Rate Derivatives Pricing
            </div>
            <div style="font-size:0.95rem;color:{muted_color};">
                Swaptions  Caps/Floors  Exotics  CVA  RV Analysis
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    # Login form - REMOVED FOR LOCAL TESTING
    
    # Features section
    st.markdown("---")
    st.markdown(f"###  Platform Features")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(
            f"""
            <div style="background:{card_bg};border:1px solid {border_color};border-radius:12px;padding:1.25rem;">
                <div style="color:{accent_color};font-weight:600;margin-bottom:0.5rem;"> Vol Surfaces</div>
                <div style="color:{muted_color};font-size:0.85rem;">
                    ATM vol grids, SABR smile modeling, 3D surface editor
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    
    with col2:
        st.markdown(
            f"""
            <div style="background:{card_bg};border:1px solid {border_color};border-radius:12px;padding:1.25rem;">
                <div style="color:{accent_color};font-weight:600;margin-bottom:0.5rem;"> Curve Analytics</div>
                <div style="color:{muted_color};font-size:0.85rem;">
                    Forward swap calculation, curve building, risk analytics
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    
    with col3:
        st.markdown(
            f"""
            <div style="background:{card_bg};border:1px solid {border_color};border-radius:12px;padding:1.25rem;">
                <div style="color:{accent_color};font-weight:600;margin-bottom:0.5rem;"> Multi-Currency</div>
                <div style="color:{muted_color};font-size:0.85rem;">
                    AUD, NZD, USD with correct market conventions
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    
    # Market conventions - Swaption + Underlying Swap by Currency
    st.markdown("---")
    st.markdown(f"###  Market Conventions by Currency")
    
    conv_tabs = st.tabs([" AUD", " NZD", " USD", " EUR", " GBP", " JPY", " CAD"])
    
    with conv_tabs[0]:  # AUD
        st.markdown("#### AUD Swaption Conventions *(AFMA IRD & IRO Conventions, June 2025)*")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical   —   LCH Cleared Swap |
            | Premium Quote | Basis points of notional face value |
            | Premium Payment | Same day as expiry |
            | Exercise | European   —   10:00am AEST |
            | Vol Quote | Normal (bp/annum) |
            | Discount Curve | AONIA OIS |
            | Expiry Style | Quarterly ≤3y, semi-annual ≥4y |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Q/Q to 3y, S/S beyond, ACT/365F |
            | Float Leg | 3M BBSW (Q/Q) or 6M BBSW (S/S) |
            | Float Index | 3M BBSW / AONIA (OIS) |
            | Spot Lag | T+1 |
            | Roll | Modified Following |
            | Business Days | Sydney (NSW bank holidays) |
            | Discounting | OIS (AONIA) in all cases |
            | Clearing | LCH Cleared Swaps (default) |
            """)

        st.markdown("**Swaption Date Conventions** *(AFMA IRO 3.12)*")
        st.markdown("""
        | | Swaptions |
        |---|---|
        | **Expiry Date** | T + Tenor (FOLL) |
        | **Start Date** | Expiry Date + Swap Spot Offset (1BD) (FOLL) |
        | **End Date** | Start Date + Underlying Tenor (MODFOLL) |
        | **Roll Date** | Underlying Start Date (MODFOLL) |
        | **Termination Option Notice** | n/a |
        """)

        st.markdown("#### Standard Transaction Size   —   Swaptions (AUD million) *(AFMA IRO 3.4)*")
        st.markdown("""
        | Expiry \\ Tenor | 1Y | 2Y | 3Y | 4Y | 5Y | 7Y | 10Y | 15Y | 20Y | 30Y |
        |---|---|---|---|---|---|---|---|---|---|---|
        | 1m | 200 | 200 | 100 | 50 | 50 | 50 | 25 | 25 | 10 | 10 |
        | 3m | 200 | 200 | 100 | 50 | 50 | 50 | 25 | 15 | 10 | 10 |
        | 6m | 200 | 100 | 100 | 50 | 50 | 50 | 25 | 15 | 10 | 10 |
        | 1y | 100 | 100 | 100 | 50 | 50 | 50 | 25 | 15 | 10 | 10 |
        | 2y | 100 | 100 | 75 | 50 | 50 | 25 | 25 | 15 | 10 | 10 |
        | 3y | 100 | 75 | 50 | 50 | 50 | 25 | 25 | 15 | 10 | 10 |
        | 4y | 75 | 75 | 50 | 50 | 50 | 25 | 25 | 15 | 10 | 10 |
        | 5y | 50 | 50 | 50 | 50 | 50 | 25 | 25 | 15 | 10 | 10 |
        | 7y | 50 | 50 | 25 | 25 | 25 | 25 | 25 | 15 | 10 | 10 |
        | 10y | 50 | 25 | 25 | 25 | 25 | 25 | 25 | 15 | 10 | 10 |
        """)
        st.caption("Bermuda Swaption minimum: AUD 10 million.")

        st.markdown("---")
        st.markdown("#### AUD Cap / Floor Conventions *(AFMA IRO Conventions, June 2025)*")
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            st.markdown("""
            **Structure**
            | Convention | Standard |
            |------------|----------|
            | Caplet/Floorlet | Series of individual options |
            | Floating Index | 3M BBSW (quarterly resets) |
            | Reset Frequency | Quarterly |
            | Day Count | ACT/365 Fixed |
            | Business Days | Sydney (NSW bank holidays) |
            | Roll | Modified Following |
            """)
        with cf2:
            st.markdown("""
            **Quotation & Premium**
            | Convention | Standard |
            |------------|----------|
            | Premium Quote | Basis points of notional face value |
            | Vol Quote | Normal vol (bp/annum) |
            | Premium Payment | T+2 |
            | Settlement Style | Non-discounted, paid in arrears |
            | ATM Reference | Forward swap rate for underlying tenor |
            | CSA Default | AUD   —   premium adj agreed post-trade |
            | FRA Yield Discounting | Must be excluded in confirm ≥ |
            """)
        with cf3:
            st.markdown("""
            **Settlement & Exercise**
            | Convention | Standard |
            |------------|----------|
            | Settlement Index | BBSW (Refinitiv "BBSW" page) |
            | Caplet Settlement | Non-discounted in arrears |
            | Payment Date | Reference date (end of accrual period) |
            | Exercise | Automatic at expiry |
            | Spot Lag | T+1 |
            | ISDA Note | Confirm "FRA Yield Discounting will not apply" ≥ |
            """)

        st.markdown("**Cap / Floor Date Conventions** *(AFMA IRO 3.12)*")
        st.markdown("""
        | | Caps / Floors |
        |---|---|
        | **Expiry Date** | n/a |
        | **Start Date** | T + Start Tenor (FOLL) + Swap Spot Offset (1BD) (FOLL) |
        | **End Date** | Start Date + Tenor (MODFOLL) |
        | **Roll Date** | Start Date (MODFOLL) |
        | **Termination Option Notice** | n/a |
        """)

        st.markdown("#### Standard Transaction Size   —   Cap/Floor Straddles (AUD million) *(AFMA IRO 3.4.2)*")
        st.markdown("""
        | Tenor | CFS Notional (A$m) |
        |-------|-------------------|
        | 1Y | 200 |
        | 2Y | 100 |
        | 3Y | 50 |
        | 4Y | 35 |
        | 5Y | 25 |
        | 6Y | 20 |
        | 7Y | 20 |
        | 8Y | 15 |
        | 9Y | 15 |
        | 10Y | 15 |
        """)
        st.caption(
            "≥ AFMA IRO Conventions (June 2025, s4.3.1): where ISDA 2021 Definitions are incorporated, "
            "FRA Yield Discounting applies by default to AUD cap/floor/collar transactions. "
            "Confirmations must explicitly state 'FRA Yield Discounting will not apply' because these products pay in arrears. "
            "In AUD interdealer CFS practice this is typically handled via explicit settlement terms in the bilateral confirm "
            "rather than the ISDA election   —   but must be verified bilaterally. "
            "**⚙️ WPO: confirm FRA Yield Discounting election approach with each client at deployment.**"
        )
        st.caption("Source: AFMA Interest Rate Options Conventions (June 2025) & AFMA Interest Rate Derivative Conventions (June 2025)")
    
    with conv_tabs[1]:  # NZD
        st.markdown("#### NZD Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical delivery |
            | Cash Settlement | Par rate (rare) |
            | Premium | Forward premium (T+2) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | BKBM/OCR curve |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Q/Q to 2y, S/S beyond, ACT/365F |
            | Float Leg | Quarterly |
            | Float Index | 3M BKBM / OCR (OIS) |
            | Spot Lag | T+2 |
            | Roll | Modified Following |
            | Calendar | Auckland/Wellington |
            """)
    
    with conv_tabs[2]:  # USD
        st.markdown("#### USD Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical delivery (standard) |
            | Cash Settlement | Par rate method (ISDA 2006) |
            | Premium | Forward premium (T+2) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | Actual swap curve |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Semi-annual, 30/360 |
            | Float Leg | Quarterly, ACT/360 |
            | Float Index | SOFR (Term or compounded) |
            | Spot Lag | T+2 |
            | Roll | Modified Following |
            | Calendar | NYC |
            """)
    
    with conv_tabs[3]:  # EUR
        st.markdown("#### EUR Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Cash-settled (standard) |
            | Cash Settlement | Par rate, ISDA 2006 |
            | Premium | Forward premium (T+2) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | Curve-based |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Annual, 30/360 |
            | Float Leg | Annual (was 6M) |
            | Float Index | ├ö├®┬╝STR |
            | Spot Lag | T+2 |
            | Roll | Modified Following |
            | Calendar | TARGET |
            """)
    
    with conv_tabs[4]:  # GBP
        st.markdown("#### GBP Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical (LCH clearable) |
            | Cash Settlement | Par rate method |
            | Premium | Forward premium (T+0) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | Curve-based |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Annual, ACT/365F |
            | Float Leg | Annual |
            | Float Index | SONIA compounded |
            | Spot Lag | T+0 |
            | Roll | Modified Following |
            | Calendar | London |
            """)
    
    with conv_tabs[5]:  # JPY
        st.markdown("#### JPY Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical or Cash |
            | Cash Settlement | Yield-settled (JSCC) |
            | Premium | Forward premium (T+2) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | JGB curve-based |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Semi-annual, ACT/365F |
            | Float Leg | Semi-annual |
            | Float Index | TONA |
            | Spot Lag | T+2 |
            | Roll | Modified Following |
            | Calendar | Tokyo |
            """)
    
    with conv_tabs[6]:  # CAD
        st.markdown("#### CAD Swaption Conventions")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **Swaption Specifics**
            | Convention | Standard |
            |------------|----------|
            | Settlement | Physical (standard) |
            | Cash Settlement | Par rate method |
            | Premium | Forward premium (T+1) |
            | Exercise | European |
            | Vol Quote | Normal (bp/annum) |
            | Annuity | CORRA curve |
            """)
        with col2:
            st.markdown("""
            **Underlying Swap**
            | Convention | Standard |
            |------------|----------|
            | Fixed Leg | Semi-annual, ACT/365F |
            | Float Leg | Semi-annual |
            | Float Index | CORRA |
            | Spot Lag | T+1 |
            | Roll | Modified Following |
            | Calendar | Toronto |
            """)


def generate_basis_matrix(ccy: str, basis_6v3: pd.DataFrame) -> pd.DataFrame:
    """Generate basis matrix - wrapper for cached version"""
    basis_tuple = tuple(basis_6v3["MaturityY"].tolist()), tuple(basis_6v3["BasisBp"].tolist())
    return _generate_basis_matrix_cached(ccy, basis_tuple)


@st.cache_data(ttl=3600, show_spinner=False)
def _generate_basis_matrix_cached(ccy: str, basis_tuple: tuple) -> pd.DataFrame:
    """Generate basis matrix interpolated across expiry/tenor grid - CACHED"""
    expiries = ["1w", "1m", "2m", "3m", "6m", "9m", "1y", "18m", "2y", "3y", "4y", "5y", "6y", "7y", "8y", "9y", "10y", "12y", "15y", "20y", "25y", "30y"]
    tenors = ["1Y", "2Y", "3Y", "4Y", "5Y", "7Y", "10Y", "12Y", "15Y", "20Y", "25Y", "30Y"]
    
    basis_x = np.array(basis_tuple[0])
    basis_y = np.array(basis_tuple[1])
    
    matrix = []
    for exp in expiries:
        exp_y = label_to_years(exp)
        row = {"Expiry": exp}
        for tenor in tenors:
            tenor_y = float(tenor[:-1])
            mid_point = exp_y + tenor_y / 2
            basis_bp = float(np.interp(mid_point, basis_x, basis_y))
            row[tenor] = round(basis_bp, 4)
        matrix.append(row)
    
    df = pd.DataFrame(matrix).set_index("Expiry")
    return df


def calculate_atm_premium_matrix(ccy: str, curve: pd.DataFrame, atm_vols: pd.DataFrame,
                                  basis_6v3: Optional[pd.DataFrame] = None):
    """Calculate ATM straddle premiums and vega matrix using real annuities from curve.
    Returns: (prem_df, vega_df)
      prem_df    —   ATM straddle forward premium in bp
      vega_df    —   Vega in $ per 1bp vol move, scaled to 100mm notional
    """
    if "Expiry" not in atm_vols.columns:
        empty = pd.DataFrame()
        return empty, empty

    expiries = atm_vols["Expiry"].tolist()
    tenors = [c for c in atm_vols.columns if c != "Expiry"]

    _ois_cb = st.session_state.get("config_basis", {}).get(ccy, {}).get("ois")
    ois_curve = _ois_cb if _ois_cb is not None else get_basis_curve(ccy, "ois")

    prem_rows = []
    vega_rows = []

    for i, exp in enumerate(expiries):
        try:
            exp_y = label_to_years(exp)
        except Exception:
            continue
        if exp_y is None or exp_y <= 0:
            continue
        prow = {"Expiry": exp}
        vrow = {"Expiry": exp}

        for tenor in tenors:
            try:
                vol_bp = atm_vols.iloc[i][tenor]
                if pd.isna(vol_bp):
                    prow[tenor] = None
                    vrow[tenor] = None
                    continue

                tenor_y = label_to_years(tenor)
                _, ann, _ = forward_and_annuity_from_curve(curve, ccy, exp_y, tenor_y, ois_curve)

                sigma_n = vol_bp / 10000.0
                sqrt_t = math.sqrt(max(exp_y, 0.001))

                # ATM straddle FORWARD premium (bp of notional)
                # spot_prem = 2*N'(0)*sigma*sqrt(T)*annuity
                # fwd_prem  = spot_prem / df(expiry)  [market convention: OIS discounted]
                xs_c = curve["MaturityY"].to_numpy().astype(float)
                ys_c = curve["ZeroRatePct"].to_numpy().astype(float) / 100.0
                df_expiry = math.exp(-float(np.interp(exp_y, xs_c, ys_c)) * exp_y)
                spot_prem_bp = 2 * 0.3989 * sigma_n * sqrt_t * ann * 10000
                fwd_prem_bp = spot_prem_bp / df_expiry if df_expiry > 0 else spot_prem_bp
                prow[tenor] = round(fwd_prem_bp, 2)

                # Vega: d(fwd_prem $) / d(vol in bp), scaled to 100mm notional
                d_spot_prem_per_bp = 2 * 0.3989 * sqrt_t * ann
                d_fwd_prem_per_bp = d_spot_prem_per_bp / df_expiry if df_expiry > 0 else d_spot_prem_per_bp
                vega_dollars = (d_fwd_prem_per_bp / 10000.0) * 100e6
                vrow[tenor] = round(vega_dollars, 0)

            except:
                prow[tenor] = None
                vrow[tenor] = None

        prem_rows.append(prow)
        vega_rows.append(vrow)

    prem_df = pd.DataFrame(prem_rows).set_index("Expiry")
    vega_df = pd.DataFrame(vega_rows).set_index("Expiry")
    return prem_df, vega_df


# Keep old cached version for backwards compat but it's not used anymore
@st.cache_data(ttl=3600, show_spinner=False)
def _calculate_atm_premium_matrix_cached(ccy: str, expiries: tuple, tenors: tuple, vol_data: tuple) -> pd.DataFrame:
    """DEPRECATED - kept for backwards compatibility"""
    matrix = []
    for i, exp in enumerate(expiries):
        exp_y = label_to_years(exp)
        row = {"Expiry": exp}
        for j, tenor in enumerate(tenors):
            try:
                vol_bp = vol_data[i][j]
                if vol_bp == -999:
                    row[tenor] = None
                    continue
                tenor_y = label_to_years(tenor)
                annuity = tenor_y * 0.85
                sigma_n = vol_bp / 10000.0
                sqrt_t = math.sqrt(max(exp_y, 0.001))
                fwd_premium = 2 * 0.3989 * sigma_n * sqrt_t * annuity
                premium_bp = fwd_premium * 10000
                row[tenor] = round(premium_bp, 2)
            except:
                row[tenor] = None
        matrix.append(row)
    df = pd.DataFrame(matrix).set_index("Expiry")
    return df


def clear_matrix_cache():
    """Clear all cached matrix data"""
    _generate_forward_matrix_cached.clear()
    _generate_basis_matrix_cached.clear()
    _calculate_atm_premium_matrix_cached.clear()
    st.session_state.pop("_fwd_ann_cache", None)


# ============================
# Main
# ============================

def main():
    st.set_page_config(
        page_title="RateEdge Options",
        layout="wide",
        page_icon="📊",
        initial_sidebar_state="expanded"
    )
    init_session()
    
    # Auto-load curves + vols from DB on first load after login
    if HAS_POSTGRES and get_db_url() and not st.session_state.get("db_auto_loaded", False):
        user_id = st.session_state.get("username", "default")
        if user_id and user_id != "default":
            _ADMIN_EMAILS = {"wpo70@icloud.com", "wpo@rateedge.au"}
            try:
                _role_conn = get_db_connection()
                if _role_conn:
                    _role_cur = _role_conn.cursor()
                    _role_cur.execute("SELECT role FROM user_roles WHERE email=%s", (user_id,))
                    _role_row = _role_cur.fetchone()
                    _role_cur.close()
                    _role_conn.close()
                    st.session_state["user_role"] = "super_admin" if user_id in _ADMIN_EMAILS else (
                        _role_row[0] if _role_row else "read_only"
                    )
            except Exception:
                if user_id in _ADMIN_EMAILS:
                    st.session_state["user_role"] = "super_admin"
            # ── Step 1: Load ATM vol surfaces from vol_history FIRST ──
            # Must happen before load_all_session_data so ATM is correct
            # when SABR params are applied
            try:
                _sc = get_db_connection()
                if _sc:
                    _cur = _sc.cursor()
                    _sl = []
                    for _cy in SUPPORTED_CURRENCIES:
                        _cur.execute("""
                            SELECT id FROM vol_history 
                            WHERE currency=%s AND atm_vols IS NOT NULL
                            ORDER BY snapshot_date DESC LIMIT 1
                        """, (_cy,))
                        _row = _cur.fetchone()
                        if _row:
                            _cur2 = _sc.cursor()
                            _cur2.execute("SELECT currency,atm_vols,sabr_alpha,sabr_beta,sabr_rho,sabr_nu,label,snapshot_date FROM vol_history WHERE id=%s",(_row[0],))
                            _srow = _cur2.fetchone()
                            _cur2.close()
                            if _srow:
                                _cc2,_av,_sa,_sb,_sr,_sn,_lbl,_dt = _srow
                                if "vol_data" not in st.session_state: st.session_state["vol_data"] = {}
                                if _cc2 not in st.session_state["vol_data"]: st.session_state["vol_data"][_cc2] = {}
                                if _av:
                                    _df = pd.DataFrame(_av["values"])
                                    if "Expiry" in _df.columns: _df = _df[["Expiry"]+[c for c in _df.columns if c!="Expiry"]]
                                    st.session_state["vol_data"][_cc2]["atm"] = _df
                                    if "vol_editor" not in st.session_state: st.session_state["vol_editor"]={"working":{},"base":{},"history":{},"future":{},"redo_stack":{}}
                                    st.session_state["vol_editor"]["base"][_cc2] = _df.copy()
                                    st.session_state["vol_editor"]["working"][_cc2] = _df.copy()
                                for _pm,_pd in [("alpha",_sa),("beta",_sb),("rho",_sr),("nu",_sn)]:
                                    if _pd and "values" in _pd:
                                        try: st.session_state["vol_data"][_cc2][_pm] = pd.DataFrame(_pd["values"])
                                        except: pass
                                _vd = st.session_state["vol_data"][_cc2]
                                if _vd.get("alpha") is None and _vd.get("atm") is not None:
                                    try:
                                        _ar = _vd["atm"].copy()
                                        _tc = [c for c in _ar.columns if c!="Expiry"]
                                        for _pp,_dv in [("beta",0.5),("rho",0.20),("nu",0.30)]:
                                            _dp = _ar[["Expiry"]].copy()
                                            for _t in _tc: _dp[_t]=_dv
                                            _vd[_pp]=_dp
                                        _da = _ar.copy()
                                        for _t in _tc: _da[_t]=_da[_t]/10000.0
                                        _vd["alpha"]=_da
                                    except: pass
                                _sl.append(f"{_cc2}:{_lbl}")
                                st.session_state[f"_vol_loaded_{_cc2}"] = True
                                _h = st.session_state.get(f"_atm_hash_{_cc2}", 0)
                                st.session_state[f"_atm_hash_{_cc2}"] = _h + 1
                    _cur.close(); _sc.close()
                    if _sl: st.session_state["_auto_load_msg"] = f"✅ Vols: {', '.join(_sl)}"
            except Exception as _ve:
                st.session_state["_auto_load_msg"] = f"⚠️ Vol load error: {_ve}"

            # ── Step 2: Load SABR/curves/spreads from user_configs ──
            try:
                _auto_loaded = load_all_session_data(user_id)
                if _auto_loaded > 0:
                    st.session_state["_auto_load_msg"] = st.session_state.get("_auto_load_msg","") + f" | Configs: {_auto_loaded}"
            except Exception as _ale:
                st.session_state["_auto_load_msg"] = st.session_state.get("_auto_load_msg","") + f" | Config load error: {_ale}"
            # Load portfolio scratchpad for this user
            try:
                _saved_port = _load_portfolio()
                if _saved_port:
                    st.session_state["portfolio"] = _saved_port
                    st.session_state["swaption_portfolio"] = [t for t in _saved_port if t.get("instrument_type","Swaption") == "Swaption"]
            except: pass
            st.session_state["db_auto_loaded"] = True

    # Sidebar for settings
    with st.sidebar:
        st.markdown(
            """
            <div style="text-align:center;padding:1rem 0;border-bottom:1px solid #334155;margin-bottom:1rem;">
                <div style="font-size:1.4rem;font-weight:700;">
                    <span style="color:#1e3a5f;">Rate</span><span style="color:#ef4444;">Edge</span>
                </div>
                <div style="font-size:0.75rem;color:#94a3b8;">Options Platform v3107j</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
        st.markdown("###  Settings")
        
        # Theme
        theme_choice = st.selectbox(
            " Theme",
            ["Dealer Dark", "Clean Light"],
            index=0 if st.session_state.get("theme_name", "Dealer Dark") == "Dealer Dark" else 1,
            key="sidebar_theme",
        )
        st.session_state["theme_name"] = theme_choice
        
        # Currency
        ccy = st.selectbox(
            " Currency",
            SUPPORTED_CURRENCIES,
            index=0,
            key="sidebar_ccy",
        )
        
        # Vol mode
        vol_mode = st.selectbox(
            " Vol Mode",
            ["Normal (bp)", "Black (lognormal)"],
            index=0,
            key="sidebar_volmode",
        )
        
        st.markdown("---")
        
        # User status
        if st.session_state.get("authenticated"):
            st.markdown(
                f"""
                <div style="background:#1e3a5f;padding:0.75rem;border-radius:8px;margin-top:0.5rem;">
                    <div style="color:#94a3b8;font-size:0.7rem;">Logged in as</div>
                    <div style="color:#22c55e;font-weight:600;">{st.session_state.get('username', 'User')}</div>
                    <div style="color:#94a3b8;font-size:0.65rem;margin-top:2px;">{st.session_state.get('user_role','read_only').upper()}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button(" Logout", key="logout_btn", use_container_width=True):
                _save_portfolio()  # Persist portfolio before logout
                st.session_state["authenticated"] = False
                st.session_state["username"] = None
                st.session_state["db_auto_loaded"] = False
                st.rerun()

            # ── Super Admin: User Role Management ────────────────────
            if is_super_admin():
                st.markdown("---")
                with st.expander("👑 User Management", expanded=False):
                    try:
                        _rm_conn = get_db_connection()
                        if _rm_conn:
                            _rm_cur = _rm_conn.cursor()
                            _rm_cur.execute("SELECT email, role FROM user_roles ORDER BY email")
                            _rm_rows = _rm_cur.fetchall()
                            _rm_cur.close()
                            _rm_conn.close()
                            if _rm_rows:
                                for _rm_email, _rm_role in _rm_rows:
                                    _rc1, _rc2 = st.columns([2, 1])
                                    with _rc1:
                                        st.caption(_rm_email)
                                    with _rc2:
                                        _new_role = st.selectbox(
                                            "Role", ["super_admin","admin","user","read_only","trainee"],
                                            index=["super_admin","admin","user","read_only","trainee"].index(_rm_role) if _rm_role in ["super_admin","admin","user","read_only","trainee"] else 2,
                                            key=f"role_{_rm_email}",
                                            label_visibility="collapsed"
                                        )
                                    if _new_role != _rm_role:
                                        if st.button(f"Save {_rm_email[:15]}", key=f"save_role_{_rm_email}"):
                                            _uc = get_db_connection()
                                            if _uc:
                                                _uc2 = _uc.cursor()
                                                _uc2.execute("UPDATE user_roles SET role=%s WHERE email=%s", (_new_role, _rm_email))
                                                _uc.commit(); _uc2.close(); _uc.close()
                                                st.success(f"✅ {_rm_email} → {_new_role}")
                                                st.rerun()
                    except Exception as _rme:
                        st.caption(f"Role mgmt error: {_rme}")
        else:
            st.warning(" Login required")
            st.caption("Use the main login page to sign in with your email")
        
        st.markdown("---")

        # Technical Support
        with st.expander("🛠️ Technical Support", expanded=False):
            _user_email = st.session_state.get("username", "")

            _issue_type = st.selectbox(
                "Issue type",
                ["Pricing Issue", "Curve / Data Issue", "Vol Surface Issue",
                 "CFS / Wedge Issue", "SOD Report Issue", "Login / Access Issue",
                 "Performance / Loading", "Other"],
                key="support_issue_type"
            )
            _severity = st.selectbox(
                "Severity",
                ["🔴 Critical — platform unusable", "🔴 High — major feature broken",
                 "🟡 Medium — partial functionality affected", "🟢 Low — cosmetic / minor"],
                key="support_severity"
            )
            _support_desc = st.text_area(
                "Describe the issue",
                placeholder="What happened? What were you pricing? Any error messages?",
                height=80, key="support_desc"
            )

            if st.button("📨 Submit Support Ticket", key="submit_support_btn", type="primary", use_container_width=True):
                if not _support_desc.strip():
                    st.warning("Please describe the issue first.")
                else:
                    _sev_label = _severity.split("—")[0].strip()
                    _subj = f"RateEdge Support — {_issue_type} [{_sev_label}]"
                    _body_html = f"""
                    <h2>RateEdge Support Ticket</h2>
                    <table>
                    <tr><td><b>User</b></td><td>{_user_email}</td></tr>
                    <tr><td><b>Issue Type</b></td><td>{_issue_type}</td></tr>
                    <tr><td><b>Severity</b></td><td>{_sev_label}</td></tr>
                    <tr><td><b>Description</b></td><td>{_support_desc.replace(chr(10), '<br>')}</td></tr>
                    </table>
                    """
                    _body_text = f"User: {_user_email}\nIssue: {_issue_type}\nSeverity: {_sev_label}\n\n{_support_desc}"
                    _sms_body  = f"RateEdge Support [{_sev_label}] from {_user_email}: {_issue_type} — {_support_desc[:120]}"

                    _sent_email = False
                    _sent_sms   = False

                    try:
                        _resend_key = st.secrets.get("RESEND_API_KEY", os.environ.get("RESEND_API_KEY", ""))
                        if _resend_key:
                            import requests as _req
                            _er = _req.post(
                                "https://api.resend.com/emails",
                                headers={"Authorization": f"Bearer {_resend_key}", "Content-Type": "application/json"},
                                json={"from": "support@rateedge.au",
                                      "to": ["wpo@rateedge.au", "wpo70@icloud.com"],
                                      "reply_to": _user_email,
                                      "subject": _subj, "html": _body_html, "text": _body_text},
                                timeout=10
                            )
                            _sent_email = _er.status_code in (200, 201)

                        # SMS via Twilio — uses env vars on Render
                        _tw_sid   = st.secrets.get("TWILIO_SID",   os.environ.get("TWILIO_SID",   "ACbb0ff7ea8f12c1386d4553977b8d8db7"))
                        _tw_token = st.secrets.get("TWILIO_TOKEN", os.environ.get("TWILIO_TOKEN", ""))
                        _tw_from  = st.secrets.get("TWILIO_FROM",  os.environ.get("TWILIO_FROM",  "+12602979976"))
                        _tw_to    = os.environ.get("TWILIO_TO", "+61478829669")
                        if _tw_sid and _tw_token:
                            import requests as _req, base64 as _b64
                            _tw_auth = _b64.b64encode(f"{_tw_sid}:{_tw_token}".encode()).decode()
                            _sr = _req.post(
                                f"https://api.twilio.com/2010-04-01/Accounts/{_tw_sid}/Messages.json",
                                headers={"Authorization": f"Basic {_tw_auth}"},
                                data={"To": _tw_to, "From": _tw_from, "Body": _sms_body},
                                timeout=10
                            )
                            _sent_sms = _sr.status_code in (200, 201)
                        if not _resend_key and not _tw_token:
                            st.warning("RESEND_API_KEY / TWILIO_TOKEN not configured on Render.")
                    except Exception as _se:
                        st.error(f"Failed to send: {_se}")

                    if _sent_email:
                        st.success("✅ Support ticket sent to wpo@rateedge.au")
                    if _sent_sms:
                        st.success("✅ SMS alert sent")
                    if not _sent_email and not _sent_sms:
                        st.info(f"📧 Email manually: wpo@rateedge.au\nSubject: {_subj}")

        st.markdown("---")
        # User Management (super_admin only)
        if is_admin():
            with st.expander("👥 User Access", expanded=False):
                st.caption("Manage user roles")
                if st.button("🔄 Load Users", key="load_users_btn"):
                    st.session_state["_user_list_loaded"] = True
                if HAS_POSTGRES and st.session_state.get("_user_list_loaded", False):
                    _ALL_ROLES = ["super_admin","admin","user","read_only","trainee"]
                    try:
                        _conn = get_db_connection()
                        if _conn:
                            _cur = _conn.cursor()
                            _cur.execute("SELECT email, role FROM user_roles ORDER BY email")
                            _users = _cur.fetchall()
                            _conn.close()
                            for _email, _role in _users:
                                _c1, _c2 = st.columns([3,2])
                                _c1.caption(_email)
                                _safe_role = _role if _role in _ALL_ROLES else "read_only"
                                _new_role = _c2.selectbox("", _ALL_ROLES,
                                    index=_ALL_ROLES.index(_safe_role),
                                    key=f"role2_{_email}", label_visibility="collapsed")
                                if _new_role != _role:
                                    try:
                                        _conn2 = get_db_connection()
                                        _cur2 = _conn2.cursor()
                                        _cur2.execute("UPDATE user_roles SET role=%s WHERE email=%s", (_new_role, _email))
                                        _conn2.commit()
                                        _conn2.close()
                                        st.rerun()
                                    except Exception:
                                        pass
                    except Exception:
                        pass
                st.caption("New users get read_only by default on first login.")

        st.markdown(
            """
            <div style="color:#64748b;font-size:0.7rem;text-align:center;">
                 2024 RateEdge Australia<br>
                <a href="https://rateedge.au" style="color:#3b82f6;">rateedge.au</a>
            </div>
            """,
            unsafe_allow_html=True,
        )

    apply_rateedge_theme(st.session_state["theme_name"])

    # Preserve auth through vol editor Apply reload
    if 'v3d_data' in st.query_params:
        st.session_state["authenticated"] = True
        # Jump directly to vol editor without showing tabs
        vol_surface_editor_tab()
        return

    # Check if authenticated
    if not st.session_state.get("authenticated"):
        show_login_page()
        return
    
    # Single session validation   —   check token still valid in DB
    try:
        if HAS_POSTGRES and st.session_state.get("session_token"):
            _vconn = get_db_connection()
            if _vconn:
                _vcur = _vconn.cursor()
                _vcur.execute("SELECT session_token FROM active_sessions WHERE email=%s",
                              (st.session_state.get("user_email",""),))
                _vrow = _vcur.fetchone()
                _vconn.close()
                if _vrow and _vrow[0] != st.session_state["session_token"]:
                    st.session_state["authenticated"] = False
                    st.session_state["session_token"] = None
                    st.warning("Your session was ended because you logged in from another device.")
                    st.stop()
    except Exception:
        pass

    # Only show tabs if authenticated
    _show_hidden = is_super_admin()

    _tab_names = [
        "🏡 Home",
        "📋 IRS / Vol Upload",
        "📏 Curves",
        "📈 FWD IRS Analysis",
        "📊 Historical VOL Analysis",
        "📊 Swaptions",
        "🔔 Caps & Floors",
        "💼 Trade Blotter",
        "⚛️ RV / Calendar",
        "🔮 Exotics",
        "📏 SOD Report",
        "✅ Vol Editor",
        "📑 Vol Export",
    ]
    _tab_funcs = [
        home_tab,
        vol_config_tab,
        curves_tab,
        fwd_analysis_tab,
        backtesting_tab,
        lambda: swaptions_tab(vol_mode),
        lambda: caps_floors_tab(vol_mode),
        portfolio_tab,
        rv_tab,
        lambda: exotics_tab(vol_mode),
        sod_report_tab,
        vol_surface_editor_tab,
        vol_export_tab,
    ]
    if _show_hidden:
        _tab_names += ["📍 Multi-CCY", "📜 Bond Options"]
        _tab_funcs += [lambda: multi_ccy_tab(vol_mode), bond_option_tab]

    tabs = st.tabs(_tab_names)
    for _ti, _tf in enumerate(_tab_funcs):
        with tabs[_ti]:
            _tf()


def sod_report_tab():
    """Start-of-Day Report   —   USD overnight moves → implied AUD vol open."""
    st.subheader("📋 Start-of-Day Report   —   USD Overnight → AUD Implied Open")
    st.caption(
        "Compares USD previous close vs the close before that. "
        "Uses USD IRS + vol moves to imply theoretical AUD vol changes at the open. "
        "Useful for time-sensitive gamma trades after large overnight USD moves."
    )

    user_id = st.session_state.get("username", "default")

    # ── Load available snapshots — cached, refresh on button ─────
    if st.button("🔄 Reload Snapshots", key="sod_reload_snaps") or "sod_snaps_usd" not in st.session_state:
        st.session_state["sod_snaps_usd"] = list_vol_snapshots(user_id, "USD") if HAS_POSTGRES else []
        st.session_state["sod_snaps_aud"] = list_vol_snapshots(user_id, "AUD") if HAS_POSTGRES else []
    _snaps_usd = st.session_state.get("sod_snaps_usd", [])
    _snaps_aud = st.session_state.get("sod_snaps_aud", [])

    if not HAS_POSTGRES:
        st.warning("Database not connected   —   SOD Report requires saved vol snapshots.")
        return

    # ── Dev Tool: Seed dummy USD snapshots for testing ────────────
    with st.expander("🛠️ Seed Test Data (Dev)", expanded=len(_snaps_usd) < 2):
        st.caption("Seeds dummy USD vol snapshots based on the currently loaded USD vol surface, "
                   "with randomised daily moves. Use this to test the SOD Report without real EOD data.")
        _seed_c1, _seed_c2, _seed_c3 = st.columns([2, 2, 2])
        with _seed_c1:
            _seed_days = st.number_input("Days of history to seed", min_value=2, max_value=90,
                                         value=45, key="sod_seed_days")
        with _seed_c2:
            _seed_vol_sigma = st.number_input("Daily vol move ┬ñ├ó (bp)", min_value=0.5, max_value=10.0,
                                              value=2.0, step=0.5, key="sod_seed_sigma")
        with _seed_c3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _seed_btn = st.button("🌱 Seed Dummy USD Snapshots", key="sod_seed_btn", type="primary")

        if _seed_btn:
            _usd_vol_data = st.session_state.get("vol_data", {}).get("USD", {})
            _usd_atm_live = _usd_vol_data.get("atm")

            if _usd_atm_live is None:
                st.error("No USD ATM vol surface loaded. Load USD data in Vol/SABR tab first.")
            else:
                import json as _json
                from datetime import datetime as _dt, timedelta as _td
                import random as _rnd

                _atm_base = _usd_atm_live.copy()
                if "Expiry" in _atm_base.columns:
                    _atm_base = _atm_base.set_index("Expiry")
                _tenor_cols = [c for c in _atm_base.columns]
                _exp_rows   = list(_atm_base.index)

                conn = get_db_connection()
                if not conn:
                    st.error("Cannot connect to database.")
                else:
                    cur = conn.cursor()
                    _seeded = 0
                    _rnd.seed(42)

                    # Walk backwards from yesterday, one snapshot per business day
                    _day = _dt.now().replace(hour=21, minute=0, second=0, microsecond=0)
                    _running = _atm_base.copy().astype(float)

                    for _d in range(int(_seed_days)):
                        _day -= _td(days=1)
                        # Skip weekends
                        if _day.weekday() >= 5:
                            _day -= _td(days=_day.weekday() - 4)

                        # Apply random daily move to each cell
                        for _e in _exp_rows:
                            for _t in _tenor_cols:
                                try:
                                    _mv = _rnd.gauss(0, _seed_vol_sigma)
                                    _running.loc[_e, _t] = max(
                                        float(_running.loc[_e, _t]) + _mv, 1.0
                                    )
                                except Exception:
                                    pass

                        _snap_records = _running.reset_index().rename(columns={"index": "Expiry"}).to_dict(orient="records")
                        _label = f"USD EOD {_day.strftime('%Y-%m-%d')} [DUMMY]"

                        try:
                            cur.execute("""
                                INSERT INTO vol_history
                                (user_id, currency, snapshot_date, label, atm_vols, notes)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, (
                                user_id, "USD", _day,
                                _label,
                                Json({"values": _snap_records}),
                                "Seeded dummy data for SOD Report testing"
                            ))
                            _seeded += 1
                        except Exception as _e:
                            pass

                    conn.commit()
                    cur.close()
                    conn.close()
                    st.success(f"✅ Seeded {_seeded} dummy USD snapshots. Reload the page to see them.")
                    st.rerun()

    if len(_snaps_usd) < 2:
        st.info(
            "Need at least 2 USD vol snapshots to compute overnight changes. "
            "Save EOD snapshots from the Vol Export tab each day."
        )
        return

    # ── Snapshot selectors ────────────────────────────────────────
    st.markdown("### 📅 Select Closing Snapshots")
    _sc1, _sc2, _sc3 = st.columns(3)

    def _snap_label(s):
        d = s["snapshot_date"]
        lbl = s["label"] or ""
        return f"{str(d)[:16]}  {lbl}"

    with _sc1:
        st.markdown("**USD   —   Previous Close (T-1 NYC)**")
        _usd_t1_opts = {_snap_label(s): s for s in _snaps_usd}
        _usd_t1_sel  = st.selectbox("USD T-1", list(_usd_t1_opts.keys()), index=0, key="sod_usd_t1")
        _usd_t1      = _usd_t1_opts[_usd_t1_sel]

    with _sc2:
        st.markdown("**USD   —   Two Sessions Ago (T-2 NYC)**")
        _usd_t2_opts = {_snap_label(s): s for s in _snaps_usd if s["id"] != _usd_t1["id"]}
        _usd_t2_sel  = st.selectbox("USD T-2", list(_usd_t2_opts.keys()), index=0, key="sod_usd_t2")
        _usd_t2      = _usd_t2_opts[_usd_t2_sel]

    with _sc3:
        st.markdown("**AUD   —   Previous Close (4:30pm Sydney)**")
        if _snaps_aud:
            _aud_opts = {_snap_label(s): s for s in _snaps_aud}
            _aud_sel  = st.selectbox("AUD prev close", list(_aud_opts.keys()), index=0, key="sod_aud")
            _aud_snap = _aud_opts[_aud_sel]
        else:
            st.warning("No AUD snapshots found.")
            _aud_snap = None

    # ── Vol mode toggle ───────────────────────────────────────────
    _vm_col, _disp_col = st.columns([2, 4])
    with _vm_col:
        _disp_mode = st.radio("Display", ["Normal vol (bp)", "Premium ($)"],
                              horizontal=True, key="sod_disp_mode")
    _show_bp = "Normal" in _disp_mode

    st.markdown("---")

    # ── Load snapshot data ────────────────────────────────────────
    _d1 = load_vol_snapshot(_usd_t1["id"])
    _d2 = load_vol_snapshot(_usd_t2["id"])

    if _d1 is None or _d2 is None:
        st.error("Failed to load USD snapshots.")
        return

    _atm1 = _d1.get("atm")  # T-1 (most recent)
    _atm2 = _d2.get("atm")  # T-2

    if _atm1 is None or _atm2 is None:
        st.error("Snapshot ATM vol data missing.")
        return

    # Normalise: ensure Expiry is index
    def _norm(df):
        if df is None: return None
        df = df.copy()
        if "Expiry" in df.columns:
            df = df.set_index("Expiry")
        return df

    _atm1 = _norm(_atm1)
    _atm2 = _norm(_atm2)

    # ── USD Vol Change Matrix ─────────────────────────────────────
    st.markdown("### 🇺🇸 USD Vol Changes   —   Overnight (T-1 close vs T-2 close)")

    _common_exp = [e for e in _atm1.index if e in _atm2.index]
    _common_ten = [c for c in _atm1.columns if c in _atm2.columns]

    if not _common_exp or not _common_ten:
        st.error("No overlapping expiry/tenor between USD snapshots.")
        return

    _usd_chg = _atm1.loc[_common_exp, _common_ten].apply(pd.to_numeric, errors="coerce") - \
               _atm2.loc[_common_exp, _common_ten].apply(pd.to_numeric, errors="coerce")

    # ── USD Premium change ────────────────────────────────────────
    _usd_curve = get_ccy_curve("USD")
    def _ensure_expiry_col(df):
        """Return DataFrame with Expiry as a plain column, not index."""
        d = df.copy()
        if d.index.name == "Expiry":
            return d.reset_index()
        if "Expiry" in d.columns:
            return d
        # numeric index — try to restore from index name or just label it
        return d.reset_index().rename(columns={"index": "Expiry"})

    _usd_prem_t1, _ = calculate_atm_premium_matrix(
        "USD", _usd_curve, _ensure_expiry_col(_atm1),
    ) if _usd_curve is not None and not _usd_curve.empty else (pd.DataFrame(), pd.DataFrame())
    _usd_prem_t2, _ = calculate_atm_premium_matrix(
        "USD", _usd_curve, _ensure_expiry_col(_atm2),
    ) if _usd_curve is not None and not _usd_curve.empty else (pd.DataFrame(), pd.DataFrame())

    _usd_prem_chg = pd.DataFrame()
    if not _usd_prem_t1.empty and not _usd_prem_t2.empty:
        _pc_exp = [e for e in _usd_prem_t1.index if e in _usd_prem_t2.index]
        _pc_ten = [c for c in _usd_prem_t1.columns if c in _usd_prem_t2.columns]
        if _pc_exp and _pc_ten:
            _usd_prem_chg = _usd_prem_t1.loc[_pc_exp, _pc_ten].astype(float) - \
                            _usd_prem_t2.loc[_pc_exp, _pc_ten].astype(float)

    # Display vol change
    st.markdown("**Vol Change (bp)**")
    st.dataframe(
        _usd_chg.style
            .background_gradient(cmap="RdYlGn", axis=None, vmin=-5, vmax=5)
            .format("{:+.2f}"),
        use_container_width=True
    )
    # Display premium change
    if not _usd_prem_chg.empty:
        st.markdown("**Fwd Premium Change (bp notional)**")
        st.dataframe(
            _usd_prem_chg.style
                .background_gradient(cmap="RdYlGn", axis=None, vmin=-3, vmax=3)
                .format("{:+.2f}"),
            use_container_width=True
        )
    st.caption(f"USD T-1: {_usd_t1_sel[:16]}  vs  T-2: {_usd_t2_sel[:16]}")

    # ── IRS Curve Change ──────────────────────────────────────────
    st.markdown("### 📈 USD IRS Curve   —   T-1 Close")
    _usd_curve_aud = get_ccy_curve("AUD")

    if _usd_curve is not None and not _usd_curve.empty:
        _fig_usd_crv = go.Figure()
        _fig_usd_crv.add_trace(go.Scatter(
            x=_usd_curve["MaturityY"], y=_usd_curve["ZeroRatePct"],
            mode="lines+markers", line=dict(color="#3b82f6", width=2),
            marker=dict(size=5), name="USD IRS (loaded)"))
        _fig_usd_crv.update_layout(
            title="USD IRS Zero Curve", xaxis_title="Maturity (Y)", yaxis_title="Rate (%)",
            template="plotly_dark", height=240, margin=dict(t=35, b=35))
        st.plotly_chart(_fig_usd_crv, use_container_width=True)
    else:
        st.info("Load USD IRS curve in Curves tab for full analysis.")

    # ── AUD Vol Snapshot ──────────────────────────────────────────
    _aud_atm = None
    if _aud_snap:
        _aud_data = load_vol_snapshot(_aud_snap["id"])
        if _aud_data:
            _aud_atm = _norm(_aud_data.get("atm"))

    # ── Implied AUD Vol Open ──────────────────────────────────────
    st.markdown("---")
    st.markdown("### 💡 Implied AUD Vol Open   —   Based on USD Overnight Moves")
    st.caption(
        "AUD vol sensitivity to USD: short-dated AUD vols (≤6m) correlate ~40-60% with USD, "
        "long-dated (≥2y) ~20-35%. Adjustments are indicative   —   verify with live market."
    )

    # Sensitivity parameters   —   editable
    with st.expander("⚙️ Sensitivity Parameters (edit if needed)", expanded=False):
        _sens_c1, _sens_c2, _sens_c3 = st.columns(3)
        with _sens_c1:
            _beta_short = st.slider("Short-end beta (≤6m, %)", 10, 90, 50, 5,
                                    key="sod_beta_short",
                                    help="% of USD vol move passed through to AUD ≤6m expiries") / 100
        with _sens_c2:
            _beta_mid = st.slider("Mid beta (1y-2y, %)", 10, 70, 35, 5,
                                  key="sod_beta_mid") / 100
        with _sens_c3:
            _beta_long = st.slider("Long-end beta (≥3y, %)", 5, 50, 20, 5,
                                   key="sod_beta_long") / 100

    def _get_beta(expiry_lbl: str) -> float:
        ey = label_to_years(expiry_lbl)
        if ey <= 0.5:   return _beta_short
        if ey <= 2.0:   return _beta_mid
        return _beta_long

    if _aud_atm is not None:
        _aud_exp = [e for e in _aud_atm.index if e in _usd_chg.index]
        _aud_ten = [c for c in _aud_atm.columns if c in _usd_chg.columns]

        if _aud_exp and _aud_ten:
            _implied_chg  = pd.DataFrame(index=_aud_exp, columns=_aud_ten, dtype=float)
            _implied_open = pd.DataFrame(index=_aud_exp, columns=_aud_ten, dtype=float)

            for _e in _aud_exp:
                _b = _get_beta(_e)
                for _t in _aud_ten:
                    try:
                        _usd_mv  = float(_usd_chg.loc[_e, _t]) if _e in _usd_chg.index and _t in _usd_chg.columns else 0.0
                        _aud_now = float(_aud_atm.loc[_e, _t])
                        _implied_chg.loc[_e, _t]  = round(_usd_mv * _b, 2)
                        _implied_open.loc[_e, _t] = round(_aud_now + _usd_mv * _b, 2)
                    except Exception:
                        pass

            # ── Compute AUD premium matrices ─────────────────────────
            _aud_curve = get_ccy_curve("AUD")
            _aud_prem_prev   = pd.DataFrame()
            _aud_prem_open   = pd.DataFrame()
            _aud_prem_chg    = pd.DataFrame()

            if _aud_curve is not None and not _aud_curve.empty:
                def _atm_df_with_expiry(idx_df):
                    """Ensure ATM df has Expiry as column, not index."""
                    d = idx_df.copy()
                    if "Expiry" not in d.columns:
                        d = d.reset_index().rename(columns={"index": "Expiry"})
                    return d

                _aud_prem_prev, _ = calculate_atm_premium_matrix(
                    "AUD", _aud_curve, _atm_df_with_expiry(_aud_atm)
                )
                _aud_prem_open, _ = calculate_atm_premium_matrix(
                    "AUD", _aud_curve, _atm_df_with_expiry(_implied_open)
                )
                if not _aud_prem_prev.empty and not _aud_prem_open.empty:
                    _pchg_exp = [e for e in _aud_prem_open.index if e in _aud_prem_prev.index]
                    _pchg_ten = [c for c in _aud_prem_open.columns if c in _aud_prem_prev.columns]
                    if _pchg_exp and _pchg_ten:
                        _aud_prem_chg = _aud_prem_open.loc[_pchg_exp, _pchg_ten].astype(float) - \
                                        _aud_prem_prev.loc[_pchg_exp, _pchg_ten].astype(float)

            # ── Display: toggle Vol vs Premium ───────────────────────
            if _show_bp:
                st.markdown("#### Implied AUD Vol Change at Open (bp)")
                st.dataframe(
                    _implied_chg.astype(float).style
                        .background_gradient(cmap="RdYlGn", axis=None, vmin=-5, vmax=5)
                        .format("{:+.2f}"),
                    use_container_width=True
                )
                st.markdown("#### Implied AUD Vol Open Level (bp)")
                st.dataframe(
                    _implied_open.astype(float).style
                        .background_gradient(cmap="RdYlGn_r", axis=None)
                        .format("{:.2f}"),
                    use_container_width=True
                )
            else:
                # Premium mode
                if not _aud_prem_chg.empty:
                    st.markdown("#### Implied AUD Fwd Premium Change at Open (bp notional)")
                    st.dataframe(
                        _aud_prem_chg.astype(float).style
                            .background_gradient(cmap="RdYlGn", axis=None, vmin=-3, vmax=3)
                            .format("{:+.2f}"),
                        use_container_width=True
                    )
                if not _aud_prem_open.empty:
                    st.markdown("#### Implied AUD Fwd Premium Open Level (bp notional)")
                    st.dataframe(
                        _aud_prem_open.astype(float).style
                            .background_gradient(cmap="RdYlGn_r", axis=None)
                            .format("{:.2f}"),
                        use_container_width=True
                    )
                if _aud_prem_chg.empty:
                    st.info("Load AUD IRS curve in Curves tab to compute premium matrices.")

            # ── Load to Vol Editor button ─────────────────────────────
            st.markdown("---")
            _load_col1, _load_col2 = st.columns([2, 4])
            with _load_col1:
                if st.button("📋 Load Implied Open → Vol Editor", key="sod_load_vol_editor", type="primary"):
                    if is_admin():
                        try:
                            # Reformat _implied_open to match vol_editor expected format
                            # Add Expiry column if not present
                            _ve_df = _implied_open.copy().reset_index()
                            if _ve_df.columns[0] != "Expiry":
                                _ve_df.columns = ["Expiry"] + list(_ve_df.columns[1:])
                            # Load into vol_editor working state (not base — so it shows as changes)
                            if "vol_editor" not in st.session_state:
                                st.session_state["vol_editor"] = {"working": {}, "base": {}, "history": {}, "future": {}, "redo_stack": {}, "view_mode": {}, "smoothing": {}, "paste_data": {}}
                            ve = st.session_state["vol_editor"]
                            # Get current surface as base
                            _current_atm = get_working_atm_surface("AUD")
                            if _current_atm is not None:
                                # Align implied open to base surface shape
                                # Fill any missing expiries from current ATM
                                _base_df = _current_atm.copy()
                                if "Expiry" not in _base_df.columns:
                                    _base_df = _base_df.reset_index()
                                    _base_df.columns = ["Expiry"] + list(_base_df.columns[1:])
                                # Merge: use implied open values where available, else current ATM
                                _merged = _base_df.copy()
                                _ve_exp_set = set(_ve_df["Expiry"].str.lower().tolist()) if "Expiry" in _ve_df.columns else set()
                                for _ri, _row in _base_df.iterrows():
                                    _exp_lbl = str(_row["Expiry"]).lower()
                                    if _exp_lbl in _ve_exp_set:
                                        _src_row = _ve_df[_ve_df["Expiry"].str.lower()==_exp_lbl]
                                        if not _src_row.empty:
                                            for _tc in _base_df.columns[1:]:
                                                if _tc in _ve_df.columns:
                                                    try:
                                                        _merged.at[_ri, _tc] = float(_src_row.iloc[0][_tc])
                                                    except Exception:
                                                        pass
                                ve["base"]["AUD"] = _base_df.copy()
                                ve["working"]["AUD"] = _merged.copy()
                                # Ensure other keys exist
                                for _k in ["history", "redo_stack", "view_mode", "smoothing", "paste_data"]:
                                    if _k not in ve: ve[_k] = {}
                                ve["history"]["AUD"] = []
                                ve["redo_stack"]["AUD"] = []
                                # Set flag so vol_editor doesn't overwrite on init
                                if "sod_loaded" not in ve:
                                    ve["sod_loaded"] = {}
                                ve["sod_loaded"]["AUD"] = True
                                st.session_state["vol_editor_auto_load"] = True
                                st.success("✅ Implied open loaded into Vol Editor. Go to Vol Editor tab to review and publish.")
                            else:
                                st.warning("Load AUD ATM surface first before loading SOD implied open.")
                        except Exception as _e:
                            st.error(f"Failed to load: {_e}")
                    else:
                        st.warning("🔒 Admin only — contact wpo@rateedge.au to request edit access.")
            with _load_col2:
                st.caption("Loads the implied AUD open vol surface into the Vol Editor as a working draft. You can review, adjust and smooth before publishing to pricing.")

            # ── Notable moves ────────────────────────────────────────
            _thresh = st.slider("Highlight moves larger than (bp)", 1, 10, 3, key="sod_thresh")
            _big_moves = []
            for _e in _aud_exp:
                for _t in _aud_ten:
                    try:
                        _mv = float(_implied_chg.loc[_e, _t])
                        if abs(_mv) >= _thresh:
                            _usd_raw = float(_usd_chg.loc[_e, _t]) if _e in _usd_chg.index and _t in _usd_chg.columns else 0.0
                            _pchg_str = ""
                            if not _aud_prem_chg.empty and _e in _aud_prem_chg.index and _t in _aud_prem_chg.columns:
                                _pchg_str = f"{float(_aud_prem_chg.loc[_e, _t]):+.2f}bp"
                            _big_moves.append({
                                "Expiry": _e, "Tenor": _t,
                                "USD >vol": f"{_usd_raw:+.2f}bp",
                                "Beta": f"{_get_beta(_e)*100:.0f}%",
                                "Implied AUD >vol": f"{_mv:+.2f}bp",
                                "AUD Prev Close": f"{float(_aud_atm.loc[_e, _t]):.2f}bp",
                                "AUD Implied Open": f"{float(_implied_open.loc[_e, _t]):.2f}bp",
                                "> Fwd Premium": _pchg_str,
                                "Alert": "🔴 LARGE MOVE" if abs(_mv) >= _thresh * 1.5 else "🟡 Notable",
                            })
                    except Exception:
                        pass

            if _big_moves:
                st.markdown(f"#### 🔔 Notable Moves (≥{_thresh}bp implied AUD change)")
                _bm_df = pd.DataFrame(_big_moves).sort_values(
                    "Implied AUD >vol",
                    key=lambda x: x.str.replace("bp","").str.replace("+","").astype(float).abs(),
                    ascending=False
                )
                st.dataframe(_bm_df, use_container_width=True, hide_index=True)

            # ── Implied AUD CFS Open ─────────────────────────────────
            # Build cumulative CFS levels using prev-day wedge spreads from session state
            # CFS(nY) = sum of [swaption_leg_prem + wedge] for each wedge in the chain
            st.markdown("---")
            st.markdown("### 📋 Implied AUD CFS Open   —   Using Previous Day Wedge Spreads")
            st.caption(
                "Cumulative CFS = running sum of (swaption leg fwd premium + wedge spread). "
                "Wedge spreads from current session state (previous close). No caplet stripping."
            )

            # CFS chain: (label, wedge_spr_key, expiry_in_matrix, tenor_in_matrix)
            _CFS_CHAIN = [
                ("1Y CFS",  "cf_spr_3m1y",  "3m",  "1Y"),
                ("2Y CFS",  "cf_spr_1y1y",  "1y",  "1Y"),
                ("3Y CFS",  "cf_spr_2y1y",  "2y",  "1Y"),
                ("4Y CFS",  "cf_spr_3y1y",  "3y",  "1Y"),
                ("5Y CFS",  "cf_spr_4y1y",  "4y",  "1Y"),
                ("7Y CFS",  "cf_spr_5y2y",  "5y",  "2Y"),
                ("10Y CFS", "cf_spr_7y3y",  "7y",  "3Y"),
                ("12Y CFS", "cf_spr_10y2y", "10y", "2Y"),
                ("15Y CFS", "cf_spr_12y3y", "12y", "3Y"),
            ]

            def _prem_lookup(df, exp, ten):
                """Case-insensitive lookup in premium DataFrame."""
                if df is None or df.empty:
                    return None
                for _e in df.index:
                    if str(_e).lower() == exp.lower():
                        for _t in df.columns:
                            if str(_t).lower() == ten.lower():
                                try:
                                    v = float(df.loc[_e, _t])
                                    return v if not math.isnan(v) else None
                                except Exception:
                                    return None
                return None

            _cfs_rows = []
            _cum_prev = 0.0
            _cum_open = 0.0
            _cfs_ok = not _aud_prem_prev.empty and not _aud_prem_open.empty

            for _cfs_lbl, _spr_key, _exp, _ten in _CFS_CHAIN:
                _spr = st.session_state.get(_spr_key, 0.0)
                _p_prev = _prem_lookup(_aud_prem_prev, _exp, _ten)
                _p_open = _prem_lookup(_aud_prem_open, _exp, _ten)

                if _p_prev is not None and _p_open is not None:
                    _leg_prev = _p_prev + _spr
                    _leg_open = _p_open + _spr
                    _cum_prev += _leg_prev
                    _cum_open += _leg_open
                    _delta = _cum_open - _cum_prev
                    _cfs_rows.append({
                        "CFS Tenor": _cfs_lbl,
                        "Swptn Leg (prev)": f"{_p_prev:.2f}",
                        "Wedge": f"{_spr:+.2f}",
                        "CFS Leg (prev)": f"{_leg_prev:.2f}",
                        "Swptn Leg (open)": f"{_p_open:.2f}",
                        "CFS Leg (open)": f"{_leg_open:.2f}",
                        "CFS Total (prev)": f"{_cum_prev:.2f}",
                        "CFS Total (open)": f"{_cum_open:.2f}",
                        "> CFS": f"{_delta:+.2f}",
                    })
                else:
                    _cfs_rows.append({
                        "CFS Tenor": _cfs_lbl,
                        "Swptn Leg (prev)": "  —  ",
                        "Wedge": f"{_spr:+.2f}",
                        "CFS Leg (prev)": "  —  ",
                        "Swptn Leg (open)": "  —  ",
                        "CFS Leg (open)": "  —  ",
                        "CFS Total (prev)": "  —  ",
                        "CFS Total (open)": "  —  ",
                        "> CFS": "  —  ",
                    })

            if _cfs_rows:
                _cfs_df = pd.DataFrame(_cfs_rows)
                st.dataframe(_cfs_df, use_container_width=True, hide_index=True)
                if not _cfs_ok:
                    st.info("Load AUD IRS curve to compute CFS premium levels.")
            else:
                st.info("CFS chain requires AUD curve and ATM vol surface to be loaded.")

            # ── Narrative Summary ────────────────────────────────────
            st.markdown("---")
            st.markdown("### 📝 SOD Summary")

            # Build narrative
            _usd_avg_chg = float(_usd_chg.values.astype(float).mean())
            _usd_max_chg = float(_usd_chg.values.astype(float).max())
            _usd_min_chg = float(_usd_chg.values.astype(float).min())
            _usd_direction = "higher" if _usd_avg_chg > 0 else "lower"

            _aud_avg_chg = float(_implied_chg.values.astype(float).mean())
            _aud_direction = "higher" if _aud_avg_chg > 0 else "lower"

            # Find biggest USD moves
            _usd_flat = _usd_chg.stack().reset_index()
            _usd_flat.columns = ["Expiry", "Tenor", "Chg"]
            _usd_flat["AbsChg"] = _usd_flat["Chg"].abs()
            _usd_flat = _usd_flat.sort_values("AbsChg", ascending=False)
            _top_usd = _usd_flat.head(3)

            # Find biggest AUD implied moves
            _aud_flat = _implied_chg.stack().reset_index()
            _aud_flat.columns = ["Expiry", "Tenor", "Chg"]
            _aud_flat["AbsChg"] = _aud_flat["Chg"].abs()
            _aud_flat = _aud_flat.sort_values("AbsChg", ascending=False)
            _top_aud = _aud_flat.head(3)

            _usd_move_str = ", ".join(
                f"{r['Expiry']}≈{r['Tenor']} {r['Chg']:+.1f}bp"
                for _, r in _top_usd.iterrows()
            )
            _aud_move_str = ", ".join(
                f"{r['Expiry']}≈{r['Tenor']} {r['Chg']:+.1f}bp"
                for _, r in _top_aud.iterrows()
            )

            # USD premium context
            _usd_prem_ctx = ""
            if not _usd_prem_chg.empty:
                _usd_avg_pchg = float(_usd_prem_chg.values.astype(float).mean())
                _usd_prem_ctx = (
                    f" In premium terms, the average USD fwd premium moved "
                    f"{'up' if _usd_avg_pchg > 0 else 'down'} "
                    f"{abs(_usd_avg_pchg):.1f}bp across the surface."
                )

            _aud_prem_ctx = ""
            if not _aud_prem_chg.empty:
                _aud_avg_pchg = float(_aud_prem_chg.values.astype(float).mean())
                _aud_prem_ctx = (
                    f" Translating to AUD premium: the implied open fwd premium moves "
                    f"an average of {abs(_aud_avg_pchg):.1f}bp "
                    f"({'higher' if _aud_avg_pchg > 0 else 'lower'}) vs yesterday's AUD close."
                )

            _narrative = f"""
**USD overnight vol session (T-2 → T-1 NYC close):** USD ATM vols moved broadly {_usd_direction} \
overnight, with the surface averaging {abs(_usd_avg_chg):.1f}bp change. \
The largest moves were in {_usd_move_str}. \
The range across the surface was {_usd_min_chg:+.1f}bp to {_usd_max_chg:+.1f}bp.{_usd_prem_ctx}

**Implied AUD open (vs yesterday's 4:30pm Sydney close):** Applying USD/AUD vol betas \
(short-end {int(_beta_short*100)}%, mid {int(_beta_mid*100)}%, long-end {int(_beta_long*100)}%), \
AUD vols are implied to open broadly {_aud_direction}, averaging {abs(_aud_avg_chg):.1f}bp change. \
Key AUD moves to watch: {_aud_move_str}.{_aud_prem_ctx}

**Tactical note:** {"Short-dated AUD gamma looks relatively more affected given the short-end beta pickup. Consider reviewing 3m-6m expiry trades before the open." if abs(_aud_avg_chg) > 1.5 else "Moves are modest   —   no urgent repricing expected at the AUD open, but monitor live market confirmation."} \
These are indicative adjustments based on observed USD/AUD correlations and should be verified against live interdealer markets at open.
"""
            st.markdown(_narrative)

            # ── Save & Download report ───────────────────────────────
            _report_lines = [
                f"RateEdge SOD Report   —   {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')} AEST",
                f"USD T-1: {_usd_t1_sel[:40]}",
                f"USD T-2: {_usd_t2_sel[:40]}",
                f"AUD prev close: {_aud_sel[:40]}",
                "",
                "USD Vol Changes (bp):",
                _usd_chg.to_string(),
            ]
            if not _usd_prem_chg.empty:
                _report_lines += ["", "USD Fwd Premium Changes (bp):", _usd_prem_chg.to_string()]
            _report_lines += [
                "",
                "Implied AUD Vol Change at Open (bp):",
                _implied_chg.astype(float).to_string(),
                "",
                "Implied AUD Vol Open Level (bp):",
                _implied_open.astype(float).to_string(),
            ]
            if not _aud_prem_chg.empty:
                _report_lines += ["", "Implied AUD Fwd Premium Change (bp):", _aud_prem_chg.to_string()]
            if _cfs_rows:
                _cfs_report_df = pd.DataFrame(_cfs_rows)[["CFS Tenor", "CFS Total (prev)", "CFS Total (open)", "> CFS"]]
                _report_lines += ["", "Implied AUD CFS Open Levels (bp fwd premium, cumulative):", _cfs_report_df.to_string(index=False)]
            _report_lines += ["", "--- SUMMARY ---", _narrative.replace("**", "").replace("\n", " ")]

            _report_text = "\n".join(_report_lines)

            # Build JSON payload for DB storage
            _report_payload = {
                "usd_chg": _usd_chg.to_dict(),
                "usd_prem_chg": _usd_prem_chg.to_dict() if not _usd_prem_chg.empty else {},
                "aud_implied_chg": _implied_chg.astype(float).to_dict(),
                "aud_implied_open": _implied_open.astype(float).to_dict(),
                "aud_prem_chg": _aud_prem_chg.to_dict() if not _aud_prem_chg.empty else {},
                "cfs_rows": _cfs_rows if _cfs_rows else [],
                "narrative": _narrative,
                "betas": {"short": _beta_short, "mid": _beta_mid, "long": _beta_long},
                "generated_at": pd.Timestamp.now().isoformat(),
            }

            st.markdown("---")
            _btn_col1, _btn_col2 = st.columns(2)
            with _btn_col1:
                st.download_button(
                    "📂 Download SOD Report",
                    _report_text.encode(),
                    f"RateEdge_SOD_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.txt",
                    "text/plain", key="sod_download"
                )
            with _btn_col2:
                if HAS_POSTGRES:
                    _sod_notes = st.text_input("Notes (optional)", value="", key="sod_save_notes", label_visibility="collapsed",
                                               placeholder="Notes (optional)")
                    if st.button("💾 Save Report to Supabase", key="sod_save_btn", type="primary", use_container_width=True):
                        _rid = save_sod_report(
                            user_id=user_id,
                            report_date=pd.Timestamp.now().date(),
                            usd_t1_label=_usd_t1_sel[:100],
                            usd_t2_label=_usd_t2_sel[:100],
                            aud_snap_label=_aud_sel[:100],
                            report_data=_report_payload,
                            notes=_sod_notes.strip()
                        )
                        if _rid:
                            st.success(f"✅ Report saved (ID: {_rid})")
                        else:
                            st.error("Failed to save report.")

            # ── Past Reports Browser ─────────────────────────────────
            if HAS_POSTGRES:
                st.markdown("---")
                with st.expander("📰 Past SOD Reports", expanded=False):
                    _past = list_sod_reports(user_id, limit=30)
                    if not _past:
                        st.info("No saved reports yet.")
                    else:
                        _past_opts = {f"{r['report_date']}   —   {r['usd_t1'][:30]} | {r['notes'] or ''}": r for r in _past}
                        _past_sel_lbl = st.selectbox("Select report", list(_past_opts.keys()), key="sod_past_sel")
                        _past_rec = _past_opts[_past_sel_lbl]
                        _pcol1, _pcol2 = st.columns(2)
                        with _pcol1:
                            if st.button("📂 Load & Display", key="sod_past_load", type="primary"):
                                _loaded = load_sod_report(_past_rec["id"])
                                if _loaded:
                                    st.session_state["sod_loaded_report"] = _loaded
                                    st.success(f"Loaded: {_past_sel_lbl}")
                        with _pcol2:
                            if st.button("🗑 Delete", key="sod_past_del"):
                                if delete_sod_report(_past_rec["id"]):
                                    st.success("Deleted.")
                                    st.rerun()

                        # Display loaded report
                        if "sod_loaded_report" in st.session_state:
                            _lr = st.session_state["sod_loaded_report"]
                            _ld = _lr.get("data", {})
                            st.markdown(f"**Report date:** {_lr['report_date']}  |  **Generated:** {str(_lr.get('created_at',''))[:16]}")
                            st.markdown(f"USD T-1: `{_lr['usd_t1']}` | USD T-2: `{_lr['usd_t2']}` | AUD: `{_lr['aud_snap']}`")

                            if _ld.get("usd_chg"):
                                st.markdown("**USD Vol Changes (bp)**")
                                st.dataframe(pd.DataFrame(_ld["usd_chg"]).style
                                    .background_gradient(cmap="RdYlGn", axis=None, vmin=-5, vmax=5)
                                    .format("{:+.2f}"), use_container_width=True)

                            if _ld.get("aud_implied_chg"):
                                st.markdown("**Implied AUD Vol Change (bp)**")
                                st.dataframe(pd.DataFrame(_ld["aud_implied_chg"]).style
                                    .background_gradient(cmap="RdYlGn", axis=None, vmin=-5, vmax=5)
                                    .format("{:+.2f}"), use_container_width=True)

                            if _ld.get("cfs_rows"):
                                st.markdown("**CFS Open Levels**")
                                st.dataframe(pd.DataFrame(_ld["cfs_rows"])[["CFS Tenor","CFS Total (prev)","CFS Total (open)","> CFS"]],
                                             use_container_width=True, hide_index=True)

                            if _ld.get("narrative"):
                                st.markdown("**Summary**")
                                st.markdown(_ld["narrative"])

                            if st.button("➖ Clear loaded report", key="sod_clear_loaded"):
                                del st.session_state["sod_loaded_report"]
                                st.rerun()
        else:
            st.warning("No overlapping expiry/tenor between AUD snapshot and USD change matrix.")
    else:
        # No AUD snapshot   —   show USD change only and instructions
        st.info(
            "No AUD vol snapshot selected. "
            "Save an AUD EOD snapshot from the Vol Export tab to see implied AUD open levels."
        )
        st.markdown("**USD Change Matrix available above.** "
                    "Load an AUD snapshot to compute implied AUD opening vols.")


def vol_export_tab():
    """Vol Export tab - Export and email vol surfaces"""
    st.subheader("📂 Vol Surface Export & Distribution")
    st.caption("Export vol surfaces to Excel and email to multiple recipients")

    user_id = st.session_state.get("username", "default")

    # Get theme colors
    is_dark = st.session_state.get("theme_name", "Dealer Dark") == "Dealer Dark"
    card_bg = "#1e293b" if is_dark else "#ffffff"
    border_color = "#334155" if is_dark else "#e2e8f0"
    
    # Two column layout
    col_left, col_right = st.columns([1.5, 1])
    
    with col_left:
        st.markdown("### 📊 Export Settings")
        
        # Currency selection - multi-select
        export_currencies = st.multiselect(
            "Select Currencies",
            SUPPORTED_CURRENCIES,
            default=[SUPPORTED_CURRENCIES[0]],
            key="export_currencies"
        )
        
        col1, col2 = st.columns(2)
        with col1:
            include_sabr = st.checkbox("Include SABR Parameters", value=True, key="export_include_sabr")
        with col2:
            include_metadata = st.checkbox("Include Metadata Sheet", value=True, key="export_metadata")
        
        # Export buttons
        st.markdown("#### 💾 Download Options")
        
        if not export_currencies:
            st.warning("Select at least one currency to export")
        else:
            # Generate all Excel files
            if st.button("📂 Generate Excel Files", key="generate_excel_btn", type="primary"):
                for ccy in export_currencies:
                    export_data = export_vol_surface_to_excel(ccy, include_sabr)
                    if export_data:
                        st.download_button(
                            label=f"📤 Download {ccy}",
                            data=export_data,
                            file_name=f"RateEdge_VolSurface_{ccy}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_vol_{ccy}"
                        )
                    else:
                        st.error(f"No data available for {ccy}")
        
        st.markdown("---")
        
        # Email distribution section
        st.markdown("### 📧 Email Distribution")
        
        # Recipients input
        recipients_text = st.text_area(
            "Email Recipients (one per line or comma-separated)",
            placeholder="john.doe@example.com\njane.smith@example.com\nteam@rateedge.au",
            height=100,
            key="email_recipients"
        )
        
        # Email subject and message
        default_subject = f"RateEdge Vol Surfaces - {pd.Timestamp.now().strftime('%Y-%m-%d')}"
        email_subject = st.text_input(
            "Email Subject",
            value=default_subject,
            key="email_subject"
        )
        
        email_message = st.text_area(
            "Email Message",
            value=f"""Hi Team,

Please find attached the latest volatility surfaces as of {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}.

Currencies: {', '.join(export_currencies) if export_currencies else 'None selected'}
SABR Parameters: {'Included' if include_sabr else 'Not included'}

Regards,
RateEdge Options Platform""",
            height=200,
            key="email_message"
        )
        
        # SMTP settings expander
        with st.expander("⚙️ Email Settings (SMTP)", expanded=False):
            st.caption("Configure your SMTP server settings")
            
            col_smtp1, col_smtp2 = st.columns(2)
            with col_smtp1:
                smtp_server = st.text_input("SMTP Server", value="smtp.office365.com", key="smtp_server")
                smtp_port = st.number_input("SMTP Port", value=587, key="smtp_port")
            with col_smtp2:
                smtp_user = st.text_input("SMTP Username", value="wpo@rateedge.au", key="smtp_user")
                # Get password from environment variable, allow UI override
                default_password = os.getenv("EMAIL_PASSWORD", "")
                smtp_password = st.text_input("SMTP Password", value=default_password, type="password", key="smtp_password")
            
            use_tls = st.checkbox("Use TLS", value=True, key="smtp_tls")
            
            st.caption("💡 For Office 365/Outlook: smtp.office365.com:587 with TLS")
            st.caption("💡 For Gmail: smtp.gmail.com:587 with TLS (requires app password)")
            if default_password:
                st.caption("✅ Password loaded from environment variable EMAIL_PASSWORD")
            else:
                st.caption("⚙️ Set EMAIL_PASSWORD environment variable in Streamlit secrets to auto-fill")
        
        # Send button
        if st.button("📧 Send Email with Attachments", key="send_email_btn", type="primary", use_container_width=True):
            if not recipients_text.strip():
                st.error("Please enter at least one email recipient")
            elif not export_currencies:
                st.error("Please select at least one currency to export")
            elif not smtp_user or not smtp_password:
                st.error("Please configure SMTP credentials in Email Settings")
            else:
                # Parse recipients
                recipients = []
                for line in recipients_text.replace(',', '\n').split('\n'):
                    email = line.strip()
                    if email and '@' in email:
                        recipients.append(email)
                
                if not recipients:
                    st.error("No valid email addresses found")
                else:
                    # Send email
                    with st.spinner(f"Sending email to {len(recipients)} recipient(s)..."):
                        success = send_vol_email(
                            recipients=recipients,
                            subject=email_subject,
                            message=email_message,
                            currencies=export_currencies,
                            include_sabr=include_sabr,
                            smtp_config={
                                'server': smtp_server,
                                'port': smtp_port,
                                'user': smtp_user,
                                'password': smtp_password,
                                'use_tls': use_tls
                            }
                        )
                    
                    if success:
                        st.success(f"✅ Email sent successfully to {len(recipients)} recipient(s)!")
                    else:
                        st.error("❌ Failed to send email. Check SMTP settings and try again.")
    
    with col_right:
        st.markdown("### 📅 Preview")
        
        # Show export summary
        if export_currencies:
            st.markdown(f"""
            <div style="background:{card_bg};border:1px solid {border_color};border-radius:10px;padding:1rem;margin:0.5rem 0;">
                <div style="font-weight:600;margin-bottom:0.5rem;">Export Summary</div>
                <div style="font-size:0.9rem;">
                    <strong>Currencies:</strong> {', '.join(export_currencies)}<br>
                    <strong>SABR Params:</strong> {'Yes' if include_sabr else 'No'}<br>
                    <strong>Metadata:</strong> {'Yes' if include_metadata else 'No'}<br>
                    <strong>Export Time:</strong> {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Show recipient list
        if recipients_text.strip():
            recipients = []
            for line in recipients_text.replace(',', '\n').split('\n'):
                email = line.strip()
                if email and '@' in email:
                    recipients.append(email)
            
            if recipients:
                st.markdown(f"""
                <div style="background:{card_bg};border:1px solid {border_color};border-radius:10px;padding:1rem;margin:1rem 0;">
                    <div style="font-weight:600;margin-bottom:0.5rem;">Recipients ({len(recipients)})</div>
                    <div style="font-size:0.85rem;">
                        {'<br>'.join(recipients)}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        # Save EOD Snapshot
        st.markdown("### 💾 Save EOD Snapshot")
        st.caption("Save current vol surface to database for SOD Report / implied open calculations")

        _eod_label = st.text_input(
            "Snapshot Label",
            value=f"EOD {pd.Timestamp.now().strftime('%Y-%m-%d')}",
            key="eod_snap_label"
        )
        _eod_notes = st.text_input("Notes (optional)", value="", key="eod_snap_notes")

        if st.button("💾 Save EOD Snapshot", key="save_eod_snap_btn", type="primary", use_container_width=True):
            if not HAS_POSTGRES:
                st.error("Database not connected   —   cannot save snapshot.")
            elif not export_currencies:
                st.error("Select at least one currency above first.")
            else:
                _saved, _failed = [], []
                for _ccy in export_currencies:
                    _sid = save_vol_snapshot(user_id, _ccy, _eod_label.strip(), _eod_notes.strip())
                    if _sid:
                        _saved.append(_ccy)
                    else:
                        _failed.append(_ccy)
                if _saved:
                    st.success(f"✅ Saved EOD snapshot for: {', '.join(_saved)}")
                if _failed:
                    st.error(f"├ö├ÿ├« Failed for: {', '.join(_failed)}   —   check vol data is loaded.")

        # Quick tips
        st.markdown(f"""
        <div style="background:{card_bg};border:1px solid {border_color};border-radius:10px;padding:1rem;margin:1rem 0;">
            <div style="font-weight:600;margin-bottom:0.5rem;">💡 Quick Tips</div>
            <div style="font-size:0.85rem;">
                • Select multiple currencies to send all in one email<br>
                • Recipients can be comma or newline separated<br>
                • SMTP settings are saved for the session<br>
                • Files are generated fresh on each send<br>
                • Save EOD Snapshot each day for next morning's implied open report
            </div>
        </div>
        """, unsafe_allow_html=True)


def send_vol_email(recipients: list, subject: str, message: str, currencies: list, include_sabr: bool, smtp_config: dict) -> bool:
    """Send email with vol surface Excel attachments"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = smtp_config['user']
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(message, 'plain'))
        
        # Generate and attach Excel files for each currency
        for ccy in currencies:
            excel_data = export_vol_surface_to_excel(ccy, include_sabr)
            if excel_data:
                filename = f"RateEdge_VolSurface_{ccy}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(excel_data)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={filename}')
                msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(smtp_config['server'], smtp_config['port'])
        if smtp_config['use_tls']:
            server.starttls()
        server.login(smtp_config['user'], smtp_config['password'])
        server.sendmail(smtp_config['user'], recipients, msg.as_string())
        server.quit()
        
        return True
        
    except Exception as e:
        st.error(f"Email error: {e}")
        return False


def show_login_page():
    """Full-page login with email OTP authentication"""
    
    st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="manage-app-button"] {display: none !important;}
    button[kind="managedApp"] {display: none !important;}
    .stAppDeployButton {display: none !important;}
    [title="Manage app"] {display: none !important;}
    .stApp {background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);}
    .stDeployButton {display: none !important;}
    .stTextInput > div > div > input {
        background: #1e293b !important;
        color: white !important;
        border: 1px solid #334155 !important;
        border-radius: 8px !important;
        padding: 12px 16px !important;
    }
    .stTextInput > div > div > input::placeholder {color: #64748b !important;}
    .stTextInput > div > div > input:focus {border-color: #dc2626 !important;}
    .stButton > button {
        background: #dc2626 !important;
        color: white !important;
        border: none !important;
        width: 100% !important;
        border-radius: 8px !important;
        padding: 12px !important;
        font-weight: 600 !important;
    }
    .stButton > button:hover {background: #b91c1c !important;}
    .stFormSubmitButton > button {
        background: #dc2626 !important;
        color: white !important;
        border: none !important;
        width: 100% !important;
        border-radius: 8px !important;
        padding: 12px !important;
        font-weight: 600 !important;
    }
    .stFormSubmitButton > button:hover {background: #b91c1c !important;}
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div style="height: 8vh;"></div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.2, 1])
    
    with col2:
        # Card with logo+title inside   —   matches IRS style
        st.markdown("""
        <div style="background:#1e293b;border-radius:16px;padding:2.5rem;border:1px solid #334155;margin-bottom:1rem;">
            <div style="text-align:center;margin-bottom:1.5rem;">
                <div style="display:inline-flex;align-items:center;gap:12px;">
                    <svg width="40" height="40" viewBox="0 0 40 40" fill="none">
                        <rect x="20" y="2" width="25" height="25" rx="3" transform="rotate(45 20 2)" fill="#dc2626"/>
                        <rect x="20" y="8" width="17" height="17" rx="2" transform="rotate(45 20 8)" fill="#1e293b"/>
                    </svg>
                    <span style="font-size:1.8rem;font-weight:700;color:white;">RateEdge</span>
                </div>
            </div>
            <div style="text-align:center;margin-bottom:0.5rem;">
                <div style="font-size:1.3rem;font-weight:700;color:white;">Options Platform</div>
                <div style="font-size:0.9rem;color:#94a3b8;margin-top:0.25rem;">Sign in with your email</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Initialize auth state
        if 'auth_step' not in st.session_state:
            st.session_state.auth_step = 'email'
        
        if st.session_state.auth_step == 'email':
            with st.form("email_form", clear_on_submit=False):
                email = st.text_input("Email", placeholder="Enter your email", key="login_email_input", label_visibility="collapsed")
                st.markdown('<p style="color: #94a3b8; font-size: 0.85rem; text-align: center; margin: -8px 0 24px 0;">We\'ll send you a verification code</p>', unsafe_allow_html=True)
                _submitted = st.form_submit_button("Send Code", use_container_width=True)
            if _submitted:
                if email and "@" in email:
                    status, data = request_otp(email.strip().lower())
                    if status == 200:
                        st.session_state.auth_email = email.strip().lower()
                        st.session_state.auth_step = 'otp'
                        st.rerun()
                    elif status == 202 or (status == 403 and data.get("error") == "access_pending"):
                        st.info("Access request submitted. You will be notified when approved.")
                    else:
                        st.error(data.get("error", "Failed to send code"))
                else:
                    st.error("Please enter a valid email address")
        
        elif st.session_state.auth_step == 'otp':
            st.markdown(f'<p style="color: #94a3b8; text-align: center; margin-bottom: 16px;">Code sent to <strong style="color: #f1f5f9;">{st.session_state.auth_email}</strong></p>', unsafe_allow_html=True)
            with st.form("otp_form", clear_on_submit=False):
                otp = st.text_input("Code", placeholder="Enter 6-digit code", max_chars=6, key="login_otp_input", label_visibility="collapsed")
                _verify = st.form_submit_button("Verify", use_container_width=True)
            if st.button("← Back", key="back_btn", use_container_width=True, type="secondary"):
                st.session_state.auth_step = 'email'
                st.rerun()
            if _verify:
                if otp and len(otp) == 6:
                    status, data = verify_otp(st.session_state.auth_email, otp)
                    if status == 200:
                        # Single session enforcement — store token in DB, invalidate old session
                        import secrets as _sec
                        _sess_token = _sec.token_hex(32)
                        try:
                            if HAS_POSTGRES:
                                _conn = get_db_connection()
                                if _conn:
                                    _cur = _conn.cursor()
                                    _cur.execute("""
                                        CREATE TABLE IF NOT EXISTS active_sessions (
                                            email TEXT PRIMARY KEY,
                                            session_token TEXT NOT NULL,
                                            created_at TIMESTAMPTZ DEFAULT NOW()
                                        )
                                    """)
                                    _cur.execute("""
                                        INSERT INTO active_sessions (email, session_token, created_at)
                                        VALUES (%s, %s, NOW())
                                        ON CONFLICT (email) DO UPDATE SET session_token=EXCLUDED.session_token, created_at=NOW()
                                    """, (st.session_state.auth_email, _sess_token))
                                    _cur.execute("""
                                        CREATE TABLE IF NOT EXISTS user_roles (
                                            email TEXT PRIMARY KEY,
                                            role TEXT NOT NULL DEFAULT 'read_only'
                                                CHECK (role IN ('admin','user','read_only','trainee')),
                                            created_at TIMESTAMPTZ DEFAULT NOW()
                                        )
                                    """)
                                    _cur.execute("""
                                        INSERT INTO user_roles (email, role)
                                        VALUES (%s, 'read_only')
                                        ON CONFLICT (email) DO NOTHING
                                    """, (st.session_state.auth_email,))
                                    _cur.execute("SELECT role FROM user_roles WHERE email=%s", (st.session_state.auth_email,))
                                    _role_row = _cur.fetchone()
                                    _ADMIN_EMAILS = {"wpo70@icloud.com", "wpo@rateedge.au"}
                                    if st.session_state.auth_email in _ADMIN_EMAILS:
                                        st.session_state["user_role"] = "super_admin"
                                    else:
                                        _db_role = _role_row[0] if _role_row else "read_only"
                                        _db_role = _db_role if _db_role in ("super_admin","super_admin","admin","user","read_only","trainee") else "read_only"
                                        st.session_state["user_role"] = _db_role
                                    _conn.commit()
                                    _conn.close()
                        except Exception:
                            pass
                        st.session_state["authenticated"] = True
                        st.session_state["username"] = st.session_state.auth_email
                        st.session_state["user_email"] = st.session_state.auth_email
                        st.session_state["session_token"] = _sess_token
                        st.session_state.auth_step = 'email'
                        st.session_state["db_auto_loaded"] = False
                        st.rerun()
                    else:
                        st.error(data.get("error", "Invalid code"))
                else:
                    st.error("Please enter the 6-digit code")
        
        st.markdown(
            "<div style='text-align:center;margin-top:1rem;color:#64748b;font-size:0.85rem'>"
            "Login issues? Email <b>wpo@rateedge.au</b>"
            "</div>",
            unsafe_allow_html=True
        )
    
    st.stop()


if __name__ == "__main__":
    main()
